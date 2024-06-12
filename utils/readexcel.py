import math
import os
import openpyxl
import allure
import pandas as pd
from configurations.config import ReadConfig as config
from utils.library import encrypte_decrypte,side_bar_to_run_for_androidtest,Testrun_mode
# The objective of this class is to Fetch campaigns,devices and components list whose execute status are marked as "Yes"
################################################### fetch_camapaigns_enviroment #############################################################################################
def fetch_camapaigns_enviroment():
    # Fetch campaigns and devices list whose execute status are marked as "Yes"
    global test_data_file_path
    keys_to_check = ['Map_View', 'Graph_View', 'Exports', 'PDF Data Export and Validation',
                     'Table Summary Export Validation', 'NW Freeze',
                     'Combined Export Data Validation', 'Individual Popup window data validation',
                     'NQC table data validation','Default Settings','Change Settings']
    androidtest = side_bar_to_run_for_androidtest(keys_to_check)
    remotetest_runvalue = Testrun_mode(value="Remote Test")
    group_runvalue = Testrun_mode(value="Group")
    scheduletest_runvalue = Testrun_mode(value="Schedule test")
    continuoustest_runvalue = Testrun_mode(value="Continuous Test")
    Default_settings_runvalue = Testrun_mode(value="Default Settings")
    Change_settings_runvalue = Testrun_mode(value="Change Settings")
    campaign_environment = []
    # Normal and Remote
    if "RUNNED".lower() != remotetest_runvalue[-1].strip().lower() and "RUNNED".lower() != scheduletest_runvalue[-1].strip().lower() and "RUNNED".lower() != continuoustest_runvalue[-1].strip().lower() and "FINISHED".lower() != remotetest_runvalue[-1].strip().lower() and "FINISHED".lower() != scheduletest_runvalue[-1].strip().lower() and "FINISHED".lower() != continuoustest_runvalue[-1].strip().lower() and "RUNNED".lower() != Change_settings_runvalue[-1].strip().lower():
        try:
            # test_data_file_path = config.test_data_path
            test_data_file_path = "C:\\RantCell_Automation_Data_and_Reports\\testdata\\Test_Data.xlsx"
        except:
            with allure.step(f"Check {test_data_file_path}"):
                print(f"Check {test_data_file_path}")
                assert False
        # If not found stop the execution else continue the execution
        campaignwb = openpyxl.load_workbook(test_data_file_path)
        campaignsheet = campaignwb["CAMPAIGNS_TOTEST"]
        campaignsheetrows = campaignsheet.max_row
        campaigns_test = []
        yes = 0
        for r in range(2, campaignsheetrows + 1):
            values = ''
            if campaignsheet.cell(row=r, column=3).value == "Yes":
                yes +=1
                for c in range(1, 6):
                    value = str(campaignsheet.cell(row=r, column=c).value).strip()
                    values = values + value + ","
                temp = values.split(",")
                campaigns_test.append(tuple(temp))
        print(yes)
        # Fetch environment details to carry out execution based on user input
        environment_test = []
        environmentwb = openpyxl.load_workbook(test_data_file_path)
        environmentsheet = environmentwb["ENVIRONMENTS_USERINPUT_LOGIN"]
        campaignsheetrows = environmentsheet.max_row
        for r in range(2, campaignsheetrows + 1):
            values = ''
            if environmentsheet.cell(row=r, column=5).value == "Yes":
                for c in range(1, 5):
                    value = environmentsheet.cell(row=r, column=c).value.strip()
                    values = values + value + ","
                temp = values.split(",")
                environment_test.append(tuple(temp))
        print(environment_test)
        campaign_environment = []
        for x in campaigns_test:
            va = ''
            for y in environment_test:
                encrypted_y_3 = encrypte_decrypte(text=y[3])
                va = x[0] + "," + x[1] + "," + x[3]+","+x[4]+","+y[0] + "," + y[1] + "," + y[2] + "," + encrypted_y_3
                temp = va.split(",")
                campaign_environment.append(tuple(temp))
        print(campaign_environment)
    elif ("RUNNED".lower() == remotetest_runvalue[-1].strip().lower() or "RUNNED".lower() == scheduletest_runvalue[-1].strip().lower() or "RUNNED".lower() == continuoustest_runvalue[-1].strip().lower() or "FINISHED".lower() == remotetest_runvalue[-1].strip().lower() or "FINISHED".lower() == scheduletest_runvalue[-1].strip().lower() or "FINISHED".lower() == continuoustest_runvalue[-1].strip().lower()) and ((androidtest == True or "Yes".lower() == group_runvalue[-1].strip().lower()) and ("RUNNED".lower() == Default_settings_runvalue[-1].strip().lower() or "WAITING LOAD".lower() == Change_settings_runvalue[-1].strip().lower() or "No".lower() == Default_settings_runvalue[-1].strip().lower() or "No".lower() == Change_settings_runvalue[-1].strip().lower())):
        if os.path.exists("C:\\RantCell_Automation_Data_and_Reports\\testdata\\Automation_data.xlsx"):
            campaignwb = openpyxl.load_workbook("C:\\RantCell_Automation_Data_and_Reports\\testdata\\Automation_data.xlsx")
            campaignsheet = None
            campaignsheet = campaignwb["AUTOMATION_DATA"]
            campaignsheetrows = campaignsheet.max_row
            campaigns_test = []
            campaign_environment = []
            yes = 0
            for r in range(2, campaignsheetrows + 1):
                values = ''
                if campaignsheet.cell(row=r, column=3).value == "Yes":
                    yes += 1
                    for c in range(1, 12):
                        value = str(campaignsheet.cell(row=r, column=c).value).strip()
                        values = values + value + ","
                    temp = values.split(",")
                    campaigns_test.append(tuple(temp))
            va = ''
            for x in campaigns_test:
                encrypted_y_8 = encrypte_decrypte(text=x[8])
                va = x[0] + "," + x[1] + "," + x[3] + "," + x[4] + "," + x[5] + "," + x[6] + "," + x[7] + "," + encrypted_y_8 + "," + x[9] + "," + x[10]
                temp = va.split(",")
                campaign_environment.append(tuple(temp))
    elif ((androidtest == True or  "Yes".lower() == group_runvalue[-1].strip().lower()) and ("Yes".lower() == Default_settings_runvalue[-1].strip().lower() or "RUNNED".lower() == Change_settings_runvalue[-1].strip().lower())):
        if os.path.exists("C:\\RantCell_Automation_Data_and_Reports\\testdata\\Automation_data.xlsx"):
            campaignwb = openpyxl.load_workbook("C:\\RantCell_Automation_Data_and_Reports\\testdata\\Automation_data.xlsx")
            campaignsheet = None
            campaignsheet = campaignwb["CHANGE AUTOMATION_DATA"]
            campaignsheetrows = campaignsheet.max_row
            campaigns_test = []
            campaign_environment = []
            yes = 0
            for r in range(2, campaignsheetrows + 1):
                values = ''
                if campaignsheet.cell(row=r, column=3).value == "Yes":
                    yes += 1
                    for c in range(1, 12):
                        value = str(campaignsheet.cell(row=r, column=c).value).strip()
                        values = values + value + ","
                    temp = values.split(",")
                    campaigns_test.append(tuple(temp))
            va = ''
            for x in campaigns_test:
                encrypted_y_8 = encrypte_decrypte(text=x[8])
                va = x[0] + "," + x[1] + "," + x[3] + "," + x[4] + "," + x[5] + "," + x[6] + "," + x[7] + "," + encrypted_y_8 + "," + x[9] + "," + x[10]
                temp = va.split(",")
                campaign_environment.append(tuple(temp))
    print(campaign_environment)
    return campaign_environment
################################################################################################################################################
def fetch_enviroment():
    # Fetch campaigns and devices list whose execute status are marked as "Yes"
    global test_data_file_path
    try:
        # test_data_file_path = config.test_data_path
        test_data_file_path = "C:\\RantCell_Automation_Data_and_Reports\\testdata\\Test_Data.xlsx"
    except:
        with allure.step(f"Check {test_data_file_path}"):
            print(f"Check {test_data_file_path}")
            assert False
    # Fetch environment details to carry out execution based on user input
    environment_test = []
    environmentwb = openpyxl.load_workbook(test_data_file_path)
    environmentsheet = environmentwb["ENVIRONMENTS_USERINPUT_LOGIN"]
    campaignsheetrows = environmentsheet.max_row
    for r in range(2, campaignsheetrows + 1):
        values = ''
        if environmentsheet.cell(row=r, column=5).value == "Yes":
            for c in range(1, 5):
                value = environmentsheet.cell(row=r, column=c).value.strip()
                values = values + value + ","
            temp = values.split(",")
            environment_test.append(tuple(temp))
    print(environment_test)
    campaign_environment = []
    va = ''
    for y in environment_test:
        encrypted_y_3 = encrypte_decrypte(text=y[3])
        va = y[0] + "," + y[1] + "," + y[2] + "," + encrypted_y_3
        temp = va.split(",")
        campaign_environment.append(tuple(temp))
    print(campaign_environment)
    return campaign_environment
###############################################################################################################################################################################
def fetch_components(strcampaignname,startcolumn,EndColumn):
    # Fetch components list based on campaign name
    # current directory
    test_data_file_path = "None"
    current_dir = os.getcwd()
    parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))

    try:
        test_data_file_path = config.test_data_path
    except Exception as e:
        with allure.step(f"Check {test_data_file_path}"):
            print(f"Check {test_data_file_path}")
            assert False

    component_list = []
    componentwb = openpyxl.load_workbook(test_data_file_path)
    componentsheet = componentwb["TC"]
    componentsheetrows = componentsheet.max_row
    componentsheetcoloumns = componentsheet.max_column
    for r in range(3, componentsheetrows + 1):
        if componentsheet.cell(row=r, column=4).value is not None and componentsheet.cell(row=r, column=4).value.strip() == strcampaignname:
            for c in range(startcolumn, EndColumn): #componentsheetcoloumns + 1):
                if componentsheet.cell(row=r, column=c).value == "Yes":
                    value = componentsheet.cell(row=2, column=c).value
                    component_list.append(value)
    print(component_list)
    return component_list
def fetch_components_datetime_query(startcolumn,EndColumn):
    # Fetch components list based on campaign name
    # current directory
    test_data_file_path = "None"
    try:
        test_data_file_path = config.test_data_path
    except Exception as e:
        with allure.step(f"Check {test_data_file_path}"):
            print(f"Check {test_data_file_path}")
            assert False

    component_list = []
    componentwb = openpyxl.load_workbook(test_data_file_path)
    componentsheet = componentwb["TC"]
    for c in range(startcolumn, EndColumn):
        # if componentsheet.cell(row=r, column=c).value == "Yes":
        value = componentsheet.cell(row=2, column=c).value
        component_list.append(value)
    print(component_list)
    return component_list

def fetch_components_for_no_yes(strcampaignname,startcolumn,EndColumn):
    # Fetch components list based on campaign name
    # current directory
    global test_data_file_path
    current_dir = os.getcwd()
    parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))

    try:
        test_data_file_path = config.test_data_path
    except Exception as e:
        with allure.step(f"Check {test_data_file_path}"):
            print(f"Check {test_data_file_path}")
            assert False

    component_list = []
    componentwb = openpyxl.load_workbook(test_data_file_path)
    componentsheet = componentwb["TC"]
    componentsheetrows = componentsheet.max_row
    componentsheetcoloumns = componentsheet.max_column
    for r in range(3, componentsheetrows + 1):
        if componentsheet.cell(row=r, column=4).value is not None and componentsheet.cell(row=r, column=4).value.strip() == strcampaignname:
            for c in range(startcolumn, EndColumn): #componentsheetcoloumns + 1):
                if componentsheet.cell(row=r, column=c).value == "" or componentsheet.cell(row=r, column=c).value == None:
                    value = componentsheet.cell(row=2, column=c).value
                    component_list.append(value)
    print(component_list)
    return component_list
###############################################################################################################################################################################

##########################################################fetch_components points #####################################################################################################################
def read_first_row( file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    first_row = sheet[1]
    values = [cell.value for cell in first_row]
    return values
def fetch_input_points():
    file_path = config.test_data_path
    sheet_name = 'TC'
    first_row_values = read_first_row(file_path, sheet_name)
    remote_test_index = first_row_values.index('Remote Test')
    map_view_index = first_row_values.index('Map View')
    graph_view_index = first_row_values.index('Graph View')
    export_index = first_row_values.index('Exports')
    loading_index = first_row_values.index('Loading')
    PDF_Export_index = first_row_values.index('PDF Export')
    END_index = first_row_values.index('END')
    return (remote_test_index+1),(map_view_index+1), (graph_view_index+1), (export_index+1), (loading_index+1),(PDF_Export_index+1),(END_index+1)
###############################################################################################################################################################################

######################################################## readcomponentstatus #######################################################################################################################
# readcomponentstatus : This method fetches status of components from COMPONENTSTATUS sheet of Excel Report.
def readcomponentstatus(path):
    failed_component_list = ''
    componentwb = openpyxl.load_workbook(path)
    componentsheet = componentwb["COMPONENTSTATUS"]
    componentsheetrows = componentsheet.max_row
    componentsheetcoloumns = componentsheet.max_column
    for r in range(2, componentsheetrows + 1):
        if str(componentsheet.cell(row=r, column=3).value).lower()=='failed':
            failed_component_list = failed_component_list + componentsheet.cell(row=r, column=3).value
    componentwb.close()
    return failed_component_list
###############################################################################################################################################################################
# fetch_camapaigns_enviroment()

def fetch_camapaigns():
    # Fetch campaigns and devices list whose execute status are marked as "Yes"
    global test_data_file_path
    try:
        # test_data_file_path = config.test_data_path
        test_data_file_path = "C:\\RantCell_Automation_Data_and_Reports\\testdata\\Test_Data.xlsx"
    except:
        with allure.step(f"Check {test_data_file_path}"):
            print(f"Check {test_data_file_path}")
            assert False
    # If not found stop the execution else continue the execution
    campaignwb = openpyxl.load_workbook(test_data_file_path)
    campaignsheet = campaignwb["CAMPAIGNS_TOTEST"]
    campaignsheetrows = campaignsheet.max_row
    campaigns_test = []
    yes = 0
    for r in range(2, campaignsheetrows + 1):
        values = ''
        if campaignsheet.cell(row=r, column=3).value == "Yes":
            yes +=1
            for c in range(1, 6):
                value = str(campaignsheet.cell(row=r, column=c).value).strip()
                values = values + value + ","
            temp = values.split(",")
            campaigns_test.append(tuple(temp))
    print(yes)
    campaign = []
    for x in campaigns_test:
        va = ''
        va = x[0] + "," + x[1] + "," + x[3]+","+x[4]
        temp = va.split(",")
        campaign.append(tuple(temp))
    print(campaign)
    return campaign