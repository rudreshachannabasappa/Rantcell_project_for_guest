import allure
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xlwings as xw
import re

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait



def updatedatapoints(path, data, componentname, Title):
    """
    Update data points in an Excel workbook.
    Args:
        path (str): The path to the Excel workbook.
        data (list): A list of data points to update in the workbook.
        componentname (str): The name of the component or section being updated.
        Title (str): The title or name of the data being updated.
    Notes:
        This function loads an Excel workbook using openpyxl, and updates data points in the 'DATAEXTRACT' worksheet.
        It iterates through the provided data, appending it to the worksheet, and then saves and closes the workbook.
        If no data is provided, it raises a ValueError and logs a failure message in the worksheet.
        Additionally, if an exception occurs during the update process, it catches the exception, logs a failure message,
        and updates the status of the component using the 'updatecomponentstatus' function.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    try:
        if len(data) > 0 :
            for r in data:
                k=1
                rows = ws.max_row
                for c in r:
                    ws.cell(row=rows + 1, column=k).value = c
                    k = k + 1
            wb.save(path)
            wb.close()
        else:
            raise ValueError("No Blob")
    except:
        rows = ws.max_row
        ws.cell(row=rows + 1, column=1).value = f"FAILED :- No Data for '{Title}' in '{str(componentname).upper()}'"
        fill_cell = PatternFill(patternType='solid', fgColor='FF0000')
        ws.cell(row=rows + 1, column=1).fill = fill_cell
        updatecomponentstatus(Title,componentname, "FAILED", "No Data", path)
        wb.save(path)
        wb.close()
def updatedatapoints3D(path, data,componentname,Title):
    """
    Update data points in an Excel workbook with support for 3D lists.
    Args:
        path (str): The path to the Excel workbook.
        data (list): A 3D list of data points to update in the workbook.
        componentname (str): The name of the component or section being updated.
        Title (str): The title or name of the data being updated.
    Notes:
        This function loads an Excel workbook using openpyxl, and updates data points in the 'DATAEXTRACT' worksheet.
        It iterates through the provided 3D data list and appends the values to the worksheet, maintaining the 3D structure.
        After updating, it saves and closes the workbook.

        If a TypeError occurs (typically indicating no data), it logs a failure message in the worksheet and updates the
        status of the component using the 'updatecomponentstatus' function.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb['DATAEXTRACT']
        # Find the next available row
        next_row = ws.max_row + 1
        wb = openpyxl.load_workbook(path)
        ws = wb['DATAEXTRACT']
        # Find the next available row
        next_row = ws.max_row + 1
        # Iterate over the 3D list and write the values to the worksheet
        for row in data:
            for inner_list in row:
                ws.append(inner_list)
        # Save the workbook
        wb.save(path)
        wb.close()
    except TypeError:
        with allure.step(f"failed steps :- No Data '{Title}' in '{componentname}'"):
            rows = ws.max_row
            ws.cell(row=rows + 1, column=1).value = "No Data"
            ws.cell(row=rows + 1, column=2).value = "FAILED"
            fill_cell = PatternFill(patternType='solid', fgColor='FF0000')
            ws.cell(row=rows + 1, column=2).fill = fill_cell
            updatecomponentstatus(Title,componentname, "FAILED", "No Data", path)
            wb.save(path)
            wb.close()
def updatedatapoints1(path, data,componentname,Title):
    """
    Update data points in an Excel workbook with support for 2D lists containing single-element sublists.
    Args:
        path (str): The path to the Excel workbook.
        data (list): A 2D list of data points to update in the workbook, where each element is a single-element sublist.
        componentname (str): The name of the component or section being updated.
        Title (str): The title or name of the data being updated.
    Notes:
        This function loads an Excel workbook using openpyxl and updates data points in the 'DATAEXTRACT' worksheet.
        It iterates through the provided 2D data list, where each element is a single-element sublist containing a string.
        The function capitalizes each string and appends it to the worksheet, maintaining the 2D structure.
        After updating, it saves the workbook.

        If a TypeError occurs (typically indicating no data), it logs a failure message in the worksheet and updates the
        status of the component using the 'updatecomponentstatus' function.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    try:
        for r in data:
            k = 1
            rows = ws.max_row
            for c in r:
                ws.cell(row=rows + 1, column=k).value = c[0].upper()
                k += 1
            rows += 1
        wb.save(path)
    except TypeError:
        with allure.step(f"failed steps :- No Data '{Title}' in '{componentname}'"):
            rows = ws.max_row
            ws.cell(row=rows + 1, column=1).value = "No Data"
            ws.cell(row=rows + 1, column=2).value = "FAILED"
            updatecomponentstatus(Title,componentname, "FAILED", "No Data", path)
            wb.save(path)
            wb.close()
def updatename(path, componentname):
    """
    Update the component name in an Excel workbook.
    Args:
        path (str): The path to the Excel workbook.
        componentname (str): The component name to be added.
    Notes:
        This function loads an Excel workbook using openpyxl and updates the component name in the 'DATAEXTRACT' worksheet.
        The provided 'componentname' is added to the worksheet and highlighted with a turquoise color (hex code: '03FCF4').
        After updating, it saves the workbook.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row
    ws.cell(row=rows + 1, column=1).value = componentname
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 1, column=1).fill = fill_cell
    wb.save(path)
    wb.close()
def Graphupdatename(path, componentname):
    """
    Graph Update the component name in an Excel workbook for a graph.
    Args:
        path (str): The path to the Excel workbook.
        componentname (str): The component name to be added for the graph.
    Notes:
        This function loads an Excel workbook using openpyxl and updates the component name in the 'DATAEXTRACT' worksheet.
        The provided 'componentname' is added to the worksheet, merged across multiple columns, and highlighted with a
        yellow color (hex code: 'FCBA03').
        The text alignment is set to left.
        After updating, it saves the workbook.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row
    ws.merge_cells('A{}:Z{}'.format(rows + 1, rows + 1))
    cell = ws.cell(row=rows + 1, column=1)
    cell.value = componentname
    cell.fill = PatternFill(patternType='solid', fgColor='FCBA03')
    cell.alignment = Alignment(horizontal='left')
    wb.save(path)
    wb.close()
def GraphViewname(path, componentname):
    """
    Update the download name for a report in an Excel workbook.
    Args:
        path (str): The path to the Excel workbook.
        componentname (str): The download name to be added for the report.
    Notes:
        This function loads an Excel workbook using openpyxl and updates the download name in the 'DATAEXTRACT' worksheet.
        The provided 'componentname' is added to the worksheet, merged across multiple columns, and highlighted with a
        lavender color (hex code: 'EE82EE').
        The text alignment is set to left.
        After making these changes, it saves the workbook.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row
    ws.merge_cells('A{}:Z{}'.format(rows + 3, rows + 3))
    cell = ws.cell(row=rows + 3, column=1)
    cell.value = componentname
    cell.fill = PatternFill(patternType='solid', fgColor='FCBA03')
    cell.alignment = Alignment(horizontal='left')
    ws.insert_rows(rows + 4, amount=1)  # Insert two blank rows
    wb.save(path)
    wb.close()
def ReportDownlaodName(path, componentname):
    """
    Update the download name for a report in an Excel workbook.
    Args:
        path (str): The path to the Excel workbook.
        componentname (str): The download name to be added for the report.
    Notes:
        This function loads an Excel workbook using openpyxl and updates the download name in the 'DATAEXTRACT' worksheet.
        The provided 'componentname' is added to the worksheet, merged across multiple columns, and highlighted with a
        specified fill color. The text alignment is set to left.
        After making these changes, it saves the workbook.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row
    ws.merge_cells('A{}:Z{}'.format(rows + 3, rows + 3))
    cell = ws.cell(row=rows + 3, column=1)
    cell.value = componentname
    cell.fill = PatternFill(patternType='solid', fgColor='EE82EE')
    cell.alignment = Alignment(horizontal='left')
    wb.save(path)
    wb.close()
def Failupdatename(path, componentname):
    """
        Update the name of a component in an Excel workbook to indicate failure.
        Args:
            path (str): The path to the Excel workbook.
            componentname (str): The name of the component to be updated.
        Notes:
            This function loads an Excel workbook using openpyxl and updates the component name in the 'DATAEXTRACT' worksheet.
            The provided 'componentname' is added to the worksheet, merged across multiple columns, and highlighted with a
            specified fill color (red). The text alignment is set to left.
            After updating, it saves the workbook.
        """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row

    ws.merge_cells('A{}:Z{}'.format(rows + 1, rows + 1))
    cell = ws.cell(row=rows + 1, column=1)
    cell.value = componentname
    cell.fill = PatternFill(patternType='solid', fgColor='FC2C03')
    cell.alignment = Alignment(horizontal='left')
    wb.save(path)
    wb.close()
def pass_updatename(path, componentname):
    """
       Update the name of a component in an Excel workbook to indicate a passing status.
       Args:
           path (str): The path to the Excel workbook.
           componentname (str): The name of the component to be updated.
       Notes:
           This function loads an Excel workbook using openpyxl and updates the component name in the 'DATAEXTRACT' worksheet.
           The provided 'componentname' is added to the worksheet, merged across multiple columns, and highlighted with a
           specified fill color (green). The text alignment is set to the left.
           After updating, it saves the workbook.
       """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row
    ws.merge_cells('A{}:Z{}'.format(rows + 1, rows + 1))
    cell = ws.cell(row=rows + 1, column=1)
    cell.value = componentname
    cell.fill = PatternFill(patternType='solid', fgColor='35FC03')
    cell.alignment = Alignment(horizontal='left')
    wb.save(path)
    wb.close()
def updateiteration(testcasename, devicename, environmentname, url, path):
    """
        Update test iteration information in an Excel workbook.
        Args:
            iternation (int): The iteration number.
            testcasename (str): The name or identifier of the test case.
            devicename (str): The name or identifier of the test device.
            environmentname (str): The name or identifier of the test environment.
            url (str): The URL associated with the test environment.
            path (str): The path to the Excel workbook.
        Notes:
            This function loads an Excel workbook using openpyxl and updates test iteration information in the 'DATAEXTRACT' worksheet.
            It adds a new row with the following information:
            - Iteration number
            - Test case name
            - Device name
            - Environment name
            - Environment URL
            Each cell is filled with a solid color for formatting.
            After updating, it saves the workbook.
        """
    wb = openpyxl.load_workbook(path)
    ws = wb['DATAEXTRACT']
    rows = ws.max_row
    ws.cell(row=rows + 2, column=1).value = "TestCase"
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 2, column=1).fill = fill_cell
    ws.cell(row=rows + 2, column=2).value = testcasename
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 2, column=2).fill = fill_cell
    ws.cell(row=rows + 2, column=3).value = "Device Name"
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 2, column=3).fill = fill_cell
    ws.cell(row=rows + 2, column=4).value = devicename
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 2, column=4).fill = fill_cell
    ws.cell(row=rows + 2, column=5).value = "Environment Name"
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 2, column=5).fill = fill_cell
    ws.cell(row=rows + 2, column=6).value = environmentname + " : " + url
    fill_cell = PatternFill(patternType='solid',fgColor='03FCF4')
    ws.cell(row=rows + 2, column=6).fill = fill_cell
    wb.save(path)
    wb.close()
def startcomponentstatus_test_case_(testcasename,path):
    """
        Mark the start of a test case in an Excel workbook that tracks component status.
        Args:
            testcasename (str): The name or identifier of the test case.
            path (str): The path to the Excel workbook.
        Notes:
            This function loads an Excel workbook using openpyxl and marks the start of a test case in the 'COMPONENTSTATUS' worksheet.
            It adds a row with a message indicating the start of the test case.
            The cell containing the message is filled with a specified solid color (tan) for formatting.
            After marking the start of the test case, it saves the workbook.
        """
    wb = openpyxl.load_workbook(path)
    ws = wb['COMPONENTSTATUS']
    rows = ws.max_row
    ws.merge_cells('A{}:Z{}'.format(rows, rows))
    cell = ws.cell(row=rows, column=1)
    cell.value = f"*************    Test_case {testcasename}  --- Starts from here  *************"
    cell.fill = PatternFill(patternType='solid', fgColor='D2B48C')
    cell.alignment = Alignment(horizontal='left')
    wb.save(path)
    wb.close()
def finishcomponentstatus_test_case_(testcasename,path):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb['COMPONENTSTATUS']
        rows = ws.max_row
        ws.merge_cells('A{}:Z{}'.format(rows + 1, rows + 1))
        cell = ws.cell(row=rows + 1, column=1)
        cell.value = f"*************    Test_case {testcasename}  --- finished/ends *************"
        cell.fill = PatternFill(patternType='solid', fgColor='D2B48C')
        cell.alignment = Alignment(horizontal='left')
        wb.save(path)
        wb.close()
    except Exception as e:
        pass
def updatecomponentstatus(Title,componentname, status, comments, path):
    """
        Mark the finish or end of a test case in an Excel workbook that tracks component status.
        Args:
            testcasename (str): The name or identifier of the test case.
            path (str): The path to the Excel workbook.
        Notes:
            This function loads an Excel workbook using openpyxl and marks the finish or end of a test case in the 'COMPONENTSTATUS' worksheet.
            It adds a row with a message indicating the end of the test case.
            The cell containing the message is filled with a specified solid color (tan) for formatting.
            After marking the end of the test case, it saves the workbook.
        """
    wb = openpyxl.load_workbook(path)
    ws = wb['COMPONENTSTATUS']
    rows = ws.max_row
    ws.cell(row=rows + 1, column=1).value = Title
    ws.cell(row=rows + 1, column=2).value = componentname
    ws.cell(row=rows + 1, column=3).value = status
    ws.cell(row=rows + 1, column=4).value = comments
    if status == "PASSED":
        fill_cell = PatternFill(patternType='solid',fgColor='35FC03')
        ws.cell(row=rows + 1, column=3).fill = fill_cell
    elif status == "FAILED":
        fill_cell = PatternFill(patternType='solid',fgColor='FC2C03')
        ws.cell(row=rows + 1, column=3).fill = fill_cell
    elif status == "IGNORED":
        fill_cell = PatternFill(patternType='solid',fgColor='FFFF98')
        ws.cell(row=rows + 1, column=3).fill = fill_cell
    elif status == "SKIPPED":
        fill_cell = PatternFill(patternType='solid',fgColor='FFA590')
        ws.cell(row=rows + 1, column=3).fill = fill_cell
    elif status == "WARNING":
        fill_cell = PatternFill(patternType='solid',fgColor='FFFF98')
        ws.cell(row=rows + 1, column=3).fill = fill_cell
    wb.save(path)
    wb.close()
def updatecomponentstatus_using_pandas_sending_data_frame_for_highlevelreport(sheet, dataframe):
    """
        Update component status information in a high-level report Excel sheet using Pandas DataFrame.
        package used pandas and xlwings
        Args:
            sheet (xlwings.Sheet): The Excel sheet to update.
            dataframe (pd.DataFrame): The Pandas DataFrame containing component status data.
        Notes:
            This function updates the component status information in the specified Excel sheet.
            It takes the provided Pandas DataFrame and appends its data to the next available row in the sheet.
            Additionally, it assigns fill colors to cells in the fourth column (column D) based on the status values.
            The status colors are defined in the 'status_colors' dictionary.
        """
    try:
        # Map status values to corresponding fill colors
        status_colors = {
            "PASSED": (53, 252, 3),
            "FAILED": (252, 44, 3),
            "IGNORED": (255, 249, 152),
            "SKIPPED": (255, 165, 144),
            "WARNING": (255, 249, 152)
        }
        # Find the last used row in the sheet
        last_row = sheet.range((sheet.cells.last_cell.row, 1)).end('up').row
        # Convert the DataFrame to a Pandas DataFrame
        df_range = sheet.range((last_row + 1, 1))  # Start appending from the next row
        df_range.value = dataframe.values  # Assign the DataFrame values to the range
        # Update cell colors based on the "Status" column
        for row in range(last_row + 1, last_row + 1 + len(dataframe)):
            status = dataframe.loc[row - last_row - 1, "Status"]
            if status in status_colors:
                rgb_color = status_colors[status]
                sheet.range((row, 4)).color = rgb_color  # Assuming Status is in column B
    except Exception as e:
        pass

def update_component_status_openpyxl(worksheet, dataframe):
    """
        Update component status information in an Excel worksheet using openpyxl and Pandas DataFrame.
        package used pandas and openpyxl
        Args:
            worksheet (openpyxl.Worksheet): The Excel worksheet to update.
            dataframe (pd.DataFrame): The Pandas DataFrame containing component status data.
        Notes:
            This function updates the component status information in the specified Excel worksheet.
            It takes the provided Pandas DataFrame and appends its data to the worksheet starting from the next available row.
            Additionally, it assigns fill colors to cells in the "Status" column (column C) based on the status values.
            The status colors are defined in the 'status_colors' dictionary.
        """
    try:
        # Map status values to corresponding fill colors
        status_colors = {
            "PASSED": PatternFill(start_color="35FC03", end_color="35FC03", fill_type="solid"),
            "FAILED": PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid"),
            "IGNORED": PatternFill(start_color="FFF998", end_color="FFF998", fill_type="solid"),
            "SKIPPED": PatternFill(start_color="FFA590", end_color="FFA590", fill_type="solid")
        }
        # Find the last used row in the sheet
        last_row = worksheet.max_row
        # Append DataFrame data to the worksheet
        for index, row in dataframe.iterrows():
            worksheet.append(row.tolist())
        # Update cell colors based on the "Status" column
        for row_index, row in enumerate(worksheet.iter_rows(min_row=last_row + 1, max_row=last_row + len(dataframe), min_col=3, max_col=3)):
            status_cell = row[0]
            status = status_cell.value
            if status in status_colors:
                status_cell.fill = status_colors[status]
    except Exception as e:
        with allure.step(f"component status {str(e)}"):
            pass
def format_workbook(file_path):
    """
        Format an Excel workbook by adjusting sheet names, adding borders to cells, and setting column widths.
        Args:
            file_path (str): The path to the Excel workbook to be formatted.
        Notes:
            This function loads an Excel workbook using openpyxl and performs the following formatting tasks:
            1. Removes the default sheet named 'Sheet' if it exists.
            2. Adds borders to cells with values in all sheets.
            3. Adjusts column widths in the 'DATAEXTRACT' and 'COMPONENTSTATUS' sheets.
            4. Saves the modified workbook.
        """
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws_data = wb['DATAEXTRACT']
    ws_status = wb['COMPONENTSTATUS']
    # Remove sheet named 'Sheet' if it exists
    if 'Sheet' in wb.sheetnames:
        sheet = wb['Sheet']
        wb.remove(sheet)
    # Add borders to cells with values in all sheets
    ws = wb['COMPONENTSTATUS']
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),right=openpyxl.styles.Side(style='thin'),top=openpyxl.styles.Side(style='thin'),bottom=openpyxl.styles.Side(style='thin'))
                cell.border = border
    # Set column widths in 'DATAEXTRACT' sheet
    ws_data.column_dimensions['A'].width = 73  # Width in characters (approximately 8 cm)
    ws_data.column_dimensions['F'].width = 45  # Width in characters (approximately 8 cm)
    # Set column widths in 'COMPONENTSTATUS' sheet
    ws_status.column_dimensions['A'].width = 22
    ws_status.column_dimensions['B'].width = 22
    ws_status.column_dimensions['C'].width = 22
    ws_status.column_dimensions['D'].width = 22
    # Save the modified workbook
    wb.save(file_path)
    wb.close()
def export_pdf_update_to_excel(datas,sheetname,componentname,path):
    """
        Export PDF update data to an Excel worksheet and format it based on component and "ENDHERE" values.
        Args:
            datas (list): A list of data rows to be exported and formatted.
            sheetname (str): The name of the destination sheet in the Excel workbook.
            componentname (str): The component name to identify and format differently.
            path (str): The path to the Excel workbook where data will be exported.
        Notes:
            This function loads an existing Excel workbook using openpyxl and appends the provided data to the specified worksheet.
            It formats rows containing the 'componentname' with bold text and a green fill color and rows containing 'ENDHERE'
            with bold text and a red fill color. Other rows are written as-is.
            After updating, it saves the modified workbook.
        """
    dest_wb = load_workbook(filename=path)
    # Get the active sheet of the destination workbook
    dest_ws = dest_wb[sheetname]
    rows = dest_ws.max_row
    # Iterate over the 'datas' list and write each row to the destination worksheet
    for datas1 in datas:
        # Increment the row index
        rows += 1
        # Write the values to the destination worksheet
        for col, value in enumerate(datas1):
            if str(componentname).strip().lower() == str(value).strip().lower():
                dest_ws.cell(row=rows, column=col + 1).value = value
                font = Font(bold=True)
                fill = PatternFill(fill_type='solid', fgColor="35FC03")
                # dest_ws.cell(row=rows, column=col + 1).value = font
                dest_ws.cell(row=rows, column=col + 1).fill = fill
            elif "ENDHERE" == value:
                dest_ws.cell(row=rows, column=col + 1).value = value
                font = Font(bold=True)
                fill = PatternFill(fill_type='solid', fgColor="FC2C03")
                # dest_ws.cell(row=rows, column=col + 1).value = font
                dest_ws.cell(row=rows, column=col + 1).fill = fill
            else:
                dest_ws.cell(row=rows, column=col + 1).value = value
    # Save the destination workbook
    dest_wb.save(path)
    dest_wb.close()

def add_headers_and_data(file_path, headers, sheet_name):
    """
        Add headers to an Excel worksheet and apply a yellow fill color to the header row.
        Args:
            file_path (str): The path to the Excel workbook.
            headers (list): A list of header values to be added.
            sheet_name (str): The name of the worksheet where headers will be added.
        Notes:
            This function loads an existing Excel workbook using openpyxl and appends the provided headers to the specified
            worksheet. It also finds the row where headers are added and applies a yellow fill color to that row.
            After updating, it saves the modified workbook.
        """
    # Load the existing workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the desired sheet
    ws = wb[sheet_name]
    # Create headers in the found empty row
    ws.append(headers)
    # Save the workbook after appending headers
    wb.save(file_path)
    # Find the row where headers are present
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value in headers:
                header_row = cell.row
                break
        if header_row is not None:
            break
    # Apply yellow color to the header row
    if header_row is not None:
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[header_row]:
            cell.fill = yellow_fill
    # Save the changes
    wb.save(file_path)
    wb.close()

############################################################################ Settings section ###############################################################################################
def extract_numerical_values(setting):
    # Use regular expression to find all numerical values, including floating-point numbers
    numerical_values = re.findall(r'-?\d+(?:\.\d+)?', setting)
    return numerical_values

def check_numeric_value(a, b):
    # Extracting numeric value from 'a'
    try:
        numeric_value_a = float(a)
    except ValueError:
        print("'a' is not a valid numeric value.")
        return False

    # Extracting numeric values from 'b'
    numeric_values_b = [float(match) for match in re.findall(r'-?\d+(?:\.\d+)?', b)]

    # Checking if numeric value of 'a' is present in 'b'
    if numeric_value_a in numeric_values_b:
        return True
    else:
        return False
def updating_settings_data_extraction_to_excel(reference,combine_data1,combine_data2,combine_data3,excelpath,sheet_name):
    # Find the maximum length among all dictionaries
    max_length = max(len(reference), len(combine_data1), len(combine_data2), len(combine_data3))

    # Function to fill missing values with None
    def fill_missing_values(dictionary):
        filled_dict = {}
        for key, value in dictionary.items():
            filled_dict[key] = value + [None] * (max_length - len(value))
        return filled_dict

    # Fill dictionaries with NaN values to match the maximum length
    reference_filled = fill_missing_values(reference)
    combine_data1_filled = fill_missing_values(combine_data1)
    combine_data2_filled = fill_missing_values(combine_data2)
    combine_data3_filled = fill_missing_values(combine_data3)

    # Create DataFrames from filled dictionaries
    df_reference = pd.DataFrame(reference_filled)
    df_combine_data1 = pd.DataFrame(combine_data1_filled)
    df_combine_data2 = pd.DataFrame(combine_data2_filled)
    df_combine_data3 = pd.DataFrame(combine_data3_filled)
    a = {"Settings Data Extraction":df_reference,"Operator Comparison Data Extraction":df_combine_data1,"Map Legend Data Extraction":df_combine_data2," PDF Data Extraction":df_combine_data3}
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_data_extract = workbook[sheet_name]
    for df_key,df in a.items():
        updating_data_of_dataframe_for_excel1(worksheet=worksheet_data_extract, df_data=df,df_key=df_key,space_required=1)
    workbook.save(excelpath)
    workbook.close()

# def updating_comparison_results_to_excel(result,excelpath,sheet_name):
#     workbook = openpyxl.load_workbook(excelpath)
#     worksheet_data_extract = workbook[sheet_name]
#     for category, results in result.items():
#         if results:
#             # Convert the list of dictionaries to a DataFrame
#             df = pd.DataFrame(results)
#             updating_data_of_dataframe_for_excel(worksheet=worksheet_data_extract, df_data=df,df_key=category)
#     workbook.save(excelpath)
#     workbook.close()
def updating_comparison_results_to_excel1(result,excelpath,sheet_name):
    df_data =[]
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_data_extract = workbook[sheet_name]
    for category, results in result.items():
        if results:
            # Convert the list of dictionaries to a DataFrame
            df = pd.DataFrame(results)
            df_data.append(df)
    combined_df = pd.concat(df_data, ignore_index=True)
    # Extract the 'Data validation' column
    data_validation_column = combined_df['Data validation']
    # Drop the 'Data validation' column
    combined_df.drop(columns=['Data validation'], inplace=True)
    # Concatenate the 'Data validation' column at the end
    combined_df['Data validation'] = data_validation_column
    updating_data_of_dataframe_for_excel1(worksheet=worksheet_data_extract, df_data=combined_df,df_key="",space_required=0)
    workbook.save(excelpath)
    workbook.close()
def updating_data_of_dataframe_for_excel1(worksheet,df_data,df_key,space_required):
    # Convert the DataFrame to a list of rows
    data = list(dataframe_to_rows(df_data, index=False, header=True))
    # Find the last row in the existing data
    last_row = worksheet.max_row
    next_available_row = last_row
    if space_required !=0:
        # Calculate the next available row for appending
        next_available_row = last_row + space_required

    # Define fill colors
    found_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green color
    not_found_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color

    if df_key != "":
        first_column_fill = PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid")
        cell=worksheet.cell(row=next_available_row, column=1, value=df_key)
        cell.fill = first_column_fill
    # Define a fill color for the first column (1st column value)
    header_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    # Iterate through rows and columns to append data to the worksheet
    for row_idx, row_data in enumerate(data, 1):  # Start from row 1
        for col_idx, cell_value in enumerate(row_data, 1):  # Start from column 1
            cell = worksheet.cell(row=next_available_row + row_idx, column=col_idx, value=cell_value)
            if row_idx == 1:
                cell.fill = header_fill
            else:
                cell_value = str(cell_value)# Apply color based on content
                if "The value is found" == cell_value:
                    cell.fill = found_fill
                elif "The value is Not Found" == cell_value:
                    cell.fill = not_found_fill
                elif "The value is found,but settings application value(reference) != excel settings values" == cell_value:
                    cell.fill = not_found_fill

def create_workbook_for_automation_data(path):
    workbook = Workbook()
    workbook.create_sheet("CHANGE AUTOMATION_DATA", 0)
    workbook.save(path)