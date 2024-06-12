from openpyxl import Workbook
def create_workbook(path):
    """
        Create a new Excel workbook with specific sheet names.
        Args:
            path (str): The path where the new Excel workbook will be saved.
        Notes:
            This function creates a new Excel workbook using openpyxl and adds multiple sheets with predefined names.
            The sheets include 'COMPONENTSTATUS', 'DATAEXTRACT', 'OPERATOR_COMPARISON', 'PDF_EXPORT', 'DATA_MATCH',
            'DATA_NOT_MATCH', 'TABLESUMMARY_DATA_NOT_MATCH', 'TABLESUMMARY_DATA_MATCH', 'CBE_vs_CE_MATCH', and
            'CBE_vs_CE_DONOT_MATCH'.
            After creating the workbook with sheets, it saves the workbook to the specified path.
        """
    workbook = Workbook()
    workbook.create_sheet("COMPONENTSTATUS", 0)
    workbook.create_sheet("DATAEXTRACT", 1)
    workbook.create_sheet("OPERATOR_COMPARISON", 2)
    workbook.create_sheet("PDF_EXPORT", 3)
    workbook.create_sheet("DATA_MATCH", 4)
    workbook.create_sheet("DATA_NOT_MATCH", 5)
    workbook.create_sheet("TABLESUMMARY_DATA_NOT_MATCH", 6)
    workbook.create_sheet("TABLESUMMARY_DATA_MATCH", 7)
    workbook.create_sheet("CBE_vs_CE_MATCH", 8)
    workbook.create_sheet("CBE_vs_CE_DONOT_MATCH", 9)
    workbook.create_sheet("IPU_vs_CE_DATA_MATCH",10)
    workbook.create_sheet("IPU_vs_CE_DATA_NOT_MATCH",11)
    workbook.create_sheet("NQC_vs_OC_DATA_MATCH", 12)
    workbook.create_sheet("NQC_vs_OC_DATA_NOT_MATCH", 13)
    workbook.create_sheet("DATA_EXTRACTION_SETTINGS", 14)
    workbook.create_sheet("RESULTS_DEFAULT_SETTINGS",15)
    workbook.create_sheet("DATA_EXTRACTION_CHANGE_SETTINGS", 16)
    workbook.create_sheet("RESULTS_CHANGE_SETTINGS", 17)
    workbook.save(path)

