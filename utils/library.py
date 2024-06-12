import csv, glob, os, time, allure,openpyxl,datetime,shutil,re
import pathlib
import random
import string
import cv2
import numpy as np
import pandas as pd
from allure_commons.types import AttachmentType
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.common import NoSuchElementException, TimeoutException, ElementClickInterceptedException, \
    NoAlertPresentException, UnexpectedAlertPresentException, StaleElementReferenceException, \
    ElementNotInteractableException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from locators.locators import select_Map_View_Components
from utils.updateexcelfile import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font,Border,Side
from selenium.webdriver import *
from configurations.config import ReadConfig as config
from itertools import zip_longest
global timeout
timeout = 60

################################################################-- LAUNCHBROWSER --########################################################################################################################################################################
# Function:launchbrowser - Launches the browser and navigates to the URL
# Parameters:
#           url:https://preproductionpro.rantcell.com/
#           title:https://preproductionpro.rantcell.com/
def launchbrowser(driver, url):
    """
    Launch a web browser and navigate to a specified URL.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        url (str): The URL to navigate to.
    Returns:
        bool: True if the browser was successfully launched and navigated to the specified URL, False otherwise.
    Notes:
        This function launches a web browser, navigates to the specified URL, maximizes the browser window, and then
        verifies if the actual URL matches the expected URL. If all these steps are successful, it returns True; otherwise,
        it returns False. Screenshots are attached to the Allure report for documentation.
    """
    try:
        with allure.step("Launch the browser and navigate to " + url):
            driver.get(url)
            driver.maximize_window()
            actualtitle = driver.current_url
            allure.attach(driver.get_screenshot_as_png(), name=f"URL : {str(url)}",attachment_type=allure.attachment_type.PNG)
            return actualtitle == url
    except Exception as e:
        with allure.step("Unable to launch the browser " + url):
            with allure.step(f"Actual URL sent from Test_Data.xlsx[ENVIRONMENTS_USERINPUT_LOGIN] : {str(url)}"): pass
            with allure.step(f"Expected URL loading from browser : {str(driver.current_url)}"): pass
            allure.attach(driver.get_screenshot_as_png(), name="URL_screenshot",attachment_type=allure.attachment_type.PNG)
            return False
##################################################################-- CLICK --##############################################################################################################################################################################
# Function:click - Clicks on particular element
# Parameters:
#           locators: (By.ID, textbox_username_id)
def click(driver, locators):
    """
    Click an element on a web page with explicit waits and handle potential click intercept exceptions.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        locators (tuple): A tuple specifying how to locate the element (locatortype, locatorProperty, elementname).
    Returns:
        bool: True if the element was successfully clicked, False otherwise.
    Notes:
        This function attempts to click an element on a web page using explicit waits and handles scenarios where the
        click may be intercepted. It first waits for the element to be visible using explicit waits and then clicks it.
        If the click is intercepted (e.g., by an overlay), it tries to move to the element and click it using ActionChains.
        If that also fails, it retries to click the element using explicit waits. If any of these attempts succeed, the
        function returns True; otherwise, it returns False.
    """
    locatortype, locatorProperty, elementname = locators[:3]
    Locators = [locatortype, locatorProperty]
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located(Locators))
        driver.find_element(locatortype, locatorProperty).click()
        return True
    except ElementClickInterceptedException:
        try:
            element = driver.find_element(locatortype, locatorProperty)
            action = ActionChains(driver)
            action.move_to_element(element).click().perform()
            return True
        except:
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(Locators)).click()
                return True
            except Exception as e:
                raise e
    except Exception as e:
        with allure.step(f"Failed to click on {elementname} element"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
            time.sleep(2)
            return False
##############################################################################################################################################################################################################################################################
def clickec(driver, locators):
    """
    Click an element on a web page with explicit waits and handling potential click intercept exceptions.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        locators (tuple): A tuple specifying how to locate the element (locatortype, locatorProperty, elementname).
    Returns:
        bool: True if the element was successfully clicked, False otherwise.
    Notes:
        This function attempts to click an element on a web page using explicit waits and handles scenarios where the
        click may be intercepted. It first waits for the element to be clickable using explicit waits. If the click is
        intercepted (e.g., by an overlay), it tries to move to the element and click it using ActionChains. If that also
        fails, it retries to click the element using explicit waits. If any of these attempts succeed, the function returns
        True; otherwise, it returns False.
    """
    locatortype, locatorProperty, elementname = locators[:3]
    Locators = locators[:2]
    try:
        element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable(Locators))
        element.click()
        return True
    except ElementClickInterceptedException:
        try:
            element = driver.find_element(locatortype, locatorProperty)
            action = ActionChains(driver)
            action.move_to_element(element).click().perform()
            return True
        except:
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(Locators)).click()
                return True
            except Exception as e:
                raise e
    except ElementNotInteractableException:
        try:
            element = driver.find_element(locatortype, locatorProperty)
            action = ActionChains(driver)
            action.move_to_element(element).click().perform()
            return True
        except:
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(Locators)).click()
                return True
            except Exception as e:
                raise e
    except Exception as e:
        with allure.step(f"Failed to click on {elementname} element"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
            return False
########################################################################################################################################################################################################################################################################
def clickecwithOutImage(driver, locators):
    """
    Click an element on a web page without an associated image.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        locators (tuple): A tuple specifying how to locate the element (locatortype, locatorProperty, elementname).
    Returns:
        bool: True if the element was successfully clicked, False otherwise.
    Notes:
        This function attempts to click an element on a web page. It first waits for the element to be clickable using
        explicit waits. If that fails, it tries to move to the element and click it using ActionChains. If that also fails,
        it retries to click the element using explicit waits. If any of these attempts succeed, the function returns True;
        otherwise, it returns False.
    """
    locatortype, locatorProperty, elementname = locators[:3]
    Locators = locators[:2]
    try:
        element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(Locators))
        element.click()
        return True
    except ElementClickInterceptedException:
        try:
            element = driver.find_element(locatortype, locatorProperty)
            action = ActionChains(driver)
            action.move_to_element(element).click().perform()
            return True
        except:
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(Locators)).click()
                return True
            except Exception as e:
                raise e
    except Exception as e:
        return False
#################################################################-- INPUTTEXT --##########################################################################################################################################################
# Function:inputtext - Enters the value in Text Edit Field
# Parameters:
#           locators: (By.ID, textbox_username_id)
#           value   : eva@rantcell.com
def inputtext(driver, locators, value):
    """
    Input text into a specified input field on a web page.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        locators (tuple): A tuple specifying how to locate the input field (locatortype, locatorProperty, elementname).
        value (str): The text value to input into the field.
    Returns:
        bool: True if text input was successful, False otherwise.
    Notes:
        This function locates and inputs text into an input field on a web page. It uses explicit waits to ensure
        the input field is visible and ready for interaction. If successful, it returns True; otherwise, it returns False.
    """
    try:
        locatortype, locatorProperty, elementname = locators
        Locators = [locatortype, locatorProperty]
        with allure.step(f"Enter value in {elementname} edit field"):
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(Locators))
            driver.find_element(*Locators).clear()
            driver.find_element(*Locators).send_keys(value)
            return True
    except Exception as e:
        with allure.step(f"Failed to enter the value in {elementname} text field element"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
        return False
#####################################################################-- VERIFYELEMENTISPRESENT --##########################################################################################################################################
def verifyelementispresent(driver, Locators):
    """Function Name: verifyelementispresent
        Purpose:
            The verifyelementispresent function is designed to verify the presence of a specific element on a web page.
             It allows you to check whether a particular element, identified by its locators, is present or not.
        Arguments:
            driver (WebDriver): This argument expects a WebDriver object, which is responsible for interacting with the web application.
                                It is used to locate and verify the presence of the element.
            Locators (tuple): This argument is a tuple that contains information about how to locate the element to be verified.
                                It includes the following elements:
            locatortype: A string representing the locator strategy (e.g., "XPath," "CSS selector") to be used to find the element.
            locatorProperty: A string representing the specific value or property that helps locate the element.
            elementName: A descriptive name or identifier for the element. This provides context for the verification.
        Returns:
            The function returns a boolean value.
            It returns True if the specified element is present on the web page and successfully located, and False if there was an issue or an exception occurred during the process.
        Notes:
            The primary purpose of this function is to verify the presence of a specific element on a web page.
        """
    locatortype, locatorProperty, elementName = Locators
    locators = [locatortype, locatorProperty]
    try:
        with allure.step(f"Verify {elementName} element is present"):
            WebDriverWait(driver, 10).until(EC.presence_of_element_located(locators))
            element = driver.find_element(*locators)
            allure.attach(element.screenshot_as_png, name=elementName, attachment_type=allure.attachment_type.PNG)
            return True
    except Exception as e:
        with allure.step(f"Failed to verify the {elementName} element"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{elementName}_screenshot", attachment_type=allure.attachment_type.PNG)
        return False
###############################################################################################################################################################################################################################################
def uncheck_listOfcampaign(driver, locators):
    """
    Function Name: uncheck_listOfcampaign
        Purpose:
                The uncheck_listOfcampaign function is designed to uncheck a checkbox element on a web page.
                It allows you to programmatically unselect a checkbox, which is a common interaction in web testing scenarios.

        Arguments:
        driver (WebDriver): This argument expects a WebDriver object, which is responsible for interacting with the web application.
                            It is used to locate and manipulate the checkbox element.
        locators (tuple): This argument is a tuple that contains information about how to locate the checkbox element.
                            It includes the following elements:
        locatortype: A string representing the locator strategy (e.g., "XPath," "CSS selector") to be used to find the checkbox.
        locatorProperty: A string representing the specific value or property that helps locate the checkbox.
        elementname: A descriptive name or identifier for the checkbox element. This provides context for the operation.
        Returns:
            The function returns a boolean value. It returns True if the checkbox was successfully unchecked, and False if there was an issue or an exception occurred during the process.
        Notes:
            The primary purpose of this function is to uncheck a checkbox element on a web page.
        The function follows these steps:
            It uses the provided locators tuple to locate the checkbox element on the web page.
            If the checkbox is currently selected (checked), it clicks on it to uncheck it.
            The function returns True to indicate a successful unchecking of the checkbox.
            If there is any issue during the process, such as the checkbox not being found or an exception occurring,the function captures a screenshot of the web page for debugging purposes and returns False to indicate the failure to uncheck the checkbox.
            This function is suitable for web testing scenarios where you need to automate interactions with checkboxes, specifically unchecking them.
            It provides a way to handle checkbox interactions in a robust manner.
            """
    elementname = None
    try:
        locatortype, locatorProperty, elementname = locators
        List_of_Campaigns_checkBox = driver.find_element(locatortype,locatorProperty)  # List of Campaigns CheckBox
        if List_of_Campaigns_checkBox.is_selected():
            List_of_Campaigns_checkBox.click()
        return True
    except Exception as e:
        with allure.step(f"Failed to uncheck the {elementname} checkbox"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
        return False
####################################################################################################################################################################################################################################################
def select_from_listbox_ECs(driver, listbox_locator, nested_locators, option_text_list, Title, path):
    """
        Purpose:
                The select_from_listbox_ECs function is designed for the purpose of interacting with listboxes on web pages and selecting
                multiple options from them. It is equipped with explicit waits to ensure that the listbox and option elements are
                visible and interactable, making it a reliable choice for automating interactions with listboxes containing nested
                option elements.

            Arguments:
                driver (WebDriver): This argument expects a WebDriver object, which is essentially the driver for the
                                    web application being automated. It allows the function to interact with and control the web page.
                listbox_locator (tuple): This argument should be provided as a tuple, representing the
                                        locator strategy (e.g., "XPath," "CSS selector") and the corresponding value
                                        to locate the listbox element on the web page. This is the element that opens
                                        up the list of options.
                nested_locators (list): To locate and interact with the individual options within the listbox,
                                        a list of dictionaries containing locators is expected. Each dictionary in the
                                        list should contain two key-value pairs: "locator by" and "locator." "Locator by"
                                        specifies the strategy to locate the option element (e.g., "XPath," "CSS selector"),
                                        while "locator" provides the specific locator value with a placeholder for the option text.
                                        The function will iterate through this list to select each option.
                option_text_list (list): This is a list of option texts that you want to select from the listbox.
                                        The function will click on each option from this list in the listbox.
                Title (str): This argument expects a string that represents the title or name of the web page or application being automated.
                            It helps in providing context for reporting and debugging purposes.
                path (str): The path to the current test case or script. This is useful for tracking the location of
                            the test case or script being executed.
            Returns:
                The function returns a tuple containing two values:
                flag: This is an indicator of success or failure. It will be set to 0 for success and 1 for failure.
                alert_text: If an alert dialog is encountered during the operation, this value will contain the text message displayed in the alert. If no alert is present, it will be None.
                Notes:
                The primary purpose of this function is to interact with listboxes on web pages and select options from them.
                It does so by clicking on each option individually.
                Explicit waits are used to ensure that the listbox and option elements are visible and can be interacted with.
                This enhances the reliability of the automation.
                The function is equipped to handle exceptions that may occur during the operation, such as ElementClickInterceptedException, TimeoutException, UnexpectedAlertPresentException, and NoAlertPresentException.
                The explicit waits and handling of exceptions make this function suitable for scenarios where web pages contain complex listboxes with nested option elements.
                The function returns a flag to indicate the success or failure of the operation and any alert text encountered during the process. This information can be valuable for reporting and debugging purposes.
        """
    option_text1 = None
    option_element = None
    alert_text = None
    option_text = None
    flag = 1
    alert_text = None
    try:
        try:
            # Wait for the listbox to be visible
            listbox_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(listbox_locator))
            # Click on the listbox to open it
            listbox_element.click()
        except (ElementClickInterceptedException or TimeoutException):
            listbox_element = driver.find_element(*listbox_locator)
            try:
                action = ActionChains(driver)
                action.move_to_element(listbox_element).click().perform()
            except:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(listbox_locator)).click()
        except Exception as e:
            raise e
        for option_text in option_text_list:
            option_text1 = option_text
            # Find the nested option element
            option_locator = None
            for locator_dict in nested_locators:
                locator = (locator_dict['locator by'], locator_dict['locator'].format(option_text))
                option_element = driver.find_element(*locator)
                break
            try:
                option_element.click()
            except ElementClickInterceptedException:
                # If the option is covered by another element, try scrolling to it first
                action = ActionChains(driver)
                action.move_to_element(option_element).click().perform()
            except Exception as e:
                raise e
        time.sleep(1.2)
        try:
            alert_text = alert_accept(driver)
        except UnexpectedAlertPresentException:
            alert_text = alert_accept(driver)
        except NoAlertPresentException:
            pass
        flag = 0
        with allure.step(f'Successfully selected options from listbox'):
            allure.attach(driver.get_screenshot_as_png(), name="listbox_screenshot",attachment_type=allure.attachment_type.PNG)
        return flag, alert_text
    except Exception as e:
        updatecomponentstatus(Title, option_text, "FAILED","Unable to locate the element/No such element found and so error in selecting " + option_text + " from listbox", path)
        raise e
################################################################################################################################################################################################################################################################################################
def alert_accept(driver):
    """
        Accept and close an alert dialog in a web page.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
        Returns:
            str: The text message displayed in the alert dialog before it was accepted and closed.
        Notes:
            This function switches the WebDriver's context to an alert dialog, retrieves the text displayed in the alert,
            accepts (clicks the OK button), and then returns the text message. It is commonly used for handling alert dialogs
            or pop-up windows in web applications where user confirmation or acknowledgment is required.
        """
    alert = driver.switch_to.alert
    alert_text = alert.text
    alert.accept()
    return alert_text
##############################################################################################################################################################################################################################################################################################
def interact_with_blobmap(driver, blobmap_locator, mapelement, elementname,Title,path):
    """
        Function Name: interact_with_blobmap
            Purpose:
                The interact_with_blobmap function is designed for interacting with web maps to locate and click on blob elements. It is particularly useful for web testing scenarios where blob elements play a significant role in the web application's functionality.
            Arguments:
                driver (WebDriver): This argument expects a WebDriver object, which is responsible for interacting with the web application.
                                    It allows the function to control the web page.
                blobmap_locator (tuple): This argument should be provided as a tuple, which includes the locator strategy (e.g., "XPath," "CSS selector")
                                        and the corresponding value. This information is used to locate the blob maps on the web page.
                mapelement (tuple): Similar to blobmap_locator, this argument is a tuple that specifies the locator strategy and value to locate the map element on the web page.
                                    The map element is the container for displaying the map.
                elementname (str): This is a descriptive name or identifier for the element that is being interacted with.
                                    It provides context for the operation.
                Title (str): The title or name of the web page or application being tested.
                            This information is useful for reporting and debugging.
                campaign (str): The name or identifier of the campaign. This can be relevant in certain testing contexts.
                device (str): The name or identifier of the device being tested. Device-specific testing may require this information.
                path (str): The path to the current test case or script. It helps track the location of the executed test case or script.
            Returns:
                The function returns a string, which is a status message indicating the result of the blob element interaction.
                Possible status messages include "blob found" and "blob not found."
            Notes:
                The primary purpose of this function is to interact with web maps and locate and click on blob elements within them.
            The function follows a series of steps:
                It attempts to click on blob elements directly on the web map.
                If the blob elements are not immediately clickable due to overlays or other issues, it handles these situations.
                If the blob elements are still not found, the function zooms in or out on the map to search for them.
                It considers scenarios where blob elements are found alongside data tables on the page.
                The function's return value provides information about whether the blob elements were successfully found or not.
                This status message is valuable for reporting the results of the interaction.
                This function is well-suited for web testing scenarios where blob elements are a key part of the web application's functionality.
                It is designed to handle various scenarios that may arise during interaction with web maps containing blob elements.
        """
    t_flag = None
    if_flag = None
    flag = None
    try:
        flag = 1
        try:
            t = None
            with allure.step("Click on the blob"):
                # Find Blobmap element
                WebDriverWait(driver, 3).until(EC.visibility_of_element_located(blobmap_locator))
                blobmaps = driver.find_elements(*blobmap_locator)
                ik = 0
                for blobmap in blobmaps:
                    if blobmap.is_displayed():
                        ik +=1
                        action = ActionChains(driver)
                        action.move_to_element(blobmap).click().perform()
                        Data_Table = driver.find_elements(*select_Map_View_Components.Data_Table)
                        for table in Data_Table:
                            if table.is_displayed():
                                t = table
                                break
                        if t != None and t.is_displayed() and ik >= 3: 
                            break                        
                flag = 0
                allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_Click on the blob",attachment_type=allure.attachment_type.PNG)
        except ElementClickInterceptedException:
            with allure.step("Click on the blob:-1"):
                # If the option is covered by another element, try scrolling to it first
                click_on_blob(driver, blobmap_locator)
                flag = 0
                allure.attach(driver.get_screenshot_as_png(),name=f"{elementname}_Click on the blob:-1",attachment_type=allure.attachment_type.PNG)
        except TimeoutException:
            # Wait until the map element is displayed
            wait = WebDriverWait(driver, 1)
            map_element = wait.until(EC.visibility_of_element_located((mapelement)))
            # Set the initial zoom level to 0
            zoom_level_OUT = 0
            zoom_level_IN = 0
            blobmaps = driver.find_elements(*blobmap_locator)
            if_flag = "findingblob"
            if len(blobmaps) == 0:
                t_flag = "findingblob_with_table"
                Data_Table = driver.find_elements(*select_Map_View_Components.Data_Table)
                if len(Data_Table) != 0 and len(blobmaps) == 0:
                    with allure.step("Zoom in/out to find blob, findingblob with table"):
                        for table in Data_Table:
                            if table.is_displayed():
                                t_flag = "blob found with table"
                                if_flag = "blob found with table"
                                e = Exception
                                raise StepFailure(e)
                        allure.attach(driver.get_screenshot_as_png(),name=f"{elementname}_Zoom in/out to find blob and is displayed",attachment_type=allure.attachment_type.PNG)
                elif len(blobmaps) == 0 and len(Data_Table) == 0:
                    if t_flag == "findingblob_with_table" or if_flag == "findingblob":
                        with allure.step("Zoom in/out to find blob, findingblob with table:-1"):
                            allure.attach(driver.get_screenshot_as_png(),name=f"{elementname}_Zoom in/out to find blob, findingblob with table:-1",attachment_type=allure.attachment_type.PNG)
                            t_flag = "table not found finding blob"
                            try:
                                time.sleep(.2)
                                cluster_blobmap = driver.find_element(*select_Map_View_Components.cluster_blobmap_locator)
                                if cluster_blobmap.is_displayed():
                                    cluster_blobmap.click()
                                    time.sleep(.2)
                                    if_flag = zoom_in_notpresentnotvisible_for_image(driver,zoom_level_IN,mapelement,blobmaps,[5,5])
                                else:
                                    time.sleep(.2)
                                    if_flag = zoom_out_notpresentnotvisible_for_image(driver,zoom_level_OUT,mapelement,blobmap_locator,blobmaps,[5])
                            except:
                                time.sleep(.2)
                                if_flag = zoom_in_notpresentnotvisible_for_image(driver, zoom_level_IN,mapelement,blobmaps,[5])
                            if if_flag == "blob not found":
                                e = Exception
                                raise StepFailure(e)
                            elif if_flag == "blob found":
                                t_flag = "findingblob_with_table"
                                Data_Table = driver.find_elements(*select_Map_View_Components.Data_Table)
                                if len(Data_Table) != 0 and len(blobmaps) == 0:
                                    for table in Data_Table:
                                        if table.is_displayed():
                                            t_flag = "blob found with table"
                                            if_flag = "blob found with table"
                                            e = Exception
                                            raise StepFailure(e)
            i = 0
            for blob in blobmaps:
                i +=1
                if not blob.is_displayed() and i == 1:
                    e = Exception
                    raise StepFailure(e)
                elif blob.is_displayed() or i ==1:
                    break
            click_on_blob(driver,blobmap_locator)
            flag = 0
            return if_flag
    except Exception as e:
        if t_flag == "blob found with table":
            statement = f"blob image is displayed and clicked on the blob image during zoom in/out so table is displayed for " + elementname + f" in {Title}"
            with allure.step(statement):
                allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                updatecomponentstatus(Title, elementname, "PASSED", "Blob Found", path)
                raise StepFailure(e)
        else:
            raise StepFailure(e)
###############################################################################################################################################################################################################################################################################################
def click_on_blob(driver,blobmap_locator):
    """
        Click on a visible blob map element on a web page.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            blobmap_locator (tuple): A tuple representing the locator strategy and value to locate blob maps.
        Returns:
            None
        Notes:
            This function finds and clicks on a visible blob map element. It uses the provided WebDriver and locator
            to locate and click the first visible blob map on the web page.
        """
    blobmaps = driver.find_elements(*blobmap_locator)
    for blob in blobmaps:
        if blob.is_displayed():
            action = ActionChains(driver)
            action.move_to_element(blob).click().perform()
            break
########################################################################################################################################################################################################################################################################################################
def zoom_out_notpresentnotvisible_for_image(driver, zoom_level_OUT, mapelement, blobmap_locator, blobmaps,loopvalue):
    if_flag = "findingblob"
    map_elements = driver.find_elements(*mapelement)
    for map_element in map_elements:
        if map_element.is_displayed():
            map_element.click()
    while len(blobmaps) == 0 and zoom_level_OUT < 22:
        driver.execute_script(f"window.scrollTo({0}, {0});")
        map = driver.find_element(*select_Map_View_Components.map_element)
        action_chains = ActionChains(driver)
        action_chains.move_to_element(map).perform()
        map_width_2 = int(map.size['width'] / 33)
        if_flag = "blob not found"
        for loopval in loopvalue:
            for i in range(-(map_width_2), map_width_2, loopval):
                try:
                    action_chains.move_to_element_with_offset(map, i, -loopval).click().perform()
                except:
                    pass
                tables = driver.find_elements(*select_Map_View_Components.Data_Table)
                if len(tables) != 0:
                    if_flag = "blob found"
                    with allure.step(f"{if_flag} i:- {i}"):
                        break
            if if_flag == "blob found":
                break
            try:
                table = driver.find_element(*select_Map_View_Components.Data_Table)
                if table.is_displayed():
                    break
            except:
                pass
            time.sleep(0.01)
            # Find the blobs again
            blobmaps = driver.find_elements(*blobmap_locator)
            # Increment the zoom level
            zoom_level_OUT += 22
        try:
            table = driver.find_element(*select_Map_View_Components.Data_Table)
            if table.is_displayed():
                break
        except:
            pass
    return if_flag
############################################################################################################################################################################################################################################################################################
def zoom_in_notpresentnotvisible_for_image(driver, zoom_level_IN, mapelement, blobmaps,loopvalue):
    if_flag = "findingblob"
    map_elements = driver.find_elements(*mapelement)
    for map_element in map_elements:
        if map_element.is_displayed():
            map_element.click()
    while len(blobmaps) == 0 and zoom_level_IN < 22:
        driver.execute_script(f"window.scrollTo({0}, {0});")
        map = driver.find_element(*select_Map_View_Components.map_element)
        action_chains = ActionChains(driver)
        action_chains.move_to_element(map).perform()
        map_width_2 = int(map.size['width'] / 33)
        if_flag = "blob not found"
        for loopval in loopvalue:
            for i in range(-(map_width_2), map_width_2, loopval):
                try:
                    action_chains.move_to_element_with_offset(map, i, -loopval).click().perform()
                except:
                    pass
                tables = driver.find_elements(*select_Map_View_Components.Data_Table)
                if len(tables) != 0:
                    if_flag = "blob found"
                    with allure.step(f"{if_flag} i:- {i}"):
                        break
            if if_flag == "blob found":
                break
            try:
                table = driver.find_element(*select_Map_View_Components.Data_Table)
                if table.is_displayed():
                    break
            except:
                pass
            zoom_level_IN += 22
        try:
            table = driver.find_element(*select_Map_View_Components.Data_Table)
            if table.is_displayed():
                break
        except:
            pass
    return if_flag
###########################################################################################################################################################################################################################################################################
def extract_table_datas(driver, table_locator,elementname,Title,path):
    """
        Extract data from an HTML table on a web page using a specified table locator.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the   web application.
            table_locator (tuple): A tuple representing the locator strategy and value to locate the table.
            elementname (str): The name or description of the specific table element being extracted.
            Title (str): The title or description of the operation being performed.
            path (str): The path to the Excel workbook where data will be appended.
        Returns:
            list: A list containing the extracted table data.
        Raises:
            StepFailure: If no data is present in the table or if an exception occurs during extraction.
        """
    headers = None
    table = None
    try:
        table = WebDriverWait(driver,2).until(EC.presence_of_element_located(table_locator))
        data = html_table_datas(driver, table)
        return data
    except Exception as e:
        updatecomponentstatus(Title,elementname, "FAILED", "No data in table",path)
        raise StepFailure(e)
################################################################################################################################################################################################################################################
def extract_table_datas2(driver, table_locator,elementname,Title,path):
    """
        Extract data from an HTML table on a web page using a list of table locators.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            table_locator (list): A list of tuples representing the locator strategy and value to locate tables.
            elementname (str): The name or description of the specific table element being extracted.
            Title (str): The title or description of the operation being performed.
            path (str): The path to the Excel workbook where data will be appended.
        Returns:
            list: A list containing the extracted table data.
        Raises:
            StepFailure: If no data is present in any of the tables or if an exception occurs during extraction.
        """
    headers = None
    table = None
    data = None
    try:
        for table in table_locator:
            data = html_table_datas(driver, table)
        return data
    except Exception as e:
        updatecomponentstatus(Title,elementname, "FAILED", "No data in table",path)
        raise StepFailure(e)
#############################################################################################################################################################################################################################################################
def extract_table_datas_span(driver, table_locator,tablename,elementname,Title,path):
    """
        Extract data from an HTML table on a web page using <span> elements.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            table_locator (tuple): A tuple representing the locator strategy and value to locate the table.
            tablename (str): The name or description of the table element.
            elementname (str): The name or description of the specific table element being extracted.
            Title (str): The title or description of the operation being performed.
            path (str): The path to the Excel workbook where data will be appended.
        Returns:
            list: A list containing the extracted table data.
        Raises:
            StepFailure: If no data is present in the table or if an exception occurs during extraction.
        """
    headers = None
    table = None
    data = None
    try:
        data = []
        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located(table_locator))
        action = ActionChains(driver)
        action.move_to_element(table).perform()
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located(table_locator))
        time.sleep(2)
        if table.is_displayed():
            data = html_table_datas_using_span(driver, table,tablename)
        return data
    except Exception as e:
        updatecomponentstatus(Title,elementname, "FAILED", f"In {tablename} for {elementname} No data in table/No table",path)
        raise StepFailure(e)
#######################################################################################################################################################################################################################################################
def extract_table_datas_span1(driver, table_locator,tablename,elementname,Title,path):
    """
        Extract data from an HTML table on a web page using <span> elements.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            table_locator (tuple): A tuple representing the locator strategy and value to locate the table.
            tablename (str): The name or description of the table element.
            elementname (str): The name or description of the specific table element being extracted.
            Title (str): The title or description of the operation being performed.
            path (str): The path to the Excel workbook where data will be appended.
        Returns:
            list: A list containing the extracted table data.
        Raises:
            StepFailure: If no data is present in the table or if an exception occurs during extraction.
        """
    headers = None
    table = None
    data = None
    try:
        table = WebDriverWait(driver,0.1).until(EC.presence_of_element_located(table_locator))
        action = ActionChains(driver)
        action.move_to_element(table).perform()
        if table.is_displayed():
            data = html_table_datas_using_span(driver, table,tablename)
        return data
    except Exception as e:
        raise StepFailure(e)
#######################################################################################################################################################################################################################################################################
def html_table_datas_using_span(driver,table,tablename):
    """
        Extract data from an HTML table on a web page using <span> elements.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            table (WebElement): The WebElement representing the HTML table.
            tablename (str): The name or description of the table element.
        Returns:
            list: A list containing the extracted table data.
        Raises:
            StepFailure: If no data is present in the table.
        """
    headers = None
    rows = table.find_elements(By.TAG_NAME, "tr")
    data1 = []
    if rows[0].find_elements(By.TAG_NAME, "th"):
        headers = [th.text for th in rows[0].find_elements(By.TAG_NAME, "th")]
        if len(headers) !=0:
            data1.append(headers)
        data = []
        for row in rows[1:]:
            row_data = []
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds)!=0:
                for i in range(len(headers)):
                    try:
                        cell_text = tds[i].find_element(By.TAG_NAME, "span").text
                    except NoSuchElementException:
                        cell_text = tds[i].text
                    row_data.append(cell_text)
                data.append(row_data)
                data1.append(row_data)
    else:
        data = []
        for row in rows:
            row_data = []
            tds = row.find_elements(By.TAG_NAME, "td")
            for td in tds:
                try:
                    span_text = td.find_element(By.TAG_NAME, "span").text
                except NoSuchElementException:
                    span_text =  td.text
                row_data.append(span_text)
            data.append(row_data)
    with allure.step(f"Extracted table data from {tablename}"):
        table_html = "<table>"
        if rows[0].find_elements(By.TAG_NAME, "th"):
            table_html += "<tr>"
            for header in headers:
                table_html += "<th>" + header + "</th>"
            table_html += "</tr>"
        for row in data:
            table_html += "<tr>"
            for cell in row:
                table_html += "<td>" + cell + "</td>"
            table_html += "</tr>"
        table_html += "</table>"
        allure.attach(table_html, "Table data", AttachmentType.HTML)
        allure.attach(driver.get_screenshot_as_png(), name="Table data", attachment_type=allure.attachment_type.PNG)
        if data1.__len__() == 0:
            e = Exception
            raise StepFailure(e)
    return data1
##############################################################################################################################################################################################################################################################
def extract_table_datas_headers(driver, table_locator,elementname,Title,path):
    """
        Extract header and subheader data from an HTML table on a web page.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            table_locator (tuple): The locator (By, value) for finding the HTML table.
            elementname (str): The name or description of the table element.
            Title (str): The title or description of the operation being performed.
            path (str): The path to the Excel workbook where data will be appended.
        Returns:
            tuple: A tuple containing headers and subheaders extracted from the table.
                   headers (list): List of header names.
                   subheaders (list): List of subheader names.
        Raises:
            StepFailure: If the table is not found or no data is present in the table.
        """
    headers = None
    table = None
    try:
        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located(table_locator))
        headers,subheaders= html_table_datas_for_headers(driver, table)
        return headers,subheaders
    except Exception as e:
        updatecomponentstatus(Title,elementname, "FAILED", "No data in table",path)
        raise StepFailure(e)
##########################################################################################################################################################################################################################################################################################
def extract_table_datas_content(driver, table_locator,elementname,Title,path):
    """
        Extract and return content data from an HTML table on a web page.
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            table_locator: Locator for finding the HTML table element.
            elementname (str): The name or description of the element being processed.
            Title (str): The title or description of the operation being performed.
            path (str): The path to the Excel workbook where data will be appended.
        Returns:
            list: A list containing the extracted content data from the HTML table.
            """
    headers = None
    table= None
    try:
        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located(table_locator))
        datacontent = html_table_datas_for_content(driver, table)
        return datacontent
    except Exception as e:
        updatecomponentstatus(Title,elementname, "FAILED", "No data in table",path)
        raise StepFailure(e)
################################################################################################################################################################################################################################################################################
def html_table_datas_for_headers(driver,table):
    """    Extract and process data from an HTML table, capturing headers and content.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        table (WebElement): The HTML table element from which to extract data.
    Returns:
        List[Dict[str, str]]: A list of dictionaries representing the extracted data. Each dictionary corresponds to a row in the table, with header values as keys and cell values as values.
        """
    headers = None
    subheaders = None
    rows = table.find_elements(By.TAG_NAME, "tr")
    if rows[0].find_elements(By.TAG_NAME, "th"):
        headers = [th.text for th in rows[0].find_elements(By.TAG_NAME, "th") if "ng-hide" not in th.get_attribute("class")]
    if rows[1].find_elements(By.TAG_NAME, "th"):
        subheaders =[th.text for th in rows[1].find_elements(By.TAG_NAME, "th") if "ng-hide" not in th.get_attribute("class")]
    return headers,subheaders
#################################################################################################################################################################################################################################################################
def html_table_datas_for_content(driver,table):
    """
        Extract and process HTML table data containing content, converting it into a structured format.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        table (WebElement): The HTML table element containing content to be processed.
    Returns:
        List[List[str]]: A list of lists representing the extracted content data from the HTML table.
        """
    headers = None
    subheaders = None
    rows = table.find_elements(By.TAG_NAME, "tr")
    datacontent =[]
    for row in rows:
        tds = row.find_elements(By.TAG_NAME, "td")
        row_data =[]
        if len(tds) != 0:
            for i in range(len(tds)):
                try:
                    cell_text = tds[i].find_element(By.TAG_NAME, "abbr").text
                except NoSuchElementException:
                    try:
                        cell_text = tds[i].find_element(By.TAG_NAME, "span").text
                    except NoSuchElementException:
                        cell_text = tds[i].text
                # Exclude cells with class "ng-binding ng-hide"
                if "ng-binding ng-hide" not in tds[i].get_attribute("class"):
                    row_data.append(cell_text)
            datacontent.append(row_data)
    return datacontent
#####################################################################################################################################################################################################################################################################
def html_table_datas(driver,table):
    """
        Extract and process HTML table data, convert it into a structured format, and attach it to an Allure report.
    Args:
        driver (WebDriver): The WebDriver object for interacting with the web application.
        table (WebElement): The HTML table element to be processed.
    Returns:
        List[Dict[str, str]]: A list of dictionaries representing the extracted table data.
    Raises:
        StepFailure: If no data is found in the HTML table.
    """
    headers = None
    rows = table.find_elements(By.TAG_NAME, "tr")
    if rows[0].find_elements(By.TAG_NAME, "th"):
        headers = [th.text for th in rows[0].find_elements(By.TAG_NAME, "th")]
        data = []
        for row in rows[1:]:
            row_data = {}
            tds = row.find_elements(By.TAG_NAME, "td")
            for i in range(len(headers)):
                row_data[headers[i]] = tds[i].text
            data.append(row_data)
    else:
        data = []
        for row in rows:
            row_data = [td.text for td in row.find_elements(By.TAG_NAME, "td")]
            data.append(row_data)
    with allure.step("Extracted table data"):
        table_html = "<table>"
        if rows[0].find_elements(By.TAG_NAME, "th"):
            table_html += "<tr>"
            for header in headers:
                table_html += "<th>" + header + "</th>"
            table_html += "</tr>"
        for row in data:
            table_html += "<tr>"
            for cell in row:
                table_html += "<td>" + cell + "</td>"
            table_html += "</tr>"
        table_html += "</table>"
        allure.attach(table_html, "Hi, Table data", AttachmentType.HTML)
        allure.attach(driver.get_screenshot_as_png(), name="Table data", attachment_type=allure.attachment_type.PNG)
        if data.__len__() == 0:
            e = Exception
            raise StepFailure(e)
    return data
###############################################################################################################################################################################################################################################
def readCSVSheet(driver, Title, txt, path, result_data,result_status,downloadfilespath):
    """
        Read and process CSV files, append data to an Excel worksheet, and handle exceptions.
        This function seems to be designed to work with export views
        Args:
            driver (WebDriver): The WebDriver object for interacting with the web application.
            Title (str): The title or description of the operation being performed.
            txt (str): The text or type of CSV file to be processed.
            path (str): The path to the Excel workbook where data will be appended.
            downloadfilespath (str): The path to the folder containing CSV files.
        Returns:
            None
        Raises:
            StepFailure: If no CSV files are found, or if errors occur during processing.
        """
    file_path=None
    headerdata =None
    datas = None
    file_name = None
    data1 = None
    CSVFILE = None
    CSVFILE = []
    try:
        list_of_files = glob.glob(downloadfilespath + "\\*.csv")
        if list_of_files.__len__() == 0:
            position = 0
            CSVFILE.insert(position,"No CSV FILE")
            e = Exception
            raise e
        elif list_of_files.__len__() != 0:
            position = 0
            CSVFILE.insert(position,"CSV file is present")
        if CSVFILE[0] == "CSV file is present":
            try:
                file_names_path = []
                file_path = ""
                # Iterate over all items in the folder
                for item in os.listdir(downloadfilespath):
                    # Check if the item is a file and has a .csv extension
                    if item.lower().endswith(".csv"):
                        file_path = os.path.join(downloadfilespath, item)
                        if "Combined Binary Export" == str(txt) and re.search("Binary_Combined", item,re.IGNORECASE):
                            file_names_path.append(file_path)
                        elif "Hand OverExport" == str(txt) and re.search("Hand Over", item, re.IGNORECASE):
                            file_names_path.append(file_path)
                        elif "Export TableSummary" == str(txt) and re.search("TableSummary", item, re.IGNORECASE):
                            file_names_path.append(file_path)
                        elif "Survey Test Export" == str(txt) and re.search("Survey", item, re.IGNORECASE):
                            file_names_path.append(file_path)
                        else:
                            file_names_path.append(file_path)
                for file_path in file_names_path:
                    # Iterate over the 'datas' list and write each row to the destination worksheet
                    df = pd.read_csv(file_path)
                    df_data = df.to_dict(orient='list')
                    file_name = os.path.basename(file_path)
                    try:
                        if len(df_data) != 0:
                            CSVFILE = [f"CSV file is not empty {str(file_name)}"]
                            updatecomponentstatus2 = status(Title, txt, "PASSED", f"Passed step :- {CSVFILE[0]}")
                            result_status.put(updatecomponentstatus2)
                            # Convert the DataFrame to an HTML table
                            html_table = df.to_html()
                            df_data[f"{str(txt)}"] = [str(file_name)] * len(df)
                            result_data.put(df_data)
                            # Attach the HTML content to the Allure report
                            allure.attach(html_table, f"Table data{str(file_name)}", AttachmentType.HTML)
                        elif len(df_data) ==0:
                            dfdata = {}
                            dfdata[f"{str(txt)}"] = [" CSV file is empty {str(file_name)}"]
                            df1 = pd.DataFrame(dfdata)
                            result_data.put(dfdata)
                            # Convert the DataFrame to an HTML table
                            html_table = df1.to_html()
                            # Attach the HTML content to the Allure report
                            allure.attach(html_table, f"Table data{str(file_name)}", AttachmentType.HTML)
                            CSVFILE = [f"CSV file is empty {str(file_name)}"]
                            e = Exception
                            raise e
                    except Exception as e:
                        continue
            except Exception as e:
                raise StepFailure(e)
    except Exception as e:
        with allure.step(f"failed step :- {CSVFILE[0]}"):
            updatecomponentstatus2 = status(Title, txt, "FAILED", f"failed step :- {CSVFILE[0]}")
            result_status.put(updatecomponentstatus2)
            dfdata1 ={}
            dfdata1[f"{str(txt)}"] = [f"failed step :- {str(file_name)} {CSVFILE[0]}"]
            result_data.put(dfdata1)
            raise StepFailure(e)
    finally:
        list_of_files = glob.glob(downloadfilespath + "\\*.csv")
        if list_of_files.__len__() != 0:
            csv_files = os.listdir(downloadfilespath + "\\")  # Get the list of CSV files in the folder
            for csv_file in list_of_files:
                csv_file_path = os.path.join(downloadfilespath + "\\", csv_file)
                xlsx_file_path = os.path.join(downloadfilespath + "\\", csv_file.rsplit('.', 1)[0] + '.xlsx')
                data = pd.read_csv(csv_file_path)  # Read the CSV file using pandas
                data.to_excel(xlsx_file_path, index=False)
                os.remove(csv_file_path)
def status(Title,component,status,comments):
    df_Values = {'Title':[Title], 'Componentname': [component],'Status':[status], 'Comments':[comments]}
    return df_Values

#################################################################################################################################################################################################################################################################################
def convert_to_csv_to_xlsx(downloadfilespath):
    """
        Convert CSV files to XLSX format and remove the original CSV files.
        Args:
            downloadfilespath (str): The path to the folder containing CSV files.
        Returns:
            None
        """
    list_of_files = glob.glob(downloadfilespath + "\\*.csv")
    if list_of_files.__len__() != 0:
        csv_files = os.listdir(downloadfilespath + "\\")  # Get the list of CSV files in the folder
        for csv_file in list_of_files:
            csv_file_path = os.path.join(downloadfilespath + "\\", csv_file)
            xlsx_file_path = os.path.join(downloadfilespath + "\\", csv_file.rsplit('.', 1)[0] + '.xlsx')
            data = pd.read_csv(csv_file_path)  # Read the CSV file using pandas
            data.to_excel(xlsx_file_path, index=False)
            os.remove(csv_file_path)
##########################################################################################################################################################################################################################################################################
def html_for_csv(data,file_name):
    """
        Create an HTML table representation of CSV data and attach it to an Allure report.
        Args:
            data (list of lists): The CSV data as a list of lists.
            file_name (str): The name of the CSV file (used for the attachment name).
        Returns:
            None
        """
    html = ['<table class="my-table">', '<style>',
            '.my-table { font-family: Arial, sans-serif; border-collapse: collapse; width: 100%; }',
            '.my-table th, .my-table td { border: 1px solid #ddd; padding: 8px; }',
            '.my-table th { background-color: #f2f2f2; }',
            '.my-table td { transition: background-color 0.3s; }',
            '.my-table td:hover { background-color: #f8f8f8; }',
            '.my-table.fade-in { animation: fade-in 1s ease-in; }',
            '@keyframes fade-in { from { opacity: 0; } to { opacity: 1; } }',
            '</style>']
    for i, sublist in enumerate(data):
        if i < 2:
            html.append('<tr>')
            html.extend([f'<th>{cell}</th>' for cell in sublist])
            html.append('</tr>')
        else:
            html.append('<tr class="fade-in">')
            html.extend([f'<td>{cell}</td>' for cell in sublist])
            html.append('</tr>')
    html.append('</table>')
    allure.attach(''.join(html), f"Table data{file_name}", AttachmentType.HTML)
#################################################################################################################################################################################################################################################
def clickEC_for_listbox(driver, locators, Title, path):
    """
        Click on an element within a listbox using Expected Conditions (EC).
        Args:
            driver: The WebDriver instance for the web page or application.
            locators (tuple): The locators (By, value) used to locate the element within the listbox.
            Title (str): The title of the test case.
            path (str): The path for saving screenshots.
        Returns:
            tuple: A tuple containing a flag (0 for success, 1 for failure) and an alert text (if any).
        """
    flag = 1
    alert_text = None
    option_text = locators[2]
    Locators = [locators[0], locators[1]]
    try:
        try:
            # Wait for the element to be clickable and click it
            element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(Locators))
            element.click()
            time.sleep(0.5)
        except ElementClickInterceptedException:
            # Handle ElementClickInterceptedException by moving to the element and clicking
            element1=driver.find_element(*Locators)
            action = ActionChains(driver)
            action.move_to_element(element1).click().perform()
        try:
            # Check for and accept any alert present after the click
            alert_text = alert_accept(driver)
        except UnexpectedAlertPresentException as e:
            alert_text = alert_accept(driver)
        except NoAlertPresentException as e:
            pass
        flag =0 # Indicate success
        return flag,alert_text
    except Exception as e:
        with allure.step("Failed to click on " + option_text + "element"):
            # Attach a screenshot to the Allure report for debugging
            allure.attach(driver.get_screenshot_as_png(), name=f"{option_text}_screenshot",attachment_type=allure.attachment_type.PNG)
            # Update the test case status and log the error message
            updatecomponentstatus(Title, option_text, "FAILED","Unable to locate the element/No such element found and so error in selecting " + option_text + " from listbox", path)
            raise e
##############################################################################################################################################################################################################################################################################
def Page_up(driver):
    """
        Scroll the web page up by simulating the PAGE UP key press and other methods.
        Args:
            driver: The WebDriver instance for the web page or application.
        Returns:
            None
        """
    try:
        actions = ActionChains(driver)
        for i in range(1, 5):
            time.sleep(1)
            actions.send_keys(Keys.PAGE_UP).perform()
            time.sleep(1)
            actions.send_keys(Keys.CONTROL + Keys.HOME).perform()
            actions.key_down(Keys.CONTROL).send_keys(Keys.HOME).key_up(Keys.CONTROL).perform()
            driver.execute_script("window.scrollBy(0, -window.innerHeight);")
            driver.execute_script("window.scrollTo(0, document.documentElement.scrollTop - 1000);")
            driver.execute_script("window.scrollTo(0, 0);")
    except Exception as e:
        print("Error occurred while performing page up:", e)
################################################################################################################################################
def Page_Down(driver):
    """
        Scroll the web page down by simulating the PAGE DOWN key press.
        Args:
            driver: The WebDriver instance for the web page or application.
        Returns:
            None
        """
    try:
        actions = ActionChains(driver)
        for i in range(1, 5):
            actions.send_keys(Keys.PAGE_DOWN).perform()
    except Exception as e:
        print("Error occurred while performing Page_Down:", e)
###############################################################################################################################################
def find_blob(driver,screenshot_path, image_paths):
    """
        Find a specific image (blob) in a given screenshot.
        Args:
            driver: The WebDriver instance for the web page or application.
            screenshot_path (str): The path to the screenshot image.
            image_paths (str): The path to the folder containing image files to search for.
        Returns:
            str: A status flag indicating whether the blob was found ("blob found") or not found ("blob not found").
            Tuple[int, int] or None: The position (x, y) of the found blob in the screenshot, or None if not found.
        """
    f_flag = "findingblob"
    loc = None
    # Get a list of image file names in the folder
    image_files = [filename for filename in os.listdir(image_paths) if filename.endswith('.jpg') or filename.endswith('.png')]
    print(str(image_files))
    for image_file in image_files:
        try:
            # Construct the full path to the image
            image_path = os.path.join(image_paths, image_file)
            # Load the screenshot and image
            screenshot = cv2.imread(screenshot_path, cv2.IMREAD_UNCHANGED)
            image = cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
            # Check if the images were loaded successfully
            if screenshot is None:
                print("Failed to load screenshot:", screenshot_path)
                return "screenshot not found", None
            if image is None:
                print("Failed to load image:", image_path)
                continue
            # Perform template matching
            result = cv2.matchTemplate(screenshot, image, cv2.TM_CCOEFF_NORMED)
            threshold = 0.8  # Adjust the threshold value as per your requirement
            loc = np.where(result >= threshold)
            i_flag = "findingblob"
            f_flag = i_flag
            # Check if any matches were found
            if len(loc[0]) > 0:
                print(f"Image {image_file} found in the screenshot.")
                i_flag = "blob found"
                f_flag = i_flag
                # Calculate the position or offset of the found blob
                blob_position = (loc[1][0], loc[0][0])  # Assuming only one match is found
                print("Blob position:", blob_position)
                x_coords = loc[0]
                y_coords = loc[1]
                break
            else:
                print(f"Image {image_file} not found in the screenshot.")
                i_flag = "blob not found"
                f_flag = i_flag
        except Exception as e:
            print(str(e))
    return f_flag
##########################################################highlevelExcelReport########################################################################################################################################################################################################
#def highlevelExcelReport(excel_report_path, high_level_excel_report_path): This function is used to update the high level excel report
# where all testcase report will be in one excel file
#Note:- 1) This high level excel report file contains "COMPONENTSTATUS","DATA_MATCH",'CBE_vs_CE_MATCH','CBE_vs_CE_DONOT_MATCH','DATA_NOT_MATCH',"TABLESUMMARY_DATA_MATCH","TABLESUMMARY_DATA_NOT_MATCH".
#       2) Don't change the Sheet format/Header format in individuals excel report (Sheet are "COMPONENTSTATUS","DATA_MATCH",'CBE_vs_CE_MATCH','CBE_vs_CE_DONOT_MATCH','DATA_NOT_MATCH',"TABLESUMMARY_DATA_MATCH","TABLESUMMARY_DATA_NOT_MATCH").
def highlevelExcelReport(excel_report_path, high_level_excel_report_path):
    """
        Update the high-level Excel report where all testcase reports are merged into one Excel file.
        Args:
            excel_report_path (str): The path to the folder containing individual testcase Excel reports.
            high_level_excel_report_path (str): The path where the high-level Excel report will be saved.
        Returns:
            None
        """
    try:
        # Get all XLSX files in the folder
        files = glob.glob(os.path.join(excel_report_path, "*.xlsx"))
        file_prefix = "Merged_Data_File"
        # Exclude files starting with "Merged_Data_File"
        filespaths = [file for file in files if not os.path.basename(file).startswith(file_prefix)]
        # Create the High Level COMPONENTSTATUS sheet in the new workbook
        workbook = Workbook()
        workbook.create_sheet("HIGH LEVEL COMPONENTSTATUS", 0)
        # workbook.create_sheet("DATA_MATCH", 1)
        # workbook.create_sheet("DATA_NOT_MATCH", 2)
        # workbook.create_sheet("TABLESUMMARY_DATA_NOT_MATCH",3)
        # workbook.create_sheet("TABLESUMMARY_DATA_MATCH", 4)
        # workbook.create_sheet("CBE_vs_CE_MATCH", 5)
        # workbook.create_sheet("CBE_vs_CE_DONOT_MATCH", 6)
        # workbook.create_sheet("IPU_vs_CE_DATA_MATCH", 7)
        # workbook.create_sheet("IPU_vs_CE_DATA_NOT_MATCH", 8)
        # workbook.create_sheet("NQC_vs_OC_DATA_MATCH", 9)
        # workbook.create_sheet("NQC_vs_OC_DATA_NOT_MATCH", 10)
        workbook.create_sheet("RESULT OF EACH MODULE", 1)
        timestamp = datetime.datetime.now().strftime("%d_%m_%Y_%H_%M")
        high_level_excel_report_paths = os.path.join(high_level_excel_report_path, f"high_level_excel_report_{timestamp}.xlsx")
        workbook.save(high_level_excel_report_paths)
        workbook.close()
        add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Different-Views",'Componentname', 'Status', 'Comments'],sheet_name="HIGH LEVEL COMPONENTSTATUS")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'Component Type', 'Data validation'],sheet_name="DATA_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'Component Type', 'Data validation'],sheet_name="DATA_NOT_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'Data validation'], sheet_name="TABLESUMMARY_DATA_NOT_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'Data validation'], sheet_name="TABLESUMMARY_DATA_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'File', 'ParameterType', 'Data validation'],sheet_name="CBE_vs_CE_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'File', 'ParameterType', 'Data validation'],sheet_name="CBE_vs_CE_DONOT_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'File',"Individual pop up headers","Individual pop up value","combine export value",'Data validation'],sheet_name="IPU_vs_CE_DATA_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths, headers=["Test-Cases","Module",'File',"Individual pop up headers","Individual pop up value","combine export value", 'Data validation'],sheet_name="IPU_vs_CE_DATA_NOT_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths,headers=["Test-Cases","Module",'File',"map view Operator","map view Operator value","calculated csv value",'Data validation'], sheet_name="NQC_vs_OC_DATA_MATCH")
        # add_headers_and_data(file_path=high_level_excel_report_paths,headers=["Test-Cases","Module",'File',"map view Operator","map view Operator value","calculated csv value",'Data validation'], sheet_name="NQC_vs_OC_DATA_NOT_MATCH")
        app = xw.App(visible=False)
        try:
            workbook = app.books.open(high_level_excel_report_paths)
        except Exception as e:
            pass
        sheet_high_level_componentstatus = workbook.sheets['HIGH LEVEL COMPONENTSTATUS']
        # sheet_cbe_vs_ce_match = workbook.sheets['CBE_vs_CE_MATCH']
        # sheet_cbe_vs_ce_donot_match = workbook.sheets['CBE_vs_CE_DONOT_MATCH']
        # sheet_data_match = workbook.sheets['DATA_MATCH']
        # sheet_data_not_match = workbook.sheets['DATA_NOT_MATCH']
        # sheet_tablesummary_data_not_match = workbook.sheets['TABLESUMMARY_DATA_NOT_MATCH']
        # sheet_tablesummary_data_match = workbook.sheets['TABLESUMMARY_DATA_MATCH']
        # sheet_ipu_vs_ce_data_not_match = workbook.sheets['IPU_vs_CE_DATA_NOT_MATCH']
        # sheet_ipu_vs_ce_data_match = workbook.sheets['IPU_vs_CE_DATA_MATCH']
        # sheet_nqc_table_data_vs_oc_data_match = workbook.sheets['NQC_vs_OC_DATA_MATCH']
        # sheet_nqc_table_data_vs_oc_data_not_match= workbook.sheets['NQC_vs_OC_DATA_NOT_MATCH']
        sheet_result = workbook.sheets["RESULT OF EACH MODULE"]
        data_high_level_componentstatus = []
        data_data_match = []
        data_data_not_match = []
        data_tablesummary_data_not_match = []
        data_tablesummary_data_match = []
        data_cbe_vs_ce_match = []
        data_cbe_vs_ce_donot_match =[]
        data_ipu_vs_ce_match = []
        data_ipu_vs_ce_donot_match = []
        data_nqc_table_data_vs_oc_data_not_match = []
        data_nqc_table_data_vs_oc_data_match = []
        data_comparison_default_settings = []
        data_comparison_change_settings = []
        for file in filespaths:
            try:
                sheets_high = ["COMPONENTSTATUS","DATA_MATCH",'CBE_vs_CE_MATCH','CBE_vs_CE_DONOT_MATCH','DATA_NOT_MATCH',"TABLESUMMARY_DATA_MATCH","TABLESUMMARY_DATA_NOT_MATCH","IPU_vs_CE_DATA_MATCH","IPU_vs_CE_DATA_NOT_MATCH","NQC_vs_OC_DATA_MATCH","NQC_vs_OC_DATA_NOT_MATCH","RESULTS_DEFAULT_SETTINGS","RESULTS_CHANGE_SETTINGS"]
                basename = os.path.basename(file)
                basename = basename.rstrip(".xlsx")
                for sheet_h in sheets_high:
                    try:
                        df = pd.read_excel(file, sheet_name=sheet_h)
                        if sheet_h == "COMPONENTSTATUS":
                            df_componentstatus = pd.read_excel(file, sheet_name='COMPONENTSTATUS', header=1)
                            try:
                               data1 = df_componentstatus.to_dict()
                               lastvalue = "None"
                               try:
                                    lastvalue = list(data1["Title"].values())[-1]
                               except Exception as e:
                                   pass
                               if re.search('finished/ends',lastvalue,re.IGNORECASE):
                                   column_order_componentstatus = ["Test-Cases", "Different-Views", 'Componentname', 'Status', 'Comments']

                                   df_componentstatus = df_componentstatus.iloc[:-1]

                                   # Rename the "Title" column to "Different-Views"
                                   df_componentstatus.rename(columns={"Title": "Different-Views"}, inplace=True)

                                   # Add a new "Test-Cases" column and set it to a desired value
                                   df_componentstatus["Test-Cases"] = [basename] * len(df_componentstatus)  # Replace "your_desired_value" with the value you want

                                   # Reorder the columns based on the specified order
                                   df_componentstatus = df_componentstatus[column_order_componentstatus]

                                   data_high_level_componentstatus.append(df_componentstatus)
                               elif not re.search('finished/ends',lastvalue,re.IGNORECASE):
                                   print('finished/ends',df_componentstatus)
                                   df_componentstatus.rename(columns={"Title": "Different-Views"}, inplace=True)
                                   df_componentstatus["Test-Cases"] = [basename] * len(df_componentstatus)
                                   column_order_componentstatus = ["Test-Cases", "Different-Views", 'Componentname', 'Status', 'Comments']
                                   # Reorder the columns based on the specified order
                                   df_componentstatus = df_componentstatus[column_order_componentstatus]
                                   # Create a new row as a dictionary
                                   new_row = {"Test-Cases": basename,
                                              "Different-Views": "Test case didn't run full - failed",
                                              'Componentname': "Test case didn't run full - failed", 'Status': 'FAILED',
                                              'Comments': "Test case didn't run full - failed"}
                                   # Create a DataFrame from new_row
                                   new_row_df = pd.DataFrame([new_row])
                                   # Concatenate new_row_df with df_componentstatus
                                   df_componentstatus = pd.concat([df_componentstatus, new_row_df], ignore_index=True)
                                   data_high_level_componentstatus.append(df_componentstatus)
                            except Exception as e:
                                pass
                        elif sheet_h == "DATA_MATCH":
                            try:
                               df_data_match = df
                               column_order_data_match =["Test-Cases","Module",'Component Type', 'Data validation']
                               df_data_match["Module"] = ["PDF Data Export and Validation"] * len(df_data_match)
                               # Add a new "Test-Cases" column and set it to a desired value
                               df_data_match["Test-Cases"] = [basename] * len(df_data_match)  # Replace "your_desired_value" with the value you want
                               # Reorder the columns based on the specified order
                               df_data_match = df_data_match[column_order_data_match]
                               data_data_match.append(df_data_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "DATA_NOT_MATCH":
                            try:
                                df_data_not_match = df
                                column_order_data_not_match = ["Test-Cases","Module",'Component Type', 'Data validation']
                                df_data_not_match["Module"] = ["PDF Data Export and Validation"] * len(df_data_not_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_data_not_match["Test-Cases"] = [basename] * len(df_data_not_match)  # Replace "your_desired_value" with the value you want

                                # Reorder the columns based on the specified order
                                df_data_not_match = df_data_not_match[column_order_data_not_match]

                                data_data_not_match.append(df_data_not_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "TABLESUMMARY_DATA_NOT_MATCH":
                            try:
                                df_tablesummary_data_not_match = df
                                column_order_tablesummary_data_not_match = ["Test-Cases","Module",'Data validation']
                                df_tablesummary_data_not_match["Module"] = ["Table Summary Export Validation"] * len(df_tablesummary_data_not_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_tablesummary_data_not_match["Test-Cases"] = [basename] * len(df_tablesummary_data_not_match)  # Replace "your_desired_value" with the value you want

                                # Reorder the columns based on the specified order
                                df_tablesummary_data_not_match = df_tablesummary_data_not_match[column_order_tablesummary_data_not_match]

                                data_tablesummary_data_not_match.append(df_tablesummary_data_not_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "TABLESUMMARY_DATA_MATCH":
                            try:
                                df_tablesummary_data_match = df
                                column_order_tablesummary_data_match = ["Test-Cases","Module",'Data validation']
                                df_tablesummary_data_match["Module"] = ["Table Summary Export Validation"] * len(df_tablesummary_data_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_tablesummary_data_match["Test-Cases"] = [basename] * len(df_tablesummary_data_match)  # Replace "your_desired_value" with the value you want

                                # Reorder the columns based on the specified order
                                df_tablesummary_data_match = df_tablesummary_data_match[column_order_tablesummary_data_match]

                                data_tablesummary_data_match.append(df_tablesummary_data_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "CBE_vs_CE_DONOT_MATCH":
                            try:
                                df_cbe_vs_ce_donot_match = df
                                column_order_cbe_vs_ce_donot_match = ["Test-Cases","Module",'File', 'ParameterType', 'Data validation']
                                df_cbe_vs_ce_donot_match["Module"] = ["Combined Export Data Validation"] * len(df_cbe_vs_ce_donot_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_cbe_vs_ce_donot_match["Test-Cases"] = [basename] * len(df_cbe_vs_ce_donot_match)  # Replace "your_desired_value" with the value you want

                                # Reorder the columns based on the specified order
                                df_cbe_vs_ce_donot_match = df_cbe_vs_ce_donot_match[column_order_cbe_vs_ce_donot_match]

                                data_cbe_vs_ce_donot_match.append(df_cbe_vs_ce_donot_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "CBE_vs_CE_MATCH":
                            try:
                                df_cbe_vs_ce_match = df
                                column_order_cbe_vs_ce_match = ["Test-Cases","Module",'File', 'ParameterType', 'Data validation']
                                df_cbe_vs_ce_match["Module"] = ["Combined Export Data Validation"] * len(df_cbe_vs_ce_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_cbe_vs_ce_match["Test-Cases"] = [basename] * len(df_cbe_vs_ce_match)  # Replace "your_desired_value" with the value you want
                                # Reorder the columns based on the specified order
                                df_cbe_vs_ce_match = df_cbe_vs_ce_match[column_order_cbe_vs_ce_match]
                                data_cbe_vs_ce_match.append(df_cbe_vs_ce_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "IPU_vs_CE_DATA_MATCH":
                            try:
                                df_ipu_vs_ce_match = df
                                column_order_ipu_vs_ce_match = ["Test-Cases","Module",'File',"Individual pop up headers","Individual pop up value","combine export value",'Data validation']
                                df_ipu_vs_ce_match["Module"] = ["Individual Popup window data validation"] * len(df_ipu_vs_ce_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_ipu_vs_ce_match["Test-Cases"] = [basename] * len(df_ipu_vs_ce_match)  # Replace "your_desired_value" with the value you want
                                # Reorder the columns based on the specified order
                                df_ipu_vs_ce_match = df_ipu_vs_ce_match[column_order_ipu_vs_ce_match]
                                data_ipu_vs_ce_match.append(df_ipu_vs_ce_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "IPU_vs_CE_DATA_NOT_MATCH":
                            try:
                                df_ipu_vs_ce_not_match = df
                                column_order_ipu_vs_ce_not_match = ["Test-Cases","Module",'File',"Individual pop up headers","Individual pop up value","combine export value",'Data validation']
                                df_ipu_vs_ce_not_match["Module"] = ["Individual Popup window data validation"] * len(df_ipu_vs_ce_not_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_ipu_vs_ce_not_match["Test-Cases"] = [basename] * len(df_ipu_vs_ce_not_match)  # Replace "your_desired_value" with the value you want
                                # Reorder the columns based on the specified order
                                df_ipu_vs_ce_not_match = df_ipu_vs_ce_not_match[column_order_ipu_vs_ce_not_match]
                                data_ipu_vs_ce_donot_match.append(df_ipu_vs_ce_not_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "NQC_vs_OC_DATA_NOT_MATCH":
                            try:
                                df_nqc_table_data_vs_oc_data_not_match = df
                                column_order_nqc_table_data_vs_oc_data_not_match = ["Test-Cases","Module",'File',"map view Operator","map view Operator value","calculated csv value",'Data validation']
                                df_nqc_table_data_vs_oc_data_not_match["Module"] = ["NQC table data validation"] * len(df_nqc_table_data_vs_oc_data_not_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_nqc_table_data_vs_oc_data_not_match["Test-Cases"] = [basename] * len(df_nqc_table_data_vs_oc_data_not_match)  # Replace "your_desired_value" with the value you want
                                # Reorder the columns based on the specified order
                                df_nqc_table_data_vs_oc_data_not_match = df_nqc_table_data_vs_oc_data_not_match[column_order_nqc_table_data_vs_oc_data_not_match]
                                data_nqc_table_data_vs_oc_data_not_match.append(df_nqc_table_data_vs_oc_data_not_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "NQC_vs_OC_DATA_MATCH":
                            try:
                                df_nqc_table_data_vs_oc_data_match = df
                                column_order_nqc_table_data_vs_oc_data_match = ["Test-Cases","Module",'File',"map view Operator","map view Operator value","calculated csv value",'Data validation']
                                df_nqc_table_data_vs_oc_data_match["Module"] = ["NQC table data validation"] * len(df_nqc_table_data_vs_oc_data_match)
                                # Add a new "Test-Cases" column and set it to a desired value
                                df_nqc_table_data_vs_oc_data_match["Test-Cases"] = [basename] * len(df_nqc_table_data_vs_oc_data_match)  # Replace "your_desired_value" with the value you want
                                # Reorder the columns based on the specified order
                                df_nqc_table_data_vs_oc_data_match = df_nqc_table_data_vs_oc_data_match[column_order_nqc_table_data_vs_oc_data_match]
                                data_nqc_table_data_vs_oc_data_match.append(df_nqc_table_data_vs_oc_data_match)
                            except Exception as e:
                                pass
                        elif sheet_h == "RESULTS_DEFAULT_SETTINGS":
                            try:
                                df = pd.read_excel(file, sheet_name='RESULTS_DEFAULT_SETTINGS', header=1)
                                df_default_settings = df
                                column_default_Settings = ["Test-Cases","Module",'File',"map view Operator","map view Operator value","calculated csv value",'Data validation']
                                df_default_settings["Module"] = ["DEFAULT SETTINGS"] * len(df_default_settings)
                                # # Add a new "Test-Cases" column and set it to a desired value
                                df_default_settings["Test-Cases"] = [basename] * len(df_default_settings)  # Replace "your_desired_value" with the value you want
                                # # Reorder the columns based on the specified order
                                # df_default_settings = df_default_settings[column_default_Settings]
                                data_comparison_default_settings.append(df_default_settings)
                            except Exception as e:
                                pass
                        elif sheet_h == "RESULTS_CHANGE_SETTINGS":
                            try:
                                df = pd.read_excel(file, sheet_name='RESULTS_CHANGE_SETTINGS', header=1)
                                df_change_settings = df
                                column_change_Settings = ["Test-Cases","Module",'File',"map view Operator","map view Operator value","calculated csv value",'Data validation']
                                df_change_settings["Module"] = ["CHANGE SETTINGS"] * len(df_change_settings)
                                # # Add a new "Test-Cases" column and set it to a desired value
                                df_change_settings["Test-Cases"] = [basename] * len(df_change_settings)  # Replace "your_desired_value" with the value you want
                                # # Reorder the columns based on the specified order
                                # df_change_settings = df_change_settings[column_change_Settings]
                                data_comparison_change_settings.append(df_change_settings)
                            except Exception as e:
                                pass

                    except Exception as e:
                        continue
            except Exception as e:
                continue
        combined_status_df = {}
        combined_cbe_vs_ce_match_df = {}
        combined_cbe_vs_ce_donot_match_df = {}
        combined_data_not_match_df = {}
        combined_data_match_df = {}
        combined_tablesummary_data_match_df = {}
        combined_tablesummary_data_not_match_df = {}
        combined_ipu_vs_ce_data_match_df = {}
        combined_ipu_vs_ce_data_not_match_df = {}
        combined_nqc_table_data_vs_oc_data_match_df = {}
        combined_nqc_table_data_vs_oc_data_not_match_df = {}
        combined_default_settings = {}
        combined_change_settings = {}
        result = []
        try:
            combined_status_df = pd.concat(data_high_level_componentstatus, ignore_index=True)
        except Exception as e:
            pass
        try:
            combined_data_match_df = pd.concat(data_data_match, ignore_index=True)
            result.append(combined_data_match_df)
        except Exception as e:
            pass
        try:
            combined_data_not_match_df = pd.concat(data_data_not_match, ignore_index=True)
            result.append(combined_data_not_match_df)
        except Exception as e:
            pass
        try:
            combined_tablesummary_data_match_df = pd.concat(data_tablesummary_data_match, ignore_index=True)
            result.append(combined_tablesummary_data_match_df)
        except Exception as e:
            pass
        try:
            combined_tablesummary_data_not_match_df = pd.concat(data_tablesummary_data_not_match, ignore_index=True)
            result.append(combined_tablesummary_data_not_match_df)
        except Exception as e:
            pass
        try:
            combined_cbe_vs_ce_donot_match_df = pd.concat(data_cbe_vs_ce_donot_match, ignore_index=True)
            result.append(combined_cbe_vs_ce_donot_match_df)
        except Exception as e:
            pass
        try:
            combined_cbe_vs_ce_match_df = pd.concat(data_cbe_vs_ce_match, ignore_index=True)
            result.append(combined_cbe_vs_ce_match_df)
        except Exception as e:
            pass
        try:
            combined_ipu_vs_ce_data_match_df = pd.concat(data_ipu_vs_ce_match, ignore_index=True)
            result.append(combined_ipu_vs_ce_data_match_df)
        except Exception as e:
            pass
        try:
            combined_ipu_vs_ce_data_not_match_df = pd.concat(data_ipu_vs_ce_donot_match, ignore_index=True)
            result.append(combined_ipu_vs_ce_data_not_match_df)
        except Exception as e:
            pass
        try:
            combined_nqc_table_data_vs_oc_data_match_df = pd.concat(data_nqc_table_data_vs_oc_data_match, ignore_index=True)
            result.append(combined_nqc_table_data_vs_oc_data_match_df)
        except Exception as e:
            pass
        try:
            combined_nqc_table_data_vs_oc_data_not_match_df = pd.concat(data_nqc_table_data_vs_oc_data_not_match, ignore_index=True)
            result.append(combined_nqc_table_data_vs_oc_data_not_match_df)
        except Exception as e:
            pass
        try:
            combined_default_settings = pd.concat(data_comparison_default_settings, ignore_index=True)
            result.append(combined_default_settings)
        except Exception as e:
            pass
        try:
            combined_change_settings = pd.concat(data_comparison_change_settings, ignore_index=True)
            result.append(combined_change_settings)
        except Exception as e:
            pass
        print("combined_status_df",combined_status_df)
        print("combined_cbe_vs_ce_match_df",combined_cbe_vs_ce_match_df)
        print("combined_cbe_vs_ce_donot_match_df",combined_cbe_vs_ce_donot_match_df)
        print("combined_data_not_match_df",combined_data_not_match_df)
        print("combined_data_match_df",combined_data_match_df)
        print("combined_tablesummary_data_match_df",combined_tablesummary_data_match_df)
        print("combined_tablesummary_data_not_match_df",combined_tablesummary_data_not_match_df)
        print("combined_ipu_vs_ce_data_match_df", combined_ipu_vs_ce_data_match_df)
        print("combined_ipu_vs_ce_data_not_match_df", combined_ipu_vs_ce_data_not_match_df)
        print("combined_ipu_vs_ce_data_match_df", combined_nqc_table_data_vs_oc_data_match_df)
        print("combined_ipu_vs_ce_data_not_match_df", combined_nqc_table_data_vs_oc_data_not_match_df)
        try:
            updatecomponentstatus_using_pandas_sending_data_frame_for_highlevelreport(sheet=sheet_high_level_componentstatus, dataframe=combined_status_df)
        except Exception as e :
            pass
        # try:
        #     update_excel_datavalidation_cbe_vs_ce(df=combined_cbe_vs_ce_match_df, sheet=sheet_cbe_vs_ce_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_cbe_vs_ce(df=combined_cbe_vs_ce_donot_match_df, sheet=sheet_cbe_vs_ce_donot_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_pdf(df=combined_data_match_df, sheet=sheet_data_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_pdf(df=combined_data_not_match_df, sheet=sheet_data_not_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_table_summary(df=combined_tablesummary_data_not_match_df, sheet=sheet_tablesummary_data_not_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_table_summary(df=combined_tablesummary_data_match_df, sheet=sheet_tablesummary_data_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_ipu_vs_ce(df=combined_ipu_vs_ce_data_match_df, sheet=sheet_ipu_vs_ce_data_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_ipu_vs_ce(df=combined_ipu_vs_ce_data_not_match_df, sheet=sheet_ipu_vs_ce_data_not_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_nqc_vs_oc(df=combined_nqc_table_data_vs_oc_data_match_df, sheet=sheet_nqc_table_data_vs_oc_data_match)
        # except:
        #     pass
        # try:
        #     update_excel_datavalidation_nqc_vs_oc(df=combined_nqc_table_data_vs_oc_data_not_match_df, sheet=sheet_nqc_table_data_vs_oc_data_not_match)
        # except:
        #     pass
        try:
            def updating_result_of_dataframe_for_excel(sheet, df):
                # Define color mapping
                color_mapping = {
                    "Same": (144, 238, 144),  # light Green
                    "Difference": (255, 150, 150),  # Light Red
                    "Different":(255, 150, 150),
                    "Key name can't find in csv": (255, 255, 0),  # Yellow
                    "Row Start": (173, 216, 230),
                    "is equal": (144, 238, 144),  # light Green
                    "is not equal": (255, 150, 150),  # Light Red
                    "Key name can't find in application": (255, 255, 0),  # Yellow
                    "Key name can't find in csv file": (255, 255, 153),  # Light Yellow
                    "Key name can't find in operator comparsion": (255, 255, 0),  # Yellow
                    "Key not present in combine_export": (255, 255, 0),  # Yellow
                    "Key not present in combine_binary_export": (255, 255, 153),  # Light Yellow
                    'STARTHERE': (210, 180, 140),# Light Brown
                    'ENDHERE': (210, 180, 140),# Light Brown
                    "The value is found" : (144, 238, 144), # Light Green
                    "The value is Not Found" : (255, 150, 150), # Light Red
                    "The value is found,but settings application value(reference) != excel settings values" : (255, 150, 150)
                }
                # Find the last non-empty row
                last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
                start_row = last_row + 1
                # Append headers
                header_range = sheet.range((start_row - 1, 1), (start_row - 1, df.shape[1]))
                header_range.value = df.columns.tolist()
                header_range.color = (255, 255, 0)
                # Insert the DataFrame into the sheet
                df_range = sheet.range((start_row, 1), (start_row + df.shape[0] - 1, df.shape[1]))
                df_range.value = df.values
                data_validation_column_index = df.columns.get_loc('Data validation')
                # Apply color formatting to the entire range
                for i, row in enumerate(df.iterrows(), start=start_row):
                    validation_cell = sheet.range(f"{chr(65 + data_validation_column_index)}{i}")
                    for keyword, color in color_mapping.items():
                        if keyword in row[1]['Data validation']:
                            validation_cell.color = color
                            for j in range(data_validation_column_index-1):
                                validation_cell1 = sheet.range(f"{chr(65+j+1)}{i}")
                                validation_cell1.color = color
            # for result_df in result:
            combinedresult_df = {}
            try:
                combinedresult_df = pd.concat(result, ignore_index=True)
                # Fill missing values with 0
                combinedresult_df = combinedresult_df.fillna(" ")
                mask = (combinedresult_df == 'STARTHERE') | (combinedresult_df == 'ENDHERE')
                # Use the mask to filter out rows
                combinedresult_df = combinedresult_df[~mask.any(axis=1)]
                if 'Data validation' in combinedresult_df.columns:
                    # Move the 'Data validation' column to the last position
                    combinedresult_df = combinedresult_df[[col for col in combinedresult_df.columns if col != 'Data validation'] + ['Data validation']]
            except Exception as e:
                pass
            updating_result_of_dataframe_for_excel(sheet=sheet_result, df=combinedresult_df)
            print("combinedresult_df",combinedresult_df)
        except:
            pass
        try:
            # Save and close the workbook
            workbook.save(high_level_excel_report_paths)
            workbook.close()
            app.quit()
        except Exception as e:
            pass
        try:
            source_file = high_level_excel_report_paths
            destination_file = os.path.join(high_level_excel_report_path, "HLD.xlsx")
            shutil.copy(source_file, destination_file)
        except Exception as e:
            pass
        try:
            hld_move_project_path(highlevelExcelReport_path = high_level_excel_report_path)
        except Exception as e:
            pass
    except Exception as e:
        pass
#################################################################################################################################################################################################################################
def update_excel_datavalidation_table_summary(df,sheet):
    """
        Update the high-level Excel report for data validation of table_summary.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        # Define color mapping
        color_mapping = {
            'STARTHERE': (210, 180, 140),  # Light Brown
            'ENDHERE': (210, 180, 140),  # Light Brown
            "Same": (0, 255, 0),  # Green
            "Difference": (255, 0, 0),  # Red
            "Key name can't find in application": (255, 255, 0),  # Yellow
            "Key name can't find in csv file": (255, 255, 153),  # Light Yellow
        }
        last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
        start_row = last_row + 1
        # Insert the DataFrame into the sheet
        df_range = sheet.range((start_row, 1))
        df_range.value = df.values
        # Apply color formatting to the entire range
        for i, row in enumerate(df.iterrows(), start=start_row):
            validation_cell = sheet.range(f"C{i}")
            for keyword, color in color_mapping.items():
                if keyword in row[1]['Data validation']:
                    validation_cell.color = color
        # Set colors for File and ParameterType columns
        range_to_color_test_case = sheet.range((start_row, 1), (start_row + len(df), 1))
        range_to_color_test_case1 = sheet.range((start_row, 2), (start_row + len(df), 2))
        range_to_color_test_case.color = (173, 216, 230)
        range_to_color_test_case1.color = (173, 216, 130)
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
###################################################################################################################################################################################################################
def update_excel_datavalidation_pdf(df,sheet):
    """
        Update the high-level Excel report for data validation of PDF.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        # Define color mapping
        color_mapping = {
            'STARTHERE': (210, 180, 140),  # Light Brown
            'ENDHERE': (210, 180, 140),  # Light Brown
            "Same": (0, 255, 0),  # Green
            "Difference": (255, 0, 0),  # Red
            "Key name can't find in operator comparsion": (255, 255, 0),  # Yellow
            "Key not present in combine_binary_export": (255, 255, 153),  # Light Yellow
        }
        last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
        start_row = last_row + 1
        # Insert the DataFrame into the sheet
        df_range = sheet.range((start_row, 1))
        df_range.value = df.values
        # Apply color formatting to the entire range
        for i, row in enumerate(df.iterrows(), start=start_row):
            validation_cell = sheet.range(f"D{i}")
            for keyword, color in color_mapping.items():
                if keyword in row[1]['Data validation']:
                    validation_cell.color = color
        # Set colors for File and ParameterType columns
        range_to_color_test_case = sheet.range((start_row, 1), (start_row + len(df), 1))
        range_to_color_component = sheet.range((start_row, 2), (start_row + len(df), 2))
        range_to_color_component1 = sheet.range((start_row, 3), (start_row + len(df), 3))
        range_to_color_test_case.color = (173, 216, 230)
        range_to_color_component.color = (255, 200, 100)  # Light Orange for ParameterType column
        range_to_color_component1.color = (255, 200, 150)
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
########################################################################################################################################################################################################################
def update_excel_datavalidation_cbe_vs_ce(df,sheet):
    """
        Update the high-level Excel report for data validation of CBE vs CE.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        # Define color mapping
        color_mapping = {
            'STARTHERE': (210, 180, 140),  # Light Brown
            'ENDHERE': (210, 180, 140),  # Light Brown
            "Same": (0, 255, 0),  # Green
            "Different": (255, 0, 0),  # Red
            "Key not present in combine_export": (255, 255, 0),  # Yellow
            "Key not present in combine_binary_export": (255, 255, 153),  # Light Yellow
        }
        last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
        start_row = last_row + 1
        # Insert the DataFrame into the sheet
        df_range = sheet.range((start_row, 1))
        df_range.value = df.values
        # Apply color formatting to the entire range
        for i, row in enumerate(df.iterrows(), start=start_row):
            validation_cell = sheet.range(f"E{i}")
            for keyword, color in color_mapping.items():
                if keyword in row[1]['Data validation']:
                    validation_cell.color = color
        # Set colors for File and ParameterType columns
        range_to_color_test_case = sheet.range((start_row, 1), (start_row + len(df), 1))
        range_to_color_file = sheet.range((start_row, 2), (start_row + len(df), 2))
        range_to_color_parameter = sheet.range((start_row, 3), (start_row + len(df), 3))
        range_to_color_parameter1 = sheet.range((start_row, 4), (start_row + len(df), 4))
        range_to_color_test_case.color = (173, 216, 280)
        range_to_color_file.color = (173, 216, 230)  # Light Blue for File column
        range_to_color_parameter.color = (255, 200, 100)  # Light Orange for ParameterType column
        range_to_color_parameter1.color = (255, 200, 150)
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
##########################################################################################################################################################################################################################################
def update_excel_datavalidation_ipu_vs_ce(df,sheet):
    """
        Update the high-level Excel report for data validation of ipu.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        # Define color mapping
        color_mapping = {
            'STARTHERE': (210, 180, 140),  # Light Brown
            'ENDHERE': (210, 180, 140),  # Light Brown
            "Same": (144, 238, 144),  #light Green
            "Difference": (255, 150, 150),# Light Red
            "Key name can't find in csv": (255, 255, 0),  # Yellow
            "Row Start": (173, 216, 230),
            "is equal": (144, 238, 144),  # light Green
            "is not equal": (255, 150, 150),# Light Red
            # "Key not present in combine_binary_export": (255, 255, 153),  # Light Yellow
        }
        last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
        start_row = last_row + 1
        # Insert the DataFrame into the sheet
        df_range = sheet.range((start_row, 1))
        df_range.value = df.values
        # Apply color formatting to the entire range
        for i, row in enumerate(df.iterrows(), start=start_row):
            validation_cell1 = sheet.range(f"D{i}")
            validation_cell2 = sheet.range(f"E{i}")
            validation_cell3 = sheet.range(f"F{i}")
            validation_cell = sheet.range(f"G{i}")
            for keyword, color in color_mapping.items():
                if keyword in row[1]['Data validation']:
                    validation_cell.color = color
                    validation_cell1.color = color
                    validation_cell2.color = color
                    validation_cell3.color = color
        # Set colors for File and ParameterType columns
        range_to_color_test_case = sheet.range((start_row, 1), (start_row + len(df), 1))
        range_to_color_component = sheet.range((start_row, 2), (start_row + len(df), 2))
        range_to_color_component1 = sheet.range((start_row,3), (start_row + len(df), 3))
        range_to_color_test_case.color = (173, 216, 230)
        range_to_color_component.color = (255, 200, 100)
        range_to_color_component1.color = (255, 200, 150)# Light Orange for ParameterType column
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
update_excel_datavalidation_nqc_vs_oc=update_excel_datavalidation_ipu_vs_ce
##########################################################################################################################################################################################################################################
def hld_move_project_path(highlevelExcelReport_path):
    """
        Move HLD (High-Level Design) Excel reports from a source directory(test data root path) to a destination directory within the project path.
        Args:
            highlevelExcelReport_path (str): The path to the folder containing High-Level Design Excel reports.
        Returns:
        None
            """
    project_path = pathlib.Path(__file__).parent.parent
    print(project_path)
    source_directory_to_transfer_files = highlevelExcelReport_path
    destination_directory_to_transfer_files = os.path.join(project_path, "HLD")
    # # Clear the existing files in the destination directory
    existing_files = os.listdir(destination_directory_to_transfer_files)
    for existing_file in existing_files:
        file_path = os.path.join(destination_directory_to_transfer_files, existing_file)
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(f"Removed: {existing_file}")
    try:
        # List all files in the source directory
        files_to_transfer = os.listdir(source_directory_to_transfer_files)
        source_path = None
        destination_path = None
        # Iterate through each file and transfer it to the destination directory
        for file_name in files_to_transfer:
            if file_name.endswith("HLD.xlsx"):
                source_path = os.path.join(source_directory_to_transfer_files, file_name)
                destination_path = os.path.join(destination_directory_to_transfer_files, file_name)
            # Check if the path is a file (not a directory) before transferring
            if (source_path != None and destination_path != None):
                if os.path.isfile(source_path):
                    shutil.copy(source_path, destination_path)
                    print(f"Transferred: {file_name}")
                    print("Transfer complete.")
    except Exception as e:
        print(f"An error occurred: {e}")
###################################################################################################################################################################################################################################################
def check_selected_and_finding_enable_and_disabled_checkboxes_(driver, checkboxes_locators):
    """
        Uncheck selected checkboxes, find enabled checkboxes, and identify disabled checkboxes within a specified container.
        Args:
            driver (WebDriver): The WebDriver instance for the web page or application.
            checkboxes_locators: The locator(s) for the container containing the checkboxes.
        Returns:
            tuple: A tuple containing two lists - a list of enabled checkboxes and a list of disabled checkboxes.
        """
    disabled_checkboxes = []
    enabled_checkboxes = []
    try:
        # Find the parent element that contains the checkboxes
        checkboxes_parent = driver.find_element(*checkboxes_locators)
        # Find all checkboxes within the parent element
        checkboxes = checkboxes_parent.find_elements(By.TAG_NAME, "label")
        # Iterate through the checkboxes
        for checkbox in checkboxes:
            checkbox_inputs = checkbox.find_elements(By.TAG_NAME, "input")
            for checkbox_input in checkbox_inputs:
                try:
                    if not checkbox_input.is_selected():
                        checkbox_input.click()
                        time.sleep(1)
                    if checkbox_input.is_enabled():
                        text = checkbox.get_attribute("outerText")
                        if text == '' or text == None:
                            text = checkbox.get_attribute("innerText")
                        if text == '' or text == None:
                            text = checkbox.text
                        if text == '' or text == None:
                            text = checkbox.get_attribute("textContent")
                        enabled_checkboxes.append(text)
                    else:
                        text = checkbox.get_attribute("outerText")
                        if text == '' or text == None:
                            text = checkbox.get_attribute("innerText")
                        if text == '' or text == None:
                            text = checkbox.text
                        if text == '' or text == None:
                            text = checkbox.get_attribute("textContent")
                        disabled_checkboxes.append(text)
                except Exception as e:
                    print(f"Error occurred while unchecking checkbox: {str(e)}")
    except Exception as e:
        print(f"Error occurred while finding checkboxes: {str(e)}")
    return enabled_checkboxes, disabled_checkboxes
#############################################################################################################################################################################################################################################################################################################
def is_numeric(value):
    """
        Check if a value is of a numeric data type (e.g., float or int).
        Args:
            value: The value to be checked.
        Returns:
            bool: True if the value is numeric, False otherwise.
        """
    return pd.api.types.is_numeric_dtype(value)
##########################################################################################################################################################################################################################################################################################
def convert_to_float(value):
    """
        Convert a value to a float if possible; otherwise, return the original value.
        Args:
            value: The value to be converted.
        Returns:
            float or original value: The converted float value or the original value if conversion is not possible.
        """
    try:
        return float(value)
    except (ValueError, TypeError):
        return value
##################################################################################################################################################################################################################################################################################
def handling_all_data_type_for_comparsion(value):
    """
        Handle and normalize various data types (string, float, int, string with numeric values) for comparison.
        Args:
            value: The value to be processed.
        Returns:
            str: The normalized and processed value for comparison.
        """
    try:
        if str(value).lower() == 'NaN'.lower() or str(value).lower() == ' '.lower() or str(value).lower() == ''.lower():
            value = 'None'
    except:
        pass
    if not value is None:
        value = convert_to_float(value)
    try:
        numeric_part = re.search(r'([a-zA-Z]+)?(\d+(\.\d+)?|\.\d+)([a-zA-Z]+)?', value)
        if numeric_part:
            prefix = numeric_part.group(1) or ''
            numeric_value = numeric_part.group(2)
            suffix = numeric_part.group(4) or ''
            return f'{prefix}{float(numeric_value)}{suffix}'.lower().strip().replace(" ", '')
        else:
            return str(value).lower().strip().replace(" ", '')
    except (ValueError, TypeError):
        return str(value).lower().strip().replace(" ", '')
#####################################################################################################################################################################################################################################################################################################
def compare_values(value1, value2):
    """
        Compare two values while handling various data types (string, float, int, string with numeric values).
        Args:
            value1: The first value for comparison.
            value2: The second value for comparison.
        Returns:
            bool: True if the values are equal, False otherwise.
        """
    try:
        if str(value1).lower() == 'NaN'.lower() or str(value1).lower() == ' '.lower() or str(value1).lower() == ''.lower() :
            value1 = 'None'
    except:
        pass
    try:
        if str(value2).lower() == 'NaN'.lower() or str(value2).lower() == ' '.lower() or str(value2).lower() == ''.lower():
            value2 = 'None'
    except:
        pass
    if not value1 is None and not value2 is None:
        value1 = convert_to_float(value1)
        value2 = convert_to_float(value2)
        value1 = handling_all_data_type_for_comparsion(value1)
        value2 = handling_all_data_type_for_comparsion(value2)
    if value1 is None and value2 is None:
        return str(value1) == str(value2)
    elif is_numeric(value1) and is_numeric(value2):
        # Compare floats with a tolerance of 1e-6 to handle small differences due to float representation
        return value1 == value2
    else:
        str_value1 = str(value1).lower().strip()
        str_value2 = str(value2).lower().strip()
        return str_value1 == str_value2
################################################################################################################################################################################################################################################################################################
def comparsion_values_in_bw_two_list(list1,list2):
    """
        Compare two lists with constraints, handling various data types in the elements.
        Args:
            list1 (list): The first list for comparison.
            list2 (list): The second list for comparison.
        Returns:
            bool: True if the lists are equal with constraints, False otherwise.
        """
    list1_normalized = [handling_all_data_type_for_comparsion(x) for x in list1]
    list2_normalized = [handling_all_data_type_for_comparsion(x) for x in list2]
    print(list2_normalized)
    print(list1_normalized)
    return all(x in list2_normalized for x in list1_normalized) and all(y in list1_normalized for y in list2_normalized)
####################################################################################################################################################################################################################
class StepFailure(Exception):
    pass
##################################################################################################################################################################################################################################
def change_the_download_path(driver,downloadpath):
    """
        Change the default download path or environment-set path during driver initialization to a user-defined path.
        Args:
            driver (WebDriver): The WebDriver instance for the web page or application.
            downloadpath (str): The user-defined download path where files will be saved.
        Returns:
            None
        """
    driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {'cmd': 'Page.setDownloadBehavior','params': {'behavior': 'allow', 'downloadPath':downloadpath}}
    driver.execute("send_command", params)
#############################################################################################################################################################################################################################
def compress_png_files(folder_path, compression_level=9, resize_width=800, resize_height=600):
    """
    Compress PNG files in a specified folder.
    Args:
        folder_path (str): The path to the folder containing PNG files to be compressed.
        compression_level (int, optional): The compression level to be applied to the PNG files (0-9, where 0 is no compression and 9 is maximum compression). Defaults to 9.
        resize_width (int, optional): The width to which the images should be resized. If None, the original width is preserved. Defaults to 800.
        resize_height (int, optional): The height to which the images should be resized. If None, the original height is preserved. Defaults to 600.
    Notes:
        This function takes a folder path as input, locates all PNG files within that folder, and compresses them.
        It checks the size of each PNG file and, if it exceeds 1000 KB, it loads the image using OpenCV (cv2), resizes
        it if width and/or height are specified, and applies PNG compression. The compressed image is then saved,
        replacing the original file. If any errors occur during this process, they are ignored, and compression
        continues for other files in the folder.
    """
    # Get a list of PNG files in the folder
    png_files = [f for f in os.listdir(folder_path) if f.endswith('.png')]
    # Compress each PNG file
    for file in png_files:
        try:
            # Construct the file paths
            input_path = os.path.join(folder_path, file)
            output_path = input_path  # Replace the original file
            # Check the file size
            file_size_kb = os.path.getsize(input_path) / 1024  # Convert bytes to kilobytes
            if file_size_kb > 1000:
                # Load the image using cv2
                image = cv2.imread(input_path)
                # Resize the image if width and/or height are specified
                if resize_width is not None or resize_height is not None:
                    if resize_width is None:
                        resize_width = image.shape[1]  # Preserve the original width ratio
                    if resize_height is None:
                        resize_height = image.shape[0]  # Preserve the original height ratio
                    image = cv2.resize(image, (resize_width, resize_height))
                # Set the compression parameters
                compression_params = [cv2.IMWRITE_PNG_COMPRESSION, compression_level]
                # Compress the image
                cv2.imwrite(output_path, image, compression_params)
        except:
            continue
#############################################################################################################################################################################################################################################
def encrypte_decrypte(text):
    result ="None"
    def en_de_special_symbols(char):
        # Define a dictionary for character mappings
        char_mappings = {
            '@': '/',
            '/': '@',
            '|': '=',
            '=': '|',
            '%': '#',
            '#': '%',
            '+': '-',
            '-': '+',
            '[': ']',
            ']': '[',
            '(': ')',
            ')': '(',
            '{': '}',
            '}': '{',
            '<': '>',
            '>': '<',
            '*': '&',
            '&': '*',
            '^': '~',
            '~': '^',
            '!': '?',
            '?': '!',
            '$': '`',
            '`': '$',
            ':': ';',
            ';': ':',
            '.': ',',
            ',': '.',
            '"': "'",
            "'": '"',
            ' ': ' ',
            "_":"_"# You can add more mappings as needed
        }
        # Use the dictionary to look up the mapped character
        return char_mappings.get(char, char)
    def en_de(value,text):
        result = []
        for char in text:
            offset = ord(char)
            if 'a' <= char <= 'z':
                result.append(value+chr(((offset - ord('a') + 13) % 26) + ord('a')))
            elif 'A' <= char <= 'Z':
                result.append(value+chr(((offset - ord('A') + 13) % 26) + ord('A')))
            elif '0' <= char <= '9':
                result.append(value+chr(((offset - ord('0') + 5) % 10) + ord('0')))  # Apply ROT5 to numerals
            else:
                char = en_de_special_symbols(char)
                result.append(value+char)  # Non-alphanumeric characters remain unchanged
        return ''.join(result)
    pattern = "XAXCXBX"
    # Check if the pattern is present in the text
    match = re.search(pattern, text)
    if match:
        # If the pattern is found, remove it
        text = re.sub(pattern, '', text)
        result = en_de(value="",text=text)
    elif not match:
        result = en_de(value=pattern,text=text)
        result +=pattern
    return result

#########################################################################################################################################################################################################################################################
def side_bar_to_run_for_androidtest(keys_to_check):
    try:
        pattern_mapping_df = pd.read_excel(config.test_data_path, sheet_name="TEST_RUN")
    except Exception as e:
        with allure.step(f"Check {config.test_data_path}"):
            print(f"Check {config.test_data_path}")
            assert False

    df = pattern_mapping_df.set_index('Module').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()

    # Check if any of the specified keys have 'Yes' in their associated values
    # keys_with_yes = [key for key in keys_to_check if 'Yes' in df.get(key, [])]
    keys_with_yes = [key for key in keys_to_check if ('Yes' in df.get(key, []) or 'RUNNED' in df.get(key, []))]

    if keys_with_yes:
        return True
    else:
        return False
def Testrun_mode(value):
    try:
        pattern_mapping_df = pd.read_excel(config.test_data_path,sheet_name="TEST_RUN")
    except Exception as e:
        with allure.step(f"Check {config.test_data_path}"):
            print(f"Check {config.test_data_path}")
            assert False
    # Select columns starting from the second column (index 1) to the last column
    pattern_mapping_df = pattern_mapping_df.iloc[:, 1:3]
    pattern_mapping = pattern_mapping_df.set_index('Module').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
    test = value.strip()
    txt =[]# Remove leading and trailing spaces from test
    for pattern, values in pattern_mapping.items():
        if pattern.lower() == test.lower():
            txt = values
            break
        else:
            txt = []
    return txt
def updating_yes_to_run(remotevalue,types_of_test):
    try:
        # text_file_update_for_remote_test("C:\RantCell_Automation_Data_and_Reports\downloads\test_execution_status.txt"
        workbook = openpyxl.load_workbook(config.test_data_path)
        campaignsheet = workbook["TEST_RUN"]
        # Iterate through rows
        campaignsheetrows = campaignsheet.max_row
        campaigns_test = []
        yes = 0
        for r in range(2, campaignsheetrows + 1):
            values = ''
            value=campaignsheet.cell(row=r, column=2).value
            if str(value).lower() == str(types_of_test).lower():
                if campaignsheet.cell(row=r, column=3).value == "Yes":
                    campaignsheet.cell(row=r, column=3).value = remotevalue
                    break
                elif campaignsheet.cell(row=r, column=3).value == "RUNNED":
                    campaignsheet.cell(row=r, column=3).value = remotevalue
                    break
                elif campaignsheet.cell(row=r, column=3).value == "FINISHED":
                    campaignsheet.cell(row=r, column=3).value = remotevalue
                    break
                elif campaignsheet.cell(row=r, column=3).value == "WAITING LOAD":
                    campaignsheet.cell(row=r, column=3).value = remotevalue
                    break
                elif campaignsheet.cell(row=r, column=3).value == "LOADING":
                    campaignsheet.cell(row=r, column=3).value = remotevalue
                    break
        workbook.save(config.test_data_path)
        workbook.close(config.test_data_path)
    except Exception as e:
        pass

def generate_random_alphabet(length):
    return ''.join(random.choice(string.ascii_lowercase) for _ in range(length))

def min_max_elements_present(locator, min_count, max_count):
    def condition(driver):
        elements = driver.find_elements(*locator)
        return min_count <= len(elements) <= max_count

    return condition

# def update_component_status_openpyxl(worksheet, dataframe):
#     try:
#         # Map status values to corresponding fill colors
#         status_colors = {
#             "PASSED": PatternFill(start_color="35FC03", end_color="35FC03", fill_type="solid"),
#             "FAILED": PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid"),
#             "IGNORED": PatternFill(start_color="FFF998", end_color="FFF998", fill_type="solid"),
#             "SKIPPED": PatternFill(start_color="FFA590", end_color="FFA590", fill_type="solid")
#         }
#         # Find the last used row in the sheet
#         last_row = worksheet.max_row
#         # Append DataFrame data to the worksheet
#         for index, row in dataframe.iterrows():
#             worksheet.append(row.tolist())
#         # Update cell colors based on the "Status" column
#         # Define the columns to update
#         columns_to_update = [5, 10, 8, 9]
#         # Iterate through each column
#         for col in columns_to_update:
#             for row_index, row in enumerate(worksheet.iter_rows(min_row=last_row + 1, max_row=last_row + len(dataframe), min_col=col,max_col=col)):
#                 status_cell = row[0]
#                 status = status_cell.value
#                 if status in status_colors:
#                     status_cell.fill = status_colors[status]
#     except Exception as e:
#         with allure.step(f"component status {str(e)}"):
#             pass

############################################################# HLD Comparison ##################################################################################
# def updating_result_of_dataframe_for_excel(sheet, df):
#     # Define color mapping
#     color_mapping = {
#         "Same": (144, 238, 144),  # light Green
#         "Difference": (255, 150, 150),  # Light Red
#         "Different": (255, 150, 150),
#         "Key name can't find in csv": (255, 255, 0),  # Yellow
#         "Row Start": (173, 216, 230),
#         "is equal": (144, 238, 144),  # light Green
#         "is not equal": (255, 150, 150),  # Light Red
#         "Key name can't find in application": (255, 255, 0),  # Yellow
#         "Key name can't find in csv file": (255, 255, 153),  # Light Yellow
#         "Key name can't find in operator comparsion": (255, 255, 0),  # Yellow
#         "Key not present in combine_export": (255, 255, 0),  # Yellow
#         "Key not present in combine_binary_export": (255, 255, 153),  # Light Yellow
#         'STARTHERE': (210, 180, 140),  # Light Brown
#         'ENDHERE': (210, 180, 140),  # Light Brown
#     }
#     # Find the last non-empty row
#     last_row = sheet.range('A' + str(sheet.cells.last_cell.row)).end('up').row
#     start_row = last_row + 1
#     # Append headers
#     header_range = sheet.range((start_row - 1, 1), (start_row - 1, df.shape[1]))
#     header_range.value = df.columns.tolist()
#     header_range.color = (255, 255, 0)
#     # Insert the DataFrame into the sheet
#     df_range = sheet.range((start_row, 1), (start_row + df.shape[0] - 1, df.shape[1]))
#     df_range.value = df.values
#     data_validation_column_index = df.columns.get_loc('Data validation1')
#     # Apply color formatting to the entire range
#     for i, row in enumerate(df.iterrows(), start=start_row):
#         validation_cell = sheet.range(f"{chr(65 + data_validation_column_index)}{i}")
#         for keyword, color in color_mapping.items():
#             if keyword in row[1]['Data validation1']:
#                 validation_cell.color = color
#                 for j in range(data_validation_column_index - 1):
#                     validation_cell1 = sheet.range(f"{chr(65 + j + 1)}{i}")
#                     validation_cell1.color = color
#     data_validation_column_index = df.columns.get_loc('Data validation2')
#     # Apply color formatting to the entire range
#     for i, row in enumerate(df.iterrows(), start=start_row):
#         validation_cell = sheet.range(f"{chr(65 + data_validation_column_index)}{i}")
#         for keyword, color in color_mapping.items():
#             if keyword in row[1]['Data validation2']:
#                 validation_cell.color = color
#                 for j in range(data_validation_column_index - 1):
#                     validation_cell1 = sheet.range(f"{chr(65 + j + 1)}{i}")
#                     validation_cell1.color = color

# def read_excel_and_filter(excel_path, sheet_name,comparing_2hld_file_value,server,exclude_values = ['SCHEDULE TEST', 'CONTINUOUS TEST', 'REMOTE TEST','GROUP']):
#     df = pd.read_excel(excel_path, sheet_name=sheet_name)
#     if sheet_name == "HIGH LEVEL COMPONENTSTATUS":
#         # Exclude rows where 'Different-Views' column has specific values
#         df = df[~df['Different-Views'].isin(exclude_values)]
#     if comparing_2hld_file_value != 'Yes':
#         df = df[df['Test-Cases'].str.contains(server)].groupby('Test-Cases').apply(lambda x: x.to_dict(orient='records')).to_dict()
#     elif comparing_2hld_file_value == 'Yes':
#         df = df.groupby('Test-Cases').apply(lambda x: x.applymap(lambda y: '' if pd.isna(y) else y).to_dict(orient='records')).to_dict()
#     return df
#
#
# def update_restructured_dict(df_data, df_data_other, restructured_dict, restructured_dict_other):
#     for key, value in df_data.items():
#         new_key = key
#         value_other = df_data_other.get(key, [{}])
#
#         for entry in value:
#             if entry.get('Different-Views') == 'Side bar menu' and entry.get('Componentname', ''):
#                 key = entry['Componentname']
#                 break
#
#         restructured_dict[key] = value
#         if key not in {'Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments'}:
#             restructured_dict_other[key] = value_other
#         else:
#             restructured_dict_other = df_data_other.copy()
#
#
# def process_restructured_dict(restructured_dict):
#     for key, value in restructured_dict.items():
#         for item in value:
#             try:
#                 item['Componentname'] = re.sub(r'\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}', '', item.get('Componentname', ''))
#                 try:
#                     item['Comments'] = replace_and_sort(re.sub(r'\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}', '', item.get('Comments', '')))
#                 except :
#                     item['Comments'] = replace_and_sort(item.get('Comments', ''))
#                     pass
#             except Exception as e:
#                 continue
def sort_restructured_dict(source_dict, target_dict, key_mapping):
    for key in source_dict:
        try:
            # Get the order from source_dict for the current key
            order_source = [key_mapping(item) for item in source_dict[key]]
            # Sort target_dict based on the order from source_dict for the current key
            target_dict[key] = sorted(target_dict[key],key=lambda x: order_source.index(key_mapping(x)) if key_mapping(x) in order_source else float('inf'))
        except Exception as e:
            continue
#
# def process_restructured_dict_files(restructured_dict):
#     for key, value in restructured_dict.items():
#         for item in value:
#             try:
#                 item['File'] = re.sub(r'\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}', '', item.get('File', ''))
#                 try:
#                     item['Data validation'] = replace_and_sort(re.sub(r'\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}', '', item.get('Data validation', '')))
#                 except :
#                     item['Data validation'] = replace_and_sort(item.get('Data validation', ''))
#                     pass
#             except Exception as e:
#                 continue


# def compare_hld(dict_1, dict_2, differences, similarities, row_in_restructured_dict_key1, row_in_restructured_dict_key2,exclude_keys, fillvalue, ignore_keys):
#     for key in dict_1.keys():
#         if key in dict_2:
#             fillvalue['Test-Cases']=key
#             rows1 = dict_1[key]
#             rows2 = dict_2[key]
#             rows1_without_test_case = []
#             for row1 in rows1:
#                 a_value = {k:v for k, v in row1.items() if 'Test-Cases'!= k}
#                 rows1_without_test_case.append(a_value)
#             rows2_without_test_case = []
#             for row2 in rows2:
#                 b_value = {k: v for k, v in row2.items() if 'Test-Cases' != k}
#                 rows2_without_test_case.append(b_value)
#
#             # Compare each row in the lists
#             row_differences = []
#             row_similarities = []
#
#             for row1, row2 in zip_longest(rows1, rows2, fillvalue=fillvalue):
#                 # Exclude 'Test_Case' from the comparison
#                 row1_without_test_case = {k: v for k, v in row1.items() if k not in ignore_keys}
#                 row2_without_test_case = {k: v for k, v in row2.items() if k not in ignore_keys}
#                 fillvalue1 = {k: v for k, v in fillvalue.items() if k not in ignore_keys}
#                 if row1_without_test_case == row2_without_test_case:
#                     row_similarities.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: row2})
#                 else:
#                     # row_differences.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: row2})
#                     if row1_without_test_case != fillvalue1 and row2_without_test_case != fillvalue1:
#                         flag_1 = 0
#                         flag_2 = 0
#                         if row1_without_test_case in rows2_without_test_case:
#                             flag_1 = 1
#                             rows2.remove(row1)
#                             row_similarities.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: row1})
#                         if row2_without_test_case in rows1_without_test_case:
#                             flag_2 = 1
#                             rows1.remove(row2)
#                             row_similarities.append({row_in_restructured_dict_key1: row2, row_in_restructured_dict_key2: row2})
#                         if flag_1 == 0 and flag_2 == 0:
#                             row_differences.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: row2})
#                         elif flag_1 == 1 and flag_2 == 0:
#                             row_differences.append({row_in_restructured_dict_key1: fillvalue, row_in_restructured_dict_key2: row2})
#                         elif flag_1 == 0 and flag_2 == 1:
#                             row_differences.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: fillvalue})
#                     else:
#                         if row2_without_test_case != fillvalue1 and row1_without_test_case == fillvalue1:
#                             if row2_without_test_case in rows1_without_test_case:
#                                 rows1.remove(row2)
#                                 row_similarities.append({row_in_restructured_dict_key1: row2, row_in_restructured_dict_key2: row2})
#                             else:
#                                 row_differences.append({row_in_restructured_dict_key1: fillvalue , row_in_restructured_dict_key2: row2})
#                         if row1_without_test_case != fillvalue1 and row2_without_test_case == fillvalue1:
#                             if row1_without_test_case in rows2_without_test_case:
#                                 rows2.remove(row1)
#                                 row_similarities.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: row1})
#                             else:
#                                 row_differences.append({row_in_restructured_dict_key1: row1, row_in_restructured_dict_key2: fillvalue})
#             # Store the differences_data and similarities_data for the current key
#             if row_differences:
#                 differences[key] = row_differences
#             if row_similarities:
#                 similarities[key] = row_similarities
#         else:
#             differences[key] = [{row_in_restructured_dict_key1: dict_1[key][1], row_in_restructured_dict_key2: fillvalue}]
#     # Identify keys present in dict_2 but not in dict_1
#     for key in dict_2.keys():
#         # Assuming 'key' is the variable containing the current key to check
#         if key not in exclude_keys:
#             if key not in dict_1:
#                 fillvalue['Test-Cases'] = key
#                 differences[key] = [{row_in_restructured_dict_key1: fillvalue,row_in_restructured_dict_key2: dict_2[key][0]}]
#
# def create_dict_within_list_for_hld_components_status(df_dict,flattened_data_dict):
#     for key, value in df_dict.items():
#         for item in value:
#             try:
#                 flattened_data_dict.append({
#                     'Test_Set': key,
#                     'Test_Case_1': item['row_in_restructured_dict1']['Test-Cases'],
#                     'Test_Case_2': item['row_in_restructured_dict2']['Test-Cases'],
#                     'Different_Views_1': item['row_in_restructured_dict1']['Different-Views'],
#                     'Different_Views_2': item['row_in_restructured_dict2']['Different-Views'],
#                     'Componentname_1': item['row_in_restructured_dict1']['Componentname'],
#                     'Componentname_2': item['row_in_restructured_dict2']['Componentname'],
#                     'Status_1': item['row_in_restructured_dict1']['Status'],
#                     'Status_2': item['row_in_restructured_dict2']['Status'],
#                     'Comments_1': item['row_in_restructured_dict1']['Comments'],
#                     'Comments_2': item['row_in_restructured_dict2']['Comments']
#                 })
#             except Exception as e:
#                 continue
# def create_dict_within_list_for_hld_result_of_each_module(df_dict,flattened_data_dict):
#     for key, value in df_dict.items():
#         for item in value:
#             try:
#                 flattened_data_dict.append({
#                     'Test_Set': key,
#                     'Test_Case_1': item['row_in_restructured_dict3']['Test-Cases'],
#                     'Test_Case_2': item['row_in_restructured_dict4']['Test-Cases'],
#                     'Module_1': item['row_in_restructured_dict3']['Module'],
#                     'Module_2': item['row_in_restructured_dict4']['Module'],
#                     'Component Type1': item['row_in_restructured_dict3']['Component Type'],
#                     'Component Type2': item['row_in_restructured_dict4']['Component Type'],
#                     'File_1': item['row_in_restructured_dict3']['File'],
#                     'File_2': item['row_in_restructured_dict4']['File'],
#                     'ParameterType1': item['row_in_restructured_dict3']['ParameterType'],
#                     'ParameterType2': item['row_in_restructured_dict4']['ParameterType'],
#                     'Individual pop up headers1': item['row_in_restructured_dict3']['Individual pop up headers'],
#                     'Individual pop up headers2': item['row_in_restructured_dict4']['Individual pop up headers'],
#                     'Individual pop up value1': item['row_in_restructured_dict3']['Individual pop up value'],
#                     'Individual pop up value2': item['row_in_restructured_dict4']['Individual pop up value'],
#                     'combine export value1': item['row_in_restructured_dict3']['combine export value'],
#                     'combine export value2': item['row_in_restructured_dict3']['combine export value'],
#                     'map view Operator value1': item['row_in_restructured_dict3']['map view Operator value'],
#                     'map view Operator value2': item['row_in_restructured_dict4']['map view Operator value'],
#                     'calculated csv value1': item['row_in_restructured_dict3']['calculated csv value'],
#                     'calculated csv value2': item['row_in_restructured_dict4']['calculated csv value'],
#                     'Data validation1': item['row_in_restructured_dict3']['Data validation'],
#                     'Data validation2': item['row_in_restructured_dict4']['Data validation'],
#                 })
#             except Exception as e:
#                 continue
# def create_dataframe(flattened_data_dict,dataframe):
#     df_dict = pd.DataFrame(flattened_data_dict)
#     dataframe.append(df_dict)
#
# def comparing_hld_result(excelhld1, excelhld2,comparing_2hld_file_value,excelpath_for_ouput,server1,server2):
#     # Initialize restructured dictionaries
#     restructured_dict1 = {}
#     restructured_dict3 = {}
#     restructured_dict2 = {}
#     restructured_dict4 = {}
#     flattened_data_dict1 = []
#     flattened_data_dict2 = []
#     flattened_data_dict3 = []
#     flattened_data_dict4 = []
#     df_data4 = None
#     df_data2 = None
#     # Compare specific rows within restructured_dict1 and restructured_dict2
#     differences = {}
#     similarities = {}
#     # Compare specific rows within restructured_dict3 and restructured_dict4
#     differences_data = {}
#     similarities_data = {}
#     if comparing_2hld_file_value != 'Yes':
#         excelhld2 = excelhld1
#     # Read and filter Excel data
#     df_data1 = read_excel_and_filter(excelhld1, "HIGH LEVEL COMPONENTSTATUS",comparing_2hld_file_value,server=server1)
#     df_data2 = read_excel_and_filter(excelhld2, "HIGH LEVEL COMPONENTSTATUS",comparing_2hld_file_value,server=server2)
#
#     df_data3 = read_excel_and_filter(excelhld1, "RESULT OF EACH MODULE",comparing_2hld_file_value,server=server1)
#     df_data4 = read_excel_and_filter(excelhld2, "RESULT OF EACH MODULE",comparing_2hld_file_value,server=server2)
#
#     # Update restructured dictionaries based on conditions
#     update_restructured_dict(df_data1, df_data3, restructured_dict1, restructured_dict3)
#     update_restructured_dict(df_data2, df_data4, restructured_dict2, restructured_dict4)
#
#     # Process restructured dictionaries
#     process_restructured_dict(restructured_dict1)
#     process_restructured_dict(restructured_dict2)
#
#     # Sort restructured_dict2 based on order from restructured_dict1
#     sort_restructured_dict(restructured_dict1, restructured_dict2, lambda x: (x['Different-Views'], x['Componentname']))
#
#     # Sort restructured_dict1 based on order from restructured_dict2
#     sort_restructured_dict(restructured_dict2, restructured_dict1, lambda x: (x['Different-Views'], x['Componentname']))
#
#     # Process restructured_dict3 by removing date and time patterns from 'File'
#     process_restructured_dict_files(restructured_dict3)
#
#     # Process restructured_dict4 by removing date and time patterns from 'File'
#     process_restructured_dict_files(restructured_dict4)
#
#     # Sort restructured_dict4 based on order from restructured_dict3
#     sort_restructured_dict(restructured_dict3, restructured_dict4, lambda x: (x['File'], x['ParameterType'], x['Individual pop up headers'], x['map view Operator']))
#
#     # Sort restructured_dict3 based on order from restructured_dict4
#     sort_restructured_dict(restructured_dict4, restructured_dict3, lambda x: (x['File'], x['ParameterType'], x['Individual pop up headers'], x['map view Operator']))
#
#     compare_hld(dict_1=restructured_dict1, dict_2=restructured_dict2, differences=differences,
#                 similarities=similarities, row_in_restructured_dict_key1='row_in_restructured_dict1',
#                 row_in_restructured_dict_key2='row_in_restructured_dict2',
#                 exclude_keys=['Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments'],
#                 fillvalue={'Test-Cases': '', "Different-Views": "Result Not found", "Componentname": 'Result Not found',
#                            'Status': 'Result Not found', 'Comments': 'Result Not found'}, ignore_keys=['Test-Cases'])
#
#     compare_hld(dict_1=restructured_dict3, dict_2=restructured_dict4, differences=differences_data,
#                 similarities=similarities_data, row_in_restructured_dict_key1='row_in_restructured_dict3',
#                 row_in_restructured_dict_key2='row_in_restructured_dict4',
#                 exclude_keys=['Test-Cases', 'Module', 'Componentname', 'Component Type', 'ParameterType',
#                               'Individual pop up headers', 'Individual pop up value', 'map view Operator',
#                               'map view Operator value', 'calculated csv value', 'Data validation', 'File',
#                               'combine export value'],
#                 fillvalue={'Test-Cases': '', 'Module': 'Result Not found', 'Component Type': 'Result Not found',
#                            'File': 'Result Not found', 'ParameterType': 'Result Not found',
#                            'Individual pop up headers': 'Result Not found',
#                            'Individual pop up value': 'Result Not found', 'combine export value': 'Result Not found',
#                            'map view Operator': 'Result Not found', 'map view Operator value': 'Result Not found',
#                            'calculated csv value': 'Result Not found', 'Data validation': 'Result Not found'},
#                 ignore_keys=['Test-Cases', 'File'])
#
#     create_dict_within_list_for_hld_components_status(df_dict=similarities, flattened_data_dict=flattened_data_dict1)
#     create_dict_within_list_for_hld_components_status(df_dict=differences, flattened_data_dict=flattened_data_dict2)
#
#     create_dict_within_list_for_hld_result_of_each_module(df_dict=similarities_data,flattened_data_dict=flattened_data_dict3)
#     create_dict_within_list_for_hld_result_of_each_module(df_dict=differences_data,flattened_data_dict=flattened_data_dict4)
#
#     dataframe_status1 = []
#     create_dataframe(flattened_data_dict=flattened_data_dict1, dataframe=dataframe_status1)
#     combined_status_df1 = pd.concat(dataframe_status1, ignore_index=True)
#
#     dataframe_status2 = []
#     create_dataframe(flattened_data_dict=flattened_data_dict2, dataframe=dataframe_status2)
#     combined_status_df2 = pd.concat(dataframe_status2, ignore_index=True)
#
#     dataframe_result1 = []
#     create_dataframe(flattened_data_dict=flattened_data_dict3, dataframe=dataframe_result1)
#     combined_result_df1 = pd.concat(dataframe_result1, ignore_index=True)
#
#     dataframe_result2 = []
#     create_dataframe(flattened_data_dict=flattened_data_dict4, dataframe=dataframe_result2)
#     combined_result_df2 = pd.concat(dataframe_result2, ignore_index=True)
#
#     workbook = Workbook()
#     workbook.create_sheet("COMPONENTSTATUS_SR", 0)
#     workbook.create_sheet("COMPONENTSTATUS_DR", 1)
#     workbook.create_sheet("RESULT OF EACH MODULE_SR", 2)
#     workbook.create_sheet("RESULT OF EACH MODULE_DR", 3)
#     workbook.save(excelpath_for_ouput)
#
#     add_headers_and_data(file_path=excelpath_for_ouput,headers=["Test_Set", "Test-Cases_1", "Test-Cases_2", "Different-Views_1", "Different-Views_2",'Componentname_1', 'Componentname_2', 'Status_1', 'Status_2', 'Comments_1','Comments_2'],sheet_name="COMPONENTSTATUS_SR")
#     add_headers_and_data(file_path=excelpath_for_ouput,headers=["Test_Set", "Test-Cases_1", "Test-Cases_2", "Different-Views_1", "Different-Views_2",'Componentname_1', 'Componentname_2', 'Status_1', 'Status_2', 'Comments_1','Comments_2'],sheet_name="COMPONENTSTATUS_DR")
#
#     workbook = openpyxl.load_workbook(excelpath_for_ouput)
#     worksheet_componentstatus1 = workbook["COMPONENTSTATUS_SR"]
#     worksheet_componentstatus2 = workbook["COMPONENTSTATUS_DR"]
#
#     if len(differences) != 0:
#         update_component_status_openpyxl(worksheet=worksheet_componentstatus2, dataframe=combined_status_df2)
#     if len(similarities) != 0:
#         update_component_status_openpyxl(worksheet=worksheet_componentstatus1, dataframe=combined_status_df1)
#     workbook.save(excelpath_for_ouput)
#     workbook.close()
#
#     app = xw.App(visible=False)
#     try:
#         workbook = app.books.open(excelpath_for_ouput)
#     except Exception as e:
#         pass
#     worksheet_result1 = workbook.sheets["RESULT OF EACH MODULE_SR"]
#     worksheet_result2 = workbook.sheets["RESULT OF EACH MODULE_DR"]
#     if len(differences_data) != 0:
#         updating_result_of_dataframe_for_excel(sheet=worksheet_result2, df=combined_result_df2)
#     if len(similarities_data) != 0:
#         updating_result_of_dataframe_for_excel(sheet=worksheet_result1, df=combined_result_df1)
#     try:
#         # Save and close the workbook
#         workbook.save(excelpath_for_ouput)
#         workbook.close()
#         app.quit()
#     except Exception as e:
#         pass
#
# def replace_and_sort(data_string):
#
#     # Check if the string contains [, { or (
#     if '{' in data_string or '(' in data_string:
#         # Replace all types of brackets with square brackets []
#         replaced_string = re.sub(r'[}\\)]', ']', data_string)
#         replaced_string = re.sub(r'[{\\(]', '[', replaced_string)
#
#         # Find all instances of square brackets [] and sort elements within them
#         sorted_string = re.sub(r'\[.*?\]', lambda x: "[" + ','.join(
#             sorted(re.findall(r"'[^']*'|[^{},\s]+", x.group().replace("[", ""))[:-1])) + "]", replaced_string)
#
#         return sorted_string
#     elif '[' in data_string:
#         replaced_string = data_string
#         sorted_string = re.sub(r'\[.*?\]', lambda x: "[" + ','.join(
#             sorted(re.findall(r"'[^']*'|[^{},\s]+", x.group().replace("[", ""))[:-1])) + "]", replaced_string)
#         return sorted_string
#     else:
#         return data_string
################################################################## HLD Comparison ###############################################################################################################################


###########################################HLD server comparision script###########################################################################################################
# Function to delete all files in a given directory
def delete_files_in_directory(directory):
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")

def move_file(source_file, destination_directory, new_file_name):

    # Ensure the destination directory exists
    os.makedirs(destination_directory, exist_ok=True)

    # Destination file path with the new name
    destination_file = os.path.join(destination_directory, new_file_name)

    # Move the file
    shutil.move(source_file, destination_file)

    print(f"File moved from {source_file} to {destination_file}")

def input(input_file, file_number):

    # Given list of partial strings
    list_of_different_views_to_ignore_N_drop = ["SCHEDULE TEST", "CONTINUOUS TEST", "REMOTE TEST", "GROUP"]

    # Load the Excel file
    wb1 = pd.ExcelFile(input_file)

    # Read each sheet into a dataframe
    dfs = {sheet_name: wb1.parse(sheet_name) for sheet_name in wb1.sheet_names}

    # Create a dictionary to store modified dataframes
    modified_dfs = {}

    # Iterate over each dataframe
    for sheet_name, df in dfs.items():
        # Check if 'Different-Views' column exists in the dataframe
        if 'Different-Views' in df.columns:
            # Drop rows where 'Different-Views' column matches specific values
            # modified_df = df[
            #     ~df['Different-Views'].isin(["SCHEDULE TEST", "CONTINUOUS TEST", "REMOTE TEST", "GROUP"])].copy()
            modified_df = df[~df['Different-Views'].isin(list_of_different_views_to_ignore_N_drop)].copy()
            modified_dfs[sheet_name] = modified_df
        else:
            # If 'Different-Views' column doesn't exist, just copy the original dataframe
            modified_dfs[sheet_name] = df.copy()

    # Save modified dataframes to a new Excel file
    output_folder = "Modified_Data"
    os.makedirs(output_folder, exist_ok=True)  # Create the output folder if it doesn't exist
    filename_prefix = f"{file_number}__"
    output_filename = os.path.join(output_folder, f"{filename_prefix}{os.path.basename(input_file).split('.')[0]}_Modified.xlsx")
    with pd.ExcelWriter(output_filename) as writer:
        for sheet_name, df in modified_dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Load the modified Excel file
    df_modified = pd.read_excel(output_filename)

    partials = ["_Dev", "_PreProd","Dow","_BT","_CCK","_vodafone","_ECSITE","_Live"]

    # Iterate through the partial strings
    for partial in partials:
        # Filter rows containing the partial string
        filtered_df = df_modified[df_modified['Test-Cases'].str.contains(partial)]
        # Check if filtered dataframe is not empty
        if not filtered_df.empty:
            # Write the dataframe to a separate Excel file
            output_filename_partial = os.path.join(output_folder, f"{filename_prefix}HLD_{partial}.xlsx")
            filtered_df.to_excel(output_filename_partial, index=False)
    wb1.close()
    # Remove the modified Excel file
    os.remove(output_filename)
    return output_folder




def concatenate_files(directory):
    # Get a list of all files in the directory
    files = os.listdir(directory)

    # Filter the files to include only those that start with "compare" and end with ".xlsx"
    compare_files = [file for file in files if file.startswith("compare") and file.endswith(".xlsx")]

    # Initialize an empty list to store DataFrames
    dfs = []

    # Read each file into a DataFrame and append it to the list
    for file in compare_files:
        df = pd.read_excel(os.path.join(directory, file))
        dfs.append(df)

    # Concatenate all DataFrames into a single DataFrame
    concatenated_df = pd.concat(dfs, ignore_index=True)

    # Save the concatenated DataFrame to a new Excel file
    concatenated_file_path = os.path.join(directory, "../concatenated_compare_files.xlsx")
    concatenated_df.to_excel(concatenated_file_path, index=False)

    print(f"All compare files concatenated and saved to: {concatenated_file_path}")
    return concatenated_df

# # Directory containing the files
# target_directory = "C:\\RantCell_Automation_Data_and_Reports\\HLD_Comparision_Input_Files"
#
# # Specify the target directory here
# source_directory= "C:\\RantCell_Automation_Data_and_Reports\\excelreport"

def comparision(files,directory,timestamped_folder):
    # copy_latest_hld_file(source_directory, target_directory)
    try:
        # Clear the 'Modified_Data' directory before processing new files
        modified_data_folder = "Modified_Data"
        delete_files_in_directory(modified_data_folder)
    except:pass

    # Process each input file
    for i, file in enumerate(files, start=1):
        output_folder = input(file, i)

    a=len(files)
    all_files=None
    if len(files) == 1:

        modified_data_folder =  os.path.join(output_folder)
        files_in_modified_data = os.listdir(modified_data_folder)
        if len(files_in_modified_data) == 2:
            path1 = os.path.join(modified_data_folder, files_in_modified_data[0])
            path2 = os.path.join(modified_data_folder, files_in_modified_data[1])
            path1  = str(path1).replace('\\', '/')
            path2  = str(path2).replace('\\', '/')

            compare = compare_HLD(path1, path2)
            new_file_name = f"compare_1.xlsx"
            # Move the compared file with the new name
            move_file(compare, "Modified_Data", new_file_name)
            result_df = pd.read_excel("Modified_Data//" + new_file_name)

        else:
            print("The generated HLD do not have two environment's data to compare")
            raise Exception
    else:
        modified_data_folder = os.path.join(output_folder)
        files_in_modified_data = os.listdir(modified_data_folder)

        # Make a copy of the list to avoid modifying it while iterating
        files_list = files_in_modified_data.copy()
        count = 1
        if len(files_list) == 2:
            path1 = os.path.join("Modified_Data", files_list[0])
            path2 = os.path.join("Modified_Data", files_list[1])
            path1 = str(path1).replace('\\', '/')
            path2 = str(path2).replace('\\', '/')

            files_list.remove(files_list[0])
            files_list.remove(files_list[0])

            compare = compare_HLD(path1, path2)
            new_file_name = f"compare_1.xlsx"
            # Move the compared file with the new name
            move_file(compare, "Modified_Data", new_file_name)
            # Directory containing the files
            directory = "Modified_Data"
            # Concatenate all compare files in the directory
            result_df = concatenate_files(directory)
        else:
            # Iterate over each file in the list
            for file in files_list:
                # Iterate over the other files in the list
                for other_file in files_list:
                    # Check if the files have the same suffix and are not the same file
                    if file[3:] == other_file[3:] and file != other_file:
                        # Remove the matching files from the list
                        files_list.remove(file)
                        files_list.remove(other_file)
                        path1 = os.path.join("Modified_Data",file)
                        path2 = os.path.join("Modified_Data", other_file)
                        compare = compare_HLD(path1, path2)
                        # Define the new file name based on your logic
                        new_file_name = f"compare_{count}.xlsx"
                        count = count + 1
                        # Move the compared file with the new name
                        move_file(compare, "Modified_Data", new_file_name)
                        # Directory containing the files
                        directory = "Modified_Data"
                        # Concatenate all compare files in the directory
                        result_df = concatenate_files(directory)
                        # Break out of the inner loop
                        break
        if files_list:
            concatenated_file_path = os.path.join("Modified_Data", "../concatenated_compare_files.xlsx")
            if os.path.exists(concatenated_file_path):
                # If the file exists, read it into all_files DataFrame
                all_files = pd.read_excel(concatenated_file_path)
            else:
                # If the file doesn't exist, initialize an empty DataFrame with predefined column names
                all_files_columns = ['Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments',
                                     'Test-Cases1', 'Different-Views1', 'Componentname1', 'Status1',
                                     'Comments1']  # Example column names
                all_files = pd.DataFrame(columns=all_files_columns)
            for unmatched_file in files_list:
                path_unmatched = os.path.join("Modified_Data", unmatched_file)
                df_unmatched = pd.read_excel(path_unmatched)

                if str(unmatched_file).startswith("1"):
                    # Select only the first 5 columns (A to E)
                    all_files_columns = all_files.columns[:5]
                    all_files_columns1 = all_files.columns[-5:]
                    new_df = all_files[all_files_columns].copy()
                    new_df1 = all_files[all_files_columns1].copy()
                    df_unmatched.columns = all_files_columns
                    combined_df = pd.concat([new_df, df_unmatched], ignore_index=True)
                    result_df = pd.merge(combined_df,new_df1, left_index=True, right_index=True, how='outer', suffixes=('_combined', '_new'))
                elif str(unmatched_file).startswith("2"):
                    # Select the last 5 columns (F to J)
                    all_files_columns = all_files.columns[-5:]
                    all_files_columns1 = all_files.columns[:5]
                    new_df = all_files[all_files_columns].copy()
                    new_df1 = all_files[all_files_columns1].copy()
                    df_unmatched.columns = all_files_columns
                    combined_df = pd.concat([new_df, df_unmatched], ignore_index=True)
                    result_df = pd.merge(new_df1, combined_df, left_index=True, right_index=True, how='outer',suffixes=('_combined', '_new'))
                else:
                    raise ValueError("Invalid prefix value")
        # print(f"All unmatched files concatenated and saved to: {concatenated_file_path}")



    new_column_names = ['Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments',
                        'Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments']

    # Update the column names in the DataFrame
    result_df.columns = new_column_names
    result_df['Observation'] = result_df.apply(
        lambda row: f"Different: {', '.join(str(x) for x in row['Status'].unique() if pd.notnull(x))}" if not (row['Status'] == row['Status'].iloc[0]).all() else 'Same', axis=1)

    dropping_duplicates=result_df.drop_duplicates()
    # Write the concatenated DataFrame to an Excel file
    concatenated_file_path = os.path.join("Modified_Data", "../concatenated_compare_files.xlsx")
    dropping_duplicates.to_excel(concatenated_file_path, index=False)
    print(dropping_duplicates)
    # below script is to create a directory and move the outpput results along with the input file to the timestamp directory
    # files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]
    output_file_path = os.path.join("Modified_Data", '../concatenated_compare_files.xlsx')
    # timestamped_folder = None
    # Check the number of files
    if len(files) == 1:
        # If only one file present, move that file
        file_to_move = os.path.join(directory, files[0])
        # timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # timestamped_folder = os.path.join(directory, "Hld_comparison_" + timestamp)
        # os.makedirs(timestamped_folder)
        # print(timestamped_folder)
        shutil.copy(file_to_move, timestamped_folder)
        shutil.move(output_file_path, timestamped_folder)
        shutil.rmtree("Modified_Data")
    elif len(files) == 2:
        # If two files present, move both files
        # timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # timestamped_folder = os.path.join(directory, "Hld_comparison_" + timestamp)
        # os.makedirs(timestamped_folder)
        # print(timestamped_folder)
        for file_name in files:
            file_to_move = os.path.join(directory, file_name)
            shutil.move(file_to_move, timestamped_folder)

        shutil.move(output_file_path, timestamped_folder)
        shutil.rmtree("Modified_Data")
    else:
        print("No files or more than two files found in the directory.")
    finding_delta_from_concatinated_file(file_path=timestamped_folder+"\\concatenated_compare_files.xlsx",fileappendpath=timestamped_folder+"\\append_.xlsx")

    merging_both_concatinated_and_appended_sheet_data(filepath=timestamped_folder+"\\concatenated_compare_files.xlsx",fileappend=timestamped_folder+"\\append_.xlsx",finalfilepath=timestamped_folder+"\\comprasion_result.xlsx")
    finalfilepath =timestamped_folder+"\\comprasion_result.xlsx"
    # Apply formatting to the Excel file
    workbook = openpyxl.load_workbook(finalfilepath)
    sheet = workbook["Sheet1"]

    # Define fill colors
    green_fill = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')
    orange_fill = PatternFill(start_color='00FFCC99', end_color='00FFCC99', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    pale_violet_fill = PatternFill(start_color='FFCC99FF', end_color='FFCC99FF', fill_type='solid')

    # Color the first 5 columns in light green
    for col in range(1, 6):
        column_letter = openpyxl.utils.get_column_letter(col)
        for row in range(1, sheet.max_row + 1):
            sheet[column_letter + str(row)].fill = green_fill

    # Color the next 5 columns in light orange
    for col in range(6, 11):
        column_letter = openpyxl.utils.get_column_letter(col)
        for row in range(1, sheet.max_row + 1):
            sheet[column_letter + str(row)].fill = orange_fill

    # # Color to 10th column
    for col in range(11, 12):
        column_letter = openpyxl.utils.get_column_letter(col)
        for row in range(1, sheet.max_row + 1):
            sheet[column_letter + str(row)].fill = pale_violet_fill

    # Color the first row from column A to H in yellow
    for col in range(1, 12):
        column_letter = openpyxl.utils.get_column_letter(col)
        cell = sheet[column_letter + '1']
        cell.fill = yellow_fill

    # Set column width for specific columns
    sheet.column_dimensions['A'].width = 49
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 17
    sheet.column_dimensions['D'].width = 7
    sheet.column_dimensions['E'].width = 55
    sheet.column_dimensions['F'].width = 16
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['H'].width = 17
    sheet.column_dimensions['i'].width = 23
    sheet.column_dimensions['j'].width = 55
    sheet.column_dimensions['k'].width = 22

    # Define border styles
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # Apply borders to all cells
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border_style

    # Save the formatted Excel file
    workbook.save(finalfilepath)
    if os.path.exists(timestamped_folder + "\\concatenated_compare_files.xlsx"):
        os.remove(timestamped_folder + "\\concatenated_compare_files.xlsx")
    if os.path.exists(timestamped_folder + "\\append_.xlsx"):
        os.remove(timestamped_folder + "\\append_.xlsx")

###########################################################################################################################
def compare_HLD(path1, path2):

    if os.path.isfile('comparison_results.xlsx'):
        os.remove('comparison_results.xlsx')

    wb1 = openpyxl.load_workbook(path1)
    wb2 = openpyxl.load_workbook(path2)

    # Get the first sheet from each workbook
    sheet1 = wb1.active
    sheet2 = wb2.active
    # Create dataframes for each sheet
    df1 = pd.DataFrame(sheet1.values)
    df2 = pd.DataFrame(sheet2.values)

    # Reset the index to set proper column headers
    df1.columns = df1.iloc[0]
    df2.columns = df2.iloc[0]
    # Drop the first row (which is now the redundant header row)
    df1 = df1.drop(0)
    df2 = df2.drop(0)
    # import re  # Move the import statement to the top

    def preprocess_componentname(value):
        # Define a regex pattern to match the date and time component
        pattern = r'_\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}\.csv'
        # Use regex to replace the date and time component with an empty string
        return re.sub(pattern, '', value)

    # Preprocess the 'Componentname' column in both dataframes
    df1['Componentname'] = df1['Componentname'].apply(preprocess_componentname)
    df2['Componentname'] = df2['Componentname'].apply(preprocess_componentname)

    # Preprocess the 'Componentname' column in both dataframes
    df1['Comments'] = df1['Comments'].apply(preprocess_componentname)
    df2['Comments'] = df2['Comments'].apply(preprocess_componentname)


    def replace_and_sort(data_string):
            # Check if the string contains [, { or (
        if '{' in data_string or '(' in data_string:
            # Replace all types of brackets with square brackets []
            replaced_string = re.sub(r'[}\\)]', ']', data_string)
            replaced_string = re.sub(r'[{\\(]', '[', replaced_string)

            # Find all instances of square brackets [] and sort elements within them
            sorted_string = re.sub(r'\[.*?\]', lambda x: "[" + ','.join(
                sorted(re.findall(r"'[^']*'|[^{},\s]+", x.group().replace("[", ""))[:-1])) + "]", replaced_string)

            return sorted_string
        elif '[' in data_string:
            replaced_string = data_string
            sorted_string = re.sub(r'\[.*?\]', lambda x: "[" + ','.join(
                sorted(re.findall(r"'[^']*'|[^{},\s]+", x.group().replace("[", ""))[:-1])) + "]", replaced_string)
            return sorted_string
        else:
            return data_string
        # Apply preprocess_comments to the 'Comments' column in both dataframes

    df1['Comments'] = df1['Comments'].apply(replace_and_sort)
    df2['Comments'] = df2['Comments'].apply(replace_and_sort)

    # Initialize an empty list to store rows
    rows = []

    # Iterate through rows of the first dataframe
    for index, row in df1.iterrows():

        # Here we are expecting to keep the actual values before modifying the below row['Test-Cases']
        row_early = row.copy()
        if 'PreProd' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_PreProd')[0]
        elif 'Dev' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_Dev')[0]
        elif 'Dow' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_Dow')[0]
        elif 'Live' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_Live')[0]
        elif 'awz' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_awz')[0]
        elif 'BT' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_BT')[0]
        elif 'CCK' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_CCK')[0]
        elif 'vodafone' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_vodafone')[0]
        elif 'ECSITE' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_ECSITE')[0]

        # Check if the row is present in the second dataframe
        matching_row_df2 = df2[df2['Test-Cases'].str.startswith(row['Test-Cases'])]

        concatenated_row_df2 = []
        if not matching_row_df2.empty:
            # Check if the rows in df2 match the same Different-Views and Componentname
            matching_rows = matching_row_df2[(matching_row_df2['Different-Views'] == row['Different-Views']) &
                                             (matching_row_df2['Componentname'] == row['Componentname']) &
                                             (matching_row_df2['Comments'] == row['Comments'])]
            if not matching_rows.empty:
                # Take the first matching row (if multiple)
                matching_row_df2 = matching_rows.iloc[0]

                # Concatenate the rows with common columns
                concatenated_row_df2 = pd.concat([row_early, matching_row_df2], ignore_index=True)
            else:
                concatenated_row_df2 = pd.concat([row_early, pd.Series([None] * len(df2.columns))], ignore_index=True)
        else:
            concatenated_row_df2 = pd.concat([row_early, pd.Series([None] * len(df2.columns))], ignore_index=True)

        # Append the concatenated row to the list
        rows.append(concatenated_row_df2)

    # Iterate through rows of the first dataframe
    for index, row in df2.iterrows():

        # Here we are expecting to keep the actual values before modifying the below row['Test-Cases']
        row_early = row.copy()
        if 'PreProd' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_PreProd')[0]
        elif 'Dev' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_Dev')[0]
        elif 'Dow' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_Dow')[0]
        elif 'Live' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_Live')[0]
        elif 'awz' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_awz')[0]
        elif 'BT' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_BT')[0]
        elif 'CCK' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_CCK')[0]
        elif 'vodafone' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_vodafone')[0]
        elif 'ECSITE' in str(row['Test-Cases']):
            row['Test-Cases'] = row['Test-Cases'].split('_ECSITE')[0]

        # Check if the row is present in the second dataframe
        matching_row_df1 = df1[df1['Test-Cases'].str.startswith(row['Test-Cases'])]

        concatenated_row_df1 = []
        if not matching_row_df1.empty:
            # Check if the rows in df2 match the same Different-Views and Componentname
            matching_rows = matching_row_df1[(matching_row_df1['Different-Views'] == row['Different-Views']) &
                                             (matching_row_df1['Componentname'] == row['Componentname']) &
                                             (matching_row_df1['Comments'] == row['Comments'])]
            if not matching_rows.empty:
                # Take the first matching row (if multiple)
                matching_row_df1 = matching_rows.iloc[0]

                # Concatenate the rows with common columns
                concatenated_row_df1 = pd.concat([matching_row_df1, row_early], ignore_index=True)
            else:
                concatenated_row_df1 = pd.concat([pd.Series([None] * len(df1.columns)),row_early], ignore_index=True)
        else:
            concatenated_row_df1 = pd.concat([row_early, pd.Series([None] * len(df1.columns))], ignore_index=True)

        # Append the concatenated row to the list
        rows.append(concatenated_row_df1)

    # Create a new dataframe from the list of rows and reset index
    result_df = pd.DataFrame(rows)
    path = pathlib.Path(__file__).parent.parent
    comparison_results_path  = os.path.join(path,'comparison_results.xlsx')
    # Save the common, difference, and concatenated full data to a new Excel file
    with pd.ExcelWriter(comparison_results_path) as writer:
        result_df.to_excel(writer, sheet_name='Concatenated_Full', index=False)

    return comparison_results_path



#################################################################################################################################################

def finding_delta_from_concatinated_file(file_path,fileappendpath):
    # Load the Excel file
    df = pd.read_excel(file_path)

    # Adjust these column names if pandas uses different names for your file
    first_test_cases_col = 'Test-Cases'
    second_test_cases_col = 'Test-Cases.1'

    # Filter rows where either the first 'Test-Cases' is empty or the second 'Test-Cases' is empty
    filtered_df = df[df[first_test_cases_col].isna() | df[second_test_cases_col].isna()]
    if not filtered_df.empty:
        # Now, you can do whatever you need with filtered_df, such as inspecting it or saving it to a new file.
        # For example, to save the filtered rows to a new Excel file:
        output_file_path = fileappendpath
        filtered_df.to_excel(output_file_path, index=False)

        # file_path = '../filtered_data_both_sides.xlsx'
        df_filtered = pd.read_excel(output_file_path)


        # Split the DataFrame into two separate DataFrames
        df_first_5 = df_filtered.iloc[:, :5]  # First 5 columns
        df_last_5 = df_filtered.iloc[:, 5:10]


        df_first_5 = df_first_5.dropna()
        df_last_5 = df_last_5.dropna()

        new_column_names = ['Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments']

        # Update the column names in the DataFrame
        df_last_5.columns = new_column_names
        import re  # Move the import statement to the top

        def preprocess_componentname(value):
            # Define a regex pattern to match the date and time component
            pattern = r'_\d{4}-\d{2}-\d{2} \d{2}_\d{2}_\d{2}\.csv'
            # Use regex to replace the date and time component with an empty string
            return re.sub(pattern, '', value)

        # Preprocess the 'Componentname' column in both dataframes
        df_first_5['Componentname'] = df_first_5['Componentname'].apply(preprocess_componentname)
        df_last_5['Componentname'] = df_last_5['Componentname'].apply(preprocess_componentname)


        # Initialize an empty list to store rows
        rows = []

        # Iterate through rows of the first DataFrame
        for index, row in df_first_5.iterrows():
            # Here we are expecting to keep the actual values before modifying the below row['Test-Cases']
            row_early = row.copy()
            if 'PreProd' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_PreProd')[0]
            elif 'Dev' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_Dev')[0]
            elif 'Dow' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_Dow')[0]
            elif 'Live' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_Live')[0]
            elif 'awz' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_awz')[0]
            elif 'BT' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_BT')[0]
            elif 'CCK' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_CCK')[0]
            elif 'vodafone' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_vodafone')[0]
            elif 'ECSITE' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_ECSITE')[0]

                # Check if the row is present in the second DataFrame
            matching_row_df2 = df_last_5[df_last_5['Test-Cases'].str.startswith(row['Test-Cases'])]

            concatenated_row_df2 = []
            if not matching_row_df2.empty:
                # Check if the rows in df2 match the same Different-Views and Componentname
                matching_rows = matching_row_df2[
                    (matching_row_df2['Different-Views'] == row['Different-Views']) &
                    (matching_row_df2['Componentname'] == row['Componentname']) &
                    (matching_row_df2['Status'] != row['Status'])]

                if not matching_rows.empty:
                    # Take the first matching row (if multiple)
                    matching_row_df2 = matching_rows.iloc[0]

                    # Concatenate the rows with common columns
                    concatenated_row_df2 = pd.concat([row_early, matching_row_df2], ignore_index=True)
                else:
                    concatenated_row_df2 = pd.concat([row_early, pd.Series([None] * len(df_last_5.columns))], ignore_index=True)
            else:
                concatenated_row_df2 = pd.concat([row_early, pd.Series([None] * len(df_last_5.columns))], ignore_index=True)

            # Append the concatenated row to the list
            rows.append(concatenated_row_df2)

        for index, row in df_last_5.iterrows():
            # Here we are expecting to keep the actual values before modifying the below row['Test-Cases']
            row_early = row.copy()
            if 'PreProd' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_PreProd')[0]
            elif 'Dev' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_Dev')[0]
            elif 'Dow' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_Dow')[0]
            elif 'Live' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_Live')[0]
            elif 'awz' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_awz')[0]
            elif 'BT' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_BT')[0]
            elif 'CCK' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_CCK')[0]
            elif 'vodafone' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_vodafone')[0]
            elif 'ECSITE' in str(row['Test-Cases']):
                row['Test-Cases'] = row['Test-Cases'].split('_ECSITE')[0]

                # Check if the row is present in the second DataFrame
            matching_row_df1 = df_first_5[df_first_5['Test-Cases'].str.startswith(row['Test-Cases'])]

            concatenated_row_df1 = []
            if not matching_row_df1.empty:
                # Check if the rows in df2 match the same Different-Views and Componentname
                matching_rows = matching_row_df1[
                    (matching_row_df1['Different-Views'] == row['Different-Views']) &
                    (matching_row_df1['Componentname'] == row['Componentname']) &
                    (matching_row_df1['Status'] != row['Status'])]

                if not matching_rows.empty:
                    # Take the first matching row (if multiple)
                    matching_row_df1 = matching_rows.iloc[0]

                    # Concatenate the rows with common columns
                    concatenated_row_df1 = pd.concat([matching_row_df1, row_early], ignore_index=True)
                else:
                    concatenated_row_df1 = pd.concat([pd.Series([None] * len(df_first_5.columns)), row_early], ignore_index=True)
            else:
                concatenated_row_df1 = pd.concat([row_early, pd.Series([None] * len(df_first_5.columns))], ignore_index=True)

            # Append the concatenated row to the list
            rows.append(concatenated_row_df1)

        result_df = pd.DataFrame(rows)
        new_column_names = ['Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments',
                                'Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments']

        # Update the column names in the DataFrame
        result_df.columns = new_column_names
        dropping_duplicates=result_df.drop_duplicates()
        # Save the appended data to a new Excel file
        appended_file_path = fileappendpath
        dropping_duplicates.to_excel(appended_file_path, index=False)
        # Load the Excel file from the specified directory
        wb = openpyxl.load_workbook(appended_file_path)
        sheet = wb['Sheet1']

        # Define fill colors
        green_fill = PatternFill(start_color='FFCCFFCC', end_color='FFCCFFCC', fill_type='solid')
        orange_fill = PatternFill(start_color='00FFCC99', end_color='00FFCC99', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        pale_violet_fill = PatternFill(start_color='FFCC99FF', end_color='FFCC99FF', fill_type='solid')

        # Color the first 5 columns in light green
        for col in range(1, 6):
            column_letter = openpyxl.utils.get_column_letter(col)
            for row in range(1, sheet.max_row + 1):
                sheet[column_letter + str(row)].fill = green_fill

        # Color the next 5 columns in light orange
        for col in range(6, 11):
            column_letter = openpyxl.utils.get_column_letter(col)
            for row in range(1, sheet.max_row + 1):
                sheet[column_letter + str(row)].fill = orange_fill

        # # Color to 10th column
        for col in range(11, 12):
            column_letter = openpyxl.utils.get_column_letter(col)
            for row in range(1, sheet.max_row + 1):
                sheet[column_letter + str(row)].fill = pale_violet_fill

        # Color the first row from column A to H in yellow
        for col in range(1, 12):
            column_letter = openpyxl.utils.get_column_letter(col)
            cell = sheet[column_letter + '1']
            cell.fill = yellow_fill

        # Set column width for specific columns
        sheet.column_dimensions['A'].width = 49
        sheet.column_dimensions['B'].width = 16
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 7
        sheet.column_dimensions['E'].width = 55
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 17
        sheet.column_dimensions['i'].width = 23
        sheet.column_dimensions['j'].width = 55
        sheet.column_dimensions['k'].width = 22

        # Define border styles
        border_style = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        # Apply borders to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style

        # Save the changes to the Excel file
        wb.save(appended_file_path)
    else:
        print("concatenated compare file does not contain empty rows")
        pass

# finding_delta_from_concatinated_file()

###############################################################################################################################################
# import os.path
#
# import pandas as pd
def merging_both_concatinated_and_appended_sheet_data(filepath,fileappend,finalfilepath):
    # Load the source Excel file, skipping empty rows.
    source_df = pd.read_excel(filepath, sheet_name='Sheet1').dropna(how='any')

    # Load the target Excel file.
    if os.path.exists(fileappend):
        target_df = pd.read_excel(fileappend, sheet_name='Sheet1')
        # Assuming you want to append source data at the end of the target file
        merged_df = pd.concat([target_df, source_df], ignore_index=True)
    else:
        merged_df = source_df

    new_column_names = ['Test-Cases', 'Different-Views', 'Componentname', 'Status', 'Comments',
                            'Test-Cases.1', 'Different-Views', 'Componentname', 'Status', 'Comments','Observation']

    # Update the column names in the DataFrame
    merged_df.columns = new_column_names
    # merged_df = pd.merge(target_df, source_df,how='inner')
    merged_df['Observation'] = merged_df.apply(
            lambda row: f"Different: {', '.join(str(x) for x in row['Status'].unique() if pd.notnull(x))}" if not (row['Status'] == row['Status'].iloc[0]).all() else 'Same', axis=1)
    # Replace 'ColumnName' with the actual column name you want to sort by.
    sorted_merged_df = merged_df.sort_values(by='Test-Cases', ascending=False)
    # Save the merged dataframe back to the target Excel file.
    sorted_merged_df.to_excel(finalfilepath, sheet_name='Sheet1', index=False, engine='openpyxl')

    print('Data merged successfully into target_updated.xlsx')
######################################################################################################################################################################################################################

def copy_and_rename(src,appendword):
    """
    Copy the file from src to the same directory with a modified filename.
    Appends '_change' before the file extension in the filename.
    """
    # Extract the directory, filename, and file extension from the source path

    directory = os.path.dirname(src)

    filename, extension = os.path.splitext(os.path.basename(src))
    # Construct the new filename with '_change' appended before the extension

    new_filename = f"{filename}_{appendword}{extension}"
    # Join the new filename with the source directory

    new_dst = os.path.join(directory, new_filename)
    # Copy the file to the new destination

    shutil.copyfile(src, new_dst)
    return new_dst
#####################################################################################################################################################################


