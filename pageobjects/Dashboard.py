import json
import re, allure
import pandas as pd
import pytest
from locators.locators import side_menu_Components, close_button,date_time
from locators.locators import Map_View_Select_and_ReadData
from locators.locators import hover, hover_over_second_graph, get_graph_data, get_secondGraph_data, Graph_View_Components,load_locators,individual_pop_table
from locators.locators import expand_tableView_verify_popUp, List_Of_Campaigns_Export_Dashboard, List_Of_Campaigns_components_Search_Box_not_visible_do_page_up,table_summary,nw_freeze,pdf_view,operator_comparison_table
from pageobjects.remote_test import *
from utils.createFolderforRantcell_automation_DataandReports import create_folder_for_downloads
from utils.readexcel import *
from pageobjects.login_logout import *
from openpyxl.utils.dataframe import dataframe_to_rows
import concurrent.futures
from configurations.config import ReadConfig as config
import queue
from decimal import Decimal, ROUND_HALF_UP
import win32com.client as win32
import calendar
def side_menu_Components_(driver, device, campaign, userid, password,excelpath):
    try:
        Variable_MobileDevice_Xpath = (By.XPATH, f"//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        protestdata_runvalue = Testrun_mode(value="Pro TestData")
        litetestdata_runvalue = Testrun_mode(value="LITE TestData")
        if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='Pro TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='LITE TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        classifier = (By.XPATH, "//span[normalize-space()='"+str(campaign)+"']", str(campaign))
        click_on_side_bar_menu_compnents(driver, device, userid, password, Variable_MobileDevice_Xpath, classifier)
        count_find_the_campaign = searching_visibility_of_campaigns_by_driver_refresh(driver, classifier, device, userid, password,Variable_MobileDevice_Xpath)
        click_on_campaigns(driver, device, classifier, campaign, count_find_the_campaign, excelpath)
    except Exception as e:
        with allure.step("Failed in side bar menu"):
            allure.attach(driver.get_screenshot_as_png(), name=f"Failed in side bar menu",attachment_type=allure.attachment_type.PNG)
            pytest.fail(str(e))
def searching_visibility_of_campaigns_by_driver_refresh(driver,classifier,device, userid, password,Variable_MobileDevice_Xpath):
    action = ActionChains(driver)
    count_find_the_campaign = 0
    try:
        classifier_elements= driver.find_elements(classifier[0],classifier[1])
        if len(classifier_elements) == 0:
            for i in range(0,3):
                try:
                    with allure.step(f"Attempt for driver refresh for campagins to load:- {i}"):
                        if i == 0:
                            driver.refresh()
                            time.sleep(5)
                            dashboard_loading(driver)
                            count_find_the_campaign = 1
                            click_on_side_bar_menu_compnents(driver, device, userid, password,Variable_MobileDevice_Xpath, classifier)
                        elif i == 1:
                            logout(driver)
                            time.sleep(1.2)
                            clickec(driver, Login_Logout.link_login)
                            time.sleep(1.2)
                            login(driver, userid, password)
                            time.sleep(1.2)
                            count_find_the_campaign = 2
                            click_on_side_bar_menu_compnents(driver, device, userid, password,Variable_MobileDevice_Xpath, classifier)
                        elif i == 2:
                            logout(driver)
                            time.sleep(1.2)
                            clickec(driver, Login_Logout.link_login)
                            time.sleep(1.2)
                            login(driver, userid, password)
                            driver.refresh()
                            time.sleep(5)
                            dashboard_loading(driver)
                            count_find_the_campaign = 3
                            click_on_side_bar_menu_compnents(driver, device, userid, password,Variable_MobileDevice_Xpath, classifier)
                        allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for driver refresh for campagins to load:- {i}",attachment_type=allure.attachment_type.PNG)
                        classifier_elements = driver.find_elements(classifier[0], classifier[1])
                        for classifier_element in classifier_elements:
                            action.move_to_element(classifier_element).perform()
                            break
                        if len(classifier_elements) != 0:
                            break
                except:
                    continue
        classifier_elements = driver.find_elements(classifier[0], classifier[1])
        for classifier_element in classifier_elements:
            action.move_to_element(classifier_element).perform()
            break
    except:
        pass
    finally:
        return count_find_the_campaign
def click_on_campaigns(driver,device,classifier,campaign,count_find_the_campaign,excelpath):
    try:
        assert clickec(driver, classifier)
        if count_find_the_campaign == 0:
            updatecomponentstatus("Side bar menu",str(campaign), "PASSED", f"{campaign} is found in {count_find_the_campaign} Attempt without Driver refresh.",excelpath)
        elif count_find_the_campaign == 1 or count_find_the_campaign == 2 or count_find_the_campaign == 3:
            updatecomponentstatus("Side bar menu",str(campaign), "FAILED", f"{campaign} is found in {count_find_the_campaign} Attempt after Driver refresh.",excelpath)
    except Exception as e:
        updatecomponentstatus("Side bar menu",str(campaign), "FAILED", f"Failed click on the {campaign} and check {campaign} is present in this {device} device",excelpath)
        if count_find_the_campaign == 1 or count_find_the_campaign == 2 or count_find_the_campaign == 3:
            updatecomponentstatus("Side bar menu", str(campaign), "FAILED",f"{campaign} is not found in {count_find_the_campaign} Attempt after Driver refresh.",excelpath)
        format_workbook(excelpath)
        raise e
    try:
        wait_for_loading_elements(driver)
    except:
        pass
    allure.attach(driver.get_screenshot_as_png(),name=f"Selected the device {str(device)} ==>> {str(campaign)} successfully",attachment_type=allure.attachment_type.PNG)

def side_bar_menu_for_work_list_campaigns(driver,userid, password,campaigns_data,excelpath):
    global campaign1, campaign2
    for i in range(len(campaigns_data)):
        # previous_device = []
        device, campaign, usercampaignsname, testgroup = campaigns_data[i]
        if compare_values(usercampaignsname, 'None'):
            campaign1 = campaign
            campaign2 = campaign
        elif not compare_values(usercampaignsname, 'None'):
            campaign1 = usercampaignsname
            campaign2 = usercampaignsname + campaign
        print(i, campaigns_data[i])
        Variable_MobileDevice_Xpath = (By.XPATH, f"//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        protestdata_runvalue = Testrun_mode(value="Pro TestData")
        litetestdata_runvalue = Testrun_mode(value="LITE TestData")
        if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH,f"//span[text()='Pro TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']",str(device))
        elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH,f"//span[text()='LITE TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']",str(device))
        classifier = (By.XPATH, "//span[normalize-space()='" + str(campaign1) + "']", str(campaign2))
        active_element = (By.XPATH,f"//ul[@class='treeview-menu style-1 menu-open']//li[@class='treeview ng-scope active']//a[normalize-space()='{str(device)}']")
        if i == 0:
            # previous_device.append(device)
            click_on_side_bar_menu_compnents(driver, device, userid, password, Variable_MobileDevice_Xpath, classifier)
            count_find_the_campaign = searching_visibility_of_campaigns_by_driver_refresh(driver, classifier, device,userid, password,Variable_MobileDevice_Xpath)
        elif i > 0:
            # if device in previous_device:
            #     previous_device.clear()
            # elif device not in previous_device:
            time.sleep(5)
            # MobileDevice_element = None
            try:
                MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0],Variable_MobileDevice_Xpath[1])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                # action = ActionChains(driver)
                # action.move_to_element(MobileDevice_element).perform()
                # simulate_screen_touch(100, 100)
            except:
                pass
            click_until_visible_or_active(driver, Variable_MobileDevice_Xpath, active_element)
            try:
                wait_for_loading_elements(driver)
            except:
                pass
            # previous_device.clear()
            # previous_device.append(device)
            search_campaigns(driver,classifier)
        # print(previous_device)
        try:
            result = clickec(driver, classifier)
            time.sleep(3)
            try:
                MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0],Variable_MobileDevice_Xpath[1])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                driver.execute_script("arguments[0].click();", MobileDevice_element)

                # action = ActionChains(driver)
                # action.move_to_element(MobileDevice_element).perform()
                # simulate_screen_touch(100, 100)
            except:
                pass
            # clickec(driver, Variable_MobileDevice_Xpath)
            # simulate_screen_touch(100, 100)
            if result == False:
                e = Exception
                raise e
        except Exception as e:
            updatecomponentstatus("Side bar menu",str(campaign), "FAILED", f"Failed click on the {campaign} and check {campaign} is present in this {device} device",excelpath)
            format_workbook(excelpath)
            # raise e
def click_until_visible_or_active(driver, Variable_MobileDevice_Xpath, active_element_xpath):
    while True:
        try:
            time.sleep(1)
            # clickec(driver, Variable_MobileDevice_Xpath)
            try:
                MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0],Variable_MobileDevice_Xpath[1])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                driver.execute_script("arguments[0].click();", MobileDevice_element)
                # action = ActionChains(driver)
                # action.move_to_element(MobileDevice_element).perform()
            except:pass
            # Wait for the active element to be visible or active
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(active_element_xpath))
            # Once the active element is visible or active, break out of the loop
            break
        except Exception as e:
            print("Exception occurred:", e)

def wait_for_loading_elements(driver):
    timeout = 10
    # Define the loading element locators
    wait = WebDriverWait(driver, timeout)
    with allure.step("Waiting for application to load complete"):
        try:
            for by, locator in load_locators.loading_locators:
                wait.until(EC.invisibility_of_element_located((by, locator)))
        except Exception as e:
            allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for application to load completely",attachment_type=allure.attachment_type.PNG)
            print("An error occurred while waiting for loading elements:", str(e))

def click_on_androidtestdata(driver):
    for i in range(0,2):
        try:
           with allure.step(f"Attempt for androidtestdata :- {i}"):
                if i > 0:
                   driver.refresh()
                   dashboard_loading(driver)
                time.sleep(1.2)
                androidtestdata_element = driver.find_elements(side_menu_Components.androidtestdata[0],side_menu_Components.androidtestdata[1])
                if len(androidtestdata_element)!=0:
                    assert clickec(driver, side_menu_Components.androidtestdata)
                    time.sleep(1.2)
                    protestdata_runvalue = Testrun_mode(value="Pro TestData")
                    if protestdata_runvalue[0].lower() == 'Yes'.lower():
                        protestdata_element = driver.find_elements(side_menu_Components.protestdata[0],side_menu_Components.protestdata[1])
                        if len(protestdata_element) != 0:
                            break
                    elif protestdata_runvalue[0].lower() == 'No'.lower():
                        litetestdata_element = driver.find_elements(side_menu_Components.litetestdata[0],side_menu_Components.litetestdata[1])
                        if len(litetestdata_element) != 0:
                            break
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for androidtestdata :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue
def click_on_protestdata(driver):
    for i in range(0,2):
        try:
            with allure.step(f"Attempt for protestdata :- {i}"):
                if i > 0:
                   driver.refresh()
                   dashboard_loading(driver)
                   click_on_androidtestdata(driver)
                time.sleep(1.2)
                protestdata_element = driver.find_elements(side_menu_Components.protestdata[0],side_menu_Components.protestdata[1])
                if len(protestdata_element) != 0:
                    assert clickec(driver, side_menu_Components.protestdata)
                    time.sleep(1.2)
                    device_element = driver.find_elements(*side_menu_Components.device_element)
                    if len(device_element) != 0:
                        break
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for protestdata :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue
def click_on_litetestdata(driver):
    for i in range(0,2):
        try:
            with allure.step(f"Attempt for litetestdata :- {i}"):
                if i > 0:
                   driver.refresh()
                   dashboard_loading(driver)
                   click_on_androidtestdata(driver)
                time.sleep(1.2)
                litetestdata_element = driver.find_elements(side_menu_Components.litetestdata[0],side_menu_Components.litetestdata[1])
                if len(litetestdata_element) != 0:
                    assert clickec(driver, side_menu_Components.litetestdata)
                    time.sleep(1.2)
                    device_element = driver.find_elements(*side_menu_Components.device_element)
                    if len(device_element) != 0:
                        break
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for litetestdata :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue
def click_on_device(driver,userid,password,Variable_MobileDevice_Xpath,device):
    protestdata_runvalue = Testrun_mode(value="Pro TestData")
    litetestdata_runvalue = Testrun_mode(value="LITE TestData")
    for i in range(0,3):
        try:
            with allure.step(f"Attempt for device :- {i}"):
                if i == 1:
                   driver.refresh()
                   dashboard_loading(driver)
                   click_on_androidtestdata(driver)
                   if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
                       click_on_protestdata(driver)
                   elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
                       click_on_litetestdata(driver)
                elif i == 2:
                   logout(driver)
                   time.sleep(1.2)
                   clickec(driver, Login_Logout.link_login)
                   time.sleep(1.2)
                   login(driver, userid, password)
                   time.sleep(1.2)
                   click_on_androidtestdata(driver)
                   if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
                       click_on_protestdata(driver)
                   elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
                       click_on_litetestdata(driver)
                time.sleep(1.2)
                device_elements = driver.find_elements(By.XPATH, f"//a[normalize-space()='{str(device)}']")
                if len(device_elements) != 0:
                    assert clickec(driver, Variable_MobileDevice_Xpath)
                    time.sleep(10)
                    camp_elements = driver.find_elements(*side_menu_Components.campaign_element)
                    if len(camp_elements) != 0:
                        break
                time.sleep(3)
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for device :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue

def click_on_side_bar_menu_compnents(driver,device,userid,password,Variable_MobileDevice_Xpath,classifier):
    classifier_element = "None"
    click_on_androidtestdata(driver)
    protestdata_runvalue = Testrun_mode(value="Pro TestData")
    litetestdata_runvalue = Testrun_mode(value="LITE TestData")
    if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
        click_on_protestdata(driver)
    elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
        click_on_litetestdata(driver)
    try:
        MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0], Variable_MobileDevice_Xpath[1])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
        # action = ActionChains(driver)
        # action.move_to_element(MobileDevice_element).perform()
        # simulate_screen_touch(100, 100)
    except:
        pass
    click_on_device(driver,userid,password,Variable_MobileDevice_Xpath,device)
    time.sleep(3.2)
    try:
        MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0], Variable_MobileDevice_Xpath[1])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
        # action = ActionChains(driver)
        # action.move_to_element(MobileDevice_element).perform()
        # simulate_screen_touch(100, 100)
    except:
        pass
    search_campaigns(driver, classifier)
    time.sleep(2.2)
    assert uncheck_listOfcampaign(driver, side_menu_Components.campaignCheckBox)
    try:
        allure.attach(driver.get_screenshot_as_png(), name=f"Successfully unselected 'List of Campaign' checkbox",attachment_type=allure.attachment_type.PNG)
    except:
        pass
def search_campaigns(driver,classifier):
    action = ActionChains(driver)
    result = True
    try:
        elements = driver.find_elements(By.TAG_NAME, "a")
        for x in elements:
            if x.text == "Show More":
                result = True
    except Exception as e:
        pass
    max_run_time = 60
    start_time = time.time()
    while result == True:
        result = False
        try:
            classifier_element = None
            classifier_elements = driver.find_elements(classifier[0], classifier[1])
            try:
                for classifier_element in classifier_elements:
                    action.move_to_element(classifier_element).perform()
                    break
            except:
                pass
            if len(classifier_elements) != 0 and classifier_element.is_displayed():
                break
        except:
            pass
        time.sleep(0.01)
        try:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.presence_of_element_located((side_menu_Components.element)))
        except Exception as e:
            pass
        elements = driver.find_elements(By.TAG_NAME, "a")
        for x in elements:
            if x.text == "Show More":
                try:
                    classifier_element = None
                    classifier_elements = driver.find_elements(classifier[0], classifier[1])
                    try:
                        for classifier_element in classifier_elements:
                            action.move_to_element(classifier_element).perform()
                            break
                    except:
                        pass
                    if len(classifier_elements) != 0 and classifier_element.is_displayed():
                        break
                except:
                    pass
                action.move_to_element(x).perform()
                x.click()
                result = True
                time.sleep(1)
                try:
                    with allure.step("classifers loading"):
                        wait_for_loading_elements(driver)
                except:
                    pass
                try:
                    classifier_element = None
                    classifier_elements = driver.find_elements(classifier[0], classifier[1])
                    try:
                        for classifier_element in classifier_elements:
                            action.move_to_element(classifier_element).perform()
                            break
                    except:
                        pass
                    if len(classifier_elements) != 0 and classifier_element.is_displayed():
                        break
                except:
                    pass
def campaigns_to_test(campaign):
    try:
        pattern_mapping_df = pd.read_excel(config.test_data_path, sheet_name="CAMPAIGNS_TOTEST")
    except Exception as e:
        with allure.step(f"Check {config.test_data_path}"):
            print(f"Check {config.test_data_path}")
            assert False
    # Convert pattern mapping to dictionary
    exclude_columns = ['DEVICES', 'EXECUTE']
    pattern_mapping = pattern_mapping_df.set_index('CAMPAIGNS').apply(lambda x: x.dropna().drop(exclude_columns).tolist(), axis=1).to_dict()
    campaign_txt = []
    tests = campaign
    test = tests.strip()  # Remove leading and trailing spaces from test
    for pattern, values in pattern_mapping.items():
        if pattern.lower() == test.lower():
            campaign_txt = values
            break
        else:
            campaign_txt = []
    return campaign_txt
def select_Map_View_Components_(driver, campaign, device,excelpath):
    # Fetch components based on the campaign/classifier "T001","T002" etc
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components(campaign, map_start_point, graph_start_point)
    Map_view(driver, tests, excelpath)
def Map_view_for_work_list_campaigns(driver,campaigns_data,excelpath):
    tests_data1 = []
    for i in range(len(campaigns_data)):
        # previous_device = []
        device, campaign, usercampaignsname, testgroup = campaigns_data[i]

        # Fetch components based on the campaign/classifier "T001","T002" etc
        remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
        tests = fetch_components(campaign, map_start_point, graph_start_point)
        tests_data1.append(tests)
    tests_data = [test_data1 for test_data in tests_data1 for test_data1 in test_data]
    tests_data = list(set(tests_data))
    print("tests_data-------------",tests_data)
    Map_view(driver, tests_data, excelpath)
def Map_view_for_datetime_query(driver,excelpath):
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components_datetime_query(map_start_point, graph_start_point)
    # tests= ['Call Test', 'Failed Call', 'Ping Test', 'Web test', 'Download Test', 'Upload Test', 'HTTP DL', 'HTTP UL', 'TCPiperfDl', 'TCPiperfUl', 'UDPiperfDl', 'UDPiperfUl', 'Sent SMS', 'Received SMS', 'Failed SMS', 'Stream Test']
    Map_view(driver, tests, excelpath)

def Map_view(driver,tests,excelpath):
    e_flag = "None"
    Notestdatafound_elements = "None"
    Title = "MAP VIEW"
    runvalue = Testrun_mode(value= "Map_View")
    if "No".lower() == runvalue[-1].strip().lower():
        runvalue = Testrun_mode(value= "PDF Data Export and Validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            Notestdatafound_elements = driver.find_elements(*select_Map_View_Components.No_test_data_element)
            i = 0
            while len(Notestdatafound_elements) != 0:
                Notestdatafound_elements = driver.find_elements(*select_Map_View_Components.No_test_data_element)
                i+=1
                if len(Notestdatafound_elements) == 0 or i==15:
                    break

            closeFullTableView_elements = driver.find_elements(close_button.closeFullTableView[0],close_button.closeFullTableView[1])
            e_flag = 0
            while len(closeFullTableView_elements) == 0:
                try:
                    clickec(driver, select_Map_View_Components.Expand_Map_View)
                    try:
                        WebDriverWait(driver,10).until(EC.visibility_of_element_located((close_button.closeFullTableView[0],close_button.closeFullTableView[1])))
                    except:
                        pass
                    e_flag += 1
                    closeFullTableView_elements = driver.find_elements(close_button.closeFullTableView[0],close_button.closeFullTableView[1])
                    if len(closeFullTableView_elements) !=0 or e_flag == 5:
                        break
                except:
                    click(driver, select_Map_View_Components.Expand_Map_View)
                    try:
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((close_button.closeFullTableView[0], close_button.closeFullTableView[1])))
                    except:
                        pass
                    e_flag += 1
                    closeFullTableView_elements = driver.find_elements(close_button.closeFullTableView[0],close_button.closeFullTableView[1])
                    if len(closeFullTableView_elements) != 0 or e_flag == 5:
                        break
            WebDriverWait(driver, 2.5).until(EC.visibility_of_element_located(select_Map_View_Components.export_selection_box))
            closeFullTableView_elements = driver.find_elements(close_button.closeFullTableView[0],close_button.closeFullTableView[1])
            if len(closeFullTableView_elements) != 0:
                try:
                    WebDriverWait(driver, 2.5).until(EC.presence_of_element_located(select_Map_View_Components.map_element_presence))
                    element = driver.find_element(*select_Map_View_Components.map_element_verify)
                    if element:
                        pass
                    else:
                        with allure.step("Map is not loaded"):
                            allure.attach(driver.get_screenshot_as_png(), name="Map is not loaded", attachment_type=allure.attachment_type.PNG)
                    WebDriverWait(driver, 1).until(EC.presence_of_element_located(select_Map_View_Components.satellite_element_presence))
                    element = driver.find_element(*select_Map_View_Components.satellite_element_verify)
                    if element:
                        pass
                    else:
                        with allure.step("Map is not loaded"):
                            allure.attach(driver.get_screenshot_as_png(), name="Map is not loaded", attachment_type=allure.attachment_type.PNG)
                except:
                    pass
                # Load pattern mapping from Excel file
                try:
                    pattern_mapping_df = pd.read_excel(config.map_view_components_excelpath)
                except Exception as e:
                    with allure.step(f"Check {config.map_view_components_excelpath}"):
                        print(f"Check {config.map_view_components_excelpath}")
                        assert False
                # Convert pattern mapping to dictionary
                pattern_mapping = pattern_mapping_df.set_index('Map view Components').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
                # Match patterns with tests
                txt = []
                if tests.__len__() == 0:
                    statement  = f"Map-View  --  Nothing is marked 'Yes' in {str(config.test_data_path)}"
                    with allure.step(f"Nothing is marked 'Yes' in {str(config.test_data_path)} for 'Map-View"):
                        updatename(excelpath, statement)
                        updatecomponentstatus("MAP VIEW", "", "FAILED", f"Nothing marked in {str(config.test_data_path)}", excelpath)
                        e = Exception
                        raise e
                else:
                    for test in tests:
                        test = test.strip()  # Remove leading and trailing spaces from test
                        for pattern, values in pattern_mapping.items():
                            if pattern.lower() == test.lower():
                                txt = values
                                break
                            else:
                                txt=[]
                        updatename(excelpath, f"MAP VIEW ==>> {test}")
                        time.sleep(0.1)
                        try:
                            try:
                                listbox = WebDriverWait(driver,0.1).until(EC.visibility_of_element_located(select_Map_View_Components.map_menu_dropdown))
                                if listbox.is_displayed():
                                    listbox_btn = WebDriverWait(driver, 1.2).until(EC.visibility_of_element_located(select_Map_View_Components.Test_Type_Dropdown))
                                    # Click on the listbox to close it
                                    listbox_btn.click()
                            except:
                                pass
                            Map_view_Search_Box_not_visible_do_page_up_(driver)
                            Map_View_Select_and_ReadData_(driver, select_Map_View_Components.Test_Type_Dropdown, select_Map_View_Components.nested_locators1, select_Map_View_Components.Call_Test_locator, txt, select_Map_View_Components.cluster_blobmap_locator, select_Map_View_Components.blobmap, select_Map_View_Components.map_element, test, select_Map_View_Components.Data_Table,Title,excelpath, test)
                        except Exception as e:
                            continue
                click_closeButton(driver)
            elif len(closeFullTableView_elements) == 0:
                statement = f"Failed to click on the expand for {Title}"
                with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"Expand_Map_View_screenshot", attachment_type=allure.attachment_type.PNG)
                    e = Exception
                    raise e
        except Exception as e:
            Notestdatafound_elements = driver.find_elements(By.XPATH,"// h3[contains(text(), 'No test data found. Please try different date and ')]")
            closeFullTableView_elements = driver.find_elements(close_button.closeFullTableView[0],close_button.closeFullTableView[1])
            if len(closeFullTableView_elements) == 0:
                statement = f"Failed to click on the expand for {Title}"
                Failupdatename(excelpath, statement)
                updatecomponentstatus(Title, "Expand_Map_View", "FAILED", statement, excelpath)
            elif e_flag == 1:
                print('select Map View Components fail')
            elif len(Notestdatafound_elements) != 0:
                statement = f"No test data found. Please try different date in Map View is present due to map didnt loaded"
                Failupdatename(excelpath, statement)
                updatecomponentstatus(Title, "No test data found. Please try different date", "FAILED", statement, excelpath)
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute", excelpath)
            pass
def click_closeButton(driver):
    try:
        Map_view_Search_Box_not_visible_do_page_up_(driver)
        clickec(driver,close_button.closeFullTableView)
    except Exception as e:
        raise e
def Map_view_Search_Box_not_visible_do_page_up_(driver):
    try:
        # Wait for the element to be visible
        driver.execute_script(f"window.scrollTo({0}, {0});")
    except:
        pass
def Map_View_Select_and_ReadData_(driver, listbox_locator,nested_locators1,Call_Test_locator, option_text_list, cluster_blobmap_locator,blobmap_locator,map_element,elementname,table,Title,excelpath, test):
    ListboxSelectstatus = "None"
    with allure.step(f"Map View Select '{elementname}' and Read Data"):
        try:
            time.sleep(1)
            try:
                wait_for_loading_elements(driver)
            except:
                pass
            l_flag = 0
            if option_text_list.__len__() == 0:
                with allure.step(f"In input data from {str(config.map_view_components_excelpath)} for 'Map-View for '{test}' in header of Map view Components column value against the 2nd row of headers of Map view in {str(config.test_data_path)} is mismatch/empty"):
                    l_flag =2
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                    e = Exception
                    raise e
            elif ["Call Test", "Call Test"] != option_text_list and ['Call Test', 'Failed Calls'] != option_text_list:
               ListboxSelectstatus ,alert_text = select_from_listbox_ECs(driver, listbox_locator, nested_locators1, option_text_list,Title,excelpath)
               l_flag =1
            elif ["Call Test", "Call Test"] == option_text_list:
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Test_Type_Dropdown_for_call_Test,Title,excelpath)
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Call_Test_locator2,Title,excelpath)
                ListboxSelectstatus,alert_text = clickEC_for_listbox(driver,Call_Test_locator,Title,excelpath)
                l_flag = 1
            elif ['Call Test', 'Failed Calls'] == option_text_list:
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Test_Type_Dropdown_for_call_Test,Title,excelpath)
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Call_Test_locator2,Title,excelpath)
                ListboxSelectstatus, alert_text= clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Failed_calls_locator,Title,excelpath)
                l_flag = 1
            waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=1,time_multiple_campaigns=8)
            try:
                wait_for_loading_elements(driver)
            except:
                pass
            if alert_text == None and l_flag ==1:
                try:
                    try:
                        Webtest = operator_comparison_table.operator_comparison_web
                        result_list = extract_table_datas_span1(driver, Webtest,"Operator comparsion table",option_text_list[-1], Title,excelpath)
                        result_list1 = []
                        try:
                            webtest2 = operator_comparison_table.operator_comparison_web_siblingtable
                            result_list1 = extract_table_datas_span1(driver, webtest2, "Operator comparsion table",option_text_list[-1], Title, excelpath)
                        except:
                            pass
                        operator_comparsion_table_data = []
                        for i in range(len(result_list[0])):
                            operator_comparsion_table_data.append([result_list[0][i], result_list[1][i]])
                        for i in range(len(result_list1[0])):
                            operator_comparsion_table_data.append([result_list1[0][i], result_list1[1][i]])
                    except:
                        operator_comparsion_table = operator_comparison_table.Operator_comparison_data
                        operator_comparsion_table_data = extract_table_datas_span1(driver,operator_comparsion_table,"Operator comparsion table",option_text_list[-1], Title,excelpath)
                except:
                    pass
                try:
                    try:
                        if len(operator_comparsion_table_data) != 0:
                            operator_comparsion_table_data.insert(0, [option_text_list[-1]])
                            operator_comparsion_table_data.append(["ENDHERE"])
                            export_pdf_update_to_excel(operator_comparsion_table_data,"OPERATOR_COMPARISON", option_text_list[-1], excelpath)
                            updatecomponentstatus("MAP VIEW", str(test), "PASSED",f"Passed step :- In Operator comparsion table for {option_text_list[-1]} data is found in table",excelpath)
                        elif len(operator_comparsion_table_data) == 0 or operator_comparsion_table_data == None:
                            e = Exception
                            raise e
                    except Exception:
                        if len(operator_comparsion_table_data) == 0 or operator_comparsion_table_data == None:
                            with allure.step(f"Failed step :- In Operator comparsion table for {option_text_list[-1]} No data in table/No table"):
                                Failupdatename(excelpath,f"Failed step :- In Operator comparsion table for {option_text_list[-1]} No data in table")
                                raise Exception
                        elif len(operator_comparsion_table_data) != 0:
                            with allure.step(f"Failed step :- In Operator comparsion table for {option_text_list[-1]} error in insert/appening data to Excel report"):
                                Failupdatename(excelpath, f"Failed step :- In Operator comparsion table for {option_text_list[-1]} error in insert/appening data to Excel report")
                                raise Exception
                except:
                    pass
                runvalue = Testrun_mode(value="Change Settings")
                List_of_Campaigns_runvalue = Testrun_mode(value="List of Campaigns")
                Datetime_query_runvalue = Testrun_mode(value="Date and Time")
                if "RUNNED".lower() != runvalue[-1].strip().lower() and "Yes".lower() != List_of_Campaigns_runvalue[-1].strip().lower() and "Yes".lower() != Datetime_query_runvalue[-1].strip().lower():
                    flag = 0
                    D_flag =0
                    try:
                        cluster_blobmap = driver.find_element(*cluster_blobmap_locator)
                        D_flag =1
                        if cluster_blobmap.is_displayed():
                            interact_with_blobmap(driver, cluster_blobmap_locator, map_element,elementname="cluster blobmap", Title=Title,path=excelpath)
                            if_flag = interact_with_blobmap(driver, blobmap_locator, map_element, elementname,Title,excelpath)
                            updatecomponentstatus("MAP VIEW", str(test), "PASSED", "Blob Found", excelpath)
                        else:
                            if_flag = interact_with_blobmap(driver, blobmap_locator, map_element, elementname,Title,excelpath)
                            updatecomponentstatus("MAP VIEW", str(test), "PASSED", "Blob Found", excelpath)
                        flag =1
                    except Exception as e:
                        if flag != 1 and D_flag ==1:
                            if_flag = "findingblob"
                            Data_Table = driver.find_elements(*select_Map_View_Components.Data_Table)
                            t_flag = "findingblob_with_table"
                            if len(Data_Table) != 0:
                                # blob image found and clicked on blob image during zoom process
                                for table in Data_Table:
                                    if table.is_displayed():
                                      t_flag ="blob found with table"
                                      if_flag = "blob found with table"
                                      e = Exception
                                      raise e
                            elif len(Data_Table)==0 and flag != 1 and D_flag == 1 and if_flag == "blob found":
                                # blob image found
                                e = Exception
                                raise e
                            elif (len(Data_Table)==0 and flag != 1 and D_flag == 1 and if_flag == "blob not found") or (len(Data_Table)==0 and flag != 1 and D_flag == 1 and if_flag == "findingblob") :
                                # blob image not found
                                e = Exception
                                raise e
                        elif flag == 0 and D_flag == 0 :
                            try:
                                if_flag= interact_with_blobmap(driver, blobmap_locator, map_element, elementname, Title,excelpath)
                                updatecomponentstatus("MAP VIEW", str(test), "PASSED", "Blob Found", excelpath)
                                flag = 1
                            except:
                                if flag == 0:
                                    if_flag = "findingblob"
                                    t_flag = "findingblob_with_table"
                                    Data_Table = driver.find_elements(*select_Map_View_Components.Data_Table)
                                    if len(Data_Table) != 0:
                                        # blob image found and clicked on blob image during zoom process
                                        for table in Data_Table:
                                            if table.is_displayed():
                                                t_flag = "blob found with table"
                                                if_flag = "blob found with table"
                                                e = Exception
                                                raise e
                                    elif len(Data_Table) == 0 and flag != 1 and if_flag == "blob found":
                                        e = Exception
                                        raise e
                                    elif (len(Data_Table) == 0 and flag != 1 and if_flag == "blob not found") or (len(Data_Table) == 0 and flag != 1 and if_flag == "findingblob"):
                                        e = Exception
                                        raise e
                    r_flag= 0
                    if flag == 1:
                        time.sleep(1.2)
                        data = extract_table_datas(driver,table,elementname,Title,excelpath)
                        input_data = data
                        # Remove empty lists
                        input_data = [item for item in input_data if item]
                        # Transpose the data
                        output_data = list(map(list, zip(*input_data)))
                        data = output_data
                        updatedatapoints(excelpath, data, elementname, Title)
                        r_flag = 1
            elif ListboxSelectstatus == 0 and alert_text != None and l_flag == 1:
                e = Exception
                with allure.step(f"failed step :- Alert Found is '{alert_text}' for Map View to select {elementname}"):
                    updatecomponentstatus(Title, elementname, "FAILED", f"Alert Found is '{alert_text}' for Map View to select {elementname}", excelpath)
                    raise e
        except Exception as e:
            print("Map View Select and Read Data fail")
            if l_flag == 0:
                statement = f"Unable to locate the element/No such element found and so error in selecting " + str(option_text_list) + " from listbox"
                with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
                    Failupdatename(excelpath, statement)
                    raise e
            elif flag == 0 and l_flag == 1 and alert_text == None and t_flag == "findingblob_with_table" and (if_flag == "blob not found") or(if_flag == "findingblob") :
                statement = f"No blob found for " + str(option_text_list) + f" in {Title}"
                with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
                    updatecomponentstatus(Title, elementname, "FAILED", "no blob found", excelpath)
                    Failupdatename(excelpath, statement)
                raise e
            elif flag == 0 and l_flag == 1 and alert_text == None and t_flag == "blob found with table" and if_flag == "blob found with table":
                statement = f"blob image is displayed and clicked on the blob image during zoom in/out so table is displayed for " + str(option_text_list) + f" in {Title}"
                try:
                    rd_flag =0
                    if t_flag == "blob found with table":
                        time.sleep(1.2)
                        Data_Table = driver.find_elements(*select_Map_View_Components.Data_Table)
                        data = extract_table_datas2(driver, Data_Table, elementname, Title, excelpath)
                        input_data = data
                        # Remove empty lists
                        input_data = [item for item in input_data if item]
                        # Transpose the data
                        output_data = list(map(list, zip(*input_data)))
                        data = output_data
                        updatedatapoints(excelpath, data, elementname, Title)
                        rd_flag = 1
                except Exception as e:
                    if rd_flag == 0 :
                        with allure.step(f"Step Failed :- No data for Map View to select {elementname}, hence read data failed"):
                            data = []
                            updatedatapoints(excelpath, data, elementname, Title)
                            raise e
                finally:
                    pass_updatename(excelpath, statement)
                    e = Exception
                    raise e
            elif flag == 1 and l_flag == 1 and r_flag == 0 and alert_text == None:
                with allure.step(f"Step Failed :- No data for Map View to select {elementname}, hence read data failed"):
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
                    data = []
                    updatedatapoints(excelpath, data, elementname, Title)
                    raise e
            elif option_text_list.__len__() == 0 and l_flag == 2:
                statement = f"In input data from {str(config.map_view_components_excelpath)} for 'Map-View for '{test}' in header of Map view Components column value against the 2nd row of headers of Map view in {str(config.test_data_path)} is mismatch/empty"
                with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
                    Failupdatename(excelpath, statement)
                    updatecomponentstatus("MAP VIEW", str(test), "FAILED", statement, excelpath)
                    raise e
            elif alert_text != None and l_flag == 1:
                statement= f"Alert Found is '{alert_text}' for Map View to select {elementname}"
                with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
                    Failupdatename(excelpath,statement)
                    raise e
def finding_google_default_red_ping(driver,screenshot_folder,image_paths,elementname,campaign,device,test,excelpath):
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    # Take a screenshot and save it to a file
    screenshot_path = screenshot_folder + "\\google_default_red_ping_image" + f'\\{elementname}_{campaign}_{device}_screenshot_{timestamp}.png'
    driver.save_screenshot(screenshot_path)
    i_flag = find_blob(driver, screenshot_path, image_paths)
    statement = f"google default red ping is found in {elementname} for this {device} for particular {campaign} "
    if i_flag == "blob found":
        updatecomponentstatus("MAP VIEW", elementname, "FAILED", statement, excelpath)
        Failupdatename(excelpath, statement)
    else:
        try:
            os.remove(screenshot_path)
        except:
            pass
def hover_(driver, test, excelpath):
    time.sleep(4)
    canvas = driver.find_element(*hover.canvas)
    # Create an instance of ActionChains
    action_chains = ActionChains(driver)
    action_chains.move_to_element(canvas).perform()
    canvas_width_2 = int(canvas.size['width']/2)
    for i in range(-(canvas_width_2), canvas_width_2, 30):
        try:
            action_chains.move_to_element_with_offset(canvas, i, -30).perform()
            if driver.find_element(*hover.Graph_Tootip_element).is_displayed():
                    break
        except:
            try:
                action_chains.move_to_element_with_offset(canvas, i, 0).perform()
                if driver.find_element(*hover.Graph_Tootip_element).is_displayed():
                    break
            except:
                action_chains.move_to_element_with_offset(canvas, i, 30).perform()
                if driver.find_element(*hover.Graph_Tootip_element).is_displayed():
                    break
    graphdataTooltipElement = driver.find_element(*hover.Graph_Tootip_element)
    graph_data = graphdataTooltipElement.text
    Graph_Dropdown_btn = driver.find_element(*hover.Dropdown_btn)
    Graph_Dropdown_button = Graph_Dropdown_btn.text
    if graph_data == None:
        graph_data = graphdataTooltipElement.get_attribute("outerText")
    if graph_data == None:
        graph_data = graphdataTooltipElement.get_attribute("innerText")
    if graph_data == None:
        allure.attach(driver.get_screenshot_as_png(), name=f"{test} ==> {Graph_Dropdown_button}", attachment_type=allure.attachment_type.PNG)
        statement = f"There is no data in Graph-View for {test} ==> {Graph_Dropdown_button}"
        Failupdatename(excelpath, statement)
        updatecomponentstatus("GRAPH VIEW", f"{test} ==> {Graph_Dropdown_button}", "FAILED", statement, excelpath)
        e = Exception
        raise e

def get_graph_data_(driver, txt, excelpath):
    Graph_Dropdown_btn = driver.find_element(*get_graph_data.Dropdown_btn)
    Graph_Dropdown_button = Graph_Dropdown_btn.text
    datas = []
    graphdataTooltipElement = driver.find_element(*get_graph_data.graphdataTooltipElement)
    if graphdataTooltipElement.is_displayed():
        with allure.step(f"Graph of '{txt}' ==> {Graph_Dropdown_button}"):
            time.sleep(0.2)
            if graphdataTooltipElement.is_displayed():
                # Extract the graph data from the tooltip element
                graph_data = graphdataTooltipElement.text
                if graph_data == None:
                    graph_data = graphdataTooltipElement.get_attribute("outerText")
                if graph_data == None:
                    graph_data = graphdataTooltipElement.get_attribute("innerText")
                if graph_data == None:
                    e = Exception
                    raise e
                allure.attach(driver.get_screenshot_as_png(), name=f"'{txt}'", attachment_type=allure.attachment_type.PNG)
                rows = graph_data.strip().split('\n')
                data = [row.split('\t') for row in rows[0:]]
                if data.__len__() == 0 or data is None:
                    e = Exception
                    raise e
                elif data.__len__() != 0:
                    Title = "Graph-View"
                    datas.append(data)
                    # Create a DataFrame with the graph data
                    df = pd.DataFrame(data)
                    # Attach the DataFrame as an HTML table to the Allure report
                    allure.attach(df.to_html(), "Graph Data", allure.attachment_type.HTML)
                    Graphupdatename(excelpath, f"{txt} ==> {Graph_Dropdown_button}")
                    updatedatapoints1(excelpath, datas, txt, Title)
                    updatecomponentstatus("GRAPH VIEW", f"{txt} ==> {Graph_Dropdown_button}", "PASSED",f"There is a data in Graph-View for {txt} ==> {Graph_Dropdown_button}",excelpath)
        if datas.__len__() == 0 or datas is None:
            allure.attach(driver.get_screenshot_as_png(), name=f"{txt} ==> {Graph_Dropdown_button}", attachment_type=allure.attachment_type.PNG)
            statement = f"There is no data in Graph-View for {txt} ==> {Graph_Dropdown_button}"
            Graphupdatename(excelpath, f"{txt} ==> {Graph_Dropdown_button}")
            Failupdatename(excelpath, statement)
            updatecomponentstatus("GRAPH VIEW", f"{txt} ==> {Graph_Dropdown_button}", "FAILED", statement, excelpath)
            e = Exception
            raise e
    else:
        with allure.step(f"Graph of '{txt}' ==> {Graph_Dropdown_button}"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{txt} ==> {Graph_Dropdown_button}", attachment_type=allure.attachment_type.PNG)
            statement = f"There is no data in Graph-View for {txt} ==> {Graph_Dropdown_button}"
            Graphupdatename(excelpath, f"{txt} ==> {Graph_Dropdown_button}")
            Failupdatename(excelpath, statement)
            updatecomponentstatus("GRAPH VIEW", f"{txt} ==> {Graph_Dropdown_button}", "FAILED", statement, excelpath)
            e = Exception
            raise e
def hover_over_second_graph_(driver, test, excelpath):
    canvas = driver.find_element(*hover_over_second_graph.canvas)
    # Create an instance of ActionChains
    action_chains = ActionChains(driver)
    action_chains.move_to_element(canvas).perform()
    canvas_width_2 = int(canvas.size['width']/2)
    for i in range(-(canvas_width_2), canvas_width_2, 30):
        try:
            action_chains.move_to_element_with_offset(canvas, i, -10).perform()
            if driver.find_element(*hover_over_second_graph.Graph_Tootip_element).is_displayed():
                break
        except:
            try:
                action_chains.move_to_element_with_offset(canvas, i, 0).perform()
                if driver.find_element(*hover_over_second_graph.Graph_Tootip_element).is_displayed():
                    break
            except:
                action_chains.move_to_element_with_offset(canvas, i, 10).perform()
                if driver.find_element(*hover_over_second_graph.Graph_Tootip_element).is_displayed():
                    break
    graphdataTooltipElement = driver.find_element(*hover_over_second_graph.Graph_Tootip_element)
    graph_data = graphdataTooltipElement.text
    Graph_Dropdown_btn = driver.find_element(*hover_over_second_graph.Dropdown_btn)
    Graph_Dropdown_button = Graph_Dropdown_btn.text
    if graph_data == None:
        graph_data = graphdataTooltipElement.get_attribute("outerText")
    if graph_data == None:
        graph_data = graphdataTooltipElement.get_attribute("innerText")
    if graph_data == None:
        allure.attach(driver.get_screenshot_as_png(), name=f"{test} ==> {Graph_Dropdown_button}", attachment_type=allure.attachment_type.PNG)
        statement = f"There is no data in Graph-View for {test} ==> {Graph_Dropdown_button}"
        Failupdatename(excelpath, statement)
        updatecomponentstatus("GRAPH VIEW", f"{test} ==> {Graph_Dropdown_button}", "FAILED", statement, excelpath)
        e = Exception
        raise e
def get_secondGraph_data_(driver, txt, excelpath):
    Graph_Dropdown_btn = driver.find_element(*get_secondGraph_data.Dropdown_btn)
    Graph_Dropdown_button = Graph_Dropdown_btn.text
    datas = []
    graphdataTooltipElement = driver.find_element(*get_secondGraph_data.graphdataTooltipElement)

    if graphdataTooltipElement.is_displayed():
        with allure.step(f"Graph of '{txt}' ==> {Graph_Dropdown_button}"):
            time.sleep(0.2)
            if graphdataTooltipElement.is_displayed():
                # Extract the graph data from the tooltip element
                graph_data = graphdataTooltipElement.text
                if graph_data == None:
                    graph_data = graphdataTooltipElement.get_attribute("outerText")
                if graph_data == None:
                    graph_data = graphdataTooltipElement.get_attribute("innerText")
                if graph_data == None:
                    e = Exception
                    raise e
                allure.attach(driver.get_screenshot_as_png(), name=f"'{txt}'", attachment_type=allure.attachment_type.PNG)
                rows = graph_data.strip().split('\n')
                data = [row.split('\t') for row in rows[0:]]
                if data.__len__() == 0 or data is None:
                    e = Exception
                    raise e
                elif data.__len__() != 0:
                    Title = "Graph-View"
                    datas.append(data)
                    # Create a DataFrame with the graph data
                    df = pd.DataFrame(data)
                    # Attach the DataFrame as an HTML table to the Allure report
                    allure.attach(df.to_html(), "Graph Data", allure.attachment_type.HTML)
                    Graphupdatename(excelpath, f"{txt} ==> {Graph_Dropdown_button}")
                    updatedatapoints1(excelpath, datas, txt, Title)
                    updatecomponentstatus("GRAPH VIEW", f"{txt} ==> {Graph_Dropdown_button}", "PASSED",f"There is a data in Graph-View for {txt} ==> {Graph_Dropdown_button}",excelpath)
        if datas.__len__() == 0 or datas is None:
            allure.attach(driver.get_screenshot_as_png(), name=f"{txt} ==> {Graph_Dropdown_button}", attachment_type=allure.attachment_type.PNG)
            statement = f"There is no data in Graph-View for {txt} ==> {Graph_Dropdown_button}"
            Graphupdatename(excelpath, f"{txt} ==> {Graph_Dropdown_button}")
            Failupdatename(excelpath, statement)
            updatecomponentstatus("GRAPH VIEW", f"{txt} ==> {Graph_Dropdown_button}", "FAILED", statement, excelpath)
            e = Exception
            raise e
    else:
        with allure.step(f"No Graph for '{txt}' ==> {Graph_Dropdown_button}"):
            allure.attach(driver.get_screenshot_as_png(), name=f"{txt} ==> {Graph_Dropdown_button}", attachment_type=allure.attachment_type.PNG)
            statement = f"There is no data in Graph-View for {txt} ==> {Graph_Dropdown_button}"
            Graphupdatename(excelpath, f"{txt} ==> {Graph_Dropdown_button}")
            Failupdatename(excelpath, statement)
            updatecomponentstatus("GRAPH VIEW", f"{txt} ==> {Graph_Dropdown_button}", "FAILED", statement, excelpath)
            e = Exception
            raise e

def Graph_View_Components_(driver, campaign, excelpath):
    # Fetch components based on the campaign/classifier "T001","T002" etc
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components(campaign, graph_start_point, export_start_point)
    Graph_view(driver, tests, excelpath)

def Graph_View_for_work_list_campaigns(driver, campaigns_data, excelpath):
    # Fetch components based on the campaign/classifier "T001","T002" etc
    tests_data1 = []
    for i in range(len(campaigns_data)):
        # previous_device = []
        device, campaign, usercampaignsname, testgroup = campaigns_data[i]
        # Fetch components based on the campaign/classifier "T001","T002" etc
        remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
        tests = fetch_components(campaign, graph_start_point, export_start_point)
        tests_data1.append(tests)
    tests_data = [test_data1 for test_data in tests_data1 for test_data1 in test_data]
    tests_data = list(set(tests_data))
    print("tests_data-------------", tests_data)
    Graph_view(driver, tests_data, excelpath)

def Graph_view_datetime_query(driver,excelpath):
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components_datetime_query(graph_start_point, export_start_point)
    Graph_view(driver, tests, excelpath)

def Graph_view(driver,tests,excelpath):
    Title = "GRAPH VIEW"
    runvalue = Testrun_mode(value="Graph_View")
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            e_flag = 0
            try:
                 try:
                     driver.execute_script(f"window.scrollTo({0}, {0});")
                 except:
                     pass
                 clickec(driver, Graph_View_Components.ExpGraph)
                 try:
                     driver.execute_script(f"window.scrollTo({0}, {0});")
                 except:
                     pass
                 WebDriverWait(driver, 10).until(EC.visibility_of_element_located(Graph_View_Components.By_closeFullView))
                 e_flag = 1
            except Exception as e:
                statement =f"Failed to click on the expand for {Title}"
                with allure.step(statement):
                    time.sleep(1.5)
                    Failupdatename(excelpath, statement)
                    updatecomponentstatus(Title, "Expand Graph-View", "FAILED", f"failed step :No graph found", excelpath)
                    allure.attach(driver.get_screenshot_as_png(), name=f"_screenshot",attachment_type=allure.attachment_type.PNG)
                    raise e

            GraphViewname(excelpath, "*************    Graph-View  --- Starts  from here  *************")

            if e_flag ==1:
                if tests.__len__() == 0:
                    statement  = f"Graph-View  --  Nothing is marked 'Yes' in {str(config.test_data_path)}"
                    with allure.step(f"Nothing is marked 'Yes' in {str(config.test_data_path)} for 'Graph-View for '{str(tests)}'"):
                        Graphupdatename(excelpath, statement)
                        Failupdatename(excelpath, f"Nothing is marked 'Yes' in {str(config.test_data_path)} for 'Graph-View for '{str(tests)}'")
                        updatecomponentstatus("GRAPH VIEW", statement, "FAILED", f"Nothing marked in {str(config.test_data_path)}", excelpath)
                        e = Exception
                        raise e
                else:
                    for test in tests:
                        test = test.strip()  # Remove leading and trailing spaces from test
                        test_1 = str(test).lower().replace("test", "").split()[0].capitalize()
                        test_2 = str(test).lower().replace("test", "").replace("iperf", "").split()[0].upper()
                        try:
                            driver.execute_script(f"window.scrollTo({0}, {0});")
                        except:
                            pass

                        try:
                            try:
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located(Graph_View_Components.drop_down_toggle))
                                driver.find_element(*Graph_View_Components.drop_down_toggle).click()
                                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(Graph_View_Components.Graph_dropdown_list_txt))
                                Graph_dropdown_list_txt = driver.find_element(*Graph_View_Components.Graph_dropdown_list_txt)

                                if 'ping' in str(test).lower():
                                    if 'ping' in str(driver.find_element(*Graph_View_Components.drop_down_toggle).text).lower():
                                        if WebDriverWait(driver, 10).until(EC.visibility_of_element_located(Graph_View_Components.Graph_dropdown_list_txt)).is_displayed():
                                            driver.find_element(*Graph_View_Components.drop_down_toggle).click()
                                            time.sleep(0.01)
                                            hover_(driver, test, excelpath)
                                            get_graph_data_(driver, test, excelpath)
                                    elif test in Graph_dropdown_list_txt.text:
                                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located(Graph_View_Components.Graph_dropdown_list_txt))
                                        driver.find_element(By.XPATH, f"//a[contains(.,'{str(test)}')]").click()
                                        hover_(driver, test, excelpath)
                                        get_graph_data_(driver, test, excelpath)

                                elif test in Graph_dropdown_list_txt.text:
                                    try:
                                        driver.find_element(By.XPATH, f"//a[contains(.,'{str(test)}')]").click()
                                        hover_(driver, test, excelpath)
                                        get_graph_data_(driver, test, excelpath)
                                    finally:
                                        try:
                                            driver.find_element(*Graph_View_Components.second_graph_position)
                                            hover_over_second_graph_(driver, test, excelpath)
                                            get_secondGraph_data_(driver, test, excelpath)
                                        except:
                                            pass
                                elif test_1 in Graph_dropdown_list_txt.text:
                                    try:
                                        driver.find_element(By.XPATH, f"//a[contains(.,'{str(test_1)}')]").click()
                                        hover_(driver, test, excelpath)
                                        get_graph_data_(driver, test, excelpath)
                                    finally:
                                        try:
                                            driver.find_element(*Graph_View_Components.second_graph_position)
                                            hover_over_second_graph_(driver, test, excelpath)
                                            get_secondGraph_data_(driver, test, excelpath)
                                        except:
                                            pass
                                elif test_2 in Graph_dropdown_list_txt.text:
                                    try:
                                        driver.find_element(By.XPATH, f"//a[contains(.,'{str(test_2)}')]").click()
                                        hover_(driver, test, excelpath)
                                        get_graph_data_(driver, test, excelpath)
                                    finally:
                                        try:
                                            driver.find_element(*Graph_View_Components.second_graph_position)
                                            hover_over_second_graph_(driver, test, excelpath)
                                            get_secondGraph_data_(driver, test, excelpath)
                                        except:
                                            pass
                                try:
                                    if 'stream' in str(test).lower():
                                        Page_up(driver)
                                except:
                                    pass
                            except:
                                try:
                                    if 'stream' in str(test).lower():
                                        Page_up(driver)
                                except:
                                    pass
                        except Exception as e:
                            try:
                                if 'stream' in str(test).lower():
                                    Page_up(driver)
                            except:
                                pass
                            with allure.step(f"failed step :No graph found"):
                                allure.attach(driver.get_screenshot_as_png(), name="Graphdata", attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus(Title, test, "FAILED", f"failed step :No graph found", excelpath)
                            continue
                Page_up(driver)
                clickec(driver, Graph_View_Components.closeFullView)
        except Exception as e:
            Page_up(driver)
            clickec(driver, Graph_View_Components.closeFullView)
            print("Graph View Components fail")
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute", excelpath)
            pass
def expand_tableView_verify_popUp_(driver):
    runvalue = Testrun_mode(value="Exports")
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            Page_Down(driver)
            time.sleep(0.5)
            clickecwithOutImage(driver, expand_tableView_verify_popUp.DataTable_Expand)
            time.sleep(1)
            List_Of_Campaigns_components_Search_Box_not_visible_do_page_up_(driver)
            with allure.step("List of Campaign's Table-View successfully expanded"):
                allure.attach(driver.get_screenshot_as_png(), name="screenshot", attachment_type=allure.attachment_type.PNG)
            clickec(driver,expand_tableView_verify_popUp.DataTable_Close)
            time.sleep(0.5)
            Page_Down(driver)
            clickec(driver, expand_tableView_verify_popUp.Campaigns_Name)
            time.sleep(1)
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(expand_tableView_verify_popUp.pop_up_data_element))
            time.sleep(3)
            with allure.step("Successfully verified device campaign Pop-Up View"):
                allure.attach(driver.get_screenshot_as_png(), name="screenshot", attachment_type=allure.attachment_type.PNG)
            time.sleep(1.2)
            clickec(driver, expand_tableView_verify_popUp.Close_Pop_Up_Data_point_Btn)
            time.sleep(1)
            Page_Down(driver)
        except:
            pass
    elif "No".lower() == runvalue[-1].strip().lower():
        pass
def specifying_download_path(driver,downloadfilespath,foldername):
    downloadpath= create_folder_for_downloads(destination_folder=downloadfilespath+foldername)
    change_the_download_path(driver,downloadpath)
    return downloadpath
def List_Of_Campaigns_Export_Dashboard_(driver,excelpath, campaign, downloadfilespath):
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components(campaign, export_start_point, load_start_point)
    Exports_view(driver, tests, excelpath, downloadfilespath)
def Exports_view_work_list_campaigns(driver,excelpath, campaigns_data, downloadfilespath):
    tests_data1 = []
    for i in range(len(campaigns_data)):
        # previous_device = []
        device, campaign, usercampaignsname, testgroup = campaigns_data[i]
        # Fetch components based on the campaign/classifier "T001","T002" etc
        remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
        tests = fetch_components(campaign, export_start_point, load_start_point)
        tests_data1.append(tests)
    tests_data = [test_data1 for test_data in tests_data1 for test_data1 in test_data]
    tests_data = list(set(tests_data))
    print("tests_data-------------", tests_data)
    Exports_view(driver, tests_data, excelpath, downloadfilespath)

def exports_view_datetime_query(driver,excelpath,test_case_downloading_files_path):
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components_datetime_query(export_start_point, load_start_point)
    Exports_view(driver, tests, excelpath, test_case_downloading_files_path)

def Exports_view(driver,tests,excelpath,downloadfilespath):
    downloadfilespath = specifying_download_path(driver, downloadfilespath, "EXPORTS")
    Title = "Exports"
    runvalue = Testrun_mode(value="Exports")
    result_data = queue.Queue()
    result_status = queue.Queue()
    if "Yes".lower() == runvalue[-1].strip().lower():
        # Fetch components based on the campaign/classifier "T001","T002" etc
        if tests == []:
            List_of_options_txts=[["Combined Export"], ["Survey Test Export"], ["Export TableSummary"],["Export As PDF"],["Combined Binary Export"],["Hand OverExport"]]
        else:
            input_list = tests
            List_of_options_txts = [[item] for item in input_list]
        try:
            ReportDownlaodName(excelpath, "*************    Download-Reports  --- Starts  from here  *************")
            for List_of_options_txt in List_of_options_txts:
                # traverse in the string
                exportname = ""
                for ele in List_of_options_txt:
                    exportname += ele
                try:
                    try:
                        if WebDriverWait(driver, 1).until(EC.visibility_of_element_located(List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown)).is_displayed():
                            listbox_btn = WebDriverWait(driver, 1).until(EC.visibility_of_element_located(List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown))
                            # Click on the listbox to close it
                            listbox_btn.click()
                    except:
                        pass
                    time.sleep(1.2)
                    List_Of_Campaigns_components(driver, List_of_options_txt, List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options, exportname, excelpath,result_data,result_status,downloadfilespath)
                except Exception as e:
                    continue
        except Exception as e:
            print('Exports_fail')
        try:
            with allure.step("updating_export_result_to_excel"):
                updating_export_result_to_excel(result_status,result_data,excelpath)
        except Exception as e:
            with allure.step("failed:- updating_export_result_to_excel"):
                pass
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute",excelpath)
            pass
def updating_export_result_to_excel(result_status,result_data,excelpath):
    dataframe_status =[]
    combined_status_df="None"
    while not result_status.empty():
        updatecomponentstatus2 = result_status.get()
        df_status = pd.DataFrame(updatecomponentstatus2)
        dataframe_status.append(df_status)
    if len(dataframe_status) !=0:
        with allure.step("result_status of export updating to excel"):
            combined_status_df = pd.concat(dataframe_status, ignore_index=True)
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_componentstatus = workbook["COMPONENTSTATUS"]
    worksheet_data_extract = workbook['DATAEXTRACT']
    if len(dataframe_status) !=0:
        update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
    while not result_data.empty():
        with allure.step("result_data of export updating to excel"):
            data = result_data.get()
            try:
                df_data = pd.DataFrame(data)
            except Exception as e:
                # Check if the values associated with keys are single scalars, and convert them to lists if needed
                for key, value in data.items():
                    if not isinstance(value, list):
                        data[key] = [value]
                df_data = pd.DataFrame(data)
            # List of columns to be moved to the front
            List_of_options_txts = [["Combined Export"], ["Survey Test Export"], ["Export TableSummary"], ["Export As PDF"],["Combined Binary Export"], ["Hand OverExport"]]
            # Find the first sublist item that is a key in df_data
            first_match = next((item[0] for item in List_of_options_txts if item[0] in df_data.columns), None)
            if first_match:
                # Reorder the columns in df_data with the first match at the beginning
                df_data = df_data[[first_match] + [col for col in df_data.columns if col != first_match]]
            updating_data_of_dataframe_for_excel(worksheet=worksheet_data_extract, df_data=df_data)
    workbook.save(excelpath)
    workbook.close()
def updating_data_of_dataframe_for_excel(worksheet,df_data):
    # Convert the DataFrame to a list of rows
    data = list(dataframe_to_rows(df_data, index=False, header=True))
    # Find the last row in the existing data and calculate the next available row for appending
    last_row = worksheet.max_row
    next_available_row = last_row + 1
    # Define a fill color for the first column (1st column value)
    first_column_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow color
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    # Iterate through rows and columns to append data to the worksheet
    for row_idx, row_data in enumerate(data, 1):  # Start from row 1
        for col_idx, cell_value in enumerate(row_data, 1):  # Start from column 1
            cell = worksheet.cell(row=next_available_row + row_idx, column=col_idx, value=cell_value)
            # Apply the color fill to the first column (assuming it's the first column)
            if col_idx == 1:
                cell.fill = first_column_fill
            # Apply the color fill to the header row (assuming it's the first row)
            if row_idx == 1:
                cell.fill = header_fill
def List_Of_Campaigns_components_Search_Box_not_visible_do_page_up_(driver):
    try:
        # Wait for the element to be visible
        wait = WebDriverWait(driver, 10)
        Search_Element = wait.until(EC.visibility_of_element_located(List_Of_Campaigns_components_Search_Box_not_visible_do_page_up.Search_Element))
        Search_Element.click()
    except:
        Page_up(driver)

def List_Of_Campaigns_components(driver,List_of_options_txt, List_Of_Campaigns_Export_Dropdown, List_Of_Campaigns_Export_Dropdown_Options, exportname, excelpath, result_data,result_status,downloadfilespath):
    Title = "Exports"
    with allure.step(f"List Of Campaigns Export '{exportname}' component"):
        try:
            Title = "Exports"
            # Navigate to the "Export As PDF" option
            if List_of_options_txt == ["Export As PDF"]:
                try:
                    # Store the original window handle
                    original_window_handle = driver.current_window_handle
                    time.sleep(1.2)
                    select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dropdown_Options, List_of_options_txt,Title, excelpath)
                    time.sleep(1.2)
                    # Switch to the new tab
                    driver.switch_to.window(driver.window_handles[1])
                    # Take a screenshot and attach it to the Allure report
                    with allure.step("Export As PDF Screenshot"):
                        allure.attach(driver.get_screenshot_as_png(), name="screenshot",attachment_type=allure.attachment_type.PNG)
                        data_df={}
                        data_df[f"{exportname}"]=["Screenshot of Export As PDF is taken. Please refer Allure report for screen-shot"]
                        result_data.put(data_df)
                        updatecomponentstatus2= status(Title, List_of_options_txt.__str__(), "PASSED",f"Passed step :- Screenshot of Export As PDF is taken'")
                        result_status.put(updatecomponentstatus2)
                    # Close the second window
                    driver.close()
                    # Switch back to the original window
                    driver.switch_to.window(original_window_handle)
                except Exception as e:
                    with allure.step(f"Step Failed :- Screenshot of Export As PDF is not taken due PDF is not present"):
                        data_df[f"{exportname}"] = [f"Step Failed :- Screenshot of Export As PDF is not taken due PDF is not present"]
                        result_data.put(data_df)
                        updatecomponentstatus2= status(Title, List_of_options_txt.__str__(), "FAILED",f"Step Failed :- Screenshot of Export As PDF is not taken due PDF is not present'")
                        result_status.put(updatecomponentstatus2)
                        raise e
            elif List_of_options_txt != ["Export As PDF"]:
                time.sleep(1.2)
                flag, alert_text = select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dropdown_Options, List_of_options_txt,Title, excelpath)
                time.sleep(4)
                if flag == 0 and alert_text == None:
                    readCSVSheet(driver, Title, exportname, excelpath,result_data,result_status,downloadfilespath)
                elif flag == 0 and alert_text != None:
                    data_alter = {}
                    data_alter[f"{exportname}"]=[f"Alert Found :- {alert_text}"]
                    result_data.put(data_alter)
                    e = Exception
                    with allure.step(f"Failed Step :- Alert Found is '{alert_text}' for List Of Campaigns Export '{exportname}' component"):
                        updatecomponentstatus2 = status(Title,exportname, "FAILED", f"failed step :- Alert Found :-'{alert_text}'")
                        result_status.put(updatecomponentstatus2)
                        raise e
        except Exception as e:
            raise e

def pdf_export_for_work_list_campaigns(driver,campaigns_data,excelpath,downloadpdfpath):
    tests_data1 = []
    tests_no_data1 =[]
    downloadpdfpath = specifying_download_path(driver, downloadpdfpath, "PDF")
    result_status = queue.Queue()
    data_difference = queue.Queue()
    data_same = queue.Queue()
    Title = "PDF Data Export and Validation"
    runvalue = Testrun_mode(value="PDF Data Export and Validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        for i in range(len(campaigns_data)):
            # previous_device = []
            device, campaign, usercampaignsname, testgroup = campaigns_data[i]
            # Fetch components based on the campaign/classifier "T001","T002" etc
            remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
            tests = fetch_components(campaign, PDF_Export_index_start_point, END_index)
            tests_no = fetch_components_for_no_yes(campaign, PDF_Export_index_start_point, END_index)
            tests_data1.append(tests)
            tests_no_data1.append(tests_no)

        tests_data = [test_data1 for test_data in tests_data1 for test_data1 in test_data]
        tests_data = list(set(tests_data))

        tests_no_data = [test_no_data1 for test_no_data in tests_no_data1 for test_no_data1 in test_no_data]
        tests_no_data = list(set(tests_no_data))
        # Remove items from tests_no_data that are in tests_data
        tests_no_data = list(filter(lambda item: item not in tests_data, tests_no_data))
        print("tests_data-------------",tests_data)
        print("tests_no_data-------------", tests_no_data)
        pdf_comparsion(driver, tests_data, Title, result_status, tests_no_data, downloadpdfpath, data_difference, data_same,date_time_query_yes="No",excelpath=excelpath)
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus2 = status(Title, "Not to execute", "SKIPPED", "You have selected No for execute")
            result_status.put(updatecomponentstatus2)
            pass
    try:
        with allure.step("update_result_of_pdf"):
            update_result_of_pdf(result_status, data_difference, data_same, excelpath)
    except Exception as e:
        with allure.step(f"failed step:- update_result_of_pdf {str(e)}"):
            pass
def pdf_export_for_date_and_time_query(driver,excelpath,downloadpdfpath):
    tests_data1 = []
    tests_no_data1 =[]
    downloadpdfpath = specifying_download_path(driver, downloadpdfpath, "PDF")
    result_status = queue.Queue()
    data_difference = queue.Queue()
    data_same = queue.Queue()
    Title = "PDF Data Export and Validation"
    runvalue = Testrun_mode(value="PDF Data Export and Validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        # Fetch components based on the campaign/classifier "T001","T002" etc
        remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
        tests = fetch_components_datetime_query(PDF_Export_index_start_point, END_index)
        tests_no = fetch_components_datetime_query(PDF_Export_index_start_point, END_index)
        tests_data1.append(tests)
        tests_no_data1.append(tests_no)
        tests_data = [test_data1 for test_data in tests_data1 for test_data1 in test_data]
        tests_data = list(set(tests_data))
        tests_no_data = [test_no_data1 for test_no_data in tests_no_data1 for test_no_data1 in test_no_data]
        tests_no_data = list(set(tests_no_data))
        # Remove items from tests_no_data that are in tests_data
        tests_no_data = list(filter(lambda item: item not in tests_data, tests_no_data))
        print("tests_data-------------",tests_data)
        print("tests_no_data-------------", tests_no_data)
        pdf_comparsion(driver, tests_data, Title, result_status, tests_no_data, downloadpdfpath, data_difference, data_same,date_time_query_yes="Yes",excelpath=excelpath)
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus2 = status(Title, "Not to execute", "SKIPPED", "You have selected No for execute")
            result_status.put(updatecomponentstatus2)
            pass
    try:
        with allure.step("update_result_of_pdf"):
            update_result_of_pdf(result_status, data_difference, data_same, excelpath)
    except Exception as e:
        with allure.step(f"failed step:- update_result_of_pdf {str(e)}"):
            pass
def pdf_export_file_with_operator_comparsion_(driver,campaign,excelpath,downloadpdfpath):
    downloadpdfpath = specifying_download_path(driver,downloadpdfpath,"PDF")
    result_status = queue.Queue()
    data_difference = queue.Queue()
    data_same= queue.Queue()
    Title = "PDF Data Export and Validation"
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components(campaign, PDF_Export_index_start_point, END_index)
    tests_no = fetch_components_for_no_yes(campaign, PDF_Export_index_start_point, END_index)
    runvalue = Testrun_mode(value="PDF Data Export and Validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        pdf_comparsion(driver,tests,Title,result_status,tests_no,downloadpdfpath,data_difference,data_same,date_time_query_yes="No",excelpath=excelpath)
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus2 = status(Title, "Not to execute", "SKIPPED", "You have selected No for execute")
            result_status.put(updatecomponentstatus2)
            pass
    try:
        with allure.step("update_result_of_pdf"):
            update_result_of_pdf(result_status, data_difference, data_same, excelpath)
    except Exception as e:
        with allure.step(f"failed step:- update_result_of_pdf {str(e)}"):
            pass
def pdf_comparsion(driver,tests,Title,result_status,tests_no,downloadpdfpath,data_difference,data_same,date_time_query_yes,excelpath):
    enabled_checkboxes = None
    disabled_checkboxes = None
    checkbox_option_text_list  = None
    pdf_files  = None
    try:
        List_of_options_txt = ["Export As PDF"]
        # Load pattern mapping from Excel file
        try:
            pattern_mapping_df = pd.read_excel(config.pdf_export_excel_path,sheet_name="pdf_components")
        except Exception as e:
            with allure.step(f"Check {config.map_view_components_excelpath}"):
                print(f"Check {config.map_view_components_excelpath}")
                assert False
        flag, alert_text = select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
        # Convert pattern mapping to dictionary
        pattern_mapping = pattern_mapping_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
        # Match patterns with tests
        txt = []
        if tests.__len__() == 0:
            statement = f"{Title}  --  Nothing is marked 'Yes' in {str(config.test_data_path)}"
            with allure.step(f"Nothing is marked 'Yes' in {str(config.test_data_path)} for {Title}"):
                updatecomponentstatus2 = status(Title, "", "FAILED",f"Nothing marked in {str(config.test_data_path)}")
                result_status.put(updatecomponentstatus2)
                e = Exception
                raise e
        else:
            enabled_txt=[]
            txts = []
            for test in tests:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        txts = values
                        enabled_txt.append(txts)
                        break
                    else:
                        txts = []
            disabled_txt = []
            disabledtxts = []
            for test in tests_no:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        disabledtxts = values
                        disabled_txt.append(disabledtxts)
                        break
                    else:
                        disabledtxts = []
            s_flag = 0
            try:
                driver.switch_to.window(driver.window_handles[1])
                s_flag = 1
            except Exception as e:
                pass
            pdf_export_checkbox = driver.find_elements(*pdf_view.parent_checkbox_pdf)
            try:
                try:
                    driver.switch_to.window(driver.window_handles[1])
                except Exception as e:
                    pass
                start_time = time.time()
                # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                max_run_time = 60
                if len(pdf_export_checkbox) == 0:
                    with allure.step("Waiting for pdf page to load"):
                        allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for pdf page to load",attachment_type=allure.attachment_type.PNG)
                        while time.time() - start_time < max_run_time:
                            try:
                                driver.switch_to.window(driver.window_handles[1])
                            except Exception as e:
                                pass
                            pdf_export_checkbox = driver.find_elements(*pdf_view.parent_checkbox_pdf)
                            # Check if the condition is met
                            if len(pdf_export_checkbox) != 0:
                                break
            except:
                pass
            List_of_Campaigns_runvalue = Testrun_mode(value="List of Campaigns")
            Date_and_Time_runvalue = Testrun_mode(value="Date and Time")
            if "Yes".lower() == List_of_Campaigns_runvalue[-1].strip().lower() or "Yes".lower() == Date_and_Time_runvalue[-1].strip().lower():
                time.sleep(30)
            else:
                time.sleep(5)
            enabled_checkboxes, disabled_checkboxes = check_selected_and_finding_enable_and_disabled_checkboxes_(driver,pdf_view.parent_checkbox_pdf)
            if date_time_query_yes !="Yes":
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    for enabled_checkboxs in enabled_txt:
                        executor.submit(process_enabled_checkbox, Title,result_status, enabled_checkboxes, enabled_checkboxs,disabled_checkboxes)
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    for disabled_checkboxs in disabled_txt:
                        executor.submit(process_disabled_checkbox,Title, result_status, disabled_checkboxs,enabled_checkboxes, disabled_checkboxes)
            elif date_time_query_yes =="Yes":
                for enabled_checkboxs in enabled_checkboxes:
                    updatecomponentstatus2 = status(Title, enabled_checkboxs, "PASSED","checkbox is enabled")
                    result_status.put(updatecomponentstatus2)
                for disabled_checkboxs in disabled_checkboxes:
                    updatecomponentstatus2 = status(Title, disabled_checkboxs, "PASSED","checkbox is disabled")
                    result_status.put(updatecomponentstatus2)
            checkbox_option_text_list = []
            geo_list_smstest = []
            failed_call_data = {}
            for test in tests:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        checkbox_option_text_list = values
                        break
                    else:
                        checkbox_option_text_list = []
                try:
                    operator_comparsion_with_by_reading_pdf_export_file(driver,checkbox_option_text_list,enabled_checkboxes,disabled_checkboxes,pdf_view.parent_checkbox_pdf,Title,test,geo_list_smstest,failed_call_data,result_status,data_difference,data_same,excelpath)
                except Exception as e:
                    continue
            try:
                if len(failed_call_data) !=0:
                    try:
                        # Initialize a dictionary to store the total values for 'Total Geo samples'
                        total_geo_samples = 0
                        # Iterate through the dictionary to calculate the total values
                        for key, value_list in failed_call_data.items():
                            for i in range(len(value_list)):
                                if len(value_list[i]) > 1 and str(value_list[i][0]).replace(" ","").lower() == str('Total Geo samples').replace(" ","").lower():
                                    total_geo_samples += int(value_list[i][1])

                        for key, value_list in failed_call_data.items():
                            for i in range(len(value_list)):
                                if len(value_list[i]) > 1 and str(value_list[i][0]).replace(" ","").lower() == str('Total Geo samples').replace(" ","").lower():
                                    value_list[i] = [str('Total Geo samples').replace(" ","").lower(), str(total_geo_samples)]
                    except:
                        pass
                    for checkbox_option_text_list3 ,Export_pdf_table_data in failed_call_data.items():
                        try:
                            checkbox_option_text_list = [checkbox_option_text_list3]
                            data_comparison_in_pdf_export(driver,Export_pdf_table_data,checkbox_option_text_list,geo_list_smstest,result_status,data_difference,data_same,excelpath)
                        except:
                            continue
            except:
               pass
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            except:
                pass
            time.sleep(5)
            runvalue = Testrun_mode(value="Change Settings")
            if "RUNNED".lower() != runvalue[-1].strip().lower():
                g_flag = 0
                try:
                    time.sleep(2)
                    clickec(driver,pdf_view.save_pdf_export)
                    with allure.step("Screenshot to verify clicked on generate pdf"):
                        allure.attach(driver.get_screenshot_as_png(), name=f"generate_pdf_screenshot",attachment_type=allure.attachment_type.PNG)
                    g_flag = 1
                except Exception as e:
                    raise e
                if g_flag == 1:
                    try:
                        Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                        # Set the start time of the loop
                        try:
                            start_time = time.time()
                            # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                            max_run_time = 60
                            if len(Generating_report) == 0:
                                with allure.step("Waiting for generate pdf to load"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for generate pdf to load",attachment_type=allure.attachment_type.PNG)
                                    while time.time() - start_time < max_run_time:
                                        Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                                        # Check if the condition is met
                                        if len(Generating_report) != 0:
                                           break
                        except:
                            pass
                        try:
                            pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                        except:
                            pass
                        try:
                            Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                            start_time = time.time()
                            # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                            max_run_time = 1800
                            if len(Generating_report) != 0:
                                with allure.step("Waiting for generate pdf for downloading"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for generate pdf for downloading",attachment_type=allure.attachment_type.PNG)
                                    while time.time() - start_time < max_run_time:
                                        Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                                        # Check if the condition is met
                                        if len(Generating_report) == 0:
                                           break
                        except:
                            pass
                        try:
                            pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                            start_time = time.time()
                            # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                            max_run_time = 60
                            if len(pdf_files) == 0:
                                with allure.step("Waiting for complete pdf download"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for complete pdf download",attachment_type=allure.attachment_type.PNG)
                                    while time.time() - start_time < max_run_time:
                                        pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                                        # Check if the condition is met
                                        if len(pdf_files) != 0:
                                           break
                        except:
                            pass
                        time.sleep(2)
                        pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                        if len(pdf_files) == 0:
                            statment ="failed step:- pdf is not downloaded"
                            with allure.step("failed step:- PDF is not downloaded"):
                                allure.attach(driver.get_screenshot_as_png(), name=f"failed step:- PDF is not downloaded",attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus2 = status(Title,"pdf download", "FAILED", statment)
                                result_status.put(updatecomponentstatus2)
                                e = Exception
                                raise e
                        elif len(pdf_files) != 0:
                            statment = "pdf is downloaded successfully"
                            with allure.step("PDF is downloaded successfully"):
                                allure.attach(driver.get_screenshot_as_png(),name=f"PDF is downloaded successfully",attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus2 = status(Title, "pdf download", "PASSED",statment)
                                result_status.put(updatecomponentstatus2)
                    except Exception as e:
                        pass
                elif g_flag == 0:
                    statment = f"failed step:- failed to click on save as pdf btn"
                    with allure.step(f"failed step:- failed to click on save as pdf btn"):
                        updatecomponentstatus2 = status(Title, "pdf download", "FAILED", statment)
                        result_status.put(updatecomponentstatus2)
                        e = Exception
                        raise e
            driver.switch_to.window(driver.window_handles[0])
    except Exception as e:
        try:
            driver.switch_to.window(driver.window_handles[0])
        except:
            pass
        pass
    finally:
        try:
            driver.switch_to.window(driver.window_handles[0])
        except:
            pass

def update_result_of_pdf(result_status,data_difference,data_same,excel_file_path):
    try:
        status = []
        df_data_difference = []
        df_data_same = []
        combined_status_df = None
        combined_data_differences = None
        combined_data_same = None
        while not result_status.empty():
            updatecomponentstatus2 = result_status.get()
            df = pd.DataFrame(updatecomponentstatus2)
            status.append(df)
        while not data_same.empty():
            datasame = data_same.get()
            df_same = pd.DataFrame(datasame)
            df_data_same.append(df_same)
        while not data_difference.empty():
            datadiffernce= data_difference.get()
            df_difference = pd.DataFrame(datadiffernce)
            df_data_difference.append(df_difference)
        if len(status) != 0:
            combined_status_df = pd.concat(status, ignore_index=True)
        if len(df_data_difference) !=0:
            combined_data_differences = pd.concat(df_data_difference, ignore_index=True)
        if len(df_data_difference) !=0:
            combined_data_same = pd.concat(df_data_same, ignore_index=True)
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet_componentstatus = workbook["COMPONENTSTATUS"]
        worksheet_data_match = workbook["DATA_MATCH"]
        worksheet_data_not_match = workbook["DATA_NOT_MATCH"]
        if len(status) != 0:
            with allure.step("update_result_of_pdf"):
                update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
        if len(df_data_difference) !=0:
            with allure.step("combined_data_differences"):
                update_excel_datavalidation_pdf_each_testcase_openpyxl(df=combined_data_differences,worksheet=worksheet_data_not_match)
        if len(df_data_difference) !=0:
            with allure.step("combined_data_same"):
                update_excel_datavalidation_pdf_each_testcase_openpyxl(df=combined_data_same,worksheet=worksheet_data_match)
        workbook.save(excel_file_path)
        workbook.close()
    except Exception as e:
        pass
def process_enabled_checkbox(Title,result_status,enabled_checkboxes,enabled_checkboxs,disabled_checkboxes):
    if any(re.fullmatch(enabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(enabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, enabled_checkboxs[0].strip(), "PASSED","In the 'TC' sheet of testdata.xlsx, a particular component is marked as 'yes' and the checkbox associated with it is enabled")
        result_status.put(updatecomponentstatus2)
    elif not any(re.fullmatch(enabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and any(re.fullmatch(enabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, enabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx, for a particular component, the checkbox is marked as 'yes,' but it is disabled")
        result_status.put(updatecomponentstatus2)
    elif not any(re.fullmatch(enabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(enabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, enabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx, the 'yes' value for a particular component does not correspond to the enabled/disabled checkbox options")
        result_status.put(updatecomponentstatus2)
def process_disabled_checkbox(Title,result_status,disabled_checkboxs,enabled_checkboxes,disabled_checkboxes):
    if not any(re.fullmatch(disabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and any(re.fullmatch(disabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, disabled_checkboxs[0].strip(), "PASSED","In the 'TC' sheet of testdata.xlsx,particular component's cell is empty and checkbox for that component is disabled.,")
        result_status.put(updatecomponentstatus2)
    elif any(re.fullmatch(disabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(disabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, disabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx,particular component's cell is empty and checkbox for that component is enabled")
        result_status.put(updatecomponentstatus2)
    elif not any(re.fullmatch(disabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(disabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, disabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx,particular component's cell is empty and does not correspond to the enabled/disabled checkbox options")
        result_status.put(updatecomponentstatus2)

def operator_comparsion_with_by_reading_pdf_export_file(driver,checkbox_option_text_list,enabled_checkboxes,disabled_checkboxes,parent_pdf_export_checkbox,Title, test,geo_list_smstest,failed_call_data,result_status,data_difference,data_same,excelpath):
    Export_pdf_table_data  = "None"
    locator  = "None"
    Export_pdf_table1 = "None"
    Export_pdf_table_data = None
    Export_pdf_table_data1 = []
    try:
        if checkbox_option_text_list.__len__() == 0:
            with allure.step(f"In input data from 'pdf_export'sheet {str(config.pdf_export_excel_path)} for 'PDF EXPORT VIEW for '{test}' in header of PDF EXPORT VIEW Components 1st column value against the 2nd row of headers of PDF EXPORT VIEW in {str(config.test_data_path)} is mismatch/empty"):
                allure.attach(driver.get_screenshot_as_png(), name=f"{test}_screenshot",attachment_type=allure.attachment_type.PNG)
                e = Exception
                raise e
        elif checkbox_option_text_list.__len__() != 0:
            try:
                if any(re.fullmatch(checkbox_option_text_list[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(checkbox_option_text_list[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
                    with allure.step(f"{Title} of {checkbox_option_text_list[0]} and it is enabled"):
                        try:
                            checkbox_option_text_locator = []
                            try:
                                pattern_mapping_locator_df = pd.read_excel(config.pdf_export_excel_path, sheet_name="pdf_locators_for_table")
                            except Exception as e:
                                with allure.step(f"Check {config.pdf_export_excel_path}"):
                                    print(f"Check {config.pdf_export_excel_path}")
                                    assert False
                            pattern_mapping_locator = pattern_mapping_locator_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(), axis=1).to_dict()
                            for pattern, values in pattern_mapping_locator.items():
                                if pattern.lower() == checkbox_option_text_list[0].lower():
                                    checkbox_option_text_locator = values
                                    break
                                else:
                                    checkbox_option_text_locator = []

                            if len(checkbox_option_text_locator) !=0:
                                Export_pdf_table = pdf_view.Export_pdf_table
                                for checkbox_option_text_locator1 in checkbox_option_text_locator:
                                    for Export_pdf_tablelocator_dict in Export_pdf_table:
                                        Export_pdf_table1 = (Export_pdf_tablelocator_dict['locator by'], Export_pdf_tablelocator_dict['locator'].format(checkbox_option_text_locator1))
                                        Export_pdf_table_data = extract_table_datas_span(driver, Export_pdf_table1,f"{Title} table of {checkbox_option_text_list[-1]}",checkbox_option_text_list[-1], Title,excelpath)
                                        Export_pdf_table_data1.append(Export_pdf_table_data)
                            elif len(checkbox_option_text_locator) ==0 or checkbox_option_text_locator == None:
                                with allure.step(f"In input data from 'pdf_locators_for_table'sheet {str(config.pdf_export_excel_path)} for 'PDF EXPORT VIEW for '{test}' in header of PDF EXPORT VIEW Components 1st column value against the 2nd row of headers of PDF EXPORT VIEW in {str(config.test_data_path)} is mismatch/empty"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{test}_screenshot",attachment_type=allure.attachment_type.PNG)
                                    e = Exception
                                    raise e
                        except:
                            pass
                        try:
                            pattern_mapping_df = pd.read_excel(config.map_view_components_excelpath,sheet_name="MAPVIEW_PDFVIEW")
                        except Exception as e:
                            with allure.step(f"Check {config.map_view_components_excelpath}"):
                                print(f"Check {config.map_view_components_excelpath}")
                                assert False
                        pattern_mapping = pattern_mapping_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(), axis=1).to_dict()
                        print(pattern_mapping)
                        test = checkbox_option_text_list[-1]
                        option_text_listpdf = []
                        # Remove leading and trailing spaces from test
                        for pattern, values in pattern_mapping.items():
                            if pattern.lower() == test.lower():
                                option_text_listpdf = values
                                break
                            else:
                                option_text_listpdf = []
                        for option_text_listpdf1,Export_pdf_table_data in zip(option_text_listpdf,Export_pdf_table_data1):
                            try:
                                d_flag = 0
                                a_flag = 0
                                try:
                                    if len(Export_pdf_table_data) != 0:
                                        Export_pdf_table_data1 = []
                                        for data in Export_pdf_table_data:
                                            Export_pdf_table_data1.append(data)
                                        Export_pdf_table_data1.insert(0, [checkbox_option_text_list[-1]])
                                        Export_pdf_table_data1.append(["ENDHERE"])
                                        Export_pdf_table_data.insert(0, [option_text_listpdf1])
                                        Export_pdf_table_data.append(["ENDHERE"])
                                        try:
                                            export_pdf_update_to_excel(Export_pdf_table_data1, "PDF_EXPORT",checkbox_option_text_list[-1], excelpath)
                                            a_flag = 1
                                        except Exception as e:
                                            if len(Export_pdf_table_data) != 0 and d_flag == 0 and a_flag == 0:
                                                with allure.step(f"Failed step :- In {Title} table for {checkbox_option_text_list[-1]}/{option_text_listpdf1} error in insert/appending data to Excel report"):
                                                    updatecomponentstatus2 = status(Title,f"{checkbox_option_text_list[-1]}/{option_text_listpdf1}","FAILED",f"Failed step :- In {Title} table for {checkbox_option_text_list[-1]}/{option_text_listpdf1} error in insert/appending data to Excel report")
                                                    result_status.put(updatecomponentstatus2)
                                        try:
                                            if not any(checkbox_option_text_list[-1] == checkboxvalue for checkboxvalue in ["CallDrop", "CallAborted","CallSetupFailure","CallNoNetwork"]):
                                                data_comparison_in_pdf_export(driver, Export_pdf_table_data,checkbox_option_text_list,geo_list_smstest, result_status,data_difference, data_same, excelpath)
                                            elif any(checkbox_option_text_list[-1] == checkboxvalue for checkboxvalue in ["CallDrop", "CallAborted","CallSetupFailure","CallNoNetwork"]):
                                                failed_call_data[checkbox_option_text_list[-1]] = Export_pdf_table_data
                                            d_flag = 1
                                        except Exception as e:
                                            raise e
                                    elif len(Export_pdf_table_data) == 0 or Export_pdf_table_data == None:
                                        e = Exception
                                        raise e
                                except Exception:
                                    if len(Export_pdf_table_data) == 0 or Export_pdf_table_data == None:
                                        with allure.step(f"Failed step :- The {Title} for {checkbox_option_text_list[-1]}/{option_text_listpdf1} does not contain any data table."):
                                            updatecomponentstatus2 =status(Title,checkbox_option_text_list[-1],"FAILED",f"Failed step :- The {Title} for {checkbox_option_text_list[-1]} does not contain any data table")
                                            result_status.put(updatecomponentstatus2)
                                            raise Exception
                            except Exception as e:
                                continue
                elif not any(re.fullmatch(checkbox_option_text_list[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and any(re.fullmatch(checkbox_option_text_list[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
                    with allure.step(f"Failed step :- {checkbox_option_text_list[-1]} checkbox is disabled"):
                        raise Exception
            except Exception as e:
                raise e
    except Exception as e:
        raise e
def data_comparison_in_pdf_export(driver,Export_pdf_table_data,checkbox_option_text_list,geo_list_smstest,result_status,data_difference,data_same,excelpath):
    c_flag= None
    s_flag= None
    Title = None
    try:
        Title = "Data_comparison b/w MAPVIEW & PDFVIEW"
        operator_comparisonsheet = pd.read_excel(excelpath, sheet_name="OPERATOR_COMPARISON")
        start_row_index = None
        # Remove spaces from strings in operator_comparisonsheet
        for i, row in operator_comparisonsheet.iterrows():
            for col in row.index:
                cell_value = str(row[col]).replace(' ', '')
                if any(re.fullmatch(cell_value.strip(),Export_pdf_enabled_checkbox_name.replace(' ', '').lower().strip(),re.IGNORECASE) for Export_pdf_enabled_checkbox_name in Export_pdf_table_data[0]):
                    start_row_index = i
                    break
            if start_row_index != None:
                break
        print(start_row_index)
        end_row_index = 0
        s_flag = 0
        if start_row_index != None:
            s_flag = 1
            for i, row in operator_comparisonsheet.iterrows():
                if i >= start_row_index:
                    for col in row.index:
                        cell_value = str(row[col]).replace(' ', '')
                        if re.fullmatch(cell_value.strip(), "ENDHERE", re.IGNORECASE):
                            end_row_index = i
                            break
                    if end_row_index != 0:
                        break
            print(end_row_index)
            # Select the desired rows based on the indices
            selected_rows = operator_comparisonsheet.iloc[start_row_index:end_row_index + 1]
            # Optional: Reset the index of the selected rows
            selected_rows = selected_rows.reset_index(drop=True)
            selected_rows_list = selected_rows.values.tolist()
            # Display the selected rows
            if len(selected_rows_list)!=0:
                # Clean Export_pdf_table_data
                Export_pdf_table_data_cleaned = [[str(item).replace(' ', '').lower() for item in sublist if item is not None and item != '' and not pd.isna(item)] for sublist in Export_pdf_table_data]
                # Clean selected_rows_list
                table_data = [[str(item).replace(' ', '').lower() for item in sublistdata if item is not None and not pd.isna(item)] for sublistdata in selected_rows_list]
                try:
                    if checkbox_option_text_list[-1] == "SmsSent" or checkbox_option_text_list[-1] == "SmsRecieve":
                        # Use list comprehension to find sublists containing the word 'Geo'
                        geo_list = [sublist for sublist in table_data if any('Geo'.lower() in str(item).strip().lower() for item in sublist)]
                        for geo in geo_list:
                            geo_list_smstest.append(geo)
                except:
                    pass
                if checkbox_option_text_list[-1].lower() == "SmsTest".lower():
                    # Create a dictionary to store the sum of values based on the "Total Geo samples" key
                    geo_sum_dict = {}
                    for item in geo_list_smstest:
                        key, value = item[0], item[1]
                        if key in geo_sum_dict:
                            geo_sum_dict[key] += int(value)
                        else:
                            geo_sum_dict[key] = int(value)
                    # Convert the dictionary back to a list of sublists
                    combined_geo_list = [[key.replace(" ", ""), str(value)] for key, value in geo_sum_dict.items()]
                    # Find the index of the sublist containing the word "Geo"
                    geo_index = next((i for i, sublist in enumerate(table_data) if any('geo' in item.lower() for item in sublist)),None)
                    if geo_index is not None:
                        # Replace the sublist with the combined_geo_list
                        table_data[geo_index] = combined_geo_list[0]
                if len(table_data) != 0:
                    # Compare the lists element-wise and find the differences
                    differences = []
                    similar = []
                    datas = []
                    c_flag = 0
                    if len(table_data) == len(Export_pdf_table_data_cleaned):
                        for i, (row_data, row_export) in enumerate(zip(table_data, Export_pdf_table_data_cleaned)):
                            for j, (data_item, export_item) in enumerate(zip(row_data, row_export)):
                                if not compare_values(data_item.lower(),export_item.lower()):
                                    differences.append(f"Difference at position ({i}, {j}): {data_item} vs {export_item}")
                                elif compare_values(data_item.lower(),export_item.lower()):
                                    similar.append(f"Same at position ({i}, {j}): {data_item} vs {export_item}")
                    elif len(table_data) != len(Export_pdf_table_data_cleaned):
                        # Create an empty dictionary
                        operator_table_data_dict1 = {}
                        operator_table_data_dict = []
                        # Loop through the data list
                        for sublist in table_data:
                            # Check if the sublist has at least two items
                            if len(sublist) >= 2:
                                key = sublist[0]
                                value = sublist[1]
                                operator_table_data_dict1[key] = value
                        operator_table_data_dict.append(operator_table_data_dict1)
                        Export_pdf_table_data_dict1 = {}
                        Export_pdf_table_data_dict = []
                        # Loop through the data list
                        for sublist in Export_pdf_table_data_cleaned:
                            # Check if the sublist has at least two items
                            if len(sublist) == 2:
                                key = sublist[0]
                                value = sublist[1]
                                Export_pdf_table_data_dict1[key] = value
                        Export_pdf_table_data_dict.append(Export_pdf_table_data_dict1)
                        for Export_pdf_table_data_item, operator_table_data_item in zip(Export_pdf_table_data_dict, operator_table_data_dict):
                            for key in Export_pdf_table_data_item:
                                try:
                                    if not compare_values( Export_pdf_table_data_item[key],operator_table_data_item[key.strip()]):
                                        differences.append(f"Difference in key value '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, '')}")
                                    elif compare_values( Export_pdf_table_data_item[key],operator_table_data_item[key.strip()]):
                                        similar.append(f"Same in key value '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, '')}")
                                except Exception as e:
                                    if not compare_values( Export_pdf_table_data_item[key],operator_table_data_item.get(key, 'Key_name_cant_find')):
                                        differences.append(f"Key name can't find in operator comparsion '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, 'Key_name_cant_find')}")
                                    elif compare_values( Export_pdf_table_data_item[key],operator_table_data_item[key.strip()]):
                                        similar.append(f"Same in key '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, '')}")
                    c_flag = 1
                    df_differences = None
                    if differences:
                        if len(similar) != 0:
                            df_differences = {
                                'Component Type': [f'Component {checkbox_option_text_list[-1]}'] * (len(similar) + len(differences)+2),
                                'Data validation': ["STARTHERE"] + similar + differences + ["ENDHERE"]
                            }
                        elif len(similar) == 0:
                            df_differences = {
                                'Component Type': [f'Component {checkbox_option_text_list[-1]}'] * (len(differences)+2),
                                'Data validation': ["STARTHERE"] + differences + ["ENDHERE"]
                            }
                        if df_differences != None:
                            data_difference.put(df_differences)
                        statment = f"Failed step:- There is difference in data when comparing map_view against pdf_view ."
                        with allure.step(statment+f" for {checkbox_option_text_list[-1]}"):
                            html_for_csv(datas,checkbox_option_text_list[-1])
                            updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "FAILED", statment,)
                            result_status.put(updatecomponentstatus2)
                            raise Exception
                    elif len(similar) != 0 and len(differences) == 0:
                        df_similar = {
                            'Component Type': [f'Component {checkbox_option_text_list[-1]}'] * (len(similar)+2),
                            'Data validation': ["STARTHERE"] + similar + ["ENDHERE"]
                        }
                        data_same.put(df_similar)
                        statment = f"There is same data when comparing map_view against pdf_view"
                        with allure.step(statment + f"for {checkbox_option_text_list[-1]}"):
                            updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "PASSED", statment)
                            result_status.put(updatecomponentstatus2)
                            html_for_csv(datas,checkbox_option_text_list[-1])
        elif start_row_index == None:
            statment = f"Failed step :- There is no data found in the Operator comparison sheet for this component as a reference."
            with allure.step(statment + f":- {checkbox_option_text_list[-1]}"):
                updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "FAILED",statment)
                result_status.put(updatecomponentstatus2)
                raise Exception
    except Exception as e:
        if c_flag == 0 and s_flag == 1:
            statment = f"Failed step :- error in comparing data, Please check the testcase excel report of 'PDF_EXPORT' sheet and locators path in 'pdf_locators_for_table' sheet for"
            with allure.step(statment + f":- {checkbox_option_text_list[-1]}"):
                updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "FAILED", statment)
                result_status.put(updatecomponentstatus2)
                raise Exception
        # Handle exceptions
        print(f"Error: {e}")
        raise e
def update_excel_datavalidation_pdf_each_testcase_openpyxl(df,worksheet):
    """
        Update the high-level Excel report for data validation of PDF.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        color_mapping = {
            'STARTHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            'ENDHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            "Same": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
            "Difference": PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid"),  # Red
            "Key name can't find in operator comparsion": PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid") ,  # Yellow
            "Key not present in combine_binary_export": PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type="solid")  # Light Yellow
        }
        # Find the last used row in the sheet
        last_row = worksheet.max_row
        start_row = last_row + 1
        # Insert the DataFrame into the worksheet
        for index, row in df.iterrows():
            worksheet.append(row.tolist())
        # Apply color formatting to the entire range
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=2, max_col=2),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        # Set colors for File and ParameterType columns
        for i in range(start_row, start_row + len(df)):
            worksheet.cell(row=i, column=1).fill = PatternFill(start_color='FFC864', end_color='FFC864',fill_type="solid")  # Light Orange for ParameterType column
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
def table_summary_(driver,downloadfilespath,excelpath):
    downloadfilespath = specifying_download_path(driver, downloadfilespath, "Table_summary")
    results_status = queue.Queue()
    data_difference = queue.Queue()
    data_same = queue.Queue()
    basename_file = None
    Title = "Table Summary Export Validation"
    runvalue = Testrun_mode(value="Table Summary Export Validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            try:
                convert_to_csv_to_xlsx(downloadfilespath)
            except:
                pass
            table_header_locator = table_summary.header_table_summary
            table_content_locator =table_summary.content_table_summary
            elementname=""
            List_of_options_txt = ["Export TableSummary"]
            dashbord_tablesummary_headers,dashbord_tablesummary_subheaders = extract_table_datas_headers(driver,table_header_locator,elementname,Title,excelpath)
            time.sleep(2)
            try:
                table_content = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(table_content_locator))
                action = ActionChains(driver)
                action.move_to_element(table_content).perform()
            except:
                pass
            dashbord_tablesummary_datacontent = extract_table_datas_content(driver, table_content_locator, elementname, Title, excelpath)
            if len(dashbord_tablesummary_datacontent) != 0:
                # Read the Excel file into a DataFrame
                df = pd.read_excel(config.table_summary_excel_path)
                flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                if alert_text == None:
                    # Transform the data into the expected output format
                    dashbord_tablesummary_data = {}
                    for _, row in df.iterrows():
                        key1 = row['Column 1']
                        key2 = row['Column 2']
                        value = row['Column 3']
                        if key1 in dashbord_tablesummary_data:
                            dashbord_tablesummary_data[key1][key2] = value
                        else:
                            dashbord_tablesummary_data[key1] = {key2: value}
                    # Convert the result to JSON format
                    output_json = json.dumps(dashbord_tablesummary_data, indent=4)
                    # Remove spaces from headers
                    dashbord_tablesummary_headers = [item.replace(' ', '') for item in dashbord_tablesummary_headers][1:]
                    # Load the JSON string into a Python dictionary object
                    output_dict = json.loads(output_json)
                    # Remove spaces from the keys in output_dict
                    output_dict = {key.replace(' ', ''): value for key, value in output_dict.items()}
                    dashbord_tablesummary_headers_list = []
                    for item in dashbord_tablesummary_headers:
                        if item in output_dict:
                            for key in output_dict[item]:
                                dashbord_tablesummary_headers_list.append(output_dict[item][key])
                    dashbord_tablesummary_headers_list = [item.replace(' ', '') for item in dashbord_tablesummary_headers_list]
                    dashbord_tablesummary_data_content =[]
                    for dashbord_tablesummary_datacontent1 in dashbord_tablesummary_datacontent:
                        dashbord_tablesummary_data_content1 = []
                        for item in dashbord_tablesummary_datacontent1[1:]:
                            # Remove empty strings from data_content
                            dashbord_tablesummary_data_content1.append(item.strip().replace(' ','_'))
                        dashbord_tablesummary_data_content.append(dashbord_tablesummary_data_content1)
                    if len(dashbord_tablesummary_headers_list) == len(dashbord_tablesummary_data_content[0]):
                        # Create a dictionary using headers as keys and data_content as values
                        dashbord_tablesummary_data = []
                        for dashbord_tablesummary_data_content_item in dashbord_tablesummary_data_content:
                            dashbord_tablesummary_data_dict = dict(zip(dashbord_tablesummary_headers_list, dashbord_tablesummary_data_content_item))
                            dashbord_tablesummary_data.append(dashbord_tablesummary_data_dict)
                        file_names_path = []
                        list_of_files = glob.glob(downloadfilespath + "\\*.csv")
                        for item in list_of_files:
                            # Check if the item is a file and has a .csv extension
                            if item.lower().endswith(".csv"):
                                if re.search("TableSummary", item, re.IGNORECASE):
                                    file_names_path.append(item)
                                else:
                                    file_names_path.append(item)
                        print(file_names_path)
                        csv_tablesummary_data = []
                        if len(file_names_path) != 0:
                            for file_names_path1 in file_names_path:
                                df = pd.read_csv(str(file_names_path1))
                                basename_file = os.path.basename(file_names_path1)
                                csv_tablesummary_headers = df.columns.tolist()
                                csv_tablesummary_headers = [item.strip() for item in csv_tablesummary_headers]
                                csv_tablesummary_values = df.values.tolist()
                                csv_tablesummary_values = [[str(item).strip().replace(' ', '_') if type(item) is str else None if math.isnan(item) else item for item in csv_tablesummary_sublist] for csv_tablesummary_sublist in csv_tablesummary_values]
                                for row in csv_tablesummary_values:
                                    row_dict = dict(zip(csv_tablesummary_headers, row))
                                    csv_tablesummary_data.append(row_dict)
                        elif len(file_names_path) == 0:
                            statment = f"Table summary csv is not download in {downloadfilespath}"
                            with allure.step(statment):
                                updatecomponentstatus2 = status(Title, List_of_options_txt[0], "FAILED", statment)
                                results_status.put(updatecomponentstatus2)
                            e = Exception
                            raise e
                        if len(csv_tablesummary_data) != 0 and len(dashbord_tablesummary_data) != 0:
                            differences = []
                            similar = []
                            datas =[]
                            csv_tablesummary_sorted_data = sorted(csv_tablesummary_data, key=lambda x: x['TestName'])
                            dashbord_tablesummary_sorted_data = sorted(dashbord_tablesummary_data, key=lambda x: x['TestName'])
                            for csv_tablesummary_data_item, dashbord_tablesummary_data_item in zip(csv_tablesummary_sorted_data, dashbord_tablesummary_sorted_data):
                                for key in csv_tablesummary_data_item:
                                    try:
                                        if not compare_values(csv_tablesummary_data_item[key],dashbord_tablesummary_data_item[key.strip()]):
                                            differences.append(f"Difference in key value '{key}': {csv_tablesummary_data_item[key]} vs {dashbord_tablesummary_data_item.get(key, '')}")
                                        elif compare_values(csv_tablesummary_data_item[key],dashbord_tablesummary_data_item[key.strip()]):
                                            similar.append(f"Same in key value '{key}': {csv_tablesummary_data_item[key]} vs {dashbord_tablesummary_data_item.get(key, '')}")
                                    except Exception as e:
                                        if not compare_values(csv_tablesummary_data_item[key],dashbord_tablesummary_data_item.get(key, 'Key_name_cant_find')):
                                            differences.append(f"Key name can't find in application dashboard '{key}': {csv_tablesummary_data_item[key]} vs {dashbord_tablesummary_data_item.get(key, 'Key_name_cant_find')}")
                                        elif compare_values(csv_tablesummary_data_item[key],dashbord_tablesummary_data_item[key.strip()]):
                                            similar.append(f"Same in key '{key}': {csv_tablesummary_data_item[key]} vs {dashbord_tablesummary_data_item.get(key, '')}")
                                for key in dashbord_tablesummary_data_item:
                                    try:
                                        if not compare_values(dashbord_tablesummary_data_item[key.strip()],csv_tablesummary_data_item[key]):
                                            pass
                                        elif compare_values(dashbord_tablesummary_data_item[key.strip()],csv_tablesummary_data_item[key]):
                                            pass
                                    except Exception as e:
                                        if not compare_values(dashbord_tablesummary_data_item[key.strip()],csv_tablesummary_data_item.get(key, 'Key_name_cant_find')):
                                            differences.append(f"Key name can't find in csv file'{key}': {dashbord_tablesummary_data_item[key]} vs {csv_tablesummary_data_item.get(key, 'Key_name_cant_find')}")
                                        elif compare_values(dashbord_tablesummary_data_item[key.strip()], csv_tablesummary_data_item[key]):
                                            pass
                            # Check if any differences were found and print them
                            df_differences = None
                            if differences:
                                if len(similar) != 0:
                                    df_differences = {
                                        'Data validation': ["STARTHERE"] + similar + differences + ["ENDHERE"]
                                    }
                                elif len(similar) == 0:
                                    df_differences = {
                                        'Data validation': ["STARTHERE"] + differences + ["ENDHERE"]
                                    }
                                if df_differences != None:
                                    data_difference.put(df_differences)
                                statment = f"Failed step:- There is a difference in tablesummary data"
                                with allure.step(statment):
                                    df = pd.DataFrame(df_differences)
                                    html_table = df.to_html()
                                    allure.attach(html_table, f"Table data{str(basename_file)}", AttachmentType.HTML)
                                    updatecomponentstatus2 = status(Title, "Data Validation", "FAILED", statment)
                                    results_status.put(updatecomponentstatus2)
                                    raise Exception
                            elif len(similar) != 0 and len(differences) == 0:
                                df_similar = {
                                    'Data validation': ["STARTHERE"] + similar + ["ENDHERE"]
                                }
                                data_same.put(df_similar)
                                statment = f"There is a same tablesummary data"
                                with allure.step(statment):
                                    df = pd.DataFrame(df_similar)
                                    html_table = df.to_html()
                                    allure.attach(html_table, f"Table data{str(basename_file)}", AttachmentType.HTML)
                                    updatecomponentstatus2 = status(Title, List_of_options_txt[0], "PASSED",statment)
                                    results_status.put(updatecomponentstatus2)
                        elif len(csv_tablesummary_data) == 0 or len(dashbord_tablesummary_data) == 0:
                            with allure.step(f"Failed steps :- Csv tablesummary data is empty"):
                                updatecomponentstatus2 = status(Title, List_of_options_txt[0], "FAILED",f"Csv tablesummary data is empty")
                                results_status.put(updatecomponentstatus2)
                                e = Exception
                                raise e
                    elif len(dashbord_tablesummary_headers_list) != len(dashbord_tablesummary_data_content):
                        with allure.step(f"Failed steps :- Total number of Dashbord tablesummary header is not equal to Total number of Dashbord tablesummary data content"):
                            updatecomponentstatus2 = status(Title, List_of_options_txt[0], "FAILED",f"Total number of Dashbord tablesummary header is not equal to Total number of Dashbord tablesummary data content")
                            results_status.put(updatecomponentstatus2)
                            e = Exception
                            raise e
                elif alert_text != None:
                    data = [[["Alert Found :-" + alert_text]]]
                    data1 = []
                    e = Exception
                    with allure.step(f"Failed Step :- Alert Found is '{alert_text}' for '{List_of_options_txt[0]}' component"):
                        updatedatapoints3D(excelpath, data, List_of_options_txt[0], Title)
                        updatedatapoints(excelpath, data1, List_of_options_txt[0], Title)
                        updatecomponentstatus2 = status(Title, List_of_options_txt[0], "FAILED",f"failed step :- Alert Found :-'{alert_text}'")
                        results_status.put(updatecomponentstatus2)
                        raise e
            elif len(dashbord_tablesummary_datacontent) == 0:
                with allure.step(f"Failed steps :- Dashbord tablesummary data content is empty"):
                    updatecomponentstatus2 = status(Title, List_of_options_txt[0], "FAILED",f"Dashbord tablesummary data content is empty")
                    results_status.put(updatecomponentstatus2)
                    e = Exception
                    raise e
        except Exception as e:
            pass
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus2 = status(Title, "Not to execute", "SKIPPED", "You have selected No for execute")
            results_status.put(updatecomponentstatus2)
            pass
    try:
        with allure.step("update_table_summary_result_to_excel"):
            update_table_summary_result_to_excel(results_status,data_difference,data_same,excelpath)
    except Exception as e:
        with allure.step(f"failed step:- update_table_summary_result_to_excel {str(e)}"):
            pass
def update_table_summary_result_to_excel(result_status,data_difference,data_same,excelpath):
    dataframe_status = []
    dataframe_difference =[]
    dataframe_same = []
    combined_status_df = "None"
    combined_difference_df ="None"
    combined_same_df = "None"
    while not result_status.empty():
        updatecomponentstatus2 = result_status.get()
        df_status = pd.DataFrame(updatecomponentstatus2)
        dataframe_status.append(df_status)
    while not data_same.empty():
        same = data_same.get()
        df_same = pd.DataFrame(same)
        dataframe_same.append(df_same)
    while not data_difference.empty():
        difference = data_difference.get()
        df_difference = pd.DataFrame(difference)
        dataframe_difference.append(df_difference)
    if len(dataframe_status) != 0:
        combined_status_df = pd.concat(dataframe_status, ignore_index=True)
    if len(dataframe_difference) != 0:
        combined_difference_df = pd.concat(dataframe_difference, ignore_index=True)
    if len(dataframe_same) !=0:
        combined_same_df = pd.concat(dataframe_same, ignore_index=True)
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_componentstatus = workbook["COMPONENTSTATUS"]
    data_matchsheet = workbook["TABLESUMMARY_DATA_MATCH"]
    data_not_matchsheet = workbook["TABLESUMMARY_DATA_NOT_MATCH"]
    if len(dataframe_status) != 0:
        update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
    if len(dataframe_difference) != 0:
        update_excel_datavalidation_table_summary_each_testcase_openpyxl(df=combined_difference_df, worksheet=data_not_matchsheet)
    if len(dataframe_same) != 0:
        update_excel_datavalidation_table_summary_each_testcase_openpyxl(df=combined_same_df, worksheet=data_matchsheet)
    workbook.save(excelpath)
    workbook.close()
def update_excel_datavalidation_table_summary_each_testcase_openpyxl(df,worksheet):
    """
        Update the high-level Excel report for data validation of PDF.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        color_mapping = {
            'STARTHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            'ENDHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            "Same": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
            "Difference": PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid"),  # Red
            "Key name can't find in application": PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid") ,  # Yellow
            "Key name can't find in csv file": PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type="solid")  # Light Yellow
        }
        # Find the last used row in the sheet
        last_row = worksheet.max_row
        start_row = last_row + 1
        # Insert the DataFrame into the worksheet
        for index, row in df.iterrows():
            worksheet.append(row.tolist())
        # Apply color formatting to the entire range
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=1, max_col=1),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        # Set colors for File and ParameterType columns
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
def combine_binary_export_nw_freeze(driver, excelpath, downloadfilespath):
    downloadfilespath = specifying_download_path(driver,downloadfilespath,"Combine_binary_nw_freezee")
    result = "None"
    values1 = "None"
    n_flag = "None"
    values_list2 = "None"
    Title = "NW Freeze"
    runvalue = Testrun_mode(value="NW Freeze")
    List_of_options_txt = ["Combined Binary Export"]
    result_status = queue.Queue()
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            try:
                convert_to_csv_to_xlsx(downloadfilespath)
            except:
                pass
            time.sleep(8)
            try:
                action = ActionChains(driver)
                listbox_element = nw_freeze.version
                action.move_to_element(listbox_element).perform()
            except:
                pass
            time.sleep(1)
            flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
            time.sleep(8)
            if alert_text == None:
                file_names_path = []
                list_of_files = glob.glob(downloadfilespath + "\\*.csv")
                try:
                    start_time = time.time()
                    max_run_time = 60
                    with allure.step("Waiting for complete csv download"):
                        allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for complete csv download",attachment_type=allure.attachment_type.PNG)
                        while time.time() - start_time < max_run_time:
                            list_of_files = glob.glob(downloadfilespath + "\\*.csv")
                            # Check if the condition is met
                            if len(list_of_files) != 0:
                                time.sleep(8)
                                break
                except:
                    pass
                # Iterate over all items in the folder
                for item in list_of_files:
                    # Check if the item is a file and has a .csv extension
                    if item.lower().endswith(".csv"):
                        filename = os.path.basename(item)
                        if re.search("Binary_Combined", filename, re.IGNORECASE) or re.search("Binary", filename, re.IGNORECASE) or re.match("Binary_Combined", filename, re.IGNORECASE) or re.match("Binary", filename, re.IGNORECASE) or re.findall("Binary", filename, re.IGNORECASE) or re.findall("Binary_Combined", filename, re.IGNORECASE)or re.finditer("Binary", filename, re.IGNORECASE) or re.finditer("Binary_Combined", filename, re.IGNORECASE):
                            file_names_path.append(item)
                if len(file_names_path) != 0:
                    for file_path in file_names_path:
                        try:
                            # Read the CSV file into a DataFrame
                            data = pd.read_csv(file_path)
                            file_name = os.path.basename(file_path)
                            Networktypetxt = ['RSSI/RSCP', 'lteRSRP', 'nrSsRSRP']
                            Networktypetesting = []
                            Networktypetesting2 = []
                            i = 0
                            for Networktypetxt1 in Networktypetxt:
                                i += 1
                                # Convert the 'RSSI/RSCP' column to a list
                                try:
                                    n_flag = 0
                                    values_list1 = data[Networktypetxt1.strip()].tolist()
                                    if len(values_list1) >= 30:
                                        n_flag = 1
                                        if len(values_list1) != 0:
                                            Networktypetesting.append(Networktypetxt1)
                                        values_list2 = []
                                        # Convert items to float and filter out specific string values
                                        filtered_values = [str(items) for items in values_list1 if (str(items).strip().lower() not in ('unknown', 'ns', 'none', '', ' ', 'nan'))]
                                        for filtered_values1 in filtered_values:
                                            try:
                                                values1 = convert_to_float(filtered_values1)
                                            except:
                                                pass
                                            values_list2.append(str(values1))
                                        if len(values_list2) > 30:
                                            if len(values_list2) != 0:
                                                Networktypetesting2.append(Networktypetxt1)
                                                window_size = 60
                                                if len(values_list2) < window_size:
                                                    window_size = len(values_list2)
                                                rolling_window = [values_list2[i:i + window_size] for i in range(len(values_list2) - window_size + 1)]
                                                window_index = 0
                                                failedresult = []
                                                passedresult = []
                                                # Iterate through each rolling window
                                                for window in rolling_window:
                                                    window_values = [str(value) for value in window if pd.notna(value)]  # Convert to string and exclude NaN values
                                                    window_set = set(window_values)  # Calculate the set of values in the window
                                                    window_index += 1
                                                    if len(window) != 0:
                                                        if len(window_set) > 1:
                                                            for window_set1 in window_set:
                                                                passedresult.append(window_set1)
                                                        else:
                                                            for window_set1 in window_set:
                                                                failedresult.append(window_set1)
                                                if len(failedresult) != 0:
                                                    statment = f"{set(failedresult)}" + f"there is a freeze data for {window_size} samples"
                                                    updatecomponentstatus2 = status(Title,(Networktypetxt1 + ' ' + file_name),"FAILED", statment)
                                                    result_status.put(updatecomponentstatus2)
                                                elif len(failedresult) == 0 and len(passedresult) != 0:
                                                    statment = f"{set(passedresult)}" + f"there is no freeze data for  {window_size} samples"
                                                    updatecomponentstatus2 = status(Title,(Networktypetxt1 + ' ' + file_name),"PASSED", statment)
                                                    result_status.put(updatecomponentstatus2)
                                            elif len(values_list2) == 0 and i == len(Networktypetxt) and len(Networktypetesting2) == 0:
                                                statment = f"None of the 'RSSI/RSCP','lteRSRP' and 'nrSsRSRP' components doesnt contains vaild value in this {file_name}"
                                                with allure.step(statment):
                                                    updatecomponentstatus2 = status(Title, file_name, "FAILED", statment)
                                                    result_status.put(updatecomponentstatus2)
                                                    e = Exception
                                                    raise e
                                        elif len(values_list2) < 30:
                                            statment = f"{len(values_list2)} sample does not satisfy the expected condition , hence {file_name} contains less than 30 sample so neglected"
                                            with allure.step(statment):
                                                updatecomponentstatus2 = status(Title,(Networktypetxt1 + ' ' + file_name),"IGNORED", statment)
                                                result_status.put(updatecomponentstatus2)
                                                e = Exception
                                                raise e
                                    elif len(values_list1) < 30:
                                        n_flag = 1
                                        if len(values_list1) != 0:
                                            Networktypetesting.append(Networktypetxt1)
                                            statment = f"{len(values_list1)} sample does not satisfy the expected condition , hence {file_name} contains less than 30 sample so neglected"
                                            with allure.step(statment):
                                                updatecomponentstatus2 = status(Title, (Networktypetxt1 + ' ' + file_name),"IGNORED", statment)
                                                result_status.put(updatecomponentstatus2)
                                                e = Exception
                                                raise e
                                except:
                                    continue
                            if len(Networktypetesting) == 0 and n_flag != 1:
                                statment = f"None of the 'RSSI/RSCP','lteRSRP' and 'nrSsRSRP' components not found in this {file_name}"
                                with allure.step(statment):
                                    updatecomponentstatus2 = status(Title, file_name, "FAILED", statment)
                                    result_status.put(updatecomponentstatus2)
                                    e = Exception
                                    raise e
                        except:
                            continue
                elif len(file_names_path) == 0:
                    statment = f"Verify file is downloaded in downloaded path where file name should contain Binary/Binary_Combined"
                    with allure.step(statment):
                        updatecomponentstatus2 = status(Title, "Check download path", "FAILED", statment)
                        result_status.put(updatecomponentstatus2)
                        e = Exception
                        raise e
        except Exception as e:
            pass
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus2 = status(Title, "Not to execute", "SKIPPED", "You have selected No for execute")
            result_status.put(updatecomponentstatus2)
            pass
    try:
        with allure.step("update_nw_frezee_result_to_excel"):
            update_nw_frezee_result_to_excel(result_status,excelpath)
    except Exception as e:
        with allure.step(f"failed step:- update_nw_frezee_result_to_excel {str(e)}"):
            pass
def update_nw_frezee_result_to_excel(result_status,excelpath):
    dataframe_status = []
    combined_status_df = "None"
    while not result_status.empty():
        updatecomponentstatus2 = result_status.get()
        df_status = pd.DataFrame(updatecomponentstatus2)
        dataframe_status.append(df_status)
    if len(dataframe_status) != 0:
        with allure.step("result_status of export upadting to excel"):
            combined_status_df = pd.concat(dataframe_status, ignore_index=True)
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_componentstatus = workbook["COMPONENTSTATUS"]
    if len(dataframe_status) != 0:
        update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
    workbook.save(excelpath)
    workbook.close()
def combine_export_vs_combine_binary_export(driver, downloadfilespath, excelpath):
    Title = "COMBINE_EXPORT_VS_COMBINE_BINARY_EXPORT"
    runvalue = Testrun_mode(value="Combined Export Data Validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        List_of_options_txts = [["Combined Binary Export"], ["Combined Export"]]
        combine_binary_export_folder_path = downloadfilespath + "\\Combine_binary_export"
        combine_export_folder_path = downloadfilespath + "\\Combine_export"
        os.mkdir(combine_binary_export_folder_path)
        os.mkdir(combine_export_folder_path)
        alert_texts = []
        for List_of_options_txt in List_of_options_txts:
            if List_of_options_txt == ["Combined Binary Export"]:
                change_the_download_path(driver, combine_binary_export_folder_path)
            elif List_of_options_txt == ["Combined Export"]:
                change_the_download_path(driver, combine_export_folder_path)
            waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=30)
            try:
                wait_for_loading_elements(driver)
            except:
                pass
            flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
            alert_texts.append(alert_text)
            time.sleep(1)
            try:
                wait_for_loading_elements(driver)
            except:
                pass
            time.sleep(8)
            try:
                change_the_download_path(driver, downloadfilespath)
            except Exception as e:
                pass
        if alert_texts[0] == None and alert_texts[1] == None:
            try:
                pattern_mapping_df_file_path = pd.read_excel(config.parameter_validation_excel_path,sheet_name="reference_name_file_path")
            except Exception as e:
                with allure.step(f"Check {config.parameter_validation_excel_path}"):
                    print(f"Check {config.parameter_validation_excel_path}")
                    assert False
            # Convert pattern mapping to dictionary
            pattern_mapping_file_path = pattern_mapping_df_file_path.set_index('Combined Binary Export').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
            combined_Binary_list_of_files1 = glob.glob(combine_binary_export_folder_path + "\\*.csv")
            combined_list_of_files1 = glob.glob(combine_export_folder_path + "\\*.csv")
            # Collect file pairs to process
            files_path = {}
            params_combines = ["IperfTest", "SpeedTest"]
            if len(combined_Binary_list_of_files1) !=0 and len(combined_list_of_files1) !=0:
                for Combined_Binary, Combined in pattern_mapping_file_path.items():
                    combined_Binary_list_of_files = [file for file in glob.glob(combine_binary_export_folder_path + "\\*.csv") if re.fullmatch(Combined_Binary, (file.split("\\")[-1]).split("_")[0], re.IGNORECASE)]
                    combined_list_of_files = [file for file in glob.glob(combine_export_folder_path + "\\*.csv") if re.fullmatch(Combined[0], (file.split("\\")[-1]).split("_")[0], re.IGNORECASE)]
                    if len(combined_Binary_list_of_files) != 0 and len(combined_list_of_files) != 0:
                        files_path[Combined_Binary]=(combined_Binary_list_of_files[0], combined_list_of_files[0])
                    elif len(combined_Binary_list_of_files) != 0 or len(combined_list_of_files) != 0:
                        if len(combined_Binary_list_of_files) != 0 and len(combined_list_of_files) == 0:
                            s = "C:\\RantCell_Automation_Data_and_Reports\\downloads\\testRun_downloads_"
                            updatecomponentstatus(Title=Title, componentname= f"{str(Combined_Binary), str(Combined)} :- combine binary {str(combined_Binary_list_of_files)},combine export {str(combined_list_of_files)}", status="FAILED",comments=f"check in this path {str(s)} of this test case folder where combine binary folder have csv file and combine export folder doesn't have csv file.csv file name should contains this word {str(Combined_Binary), str(Combined) }as mentioned in {config.parameter_validation_excel_path} sheet_name ='reference_name_file_path'",path=excelpath)
                        if len(combined_Binary_list_of_files) == 0 and len(combined_list_of_files) != 0:
                            if any(re.fullmatch(params_combine,(combined_list_of_files[0].split("\\")[-1]).split("_")[0],re.IGNORECASE) for params_combine in params_combines):
                                i = 2
                                try:
                                    df = pd.read_csv(combined_list_of_files[0])
                                    i = 0
                                    data = df['TestType'].str.contains(Combined[1])
                                    i = 1
                                    data = set(data)
                                    if (True in data or "True" in data):
                                        s = "C:\RantCell_Automation_Data_and_Reports\downloads\testRun_downloads_"
                                        updatecomponentstatus(Title=Title, componentname=f"{str(Combined_Binary), str(Combined)} :- combine binary {str(combined_Binary_list_of_files)},combine export {str(combined_list_of_files)}",status="FAILED",comments=f"check in this path {str(s)} of this test case folder where combine export folder have csv file and combine binary folder doesn't have csv file.csv file name should contains this word {str(Combined_Binary), str(Combined)}as mentioned in {config.parameter_validation_excel_path} sheet_name ='reference_name_file_path'",path=excelpath)
                                except Exception as e:
                                    if i == 0 and i != 1:
                                        updatecomponentstatus(Title=Title,componentname=f"{str(Combined_Binary), str(Combined)} :- combine binary {str(combined_Binary_list_of_files)},combine export {str(combined_list_of_files)}",status="FAILED",comments=f"To verify the combine binary contains csv file of {str(Combined_Binary), str(Combined)}, combine export doesn't contain the parameter called 'TestType'",path=excelpath)
                            elif not any(re.fullmatch(params_combine,(combined_list_of_files[0].split("\\")[-1]).split("_")[0],re.IGNORECASE) for params_combine in params_combines):
                                s = "C:\RantCell_Automation_Data_and_Reports\downloads\testRun_downloads_"
                                updatecomponentstatus(Title=Title,componentname=f"{str(Combined_Binary), str(Combined)} :- combine binary {str(combined_Binary_list_of_files)},combine export {str(combined_list_of_files)}",status="FAILED",comments=f"check in this path {str(s)} of this test case folder where combine export folder have csv file and combine binary folder doesn't have csv file.csv file name should contains this word {str(Combined_Binary), str(Combined)}as mentioned in {config.parameter_validation_excel_path} sheet_name ='reference_name_file_path'",path=excelpath)
                    # Create a thread-safe queue for collecting results
                result_queue = queue.Queue()
                if len(files_path)!=0:
                    for TypeofTest, pair in files_path.items():
                        file_path = "C:\\RantCell_Automation_Data_and_Reports\\testdata\\Parameter_validation.xlsx"
                        para_values = extract_values_based_on_call_condition(file_path, TypeofTest)
                        with concurrent.futures.ThreadPoolExecutor() as executor:
                            for TypeofValidation, parametertypes in para_values.items():
                                for parametertype in parametertypes:
                                    combine_export_parameter = parametertype['Combine export']
                                    combine_binary_export_parameter = parametertype['Combine binary export']
                                    executor.submit(process_file_pair, pair,str(parametertype),combine_export_parameter,combine_binary_export_parameter,params_combines, pattern_mapping_file_path,TypeofValidation,result_queue)
                    process_results(driver,result_queue, excelpath,Title)
                elif len(files_path) != 0:
                    s = "C:\RantCell_Automation_Data_and_Reports\downloads\testRun_downloads_"
                    updatecomponentstatus(Title=Title,componentname="",status="FAILED",comments=f"check in this path {str(s)} of this test case folder where combine binary and combine export folder contains csv files and file names are matching with particular words as mentioned in {config.parameter_validation_excel_path} sheet_name ='reference_name_file_path' ",path=excelpath)
            elif len(combined_Binary_list_of_files1) == 0 or len(combined_list_of_files1) == 0:
                s = "C:\\RantCell_Automation_Data_and_Reports\\downloads\\testRun_downloads_"
                updatecomponentstatus(Title=Title,componentname="",status="FAILED",comments=f"check in this path {str(s)} of this test case folder where combine binary and combine export folder contains csv files",path=excelpath)
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute",excelpath)
            pass
def waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign,time_multiple_campaigns):
    List_of_Campaigns_runvalue = Testrun_mode(value="List of Campaigns")
    Date_and_Time_runvalue = Testrun_mode(value="Date and Time")
    if "Yes".lower() == List_of_Campaigns_runvalue[-1].strip().lower() or "Yes".lower() == Date_and_Time_runvalue[-1].strip().lower():
        time.sleep(time_multiple_campaigns)
    else:
        time.sleep(time_single_campaign)
def extract_values_based_on_call_condition(file_path, call_condition):
    # Read the CSV file into a DataFrame
    df = pd.read_excel(file_path,sheet_name="Logic_parameters")
    filtered_rows = df.groupby(call_condition).apply(lambda x: x[["Combine export", "Combine binary export"]].to_dict('records')).to_dict()
    return filtered_rows

def process_file_pair(file_pair,parametertype,combine_export_parameter,combine_binary_export_parameter,params_combines, pattern_mapping_file_path,TypeofValidation,result_queue):
    combine_binary_export_value = {}
    combine_export_value ={}
    combine_binary_file_path, combined_export_file_path = file_pair
    file_name_combine_binary_export = os.path.basename(combine_binary_file_path)
    file_name_combine_export = os.path.basename(combined_export_file_path)
    file_names = []
    result_dict = []
    file_names.append(str("file_name " + " [ " + file_name_combine_binary_export + " ] " + " v/s " + " [ " + file_name_combine_export + " ] "))
    combine_binary_export_result = 'None'
    combine_export_result = 'None'
    try:
        combine_binary_export_value = read_csv_based_on_iteration_for_combine_binary_export(combine_binary_file_path,TypeofValidation ,combine_binary_export_parameter)
        statement = "Parametertype found in csv in combine_binary_export"
        combine_binary_export_result = statement
    except:
        statement = "Parametertype not found in csv in combine_binary_export"
        combine_binary_export_result = statement
        pass
    if not any(re.fullmatch(params_combine, (combined_export_file_path.split("\\")[-1]).split("_")[0], re.IGNORECASE) for params_combine in params_combines):
        try:
            combine_export_value = read_csv_based_on_iteration_for_combine_export(combined_export_file_path,combine_export_parameter)
            statement = "Parametertype found in csv in combine_export"
            combine_export_result = statement
        except:
            statement = "Parametertype not found in csv in combine_export"
            combine_export_result = statement
            pass
    elif any(re.fullmatch(params_combine, (combined_export_file_path.split("\\")[-1]).split("_")[0], re.IGNORECASE) for params_combine in params_combines):
        for Combined_Binary, Combined in pattern_mapping_file_path.items():
            if re.fullmatch((combined_export_file_path.split("\\")[-1]).split("_")[0], Combined[0],re.IGNORECASE) and re.fullmatch((combine_binary_file_path.split("\\")[-1]).split("_")[0],Combined_Binary, re.IGNORECASE):
                try:
                    combine_export_value = read_csv_based_on_iteration_and_testtype_for_combine_export(combined_export_file_path, Combined[1], combine_export_parameter)
                    statement = "Parametertype found in csv in combine_export"
                    combine_export_result = statement
                    break
                except:
                    statement = "Parametertype not found in csv in combine_export"
                    combine_export_result = statement
                    pass
    if len(combine_binary_export_value) != 0 and len(combine_export_value) != 0:
        # Compare values for each key
        comparison_result = {}
        for key in combine_binary_export_value.keys():
            if key in combine_export_value:
                combine_binary_values = combine_binary_export_value[key]
                combine_export_values = combine_export_value[key]
                combine_binary_values = [str(value).replace(' ', '').lower() for value in combine_binary_values]
                combine_export_values = [str(value).replace(' ', '').lower() for value in combine_export_values]
                if len(combine_binary_values) >= len(combine_export_values):
                    find_values_of_unknown_and_none = ['unknown','ns','none','nan']
                    # Replace specific values with 'unknown'
                    combine_binary_values = ['unknown' if str(value).replace(' ', '').lower() in find_values_of_unknown_and_none else str(value).replace(' ', '').lower() for value in combine_binary_values]
                    combine_export_values = ['unknown' if str(value).replace(' ', '').lower() in find_values_of_unknown_and_none else str(value).replace(' ', '').lower() for value in combine_export_values]
                    # combine_binary_values = list(set(combine_binary_values))
                result = comparsion_values_in_bw_two_list(list1=combine_binary_values, list2=combine_export_values)
                if result == True:
                    comparison_result[key] = f"Same combine_binary_export {combine_binary_export_value[key]}:- combine_export {combine_export_value[key]}"
                elif result == False:
                    comparison_result[key] = f"Different combine_binary_export {combine_binary_export_value[key]}:- combine_export {combine_export_value[key]}"
            else:
                comparison_result[key] = f"Key not present in combine_export {key} v/s Key present in combine_binary_export {combine_binary_export_value[key]}"
        for key in combine_export_value.keys():
            if key in combine_binary_export_value:
                pass
            else:
                comparison_result[key] = f"Key not present in combine_binary_export {key} v/s Key present in combine_export {combine_export_value[key]} "
        # Convert the result_dict values to string
        result_dict = [str(str(key) + ": " + value) for key, value in comparison_result.items()]
        result_dict.insert(0, "STARTHERE")
        result_dict.append("ENDHERE")
        # Enqueue the result for later processing
        result_queue.put((TypeofValidation,parametertype, result_dict, file_names,combine_export_result,combine_binary_export_result))
    elif len(combine_binary_export_value) == 0 or len(combine_export_value) == 0:
        result_queue.put((TypeofValidation,parametertype, result_dict, file_names, combine_export_result, combine_binary_export_result))

###################################################################Common methods for combine and combine binary##################################
def read_csv_based_on_iteration(file_path,parametertype):
    pattern_mapping_df = pd.read_csv(file_path)
    # Create the dictionary
    pattern_mapping = pattern_mapping_df.groupby('Iteration')[parametertype].apply(list).to_dict()
    # Replace NaN with 'None' string
    for key, value in pattern_mapping.items():
        pattern_mapping[key] = ['None' if isinstance(item, float) and math.isnan(item) else item for item in value]
    # Create a new dictionary with unique lists
    return pattern_mapping
###################################################################################################################################################

#############################################################Combine binary export methods###############################################################
def read_csv_based_on_iteration_for_combine_binary_export(file_path,TypeofValidation,parametertype):
    output_dict = None
    data_dict = read_csv_based_on_iteration(file_path, parametertype)
    if "Unique".lower().strip() == TypeofValidation.lower().strip():
        output_dict = {key: list(set(value)) for key, value in data_dict.items()}
    elif "Avg".lower().strip() == TypeofValidation.lower().strip():
        output_dict = Avg_based_on_iteration_dic_list_type(data_dict)
    elif "Max".lower().strip() == TypeofValidation.lower().strip():
        output_dict = max_based_on_iteration_dic_list_type(data_dict)
    elif "Min".lower().strip() == TypeofValidation.lower().strip():
        output_dict = min_based_on_iteration_dic_list_type(data_dict)
    elif "mb_to_bytes".lower().strip() == TypeofValidation.lower().strip():
        input_dict = {key: list(set(value)) for key, value in data_dict.items()}
        output_dict = convert_mb_to_bytes_taking_from_url(input_data=input_dict)
    elif "sum function".lower().strip() == TypeofValidation.lower().strip():
        output_dict = calculate_sum(data=data_dict)
    elif "Last_Data_each_iteration".lower().strip() == TypeofValidation.lower().strip():
        output_dict  = get_last_data_from_a_list(input_data=data_dict)
    elif "Sum_NR_By_DR".lower().strip() == TypeofValidation.lower().strip():
        output_dict  = sum_of_nr_vs_sum_of_dr_and_percentage(input_data=data_dict)
    return output_dict
#############################################################Combine export methods###############################################################
def read_csv_based_on_iteration_for_combine_export(file_path, parametertype):
    data_dict = "None"
    data_dict = read_csv_based_on_iteration(file_path, parametertype)
    output_data = {}
    for key, value in data_dict.items():
        try:
            cleaned_values = re.split(r"[,;']", str(value[0]).replace('"', '').replace("'", ''))
            # Remove empty strings from the result
            cleaned_values = [val for val in cleaned_values if val.strip()]
            # cleaned_values = str(value[0]).replace('"', '').split(', ' or ';')
            output_data[key] = cleaned_values
        except:
            continue
    return output_data
def read_csv_based_on_iteration_and_testtype_for_combine_export(file_path,valueshouldcontains,parametertype):
    data_dict = "None"
    data_dict = read_csv_based_on_Testtype_and_iteration_combine_export(file_path,valueshouldcontains,parametertype)
    output_data = {}
    for key, value in data_dict.items():
        try:
            cleaned_values = re.split(r"[,;']", str(value[0]).replace('"', '').replace("'", ''))
            # Remove empty strings from the result
            cleaned_values = [val for val in cleaned_values if val.strip()]
            # cleaned_values = str(value[0]).replace('"', '').split(', ')
            output_data[key] = cleaned_values
        except:
            continue
    return output_data
def read_csv_based_on_Testtype_and_iteration_combine_export(file_path,valueshouldcontains,parametertype):
    # Read the CSV file into a DataFrame
    pattern_mapping_df = pd.read_csv(file_path)
    # Filter rows based on Testtype values containing 'Download'
    download_rows = pattern_mapping_df[pattern_mapping_df['TestType'].str.contains(valueshouldcontains)]
    # Create the dictionary
    pattern_mapping = download_rows.groupby('Iteration')[parametertype].apply(list).to_dict()
    # Replace NaN with 'None' string
    for key, value in pattern_mapping.items():
        pattern_mapping[key] = ['None' if isinstance(item, float) and math.isnan(item) else item for item in value]
    return pattern_mapping
###############################################################################################################################################
def process_results(driver,result_queue, excelpath,Title):
    with allure.step(f"Updating for excel Time Start {str(result_queue)}"):
        allure.attach(driver.get_screenshot_as_png(), name="Execution Time", attachment_type=allure.attachment_type.PNG)
        try:
            updating_for_excel(driver, Title,result_queue, excelpath)
        except Exception as e:
            with allure.step(f"process_results {str(e)}"):
                pass
def updating_for_excel(driver, Title, result_queue, excel_file_path):
    TypeofValidation = None
    parametertype = None
    result_dict = None
    file_names = None
    combine_export_result = None
    combine_binary_export_result = None
    app = None
    sheet_componentstatus = None
    sheet_cbe_vs_ce_match = None
    sheet_cbe_vs_ce_donot_match = None
    workbook = None
    combined_passed_df = None
    combined_match_df = None
    combined_mismatch_df = None
    # Create lists to store data for separate sheets
    match_data_frames = []
    mismatch_data_frames = []
    status = []
    try:
        while not result_queue.empty():
            try:
                try:
                    item = result_queue.get()
                    TypeofValidation, parametertype, result_dict, file_names, combine_export_result, combine_binary_export_result = item
                    # data_validation_sublist
                except Exception as e:
                    with allure.step(f"item {str(e)}"):
                        pass
                # Update status based on conditions
                if not re.search("not found", combine_export_result, re.IGNORECASE):
                    try:
                        export_status = "PASSED"
                        df_parameterfound_export = {'Title': [Title], 'Componentname': [str(file_names) + ":- " + parametertype + " - " + TypeofValidation],'Status': [export_status], 'Comments': [combine_export_result]}
                        df_parameter_foundexport = pd.DataFrame(df_parameterfound_export)
                        status.append(df_parameter_foundexport)
                    except Exception as e:
                        with allure.step(f"export_status PASSED {str(e)}"):
                            pass
                elif re.search("not found", combine_export_result, re.IGNORECASE):
                    try:
                        export_status = "FAILED"
                        df_parameter_notfoundexport = {'Title': [Title], 'Componentname': [str(file_names) + ":- " + parametertype + " - " + TypeofValidation],'Status': [export_status], 'Comments': [combine_export_result]}
                        df_parameter_not_foundexport = pd.DataFrame(df_parameter_notfoundexport)
                        status.append(df_parameter_not_foundexport)
                    except Exception as e:
                        with allure.step(f"export_status FAILED {str(e)}"):
                            pass
                if not re.search("not found", combine_binary_export_result, re.IGNORECASE):
                    try:
                        binary_export_status = "PASSED"
                        df_parameter_foundbinary_export = {'Title': [Title], 'Componentname': [
                            str(file_names) + ":- " + parametertype + " - " + TypeofValidation],
                                                           'Status': [binary_export_status],
                                                           'Comments': [combine_binary_export_result]}
                        df_parameter_foundbinary_export = pd.DataFrame(df_parameter_foundbinary_export)
                        status.append(df_parameter_foundbinary_export)
                    except Exception as e:
                        with allure.step(f"binary_export_status PASSED {str(e)}"):
                            pass
                elif re.search("not found", combine_binary_export_result, re.IGNORECASE):
                    try:
                        binary_export_status = "FAILED"
                        df_parameter_notfoundbinary_export = {'Title': [Title], 'Componentname': [
                            str(file_names) + ":- " + parametertype + " - " + TypeofValidation],
                                                              'Status': [binary_export_status],
                                                              'Comments': [combine_binary_export_result]}
                        df_parameter_not_foundbinary_export = pd.DataFrame(df_parameter_notfoundbinary_export)
                        status.append(df_parameter_not_foundbinary_export)
                    except Exception as e:
                        with allure.step(f"binary_export_status FAILED {str(e)}"):
                            pass
                # Determine if the data should go to match or mismatch
                if len(result_dict) != 0:
                    if any("Different" in item for item in result_dict) or any("Key not present" in item for item in result_dict):
                        if len(result_dict) != 0:
                            try:
                                statement = "Values are different b/w combine binary export v/s combine export wrt 'Iteration'"
                                df_Values_are_different = {'Title': [Title], 'Componentname': [
                                    str(file_names) + ":- " + parametertype + " - " + TypeofValidation],
                                                           'Status': ["FAILED"], 'Comments': [statement]}
                                df_Values_different = pd.DataFrame(df_Values_are_different)
                                status.append(df_Values_different)
                            except Exception as e:
                                with allure.step(f"result_dict different {str(e)}"):
                                    pass
                        # Create a DataFrame for the current item
                        try:
                            df_mismatch_data_frames = {
                                'File': [file_names[0]] * len(result_dict),
                                'ParameterType': [f'ParameterType {parametertype} - {TypeofValidation}'] * len(
                                    result_dict),
                                'Data validation': result_dict
                            }
                            df = pd.DataFrame(df_mismatch_data_frames)
                            mismatch_data_frames.append(df)  # Collect each DataFrame
                        except Exception as e:
                            with allure.step(f"df_mismatch_data_frames {str(e)}"):
                                pass
                    elif not any("Different" in item for item in result_dict) or any(
                            "Key not present" in item for item in result_dict):
                        if len(result_dict) != 0:
                            try:
                                statement = "Values are same b/w combine binary export v/s combine export wrt 'Iteration'"
                                df_Values_are_same = {'Title': [Title], 'Componentname': [
                                    str(file_names) + ":- " + parametertype + " - " + TypeofValidation],
                                                      'Status': ["PASSED"], 'Comments': [statement]}
                                df_Values_same = pd.DataFrame(df_Values_are_same)
                                status.append(df_Values_same)
                            except Exception as e:
                                with allure.step(f"result_dict same {str(e)}"):
                                    pass
                        # Create a DataFrame for the current item
                        try:
                            df_match_data_frames = {
                                'File': [file_names[0]] * len(result_dict),
                                'ParameterType': [f'ParameterType {parametertype} - {TypeofValidation}'] * len(result_dict),
                                'Data validation': result_dict
                            }
                            df = pd.DataFrame(df_match_data_frames)
                            match_data_frames.append(df)  # Collect each DataFrame
                        except Exception as e:
                            with allure.step(f"df_match_data_frames {str(e)}"):
                                pass
            except Exception as e:
                with allure.step(f"while 1 {str(e)}"):
                    pass
        try:
            if len(status) != 0:
                # Concatenate all DataFrames for match and mismatch
                try:
                    combined_passed_df = pd.concat(status, ignore_index=True)
                    try:
                        # Convert the DataFrame to an HTML string
                        html_table = combined_passed_df.to_html(index=False, escape=False)
                        with allure.step(f"COMPONENTSTATUS"):
                            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
                    except Exception as e:
                        with allure.step(f"COMPONENTSTATUS {str(e)}"):
                            pass
                except Exception as e:
                    with allure.step(f"status 1{str(e)}"):
                        pass
            if len(match_data_frames) != 0:
                try:
                    # Concatenate all DataFrames for match and mismatch
                    combined_match_df = pd.concat(match_data_frames, ignore_index=True)
                    try:
                        # Convert the DataFrame to an HTML string
                        html_table = combined_match_df.to_html(index=False, escape=False)
                        with allure.step(f"CBE_vs_CE_MATCH"):
                            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
                    except Exception as e:
                        with allure.step(f"CBE_vs_CE_MATCH {str(e)}"):
                            pass
                except Exception as e:
                    with allure.step(f"match_data_frames 1{str(e)}"):
                        pass
            if len(mismatch_data_frames) != 0:
                try:
                    combined_mismatch_df = pd.concat(mismatch_data_frames, ignore_index=True)
                    try:
                        # Convert the DataFrame to an HTML string
                        html_table = combined_mismatch_df.to_html(index=False, escape=False)
                        with allure.step(f"CBE_vs_CE_DONOT_MATCH"):
                            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
                    except Exception as e:
                        with allure.step(f"CBE_vs_CE_DONOT_MATCH {str(e)}"):
                            pass
                except Exception as e:
                    with allure.step(f"mismatch_data_frames 1 {str(e)}"):
                        pass
        except Exception as e:
            with allure.step(f" condition {str(e)}"):
                pass
        a = "None"
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(excel_file_path)
            # Select the active sheet or specify the sheet by name if needed
            worksheet_cbe_vs_ce_donot_match = workbook["CBE_vs_CE_DONOT_MATCH"]
            worksheet_cbe_vs_ce_match = workbook["CBE_vs_CE_MATCH"]
            worksheet_componentstatus = workbook["COMPONENTSTATUS"]
            if len(status) != 0:
                update_component_status_openpyxl(worksheet=worksheet_componentstatus,dataframe=combined_passed_df)
            if len(match_data_frames) != 0:
                update_excel_openpyxl(df=combined_match_df, worksheet=worksheet_cbe_vs_ce_match)
            if len(mismatch_data_frames) != 0:
                update_excel_openpyxl(df=combined_mismatch_df, worksheet=worksheet_cbe_vs_ce_donot_match)
                # Save the updated workbook
            workbook.save(excel_file_path)
            workbook.close()
        except Exception as e:
            with allure.step(f"failed step:- openpyxl {str(e)}"):
                pass
    except Exception as e:
        with allure.step(f" update {str(e)}"):
            pass
def update_excel_openpyxl(df, worksheet):
    try:
        # Define color mapping
        color_mapping = {
            'STARTHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            'ENDHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            "Same": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
            "Different": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
            "Key not present in combine_export": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow
            "Key not present in combine_binary_export": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Light Yellow
        }
        # Find the last used row in the sheet
        last_row = worksheet.max_row
        start_row = last_row + 1
        # Insert the DataFrame into the worksheet
        for index, row in df.iterrows():
            worksheet.append(row.tolist())
        # Apply color formatting to the entire range
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=3, max_col=3), start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        # Set colors for File and ParameterType columns
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=1, max_col=2), start=start_row):
            file_cell, parameter_cell = row
            file_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue for File column
            parameter_cell.fill = PatternFill(start_color="FFC864", end_color="FFC864", fill_type="solid")  # Light Orange for ParameterType column
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass
def Avg_based_on_iteration_dic_list_type(data):
    result = {}
    for key, values in data.items():
        valid_values = [val for val in values if isinstance(val, (int, float))]
        if valid_values:
            average = sum(valid_values) / len(valid_values)
            result[key] = [average]
        else:
            result[key] = list(set(values))
    return result

def max_based_on_iteration_dic_list_type(data):
    result = {}
    for key, values in data.items():
        valid_values = [val for val in values if isinstance(val, (int, float))]
        if len(valid_values) > 0:
            max_val = max(valid_values)
            result[key] = [max_val]
        else:
            result[key] = list(set(values))
    return result
def min_based_on_iteration_dic_list_type(data):
    result = {}
    for key, values in data.items():
        valid_values = [val for val in values if isinstance(val, (int, float))]
        if len(valid_values) > 0:
            min_val = min(valid_values)
            result[key] = [min_val]
        else:
            result[key] = list(set(values))
    return result
def convert_mb_to_bytes_taking_from_url(input_data):
    output_data = {}
    for key, urls in input_data.items():
        sizes_bytes = []
        for url in urls:
            match = re.search(r'(\d+)mb', url, re.IGNORECASE)
            if match:
                bytes_value = convert_mb_to_bytes(match)
                sizes_bytes.append(str(bytes_value))
            else:
                sizes_bytes.append(str(f"Error: No 'mb' value found in the URL: {url}"))
                print(f"Error: No 'mb' value found in the URL: {url}")
        output_data[key] = sizes_bytes
    return output_data

def convert_mb_to_bytes(match):
    mb_value = int(match.group(1))
    bytes_value = mb_value * (1024 ** 2)
    return bytes_value
def calculate_sum(data):
    sums = {}
    for key, values in data.items():
        sum_values = sum(float(val) for val in values)
        sums[key] = [sum_values]
    return sums
def get_last_data_from_a_list(input_data):
    output_data ={}
    for key, urls in input_data.items():
        try:
            output_data[key] =[urls[-1]]
        except:
            continue
    return output_data
def sum_of_nr_vs_sum_of_dr_and_percentage(input_data):
    output_data = {}
    for key, values in input_data.items():
        total_nr = 0
        total_dr = 0
        for value in values:
            nr, dr = value.split()[0].split('/')
            total_nr += int(nr)
            total_dr += int(dr)
        percentage = (total_nr / total_dr) * 100
        output_data[key] = [f"{total_nr}/{total_dr} ({percentage:.2f}%)"]
    return output_data

def individual_popup_table_vs_ce(driver,downloadfilespath,excelpath):
    Title="Individual Pop Up"
    alert_text = None
    combine_export_csv = None
    result_same = queue.Queue()
    result_difference = queue.Queue()
    result_status = queue.Queue()
    runvalue = Testrun_mode(value="Individual Popup window data validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            List_of_options_txts = [["Combined Export"]]
            combine_export_folder_path = downloadfilespath + "\\Combine_export"
            if os.path.exists(combine_export_folder_path):
                combine_export_csv = [f for f in os.listdir(combine_export_folder_path) if f.endswith(".csv")]
                if len(combine_export_csv) == 0:
                    for List_of_options_txt in List_of_options_txts:
                        flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                        time.sleep(3)
                        combine_export_csv = [f for f in os.listdir(combine_export_folder_path) if f.endswith(".csv")]
            elif not os.path.exists(combine_export_folder_path):
                os.mkdir(combine_export_folder_path)
                for List_of_options_txt in List_of_options_txts:
                    if List_of_options_txt == ["Combined Export"]:
                        change_the_download_path(driver, combine_export_folder_path)
                    time.sleep(3)
                    try:
                        wait_for_loading_elements(driver)
                    except:
                        pass
                    flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                    time.sleep(3)
                    combine_export_csv = [f for f in os.listdir(combine_export_folder_path) if f.endswith(".csv")]
                change_the_download_path(driver, downloadfilespath)
            if alert_text == None and len(combine_export_csv) != 0:
                df_testtype = None
                try:
                    df_testtype = pd.read_excel(config.individual_popup_excel_path,sheet_name="Individual pop up")
                except Exception as e:
                    with allure.step(f"Check {config.individual_popup_excel_path}"):
                        print(f"Check {config.individual_popup_excel_path}")
                        assert False
                Test_types = df_testtype["Individual pop up"].tolist()
                csv_testtype = df_testtype.set_index('Individual pop up').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
                individual_pop_loadcamp_element = driver.find_elements(*individual_pop_table.loaderCamp[:2])

                for i in range(len(individual_pop_loadcamp_element)):
                    i+=1
                    dataofinidividualpoptesttype = {}
                    test_data_of_campaigns = []
                    dataofinidividualpoptesttype_list = []
                    individual_pop_table_loaderCamp = (By.XPATH,f"//tr[{i}]//*[@id='loaderCamp']/abbr/a","individual_pop_table_loaderCamp")
                    test_name_xpath = (By.XPATH,f"//tr[{i}]//*[@id='loaderCamp']/following-sibling::td[1]")
                    Device_name_xpath = (By.XPATH,f"//tr[{i}]//*[@id='loaderCamp']/following-sibling::td[2]")
                    individual_pop_table_loaderCamp_element = driver.find_element(*individual_pop_table_loaderCamp[:2])
                    individual_pop_table_loaderCamp_name = individual_pop_table_loaderCamp_element.text
                    test_name_element = driver.find_element(*test_name_xpath)
                    test_name = test_name_element.text
                    Device_name_element = driver.find_element(*Device_name_xpath)
                    Device_name = Device_name_element.text
                    operator_name = re.sub(r'\D', '', individual_pop_table_loaderCamp_name)
                    test_data_of_campaigns = [operator_name,test_name,Device_name]
                    waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=1,time_multiple_campaigns=8)
                    extract_data_from_individual_pop_table(driver=driver,Title=Title,dataofinidividualpoptesttype = dataofinidividualpoptesttype,Test_types= Test_types,result_status=result_status,individual_pop_table_loaderCamp=individual_pop_table_loaderCamp)
                    dataofinidividualpoptesttype_list.append(dataofinidividualpoptesttype)
                    print("dataofinidividualpoptesttype_list-------------------", dataofinidividualpoptesttype_list)
                    for dataofinidividualpoptesttype_dict in dataofinidividualpoptesttype_list:
                        for testtype, dataoftesttypeindividualpop in dataofinidividualpoptesttype_dict.items():
                            try:
                                testtypetxt = []
                                test = testtype.strip()  # Remove leading and trailing spaces from test
                                for pattern, values in csv_testtype.items():
                                    if pattern.lower() == test.lower():
                                        testtypetxt = values
                                        break
                                    else:
                                        testtypetxt = []
                                # combine_export_vs_individual_pop(combine_export_folder_path, testtype, testtypetxt,dataoftesttypeindividualpop,result_same, result_difference, result_status)
                                with concurrent.futures.ThreadPoolExecutor() as executor:
                                    executor.submit(combine_export_vs_individual_pop, combine_export_folder_path,testtype, testtypetxt, dataoftesttypeindividualpop,test_data_of_campaigns,result_same,result_difference, result_status)
                            except Exception as e:
                                continue
                    try:
                        clickec(driver=driver, locators=individual_pop_table.Close_button_of_ipu)
                    except Exception as e:
                        pass

                try:
                    with allure.step("update individualpopup vs combine export result to excel"):
                        update_individualpopup_result_to_excel(result_status=result_status, data_difference=result_difference, data_same=result_same, excelpath=excelpath)
                except Exception as e:
                    pass
        except Exception as e:
            pass
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute", excelpath)
            pass
def extract_data_from_individual_pop_table(driver,Title,dataofinidividualpoptesttype,Test_types,result_status,individual_pop_table_loaderCamp):
    dict_list = []
    try:
        # //tr[2]//*[@id="loaderCamp"]/abbr/a
        click(driver,individual_pop_table_loaderCamp)
        time.sleep(1)
        try:
            wait_for_loading_elements(driver)
        except:
            pass
        # Find the "Ping Test" table by locating the associated h4 element
        test_types_h4 = driver.find_elements(*individual_pop_table.Tableview_header)
        for test_type_h4 in test_types_h4:
            test_value = test_type_h4.text
            # Remove spaces from the input string and make it case-sensitive
            formatted_string = test_value.replace(" ", "")
            # Test_types = ["PingTest", "DownloadTest", "UploadTest", "WebTest", "CallTest", "HTTPSpeedDownloadTest","HTTPSpeedUploadTest", "SMSTest", "TCPiPerfDownloadTest", "TCPiPerfUploadTest"]
            # Use regular expressions to match the test type name followed by a colon
            matches = [test_type for test_type in Test_types if compare_values(test_type, formatted_string[:len(test_type)])]
            # Check if the pattern was found
            if matches:
                pass
                # Find the table that follows the h4 element
                try:
                    with allure.step(f"{matches[0]}"):
                        test_tables_headers = test_type_h4.find_elements(*individual_pop_table.Tableview_siblingtable_header)
                        test_tables_content = test_type_h4.find_elements(*individual_pop_table.Tableview_siblingtable_content)
                        action = ActionChains(driver)
                        action.move_to_element(test_tables_content[0]).perform()
                        allure.attach(driver.get_screenshot_as_png(), name=f"{matches[0]}",attachment_type=allure.attachment_type.PNG)
                except Exception as e:
                    pass
                try:
                    # Extract the table headers
                    headers = [th.text.strip() for th in test_tables_headers[0].find_elements(By.TAG_NAME, "th") if "ng-hide" not in th.get_attribute("class")]
                    # Extract the table data
                    data = []
                    for row in test_tables_content[0].find_elements(By.TAG_NAME, "tr"):  # Skip the header row
                        row_data = [td.text.strip() for td in row.find_elements(By.TAG_NAME, "td") if "ng-hide" not in td.get_attribute("class")]
                        data.append(row_data)
                    dict_list = [dict(zip(headers, row)) for row in data]
                    # dict_list = pd.DataFrame(dict_list)
                    # dict_list = dict_list.groupby('Iteration').apply(lambda x: x.to_dict(orient='records')).to_dict()
                    dataofinidividualpoptesttype[str(matches[0])] = dict_list
                except Exception as e:
                    pass
            else:
                r_status = status(Title=Title, component=f"{formatted_string}", status="FAILED",comments=f"'{formatted_string}' doesn't contains the words i'e mentioned in the file '{config.individual_popup_excel_path}',sheet_name= 'Individual pop up'")
                result_status.put(r_status)
    except:
        pass

def combine_export_vs_individual_pop(combine_export_folder_path,testtype,testtypetxt,dataoftesttypeindividualpop,test_data_of_campaigns,result_same,result_difference,result_status):
    dataoftesttypecombineexport =[]
    comparsion_result = []
    Title = "Individual Pop Up v/s Combine export"
    try:
        params_combines = ["IperfTest", "SpeedTest"]
        operator_name, test_name, Device_name = test_data_of_campaigns
        combined_list_of_files = [file for file in glob.glob(combine_export_folder_path + "\\*.csv") if re.fullmatch(testtypetxt[0], (file.split("\\")[-1]).split("_")[0], re.IGNORECASE)]
        if len(combined_list_of_files) != 0:
            print("test_data_of_campaigns----",test_data_of_campaigns)
            df_ce = pd.read_csv(combined_list_of_files[0])
            # df_ce['OperatorName'] = df_ce['OperatorName'].astype(str)
            # df_ce['NetworkOperator'] = df_ce['NetworkOperator'].astype(str)
            # df_ce = df_ce[df_ce['OperatorName'].str.contains(operator_name, na=False) | df_ce['NetworkOperator'].str.contains(operator_name, na=False)]
            # df_ce = df_ce[df_ce['DeviceName'].str.contains(Device_name, na=False)]
            # df_ce = df_ce[df_ce['UserTestName'].str.contains(test_name, na=False)]
            if not any(re.fullmatch(params_combine, (combined_list_of_files[0].split("\\")[-1]).split("_")[0],re.IGNORECASE) for params_combine in params_combines):
                dataoftesttypecombineexport = df_ce.to_dict(orient='records')
            elif any(re.fullmatch(params_combine, (combined_list_of_files[0].split("\\")[-1]).split("_")[0],re.IGNORECASE) for params_combine in params_combines):
                dataoftesttypecombineexport = df_ce[df_ce['TestType'].str.contains(testtypetxt[1])]
                dataoftesttypecombineexport = dataoftesttypecombineexport.to_dict(orient='records')
            # dataoftesttypecombineexport = df_ce.groupby('Iteration').apply(lambda x: x.to_dict(orient='records')).to_dict()

        if len(combined_list_of_files) != 0 and len(dataoftesttypeindividualpop) != 0 and len(dataoftesttypecombineexport) != 0:
            flag_comparsion_result = True
            file_name =(combined_list_of_files[0].split("\\")[-1])
            comparsion_result.append({"File": f"{testtype} :- {file_name}","Individual pop up headers":"STARTHERE","Individual pop up value":"STARTHERE","combine export value":"STARTHERE","Data validation": f"STARTHERE"})
            i = 0
            for dataoftesttypecombineexport1 , dataoftesttypeindividualpop1 in zip(dataoftesttypecombineexport,dataoftesttypeindividualpop):
                i +=1
                comparsion_result.append({"File": f"{testtype} :- {file_name}","Individual pop up headers":f"{i} Row Start","Individual pop up value":f"{i} Row Start","combine export value":f"{i} Row Start", "Data validation": f"{i} Row Start"})
                for key in dataoftesttypeindividualpop1.keys():
                    if not re.match("Latitude",key,re.IGNORECASE):
                        try:
                            if not compare_values(dataoftesttypeindividualpop1[key],dataoftesttypecombineexport1[key.strip()]):
                                comparsion_result.append({"File":f"{testtype} :- {file_name}","Individual pop up headers":key,"Individual pop up value":dataoftesttypeindividualpop1[key],"combine export value":dataoftesttypecombineexport1.get(key, ''),"Data validation":f"Difference in key value '{key}': individual popup = '{dataoftesttypeindividualpop1[key]}' vs combine_export_csv = '{dataoftesttypecombineexport1.get(key, '')}'."})
                                flag_comparsion_result = False
                            elif compare_values(dataoftesttypeindividualpop1[key],dataoftesttypecombineexport1[key.strip()]):
                                comparsion_result.append({"File":f"{testtype} :- {file_name}","Individual pop up headers":key,"Individual pop up value":dataoftesttypeindividualpop1[key],"combine export value":dataoftesttypecombineexport1.get(key, ''),"Data validation":f"Same in key value '{key}': individual popup = '{dataoftesttypeindividualpop1[key]}' vs combine_export_csv = '{dataoftesttypecombineexport1.get(key, '')}'."})
                        except Exception as e:
                            if not compare_values(dataoftesttypeindividualpop1[key],dataoftesttypecombineexport1.get(key, 'Key_name_cant_find_in_combine_export_csv')):
                                comparsion_result.append({"File":f"{testtype} :- {file_name}","Individual pop up headers":key,"Individual pop up value":dataoftesttypeindividualpop1[key],"combine export value":dataoftesttypecombineexport1.get(key, 'Key_name_cant_find_in_combine_export_csv'),"Data validation":f"Key name can't find in csv'{key}': individual popup = '{dataoftesttypeindividualpop1[key]}' vs combine_export_csv =  '{dataoftesttypecombineexport1.get(key, 'Key_name_cant_find_in_combine_export_csv')}'."})
                                flag_comparsion_result = False
                            elif compare_values(dataoftesttypeindividualpop1[key],dataoftesttypecombineexport1[key.strip()]):
                                comparsion_result.append({"File":f"{testtype} :- {file_name}","Individual pop up headers":key,"Individual pop up value":dataoftesttypeindividualpop1[key],"combine export value":dataoftesttypecombineexport1.get(key, ''),"Data validation":f"Same in key value '{key}': individual popup = '{dataoftesttypeindividualpop1[key]}' vs combine_export_csv =  '{dataoftesttypecombineexport1.get(key, '')}'"})
            comparsion_result.append({"File": f"{testtype} :- {file_name}","Individual pop up headers":"ENDHERE","Individual pop up value":"ENDHERE","combine export value":"ENDHERE", "Data validation": f"ENDHERE"})
            if flag_comparsion_result == True:
                statement = "Values are same b/w individual popup v/s combine export"
                result_same.put(comparsion_result)
                r_status = status(Title=Title,component=f"{operator_name, test_name, Device_name} ==>{testtype} :- {file_name}",status="PASSED",comments=statement)
                result_status.put(r_status)
            elif flag_comparsion_result == False:
                statement = "Values are difference b/w individual popup v/s combine export"
                result_difference.put(comparsion_result)
                r_status = status(Title=Title,component=f"{operator_name, test_name, Device_name} ==> {testtype} :- {file_name}",status="FAILED",comments=statement)
                result_status.put(r_status)
        elif len(combined_list_of_files) == 0 and len(dataoftesttypeindividualpop)!= 0 and len(dataoftesttypecombineexport) == 0:
            r_status = status(Title=Title, component=f"{operator_name, test_name, Device_name} ==> {testtype} :- []",status="FAILED", comments=f"There is no csv is found where file name should contain {testtype} but individual pop up contains a data and lenght is {str(len(dataoftesttypeindividualpop))}")
            result_status.put(r_status)
        elif len(combined_list_of_files) != 0 and len(dataoftesttypeindividualpop) == 0 and len(dataoftesttypecombineexport) != 0:
            file_name = (combined_list_of_files[0].split("\\")[-1])
            r_status = status(Title=Title, component=f"{operator_name, test_name, Device_name} ==> {testtype} :- {file_name}",status="FAILED", comments=f"There is csv is found where csv file as data and lenght is {str(len(dataoftesttypecombineexport))} but individual pop up doesn't contains a data")
            result_status.put(r_status)
        elif len(combined_list_of_files) != 0 and len(dataoftesttypeindividualpop) != 0 and len(dataoftesttypecombineexport) == 0:
            file_name = (combined_list_of_files[0].split("\\")[-1])
            r_status = status(Title=Title, component=f"{operator_name, test_name, Device_name} ==> {testtype} :- {file_name}",status="FAILED", comments=f"There is csv is found where csv file doesnt contains data but individual pop up contains a data and lenght is {str(len(dataoftesttypeindividualpop))}")
            result_status.put(r_status)
    except Exception as e:
        pass

def update_individualpopup_result_to_excel(result_status,data_difference,data_same,excelpath):
    dataframe_status = []
    dataframe_difference =[]
    dataframe_same = []
    combined_status_df = "None"
    combined_difference_df ="None"
    combined_same_df = "None"
    while not result_status.empty():
        updatecomponentstatus2 = result_status.get()
        df_status = pd.DataFrame(updatecomponentstatus2)
        dataframe_status.append(df_status)
    while not data_same.empty():
        same = data_same.get()
        df_same = pd.DataFrame(same)
        dataframe_same.append(df_same)
    while not data_difference.empty():
        difference = data_difference.get()
        df_difference = pd.DataFrame(difference)
        dataframe_difference.append(df_difference)
    if len(dataframe_status) != 0:
        combined_status_df = pd.concat(dataframe_status, ignore_index=True)
    if len(dataframe_difference) != 0:
        statement = "Values are difference b/w individual popup v/s combine export"
        with allure.step(statement):
            combined_difference_df = pd.concat(dataframe_difference, ignore_index=True)
            html_table = combined_difference_df.to_html(index=False, escape=False)
            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
    if len(dataframe_same) !=0:
        statement = "Values are same b/w individual popup v/s combine export"
        with allure.step(statement):
            combined_same_df = pd.concat(dataframe_same, ignore_index=True)
            html_table = combined_same_df.to_html(index=False, escape=False)
            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_componentstatus = workbook["COMPONENTSTATUS"]
    data_matchsheet = workbook["IPU_vs_CE_DATA_MATCH"]
    data_not_matchsheet = workbook["IPU_vs_CE_DATA_NOT_MATCH"]
    if len(dataframe_status) != 0:
        update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
    if len(dataframe_difference) != 0:
        update_excel_datavalidation_individualpopup_each_testcase_openpyxl(df=combined_difference_df, worksheet=data_not_matchsheet)
    if len(dataframe_same) != 0:
        update_excel_datavalidation_individualpopup_each_testcase_openpyxl(df=combined_same_df, worksheet=data_matchsheet)
    workbook.save(excelpath)
    workbook.close()
def update_excel_datavalidation_individualpopup_each_testcase_openpyxl(df,worksheet):
    """
        Update the high-level Excel report for data validation of individualpopup.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        color_mapping = {
            'STARTHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            'ENDHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            "Same": PatternFill(start_color="C2FFAD", end_color="C2FFAD", fill_type="solid"),  # Green
            "Difference": PatternFill(start_color='FF9999', end_color='FF9999', fill_type="solid"),   #light Red
            "Key name can't find in csv": PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid") , # Yellow
            "Row Start": PatternFill(start_color='ADD8E8', end_color='ADD8E9', fill_type="solid"),
            "is equal": PatternFill(start_color="C2FFAD", end_color="C2FFAD", fill_type="solid"),  # light Green
            "is not equal": PatternFill(start_color='FF9999', end_color='FF9999', fill_type="solid"),  # Red
            # "Key name can't find in csv file": PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type="solid")  # Light Yellow
        }
        # Find the last used row in the sheet
        last_row = worksheet.max_row
        start_row = last_row + 1
        # Insert the DataFrame into the worksheet
        for index, row in df.iterrows():
            worksheet.append(row.tolist())
        # Apply color formatting to the entire range
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=2, max_col=2),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=3, max_col=3),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=4, max_col=4),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=5, max_col=5),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        # Set colors for File columns
        for i in range(start_row, start_row + len(df)):
            worksheet.cell(row=i, column=1).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass

def download_combine_export_csv_and_combine_binary_export(Title,driver,cbe_flag,ce_flag,downloadfilespath,combine_binary_export_folder_path,combine_export_folder_path,excelpath):
    try:
        List_of_options_txts = [["Combined Binary Export"],["Combined Export"]]
        combine_binary_export_csv = []
        combine_export_csv = []
        alert_text = None
        for List_of_options_txt in List_of_options_txts:
            if List_of_options_txt == ["Combined Binary Export"]:
                if os.path.exists(combine_binary_export_folder_path):
                    combine_binary_export_csv = [f for f in os.listdir(combine_binary_export_folder_path) if f.endswith(".csv")]
                    if len(combine_binary_export_csv) == 0:
                        change_the_download_path(driver, combine_binary_export_folder_path)
                        flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                        waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=10)
                        combine_binary_export_csv = [f for f in os.listdir(combine_binary_export_folder_path) if f.endswith(".csv")]
                        change_the_download_path(driver, downloadfilespath)
                elif not os.path.exists(combine_binary_export_folder_path):
                    os.mkdir(combine_binary_export_folder_path)
                    if List_of_options_txt == ["Combined Binary Export"]:
                        change_the_download_path(driver, combine_binary_export_folder_path)
                    waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=10)
                    try:
                        wait_for_loading_elements(driver)
                    except:
                        pass
                    flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                    waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=10)
                    combine_export_csv = [f for f in os.listdir(combine_binary_export_folder_path) if f.endswith(".csv")]
                    change_the_download_path(driver, downloadfilespath)
                if alert_text == None and len(combine_binary_export_csv) != 0:
                    cbe_flag = True
            elif List_of_options_txt == ["Combined Export"]:
                if os.path.exists(combine_export_folder_path):
                    combine_export_csv = [f for f in os.listdir(combine_export_folder_path) if f.endswith(".csv")]
                    if len(combine_export_csv) == 0:
                        change_the_download_path(driver, combine_export_folder_path)
                        flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                        waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=10)
                        combine_export_csv = [f for f in os.listdir(combine_export_folder_path) if f.endswith(".csv")]
                        change_the_download_path(driver, downloadfilespath)
                elif not os.path.exists(combine_export_folder_path):
                    os.mkdir(combine_export_folder_path)
                    if List_of_options_txt == ["Combined Export"]:
                        change_the_download_path(driver, combine_export_folder_path)
                    waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=10)
                    try:
                        wait_for_loading_elements(driver)
                    except:
                        pass
                    flag, alert_text = select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                    waiting_loading_element_while_multiple_campaigns_or_single_campaign(time_single_campaign=3,time_multiple_campaigns=10)
                    combine_export_csv = [f for f in os.listdir(combine_export_folder_path) if f.endswith(".csv")]
                    change_the_download_path(driver,downloadfilespath)
                if alert_text == None and len(combine_export_csv) != 0:
                    ce_flag = True
    except Exception as e:
        pass
def Nqc(driver,downloadfilespath,excelpath):
    Title ="NQC table data vs operator comparsion"
    result_status = queue.Queue()
    result_same = queue.Queue()
    result_Difference = queue.Queue()
    cbe_flag = False
    ce_flag = False
    combine_binary_export_folder_path = downloadfilespath + "\\Combine_binary_export"
    combine_export_folder_path = downloadfilespath + "\\Combine_export"
    runvalue = Testrun_mode(value="NQC table data validation")
    if "Yes".lower() == runvalue[-1].strip().lower():
        try:
            download_combine_export_csv_and_combine_binary_export(Title, driver, cbe_flag, ce_flag, downloadfilespath,combine_binary_export_folder_path, combine_export_folder_path,excelpath)
            combined_list = [file for file in glob.glob(combine_export_folder_path + "\\*.csv")]
            combine_binary_list = [file for file in glob.glob(combine_binary_export_folder_path + "\\*.csv")]
            if len(combine_binary_list) != 0 and len(combined_list) != 0:
                operator_comparisonsheet = pd.read_excel(excelpath, sheet_name="OPERATOR_COMPARISON")
                start_row_index = None
                # testtypes = ["Ping Test","Download Test","Upload Test","HTTP Speed Download Test","HTTP Speed Upload Test","TCPiPerfDL","TCPiPerfUL","Sent Sms","Received Sms","Failed Sms","WebTest"]
                try:
                    df_testtype = pd.read_excel(config.nqc_testdata_excel_path, sheet_name="Nqc_table_data")
                except Exception as e:
                    with allure.step(f"Check {config.nqc_testdata_excel_path}"):
                        print(f"Check {config.nqc_testdata_excel_path}")
                        assert False
                testtypes = df_testtype["NQC table data"].tolist()
                csv_testtype = df_testtype.set_index('NQC table data').apply(lambda x: x.dropna().tolist(), axis=1).to_dict()
                for testtype in testtypes:
                    my_dict_list =[]
                    # Remove spaces from strings in operator_comparisonsheet
                    for i, row in operator_comparisonsheet.iterrows():
                        for col in row.index:
                            cell_value = str(row[col]).replace(' ', '')
                            if re.fullmatch(cell_value.strip(),testtype.replace(' ', '').lower().strip(),re.IGNORECASE):
                                start_row_index = i
                                end_row_index = 0
                                s_flag = 0
                                if start_row_index != None:
                                    print(testtype)
                                    s_flag = 1
                                    for i, row in operator_comparisonsheet.iterrows():
                                        if i >= start_row_index:
                                            for col in row.index:
                                                cell_value = str(row[col]).replace(' ', '')
                                                if re.fullmatch(cell_value.strip(), "ENDHERE", re.IGNORECASE):
                                                    end_row_index = i
                                                    break
                                            if end_row_index != 0:
                                                break
                                    print(end_row_index)
                                    # Select the desired rows based on the indices
                                    selected_rows = operator_comparisonsheet.iloc[start_row_index:end_row_index + 1]
                                    # Optional: Reset the index of the selected rows
                                    selected_rows = selected_rows.reset_index(drop=True)
                                    selected_rows_list = selected_rows.values.tolist()
                                    filtered_selected_rows_list = [[val for val in row if not pd.isna(val)] for row in selected_rows_list]
                                    if len(filtered_selected_rows_list[1]) > 2:
                                        for i in range(len(filtered_selected_rows_list[1])-1):
                                            my_dict = {}
                                            my_dict = {item[0]: item[i+1] for item in filtered_selected_rows_list[1:len(selected_rows_list) - 1]}
                                            my_dict_list.append(my_dict)
                                    elif len(filtered_selected_rows_list[1]) == 2:
                                        my_dict = {item[0]: item[1] for item in filtered_selected_rows_list[1:len(selected_rows_list) - 1]}
                                        my_dict_list.append(my_dict)
                                    # print("filtered_selected_rows_list-----",filtered_selected_rows_list)
                    #              [2:len(selected_rows_list)-1]
                    for my_dict in my_dict_list:
                        if len(my_dict) !=0:
                            testtypetxt = []
                            test = testtype.strip()  # Remove leading and trailing spaces from test
                            for pattern, values in csv_testtype.items():
                                if pattern.lower() == test.lower():
                                    testtypetxt = values
                                    break
                                else:
                                    testtypetxt = []
                            params_combines = ["IperfTest", "SpeedTest"]
                            combined_list_of_files = [file for file in glob.glob(combine_export_folder_path + "\\*.csv") if re.fullmatch(testtypetxt[1], (file.split("\\")[-1]).split("_")[0],re.IGNORECASE)]
                            combine_binary_list_of_files = [file for file in glob.glob(combine_binary_export_folder_path + "\\*.csv") if re.fullmatch(testtypetxt[0], (file.split("\\")[-1]).split("_")[0],re.IGNORECASE)]
                            dataoftesttypecombinebinaryexport = None
                            dataoftesttypecombineexport = None
                            if len(combined_list_of_files) != 0 and len(combine_binary_list_of_files) !=0:
                                df_ce = pd.read_csv(combined_list_of_files[0])
                                df_cbe = pd.read_csv(combine_binary_list_of_files[0])
                                if not any(re.fullmatch(params_combine,(combined_list_of_files[0].split("\\")[-1]).split("_")[0],re.IGNORECASE) for params_combine in params_combines):
                                    dataoftesttypecombineexport =  df_ce
                                    dataoftesttypecombinebinaryexport = df_cbe
                                elif any(re.fullmatch(params_combine,(combined_list_of_files[0].split("\\")[-1]).split("_")[0],re.IGNORECASE) for params_combine in params_combines):
                                    dataoftesttypecombineexport = df_ce[df_ce['TestType'].str.contains(testtypetxt[2])]
                                    dataoftesttypecombinebinaryexport = df_cbe
                            if str(testtype).replace(" ","").lower() == "Ping Test".replace(" ","").lower():
                                Ping_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict, result_same, result_status, result_Difference)
                            elif str(testtype).replace(" ","").lower() == "Download Test".replace(" ","").lower():
                                Download_test_or_upload_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
                            elif str(testtype).replace(" ","").lower() == "Upload Test".replace(" ","").lower():
                                Download_test_or_upload_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
                            elif str(testtype).replace(" ","").lower() == "HTTP Speed Download Test".replace(" ","").lower():
                                httpdownload_test_or_httpupload_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
                            elif str(testtype).replace(" ","").lower() == "HTTP Speed Upload Test".replace(" ","").lower():
                                httpdownload_test_or_httpupload_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "TCPiPerfDL".replace(" ", "").lower():
                                TCPiPerfDownloadTest_or_TCPiPerfUploadTest(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport, my_dict, result_Difference, result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "TCPiPerfUL".replace(" ", "").lower():
                                TCPiPerfDownloadTest_or_TCPiPerfUploadTest(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport, my_dict, result_Difference, result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "UDPiPerfDL".replace(" ", "").lower():
                                UDPiPerfDownloadTest_or_UDPiPerfUploadTest(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport, my_dict, result_Difference, result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "UDPiPerfUL".replace(" ", "").lower():
                                UDPiPerfDownloadTest_or_UDPiPerfUploadTest(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport, my_dict, result_Difference, result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "Sent Sms".replace(" ", "").lower() or str(testtype).replace(" ", "").lower() == "Received Sms".replace(" ", "").lower() or str(testtype).replace(" ", "").lower() == "Failed Sms".replace(" ", "").lower():
                                sms_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "WebTest".replace(" ", "").lower():
                                web_test(Title, testtype, dataoftesttypecombinebinaryexport,dataoftesttypecombineexport, my_dict, result_Difference, result_same, result_status)
                            elif str(testtype).replace(" ", "").lower() == "StreamTest".replace(" ", "").lower():
                                Stream_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
                            elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower() or str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                                call_test_and_failed_calls(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status)
            elif len(combine_binary_list) != 0 or len(combined_list) != 0:
                if len(combine_binary_list) == 0 and len(combined_list) != 0:
                    r_result = status(Title=Title,component="combine_binary_csv_files", status="FAILED",comments=f'combine_binary_csv_files is not present in {combine_binary_export_folder_path}.')
                    result_status.put(r_result)
                elif len(combine_binary_list) != 0 and len(combined_list) == 0:
                    r_result = status(Title=Title, component="combine_export_csv_files", status="FAILED",comments=f'combine_export_csv_files is not present in {combine_export_folder_path}.')
                    result_status.put(r_result)
                elif len(combine_binary_list) == 0 and len(combined_list) == 0:
                    r_result = status(Title=Title, component="combine_export_csv_files and combine_binary_csv_files", status="FAILED",comments=f'combine_export_csv_files and combine_binary_csv_files is not present in {combine_export_folder_path} and {combine_binary_export_folder_path}.')
                    result_status.put(r_result)
        except Exception as e:
                pass
    elif "No".lower() == runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute", excelpath)
            pass
    try:
        with allure.step("update_NQC_table_data_result_to_excel"):
            update_NQC_table_data_result_to_excel(result_status, result_Difference, result_same, excelpath)
    except Exception as e:
        with allure.step("failed to update_NQC_table_data_result_to_excel"):
            pass

def update_NQC_table_data_result_to_excel(result_status,data_difference,data_same,excelpath):
    dataframe_status = []
    dataframe_difference =[]
    dataframe_same = []
    combined_status_df = "None"
    combined_difference_df ="None"
    combined_same_df = "None"
    while not result_status.empty():
        updatecomponentstatus2 = result_status.get()
        df_status = pd.DataFrame(updatecomponentstatus2)
        dataframe_status.append(df_status)
    while not data_same.empty():
        same = data_same.get()
        df_same = pd.DataFrame(same)
        dataframe_same.append(df_same)
    while not data_difference.empty():
        difference = data_difference.get()
        df_difference = pd.DataFrame(difference)
        dataframe_difference.append(df_difference)
    if len(dataframe_status) != 0:
        combined_status_df = pd.concat(dataframe_status, ignore_index=True)
    if len(dataframe_difference) != 0:
        statement = "Values are difference b/w NQC table data vs operator comparsion"
        with allure.step(statement):
            combined_difference_df = pd.concat(dataframe_difference, ignore_index=True)
            html_table = combined_difference_df.to_html(index=False, escape=False)
            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
    if len(dataframe_same) !=0:
        statement = "Values are same b/w NQC table data vs operator comparsion"
        with allure.step(statement):
            combined_same_df = pd.concat(dataframe_same, ignore_index=True)
            html_table = combined_same_df.to_html(index=False, escape=False)
            allure.attach(html_table, name="HTML Table", attachment_type=allure.attachment_type.HTML)
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_componentstatus = workbook["COMPONENTSTATUS"]
    data_matchsheet = workbook["NQC_vs_OC_DATA_MATCH"]
    data_not_matchsheet = workbook["NQC_vs_OC_DATA_NOT_MATCH"]
    if len(dataframe_status) != 0:
        update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
    if len(dataframe_difference) != 0:
        update_excel_datavalidation_NQC_table_data_each_testcase_openpyxl(df=combined_difference_df, worksheet=data_not_matchsheet)
    if len(dataframe_same) != 0:
        update_excel_datavalidation_NQC_table_data_each_testcase_openpyxl(df=combined_same_df, worksheet=data_matchsheet)
    workbook.save(excelpath)
    workbook.close()


update_excel_datavalidation_NQC_table_data_each_testcase_openpyxl = update_excel_datavalidation_individualpopup_each_testcase_openpyxl

def Ping_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_same,result_status,result_Difference):
    flag_difference = False
    compared_data = []
    r_result = "None"
    file_type = None
    operator_name = my_dict['Latency']
    compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
    for key, value in my_dict.items():
        value = str(value).replace(" ","").lower()
        operator_name = re.sub(r'\D', '', operator_name)
        dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
        dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
        dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

        dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
        dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
        count = "None"
        percentage_key = "None"
        comprasion_data ="None"
        # Extract the range and count from the value
        if re.search("count",value,re.IGNORECASE):
            percentage_key, count_info = value.split('(count:')
            percentage_key = str(percentage_key).replace("%","")
            count, _ = count_info.split(')')
        key =  str(key).replace(" ","").lower()
        # Calculate the range count based on the key's range
        if re.search("-",key,re.IGNORECASE):
            try:
                df_data = []
                file_type = "CBE:- "
                parameter = "RTTAvg"
                try:
                    df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                except Exception as e:
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                    compared_data.append(comprasion_data)
                    raise e
                start_range, end_range = key.split('-')
                start_range = extract_numeric_value(text=start_range)
                end_range = extract_numeric_value(text=end_range)
                # start_range = int(start_range.replace('ms', '').strip())
                # end_range = int(end_range.replace('ms', '').strip())
                # Filter df_data based on the range
                filtered_df_data = [v for v in df_data if start_range <= v <= end_range]
                # Calculate the percentage for this range
                percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                formatted_percentage = f"{percentage:.4f}"
                count1 = len(filtered_df_data)
                if compare_values(count1,count) and compare_values(formatted_percentage,percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation":f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                elif not compare_values(count1,count) and not compare_values(formatted_percentage,percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    flag_difference = True
                elif not compare_values(count1,count) and compare_values(formatted_percentage,percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {formatted_percentage} but count:{count} is not equal to calculated count:{count1}."}
                    flag_difference = True
                elif compare_values(count1,count) and not compare_values(formatted_percentage,percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                    flag_difference = True
            except Exception as e:
                continue
        elif re.search('>=',key,re.IGNORECASE):
            try:
                df_data = []
                file_type = "CBE:- "
                parameter = "RTTAvg"
                try:
                    df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                except Exception as e:
                    comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                    compared_data.append(comprasion_data)
                    raise e
                # _,start_range = key.split('>=')
                start_range = extract_numeric_value(text=key)
                # start_range = int(start_range.replace('ms', '').strip())
                filtered_df_data = [v for v in df_data if start_range <= v]
                percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                formatted_percentage = f"{percentage:.4f}"
                count1 = len(filtered_df_data)
                if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    flag_difference = True
                elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {percentage_key} but count:{count} is not equal to calculated count:{count1}."}
                    flag_difference = True
                elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                    flag_difference = True
            except Exception as e:
                continue
        elif re.search("Geo".lower(),key,re.IGNORECASE):
            try:
                df_data = []
                file_type = "CBE:- "
                parameter = None
                try:
                    parameter = "Iteration"
                    df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                except Exception as e:
                    comprasion_data = {"File": file_type + testtype,"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                    compared_data.append(comprasion_data)
                    raise e
                # Compare Total Geo samples with the length of df_data
                total_geo_samples = value
                if compare_values(total_geo_samples,len(df_data)) :
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                elif not compare_values(total_geo_samples,len(df_data)):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                    flag_difference = True
            except Exception as e:
                continue
        elif re.search("Highest".lower(), key, re.IGNORECASE):
            try:
                df_data = []
                file_type = "CE:- "
                parameter = "RTTMax(ms)"
                try:
                    df_data = dataoftesttypecombineexport[parameter].tolist()
                except Exception as e:
                    comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                    compared_data.append(comprasion_data)
                    raise e
                # Compare Highest Latency (ms) with the highest value in df_data
                highest_latency_ms = value
                if compare_values(highest_latency_ms, max(df_data)) :
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{max(df_data)}","Data validation": f"{key}:{highest_latency_ms} is equal to calculated {key}:{max(df_data)}"}
                    # flag_difference = True
                elif not compare_values(highest_latency_ms, max(df_data)) :
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{max(df_data)}","Data validation": f"{key}:{highest_latency_ms} is not equal to calculated {key}:{max(df_data)}"}
                    flag_difference = True
            except Exception as e:
                continue
        elif re.search("Lowest".lower(), key, re.IGNORECASE):
            try:
                df_data = []
                file_type = "CE:- "
                parameter = "RTTMin(ms)"
                try:
                    df_data = dataoftesttypecombineexport[parameter].tolist()
                except Exception as e:
                    comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                    compared_data.append(comprasion_data)
                    raise e
                # Compare Lowest Latency (ms) with the lowest value in df_data
                lowest_latency_ms = value
                df_data_values = [value for value in df_data if value >= 0]
                if compare_values(lowest_latency_ms,  min(df_data_values)):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{min(df_data_values)}","Data validation": f"{key}:{lowest_latency_ms} is equal to calculated {key}:{min(df_data_values)}"}
                elif not compare_values(lowest_latency_ms,  min(df_data_values)) :
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{min(df_data_values)}","Data validation": f"{key}:{lowest_latency_ms} is not equal to calculated {key}:{min(df_data_values)}"}
                    flag_difference = True
            except Exception as e:
                continue
        elif re.search('DroppedPackets'.lower(),key , re.IGNORECASE):
            try:
                df_data = []
                file_type = "CBE:- "
                parameter = "RTTAvg"
                try:
                    df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                except Exception as e:
                    comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                    compared_data.append(comprasion_data)
                    raise e
                # Filter and extract only the negative values
                negative_values = [value for value in df_data if value < 0]
                percentage = (len(negative_values) / len(df_data)) * 100 if len(negative_values) > 0 else 0
                formatted_percentage = f"{percentage:.4f}"
                count1 = len(negative_values)
                if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    flag_difference = True
                elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {percentage_key} but count:{count} is not equal to calculated count:{count1}."}
                    flag_difference = True
                elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                    comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                    flag_difference = True
            except Exception as e:
                continue
        else:
            comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}
        if comprasion_data != "None":
            compared_data.append(comprasion_data)
    compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
    if flag_difference == False:
        r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
        result_same.put(compared_data)
        result_status.put(r_result)
    elif flag_difference == True:
        r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
        result_Difference.put(compared_data)
        result_status.put(r_result)
    print(r_result)
    print(compared_data)
def Download_test_or_upload_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status):
    try:
        flag_difference = False
        compared_data = []
        r_result = "None"
        file_type = None
        operator_name = my_dict['Download Speed'or 'Upload Speed' or 'Http Download Speed' or 'Http Upload Speed']
        compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
        for key, value in my_dict.items():
            value = str(value).replace(" ", "").lower()
            operator_name = re.sub(r'\D', '', operator_name)
            dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
            dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
            dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

            dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
            dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
            df_data = "None"
            count = "None"
            percentage_key = "None"
            comprasion_data = "None"
            # Extract the range and count from the value
            if re.search("count", value, re.IGNORECASE):
                percentage_key, count_info = value.split('(count:')
                percentage_key = str(percentage_key).replace("%", "")
                count, _ = count_info.split(')')
            key = str(key).replace(" ", "").lower()
            if re.search('above', key, re.IGNORECASE) or re.search('below', key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # _, start_range = key.split('above')
                    start_range = extract_numeric_value(text=key)
                    # start_range = int(start_range.replace('ms', '').strip())
                    filtered_df_data = []
                    if re.search('above', key, re.IGNORECASE):
                        filtered_df_data = [v for v in df_data if start_range < v]
                    elif re.search('below', key, re.IGNORECASE):
                        filtered_df_data = [v for v in df_data if start_range > v]
                    percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                    formatted_percentage = f"{percentage:.4f}"
                    count1 = len(filtered_df_data)
                    if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                    elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {formatted_percentage} but count:{count} is not equal to calculated count:{count1}."}
                        flag_difference = True
                    elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                except Exception as e:
                    continue
            # Calculate the range count based on the key's range
            elif re.search("-", key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    start_range, end_range = key.split('-')
                    start_range = extract_numeric_value(text=start_range)
                    end_range = extract_numeric_value(text=end_range)
                    # start_range = float(start_range.replace('ms', '').strip())
                    # end_range = float(end_range.replace('ms', '').strip())
                    # Filter df_data based on the range
                    filtered_df_data = [v for v in df_data if start_range <= v <= end_range]
                    # Calculate the percentage for this range
                    percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                    formatted_percentage = f"{percentage:.4f}"
                    count1 = len(filtered_df_data)
                    if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                    elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {formatted_percentage} but count:{count} is not equal to calculated count:{count1}."}
                        flag_difference = True
                    elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Average".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Average(mbps) with the lowest value in df_data
                    Lowest_Recorded_mbps = round(float(value),2)
                    df_data_values = [value1 for value1 in df_data if value1 >= 0]
                    average = sum(df_data_values) / len(df_data_values)
                    average = f"{average:.2f}"
                    # Original number
                    average = Decimal(average)
                    # Round off to two decimal places using ROUND_HALF_UP rounding method
                    average = average.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    if compare_values(Lowest_Recorded_mbps,average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Lowest_Recorded_mbps} is equal to calculated {key}:{average}"}
                    elif not compare_values(Lowest_Recorded_mbps, min(df_data_values)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Lowest_Recorded_mbps} is not equal to calculated {key}:{average}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Peak".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Peak Speed (mbps) with the highest value in df_data
                    Peak_Speed_mbps = value
                    Peak_Speed = max(df_data)
                    if compare_values(Peak_Speed_mbps,Peak_Speed):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Peak_Speed}","Data validation": f"{key}:{Peak_Speed_mbps} is equal to calculated {key}:{Peak_Speed}"}
                    elif not compare_values(Peak_Speed_mbps, Peak_Speed):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Peak_Speed}","Data validation": f"{key}:{Peak_Speed_mbps} is not equal to calculated {key}:{Peak_Speed}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Lowest".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Lowest Recorded (mbps) with the lowest value in df_data
                    Lowest_Recorded_mbps = value
                    Lowest_Speed = min(df_data)
                    if compare_values(Lowest_Recorded_mbps, Lowest_Speed):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Lowest_Speed}","Data validation": f"{key}:{Lowest_Recorded_mbps} is equal to calculated {key}:{Lowest_Speed}"}
                    elif not compare_values(Lowest_Recorded_mbps, Lowest_Speed):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Lowest_Speed}","Data validation": f"{key}:{Lowest_Recorded_mbps} is not equal to calculated {key}:{Lowest_Speed}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Geo".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    if compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            else:
                comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}
            if comprasion_data != "None":
                compared_data.append(comprasion_data)
        compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
        if flag_difference == False:
            r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
            result_same.put(compared_data)
            result_status.put(r_result)
        elif flag_difference == True:
            r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
            result_Difference.put(compared_data)
            result_status.put(r_result)
        print(r_result)
        print(compared_data)
    except Exception as e:
        pass

httpdownload_test_or_httpupload_test = Download_test_or_upload_test

def TCPiPerfDownloadTest_or_TCPiPerfUploadTest(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status):
    try:
        flag_difference = False
        compared_data = []
        r_result = "None"
        operator_name = my_dict['TCPiPerfDownloadTest' or 'TCPiPerfUploadTest' or 'UDPiPerfDownloadTest' or 'UDPiPerfUploadTest']
        compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
        for key, value in my_dict.items():
            value = str(value).replace(" ", "").lower()
            operator_name = re.sub(r'\D', '', operator_name)
            dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
            dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
            dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

            dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
            dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
            df_data = "None"
            comprasion_data = "None"
            key = str(key).replace(" ", "").lower()
            if compare_values("AvgBitRate(Mbps)".lower(), key):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Average(mbps) with the lowest value in df_data
                    Avg_mbps = f"{value:.2f}"
                    df_data_values = [value1 for value1 in df_data if value1 >=0]
                    average = sum(df_data_values) / len(df_data)
                    average = f"{average:.2f}"
                    if compare_values(Avg_mbps, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Avg_mbps} is equal to calculated {key}:{average}"}
                    elif not compare_values(Avg_mbps, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Avg_mbps} is not equal to calculated {key}:{average}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Peak".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "BitRate(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Peak Speed (mbps) with the highest value in df_data
                    Peak_Speed_mbps = value
                    Peak_Speed_mbps = Decimal(Peak_Speed_mbps)
                    # Round off to two decimal places using ROUND_HALF_UP rounding method
                    Peak_Speed_mbps = Peak_Speed_mbps.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    Peak_Speed = max(df_data)
                    Peak_Speed = Decimal(Peak_Speed)
                    # Round off to two decimal places using ROUND_HALF_UP rounding method
                    Peak_Speed = Peak_Speed.quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    if compare_values(Peak_Speed_mbps,Peak_Speed ):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Peak_Speed}","Data validation": f"{key}:{Peak_Speed_mbps} is equal to calculated {key}:{Peak_Speed}"}
                    elif not compare_values(Peak_Speed_mbps, Peak_Speed):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Peak_Speed}","Data validation": f"{key}:{Peak_Speed_mbps} is not equal to calculated {key}:{Peak_Speed}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("TransferredData".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "TransferData"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Peak Speed (mbps) with the highest value in df_data
                    total_Transferred_mbps = f"{float(value):.2f}"
                    sum_df_data = sum(df_data)
                    sum_df_data = f"{float(sum_df_data):.2f}"
                    if compare_values(total_Transferred_mbps,sum_df_data):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{sum_df_data}", "Data validation": f"{key}:{total_Transferred_mbps} is equal to calculated {key}:{sum_df_data}"}
                    elif not compare_values(total_Transferred_mbps, sum_df_data):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{sum_df_data}","Data validation": f"{key}:{total_Transferred_mbps} is not equal to calculated {key}:{sum_df_data}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("TotalIterations".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}', "map view Operator": key,"map view Operator value": f"{value}","calculated csv value": f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e

                    df_data = set(list(df_data))
                    # Compare Peak Speed (mbps) with the highest value in df_data
                    TotalIterations = value
                    if compare_values(TotalIterations, max(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{max(df_data)}",
                                           "Data validation": f"{key}:{TotalIterations} is equal to calculated {key}:{max(df_data)}"}
                    elif not compare_values(TotalIterations, max(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{max(df_data)}",
                                           "Data validation": f"{key}:{TotalIterations} is not equal to calculated {key}:{max(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Hostname".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "IP Address"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    df_data = list(set(df_data))
                    # Compare Peak Speed (mbps) with the highest value in df_data
                    Hostname = value
                    if compare_values(Hostname, df_data[0]):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data[0]}","Data validation": f"{key}:{Hostname} is equal to calculated {key}:{df_data[0]}"}
                    elif not compare_values(Hostname,df_data[0]):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data[0]}","Data validation": f"{key}:{Hostname} is not equal to calculated {key}:{df_data[0]}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Geo".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    if compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("UserDefinedBandwidth(Mbps)".lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "UserDefined Bandwidth(Mb)"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    calculate_bandwidth = list(set(df_data))
                    if compare_values(total_geo_samples,calculate_bandwidth[0]):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{calculate_bandwidth[0]}","Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{calculate_bandwidth[0]}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{calculate_bandwidth[0]}","Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{calculate_bandwidth[0]}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("AvgLossDatagram".lower(), key):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "LossDatagram"
                    average = None
                    Avglossdatagram = value
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                        def nr_dr_and_percentage(input):
                            # Extract numerical values and percentages
                            numerical_values_nr = [float(entry.split('/')[0]) for entry in input]
                            numerical_values_dr = [float((entry.split('/')[1]).split('(')[0]) for entry in input]
                            percentages = [float(entry.split('(')[1].strip('%)')) / 100 for entry in input]
                            numerical_values_nr1 = [nr for nr in numerical_values_nr if nr > 0]
                            numerical_values_dr1 = [dr for dr in numerical_values_dr if dr > 0]
                            percentages1 = [pr for pr in percentages if pr > 0]
                            # Calculate the average for numerical values and percentages, handling division by zero
                            average_numerical_nr = sum(numerical_values_nr1)
                            average_percentage = sum(percentages1) / len(percentages1) if len(percentages1) > 0 else float('nan')
                            average_numerical_dr = sum(numerical_values_dr1)
                            # Format the results back into the original formats
                            average_numerical_nr1 = int(average_numerical_nr)
                            average_numerical_dr1 = int(average_numerical_dr)
                            average_percentage = f"({average_percentage:.2%})"

                            format_output = f"{average_numerical_nr1}/{average_numerical_dr1}{average_percentage}"
                            print(format_output)
                            return format_output

                        average = nr_dr_and_percentage(input=df_data)
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Average(mbps) with the lowest value in df_data
                    if compare_values(Avglossdatagram, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Avglossdatagram} is equal to calculated {key}:{average}"}
                    elif not compare_values(Avglossdatagram, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Avglossdatagram} is not equal to calculated {key}:{average}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("AvgJitter".lower(), key):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "Jitter"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Average(mbps) with the lowest value in df_data
                    Avg_mbps = f"{value:.2f}"
                    df_data_values = [extract_numeric_value(text=value1) for value1 in df_data if value1 >= 0]
                    average = sum(df_data_values) / len(df_data)
                    average = f"{average:.2f}"
                    if compare_values(Avg_mbps, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Avg_mbps} is equal to calculated {key}:{average}"}
                    elif not compare_values(Avg_mbps, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{Avg_mbps} is not equal to calculated {key}:{average}"}
                        flag_difference = True
                except Exception as e:
                    continue
            else:
                comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}
            if comprasion_data != "None":
                compared_data.append(comprasion_data)
        compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
        if flag_difference == False:
            r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
            result_same.put(compared_data)
            result_status.put(r_result)
        elif flag_difference == True:
            r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
            result_Difference.put(compared_data)
            result_status.put(r_result)
        print(r_result)
        print(compared_data)
    except Exception as e:
        pass
UDPiPerfDownloadTest_or_UDPiPerfUploadTest = TCPiPerfDownloadTest_or_TCPiPerfUploadTest
def extract_numeric_value(text):
    # Replace non-numeric characters with whitespace and leading/trailing whitespaces
    numeric_part = ''.join(filter(lambda x: x.isdigit() or x in '. ', text)).strip()

    if numeric_part:
        return float(numeric_part)
    else:
        return None
def sms_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status):
    try:
        flag_difference = False
        compared_data = []
        r_result = "None"
        operator_name = my_dict['SMS']
        compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
        for key, value in my_dict.items():
            value = str(value).replace(" ", "").lower()
            operator_name = re.sub(r'\D', '', operator_name)
            dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
            dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
            dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

            dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
            dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
            df_data = "None"
            comprasion_data = "None"
            key = str(key).replace(" ", "").lower()
            if re.search("Total SMS sent success".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "SmsSentStatus"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Success".lower()]
                    total_success_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_success_count}"}
                    elif not compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_success_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total SMS sent Failed".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "SmsSentStatus"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent Failed with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() != "Success".lower()]
                    total_notsuccess_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_notsuccess_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_notsuccess_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_notsuccess_count}"}
                    elif not compare_values(SmsSentStatus, total_notsuccess_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_notsuccess_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_notsuccess_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total SMS Sent/Received Attempt".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS Sent/Received Attempt with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = df_data
                    total_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_count}"}
                    elif not compare_values(SmsSentStatus, total_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total SMS Received Success".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "SmsReceivedStatus"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS Received Success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Success".lower()]
                    total_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_count}"}
                    elif not compare_values(SmsSentStatus, total_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total SMS Received Failed".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "SmsReceivedStatus"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e                # Compare Total SMS Received Failed with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ", "").lower() != "Success".lower()]
                    total_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_count}"}
                    elif not compare_values(SmsSentStatus, total_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Highest".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "SmsSent/Received Duration"
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Sms".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                        else:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Highest with the lowest value in df_data
                    SmsSentStatus = f"{float(value):.2f}"
                    df_data_values = max(df_data)
                    df_data_values = f"{float(df_data_values):.2f}"
                    if compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_values}"}
                    elif not compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_values}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Lowest".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "SmsSent/Received Duration"
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Sms".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                        else:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                        # Compare Lowest with the lowest value in df_data
                    SmsSentStatus = f"{float(value):.2f}"
                    df_data_values = min(df_data)
                    df_data_values = f"{float(df_data_values):.2f}"
                    if compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_values}"}
                    elif not compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_values}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Geo".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    if compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            else:
                comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}
            if comprasion_data != "None":
                compared_data.append(comprasion_data)
        compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
        if flag_difference == False:
            r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
            result_same.put(compared_data)
            result_status.put(r_result)
        elif flag_difference == True:
            r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
            result_Difference.put(compared_data)
            result_status.put(r_result)
        print(r_result)
        print(compared_data)
    except Exception as e:
        pass

def web_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status):
    try:
        flag_difference = False
        compared_data = []
        r_result = "None"
        operator_name = my_dict['Operator']
        compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
        for key, value in my_dict.items():
            df_data_values = []
            value = str(value).replace(" ", "").lower()
            operator_name = re.sub(r'\D', '', operator_name)
            dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
            dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
            dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

            dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
            dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
            df_data = "None"
            comprasion_data = "None"
            file_type = None
            parameter = None
            key = str(key).replace(" ", "").lower()
            if re.search("URL".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "WebtestHost"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = list(set(df_data))
                    for df_data_value in df_data_values:
                        if compare_values(SmsSentStatus,df_data_value):
                            comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_value}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_value}"}
                        elif not compare_values(SmsSentStatus, df_data_value):
                            comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_value}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_value}"}
                            flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total Test Success".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "OK".lower()]
                    total_success_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_success_count}"}
                    elif not compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_success_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Aborted Test".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Aborted".lower()]
                    total_success_count = len(df_data_values)
                    if compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_success_count}"}
                    elif not compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_success_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total Test Attempted".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                        # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    total_success_count = len(df_data)
                    if compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_success_count}"}
                    elif not compare_values(SmsSentStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_success_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Total Test Failed".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                        df_data = df_data1[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Failed".lower()]
                    Total_Test_Attempted_count = len(df_data_values)
                    if compare_values(SmsSentStatus, Total_Test_Attempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Total_Test_Attempted_count}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{Total_Test_Attempted_count}"}
                    elif not compare_values(SmsSentStatus, Total_Test_Attempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{Total_Test_Attempted_count}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{Total_Test_Attempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("avgwebresponsetime(sec)".replace(" ", "").lower(), key) or compare_values("avgtotalwebpageloadtime(sec)".replace(" ", "").lower(), key) or compare_values("AvgVisibleTime(sec)".replace(" ", "").lower(), key) or compare_values("AvgPageLoadingTime(sec)".replace(" ", "").lower(),key) or compare_values("AvgWebPageSize(Mb)".replace(" ", "").lower(),key):
                try:
                    file_type = None
                    parameter = None
                    if compare_values("avgwebresponsetime(sec)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "WebResponseTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("AvgTotalWebPageLoadTime(sec)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "TotalWebPageLoadTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("AvgVisibleTime(sec)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "VisibleTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("AvgPageLoadingTime(sec)".replace(" ", "").lower(),key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "PageLoadingTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("AvgWebPageSize(Mb)".replace(" ", "").lower(),key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "WebPageSize(Mb)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    df_data_values = [value1 for value1 in df_data if value1 >= 0]
                    average = sum(df_data_values) / len(df_data)
                    average = f"{average:.2f}"
                    if compare_values(SmsSentStatus, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{average}"}
                    elif not compare_values(SmsSentStatus, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{average}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Highest".replace(" ", "").lower(), key, re.IGNORECASE):
                parameter = None
                try:
                    if compare_values("HighestLoadingTime(sec)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "PageLoadingTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("HighestWebResponseTime(sec)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "WebResponseTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    # Compare Highest with the lowest value in df_data
                    SmsSentStatus = f"{float(value):.2f}"
                    df_data_values = max(df_data)
                    df_data_values = f"{float(df_data_values):.2f}"
                    if compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_values}"}
                    elif not compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}","Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_values}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Lowest".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    if compare_values("LowestLoadingTime(sec))".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "PageLoadingTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("LowestWebResponseTime(sec)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "WebResponseTime(s)"
                        try:
                            df_data1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = df_data1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    # Compare Lowest with the lowest value in df_data
                    SmsSentStatus = f"{float(value):.2f}"
                    df_data_values = min(df_data)
                    df_data_values = f"{float(df_data_values):.2f}"
                    if compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_values}"}
                    elif not compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_values}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Good Scoring".replace(" ", "").lower(), key, re.IGNORECASE) or re.search("Fair Scoring".replace(" ", "").lower(), key, re.IGNORECASE) or re.search("Poor Scoring".replace(" ", "").lower(), key, re.IGNORECASE):
                df_data_values = []
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Webpage Performance Rating"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    SmsSentStatus = value
                    if re.search("Good Scoring".replace(" ", "").lower(), key, re.IGNORECASE):
                        df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Good".lower()]
                    elif re.search("Fair Scoring".replace(" ", "").lower(), key, re.IGNORECASE):
                        df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Fair".lower()]
                    elif re.search("Poor Scoring".replace(" ", "").lower(), key, re.IGNORECASE):
                        df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Poor".lower()]
                    total_count_Scoring = len(df_data_values)
                    if compare_values(SmsSentStatus, total_count_Scoring):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Scoring}",
                                           "Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{total_count_Scoring}"}
                    elif not compare_values(SmsSentStatus, total_count_Scoring):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Scoring}",
                                           "Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{total_count_Scoring}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Geo".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    if compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}",
                                           "Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}",
                                           "Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            else:
                comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}
            if comprasion_data != "None":
                compared_data.append(comprasion_data)
        compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
        if flag_difference == False:
            r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
            result_same.put(compared_data)
            result_status.put(r_result)
        elif flag_difference == True:
            r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
            result_Difference.put(compared_data)
            result_status.put(r_result)
        print(r_result)
        print(compared_data)
    except Exception as e:
        pass

def Stream_test(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status):
    try:
        flag_difference = False
        compared_data = []
        r_result = "None"
        file_type = "None"
        parameter = "None"
        operator_name = my_dict['Stream Test']
        compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
        for key, value in my_dict.items():
            value = str(value).replace(" ", "").lower()
            operator_name = re.sub(r'\D', '', operator_name)
            dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
            dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
            dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

            dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
            dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
            df_data = "None"
            count = "None"
            percentage_key = "None"
            comprasion_data = "None"
            df_data_values = None
            # Extract the range and count from the value
            if re.search("count", value, re.IGNORECASE):
                percentage_key, count_info = value.split('(count:')
                percentage_key = str(percentage_key).replace("%", "")
                count, _ = count_info.split(')')
            key = str(key).replace(" ", "").lower()
            if re.search("Video QoE Rating".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "PerformanceRating"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    if compare_values("VideoQoERating(Good)".replace(" ", "").lower(), key):
                        df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Good".lower()]
                    elif compare_values("VideoQoERating(Fair)".replace(" ", "").lower(), key):
                        df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Fair".lower()]
                    elif compare_values("VideoQoERating(Poor)".replace(" ", "").lower(), key):
                        df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "Poor".lower()]
                    total_count_Scoring = len(df_data_values)
                    # Compare Total SMS sent success with the lowest value in df_data
                    total_countStatus = value
                    if compare_values(total_countStatus, total_count_Scoring):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Scoring}",
                                           "Data validation": f"{key}:{total_countStatus} is equal to calculated {key}:{total_count_Scoring}"}
                    elif not compare_values(total_countStatus, total_count_Scoring):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Scoring}",
                                           "Data validation": f"{key}:{total_countStatus} is not equal to calculated {key}:{total_count_Scoring}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalStreamTestAttempted".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    total_count_Attempted = len(df_data)
                    # Compare Total SMS sent success with the lowest value in df_data
                    total_count_Attempted_Status = value
                    if compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}",
                                           "Data validation": f"{key}:{total_count_Attempted_Status} is equal to calculated {key}:{total_count_Attempted}"}
                    elif not compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}",
                                           "Data validation": f"{key}:{total_count_Attempted_Status} is not equal to calculated {key}:{total_count_Attempted}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalStreamTestSuccess".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ", "").lower() == "OK".lower()]
                    total_count_Attempted = len(df_data_values)
                    # Compare Total SMS sent success with the lowest value in df_data
                    total_count_Attempted_Status = value
                    if compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is equal to calculated {key}:{total_count_Attempted}"}
                    elif not compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is not equal to calculated {key}:{total_count_Attempted}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalStreamTestFailed".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ", "").lower() != "OK".lower()]
                    total_count_Attempted = len(df_data_values)
                    # Compare Total SMS sent success with the lowest value in df_data
                    total_count_Attempted_Status = value
                    if compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is equal to calculated {key}:{total_count_Attempted}"}
                    elif not compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is not equal to calculated {key}:{total_count_Attempted}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Highest".replace(" ", "").lower(), key, re.IGNORECASE) or re.search("Peak".replace(" ", "").lower(), key, re.IGNORECASE):
                parameter = None
                file_type = None
                try:
                    if compare_values("HighestLaunchTime(s)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "LaunchTime(s)"
                        try:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("HighestLoadTime(s)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "LoadTime(s)"
                        try:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("HighestStalledTime(s)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "StalledTime(s)"
                        try:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("PeakDLdatarate(mbps)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CBE:- "
                        parameter = "DL(Mbps)"
                        try:
                            df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    max_value = max(df_data)
                    # Compare Total SMS sent success with the lowest value in df_data
                    max_value_oc = value
                    if compare_values(max_value_oc, max_value):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{max_value}","Data validation": f"{key}:{max_value_oc} is equal to calculated {key}:{max_value}"}
                    elif not compare_values(max_value_oc, max_value):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{max_value}","Data validation": f"{key}:{max_value_oc} is not equal to calculated {key}:{max_value}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Lowest".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    if compare_values("LowestLaunchTime(s)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "LaunchTime(s)"
                        try:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("LowestLoadTime(s)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "LoadTime(s)"
                        try:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("LowestStalledTime(s)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CE:- "
                        parameter = "StalledTime(s)"
                        try:
                            df_data = dataoftesttypecombineexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                            compared_data.append(comprasion_data)
                            raise e
                    elif compare_values("LowestDLdatarate(mbps)".replace(" ", "").lower(), key):
                        df_data = []
                        file_type = "CBE:- "
                        parameter = "DL(Mbps)"
                        try:
                            df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                        except Exception as e:
                            comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine binary export'}
                            compared_data.append(comprasion_data)
                            raise e
                    min_value = min(df_data)
                    # Compare Total SMS sent success with the lowest value in df_data
                    min_value_oc = value
                    if compare_values(min_value_oc, min_value):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{min_value}","Data validation": f"{key}:{min_value_oc} is equal to calculated {key}:{min_value}"}
                    elif not compare_values(min_value_oc, min_value):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{min_value}","Data validation": f"{key}:{min_value_oc} is not equal to calculated {key}:{min_value}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("AverageDLdatarate(mbps)".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "DL(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    DL_data_rate = f"{float(value):.2f}"
                    df_data_values = [value1 for value1 in df_data if value1 >= 0]
                    average = sum(df_data_values) / len(df_data)
                    average = f"{average:.2f}"
                    if compare_values(DL_data_rate, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{DL_data_rate} is equal to calculated {key}:{average}"}
                    elif not compare_values(DL_data_rate, average):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{average}","Data validation": f"{key}:{DL_data_rate} is not equal to calculated {key}:{average}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search('above', key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "DL(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # _, start_range = key.split('above')
                    start_range = extract_numeric_value(text=key)
                    # start_range = int(start_range.replace('ms', '').strip())
                    filtered_df_data = []
                    filtered_df_data = [v for v in df_data if start_range < v]
                    percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                    formatted_percentage = f"{percentage:.4f}"
                    count1 = len(filtered_df_data)
                    if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                    elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {formatted_percentage} but count:{count} is not equal to calculated count:{count1}."}
                        flag_difference = True
                    elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                except Exception as e:
                    continue
            # Calculate the range count based on the key's range
            elif re.search("-", key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "DL(Mbps)"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    start_range, end_range = key.split('-')
                    start_range = extract_numeric_value(text=start_range)
                    end_range = extract_numeric_value(text=end_range)
                    # start_range = float(start_range.replace('ms', '').strip())
                    # end_range = float(end_range.replace('ms', '').strip())
                    # Filter df_data based on the range
                    filtered_df_data = [v for v in df_data if start_range <= v <= end_range]
                    # Calculate the percentage for this range
                    percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                    formatted_percentage = f"{percentage:.4f}"
                    count1 = len(filtered_df_data)
                    if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                    elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {formatted_percentage} but count:{count} is not equal to calculated count:{count1}."}
                        flag_difference = True
                    elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}","Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Geo".lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CBE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine binary export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    if compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}","Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalStalledTime(s)".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "StalledTime(s)"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    total_count_Attempted = sum(df_data)
                    # Compare Total SMS sent success with the lowest value in df_data
                    total_count_Attempted_Status = value
                    if compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is equal to calculated {key}:{total_count_Attempted}"}
                    elif not compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is not equal to calculated {key}:{total_count_Attempted}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalStalledCount(s)".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "StalledCount"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    total_count_Attempted = sum(df_data)
                    # Compare Total SMS sent success with the lowest value in df_data
                    total_count_Attempted_Status = value
                    if compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is equal to calculated {key}:{total_count_Attempted}"}
                    elif not compare_values(total_count_Attempted_Status, total_count_Attempted):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_count_Attempted}","Data validation": f"{key}:{total_count_Attempted_Status} is not equal to calculated {key}:{total_count_Attempted}"}
                        flag_difference = True
                except Exception as e:
                    continue
            else:
                comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}
            if comprasion_data != "None":
                compared_data.append(comprasion_data)
        compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
        if flag_difference == False:
            r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
            result_same.put(compared_data)
            result_status.put(r_result)
        elif flag_difference == True:
            r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
            result_Difference.put(compared_data)
            result_status.put(r_result)
        print(r_result)
        print(compared_data)
    except Exception as e:
        pass

def call_test_and_failed_calls(Title,testtype,dataoftesttypecombinebinaryexport,dataoftesttypecombineexport,my_dict,result_Difference,result_same,result_status):
    try:
        flag_difference = False
        compared_data = []
        r_result = "None"
        operator_name =my_dict['Call Setup Time']
        compared_data.append({"File": testtype,"map view Operator":"STARTHERE","map view Operator value":f"STARTHERE","calculated csv value":f"STARTHERE", "Data validation": "STARTHERE"})
        for key, value in my_dict.items():
            value = str(value).replace(" ", "").lower()
            operator_name = re.sub(r'\D', '', operator_name)
            dataoftesttypecombineexport['OperatorName'] = dataoftesttypecombineexport['OperatorName'].astype(str)
            dataoftesttypecombineexport['NetworkOperator'] = dataoftesttypecombineexport['NetworkOperator'].astype(str)
            dataoftesttypecombineexport = dataoftesttypecombineexport[dataoftesttypecombineexport['OperatorName'].str.contains(operator_name, na=False) | dataoftesttypecombineexport['NetworkOperator'].str.contains(operator_name, na=False)]

            dataoftesttypecombinebinaryexport['NetworkOperator'] = dataoftesttypecombinebinaryexport['NetworkOperator'].astype(str)
            dataoftesttypecombinebinaryexport = dataoftesttypecombinebinaryexport[ dataoftesttypecombinebinaryexport['NetworkOperator'].str.contains(operator_name, na=False)]
            count = "None"
            percentage_key = "None"
            comprasion_data = "None"
            df_data = "None"
            # Extract the range and count from the value
            if re.search("count", value, re.IGNORECASE):
                percentage_key, count_info = value.split('(count:')
                percentage_key = str(percentage_key).replace("%", "")
                count, _ = count_info.split(')')
            key = str(key).replace(" ", "").lower()
            if re.search("to", key, re.IGNORECASE) and not re.search("total", key, re.IGNORECASE) and not re.search("call", key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "CallSetupTime(s)"
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("NO_NETWORK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    start_range, end_range = key.split('to')
                    start_range = extract_numeric_value(text=start_range)
                    end_range = extract_numeric_value(text=end_range)
                    # start_range = int(start_range.replace('ms', '').strip())
                    # end_range = int(end_range.replace('ms', '').strip())
                    # Filter df_data based on the range
                    filtered_df_data = [v for v in df_data if start_range <= v <= end_range]
                    # Calculate the percentage for this range
                    percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                    formatted_percentage = f"{percentage:.4f}"
                    count1 = len(filtered_df_data)
                    if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                    elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {formatted_percentage} but count:{count} is not equal to calculated count:{count1}."}
                        flag_difference = True
                    elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search('>', key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "CallSetupTime(s)"
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("NO_NETWORK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # _,start_range = key.split('>=')
                    start_range = extract_numeric_value(text=key)
                    # start_range = int(start_range.replace('ms', '').strip())
                    filtered_df_data = [v for v in df_data if start_range < float(v)]
                    percentage = (len(filtered_df_data) / len(df_data)) * 100 if len(filtered_df_data) > 0 else 0
                    formatted_percentage = f"{percentage:.4f}"
                    count1 = len(filtered_df_data)
                    if compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                    elif not compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} and percentage: {percentage_key} is not equal to calculated count:{count1} and calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                    elif not compare_values(count1, count) and compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = percentage: {percentage_key} is equal to calculated percentage: {percentage_key} but count:{count} is not equal to calculated count:{count1}."}
                        flag_difference = True
                    elif compare_values(count1, count) and not compare_values(formatted_percentage, percentage_key):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{count1} and {formatted_percentage}",
                                           "Data validation": f"{key} = count:{count} is equal to calculated count:{count1} but percentage: {percentage_key} is not equal to calculated percentage: {formatted_percentage}."}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Highest".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "CallSetupTime(s)"
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Highest with the lowest value in df_data
                    SmsSentStatus = f"{float(value):.2f}"
                    df_data_values = max(df_data)
                    df_data_values = f"{float(df_data_values):.2f}"
                    if compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_values}"}
                    elif not compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_values}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Lowest".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "CallSetupTime(s)"
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data = dataoftesttypecombineexport1[parameter].tolist()
                            # df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Lowest with the lowest value in df_data
                    SmsSentStatus = f"{float(value):.2f}"
                    df_data_values = min(df_data)
                    df_data_values = f"{float(df_data_values):.2f}"
                    if compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is equal to calculated {key}:{df_data_values}"}
                    elif not compare_values(SmsSentStatus, df_data_values):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{df_data_values}",
                                           "Data validation": f"{key}:{SmsSentStatus} is not equal to calculated {key}:{df_data_values}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("Geo", key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Iteration"
                    try:
                        df_data = dataoftesttypecombinebinaryexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total Geo samples with the length of df_data
                    total_geo_samples = value
                    if compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}",
                                           "Data validation": f"{key}:{total_geo_samples} is equal to calculated {key}:{len(df_data)}"}
                    elif not compare_values(total_geo_samples, len(df_data)):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{len(df_data)}",
                                           "Data validation": f"{key}:{total_geo_samples} is not equal to calculated {key}:{len(df_data)}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.fullmatch("TotalCallsSuccess".replace(" ", "").lower(), key,re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    CallsSuccessStatus = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "OK".lower()]
                    total_success_count = len(df_data_values)
                    if compare_values(CallsSuccessStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}",
                                           "Data validation": f"{key}:{CallsSuccessStatus} is equal to calculated {key}:{total_success_count}"}
                    elif not compare_values(CallsSuccessStatus, total_success_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_success_count}",
                                           "Data validation": f"{key}:{CallsSuccessStatus} is not equal to calculated {key}:{total_success_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalCallsAttempted".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = df_data
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("TotalCallsFailed".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() != "OK".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("DroppedCalls".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "DROPPED".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("AbortedCalls".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "ABORTED".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("CallSetupFailed".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "CALL_SETUP_FAILURE".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.search("No Network".replace(" ", "").lower(), key, re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "Status"
                    try:
                        df_data = dataoftesttypecombineexport[parameter].tolist()
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower() or str(value1).replace(" ","").lower() == "NO_NETWORK".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.fullmatch("AttemptedVoicecallon2G".replace(" ", "").lower(), key,re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "VoiceCalltype"
                    df_data_calltype = None
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data_calltype = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("2G" or "Vo2G")]
                            df_data_calltype = df_data_calltype[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            dataoftesttypecombineexport2 = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("2G" or "Vo2G")]
                            df_data_calltype = dataoftesttypecombineexport2[parameter].tolist()
                            # dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['VoiceCalltype'].str.contains("2G" or "Vo2G")]
                            # df_data_calltype = dataoftesttypecombineexport1[parameter].tolist()
                        df_data = df_data_calltype
                        # for value_calltype in df_data_calltype:
                        #     cleaned_values = str(value_calltype).replace('"', '').split(', ')
                        #     for cleaned_value in cleaned_values:
                        #         df_data.append(cleaned_value)
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "2G".lower() or str(value1).replace(" ","").lower() == "Vo2G".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif re.fullmatch("AttemptedVoicecallon3G".replace(" ", "").lower(), key,re.IGNORECASE):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "VoiceCalltype"
                    df_data_calltype = None
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data_calltype = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("3G" or "Vo3G")]
                            df_data_calltype = df_data_calltype[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            dataoftesttypecombineexport2 = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("3G" or "Vo3G")]
                            df_data_calltype = dataoftesttypecombineexport2[parameter].tolist()
                            # dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['VoiceCalltype'].str.contains("3G")]
                            # df_data_calltype = dataoftesttypecombineexport1[parameter].tolist()
                        df_data = df_data_calltype
                        # for value_calltype in df_data_calltype:
                        #     cleaned_values = str(value_calltype).replace('"', '').split(', ')
                        #     for cleaned_value in cleaned_values:
                        #         df_data.append(cleaned_value)
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "3G".lower() or str(value1).replace(" ","").lower() == "Vo3G".lower() ]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}: {TotalCallsAttempted} is equal to calculated {key}: {total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}: {TotalCallsAttempted} is not equal to calculated {key}: {total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("AttemptedVoicecallon4G(VoLTE)".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "VoiceCalltype"
                    df_data_calltype = None
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data_calltype = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("VoLTE")]
                            df_data_calltype = df_data_calltype[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['Status'].str.contains("OK")]
                            dataoftesttypecombineexport2 = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("VoLTE")]
                            df_data_calltype = dataoftesttypecombineexport2[parameter].tolist()
                        df_data = df_data_calltype
                        # for value_calltype in df_data_calltype:
                        #     cleaned_values = str(value_calltype).replace('"', '').split(', ')
                        #     for cleaned_value in cleaned_values:
                        #         df_data.append(cleaned_value)
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "VoLTE".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            # elif re.search("Attempted Voice call on 2G/3G".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("Attempted Voice call on 3G/2G".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("Successful CSFB call to 2G".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("Successful CSFB call to 3G".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("SRVCC call to 3G".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("SRVCC call to 2G".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("CSFB call to 2G Failed".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            # elif re.search("CSFB call to 3G Failed".replace(" ", "").lower(), key, re.IGNORECASE):
            #     df_data = dataoftesttypecombineexport["Status"].tolist()
            #     # Compare Total SMS sent success with the lowest value in df_data
            #     TotalCallsAttempted = value
            #     df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NO NETWORK".lower()]
            #     total_CallsAttempted_count = len(df_data_values)
            #     if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
            #     elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
            #         comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,
            #                            "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
            #         flag_difference = True
            elif compare_values("NotDetermined(Voice CallType)".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "CallType"
                    df_data_calltype = None
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data_calltype = dataoftesttypecombineexport1[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            df_data_calltype = dataoftesttypecombineexport[parameter].tolist()
                        for value_calltype in df_data_calltype:
                            cleaned_values = str(value_calltype).replace('"', '').split(', ')
                            for cleaned_value in cleaned_values:
                                df_data.append(cleaned_value)
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "NS".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("VoiceCallonvo-Wifi".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = "VoiceCalltype"
                    df_data_calltype = None
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data_calltype = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("vo-Wifi")]
                            df_data_calltype = df_data_calltype[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['VoiceCalltype'].str.contains("vo-Wifi")]
                            df_data_calltype = dataoftesttypecombineexport1[parameter].tolist()
                        df_data = df_data_calltype
                        # for value_calltype in df_data_calltype:
                        #     cleaned_values = str(value_calltype).replace('"', '').split(', ')
                        #     for cleaned_value in cleaned_values:
                        #         df_data.append(cleaned_value)
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "vo-Wifi".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            elif compare_values("VoiceCallonVoNR".replace(" ", "").lower(), key):
                try:
                    df_data = []
                    file_type = "CE:- "
                    parameter = 'VoiceCalltype'
                    df_data_calltype = None
                    try:
                        if str(testtype).replace(" ", "").lower() == "Failed Calls".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[~dataoftesttypecombineexport['Status'].str.contains("OK")]
                            df_data_calltype = dataoftesttypecombineexport1[dataoftesttypecombineexport1['VoiceCalltype'].str.contains("VoNR")]
                            df_data_calltype = df_data_calltype[parameter].tolist()
                        elif str(testtype).replace(" ", "").lower() == "CallTest".replace(" ", "").lower():
                            dataoftesttypecombineexport1 = dataoftesttypecombineexport[dataoftesttypecombineexport['VoiceCalltype'].str.contains("VoNR")]
                            df_data_calltype = dataoftesttypecombineexport1[parameter].tolist()
                        # for value_calltype in df_data_calltype:
                        #     cleaned_values = str(value_calltype).replace('"', '').split(', ')
                        #     for cleaned_value in cleaned_values:
                        #         df_data.append(cleaned_value)
                        df_data = df_data_calltype
                    except Exception as e:
                        comprasion_data = {"File": f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f'{key}:- parameter "{parameter}" name cant find in csv combine export'}
                        compared_data.append(comprasion_data)
                        raise e
                    # Compare Total SMS sent success with the lowest value in df_data
                    TotalCallsAttempted = value
                    df_data_values = [value1 for value1 in df_data if str(value1).replace(" ","").lower() == "VoNR".lower()]
                    total_CallsAttempted_count = len(df_data_values)
                    if compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is equal to calculated {key}:{total_CallsAttempted_count}"}
                    elif not compare_values(TotalCallsAttempted, total_CallsAttempted_count):
                        comprasion_data = {"File":f'({file_type + testtype}):- {parameter}',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"{total_CallsAttempted_count}",
                                           "Data validation": f"{key}:{TotalCallsAttempted} is not equal to calculated {key}:{total_CallsAttempted_count}"}
                        flag_difference = True
                except Exception as e:
                    continue
            else:
                comprasion_data = {"File": f'({" ? " + testtype}):- " ? "',"map view Operator":key,"map view Operator value":f"{value}","calculated csv value":f"Not calculated","Data validation": f"{key} is not satisfied any condition, check key name."}

            if comprasion_data != "None":
                compared_data.append(comprasion_data)
        compared_data.append({"File":testtype,"map view Operator":"ENDHERE","map view Operator value":f"ENDHERE","calculated csv value":f"ENDHERE","Data validation": "ENDHERE"})
        if flag_difference == False:
            r_result = status(Title=Title, component=testtype, status="PASSED", comments="Values are same")
            result_same.put(compared_data)
            result_status.put(r_result)
        elif flag_difference == True:
            r_result = status(Title=Title, component=testtype, status="FAILED", comments="Values are different")
            result_Difference.put(compared_data)
            result_status.put(r_result)
        print(r_result)
        print(compared_data)
    except Exception as e:
        pass

########################################################################################################################################################
def click_load_more(driver, load_more_button_xpath):
    while True:
        try:
            load_more_button = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.XPATH, load_more_button_xpath)))
            load_more_button.click()
            time.sleep(20)
        except:
            print("Load more button not found or not clickable.")
            break
###########################################################################################################################################################################
def navigate_to_date(driver, start_date, end_date):
    # Extract year, month, and day from start_date
    start_year = start_date.year
    start_month = start_date.month
    start_day = start_date.day

    # Extract year, month, and day from end_date
    end_year = end_date.year
    end_month = end_date.month
    end_day = end_date.day

    # Click on the button to open the date picker
    date_picker_button_xpath = (By.XPATH,"//li[@id='calendar']//a[contains(text(),'Custom Date')]","date_picker_button_xpath")
    click(driver,date_picker_button_xpath)

    left_calender_previous_btn = (By.XPATH,"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//th[@class ='prev available']","left_calender_previous_btn")
    right_calender_forward_btn = (By.XPATH,"//div[@class='calendar first right']//th[@class='next available']","right_calender_forward_btn")
    start_months_difference = current_difference_in_months(start_year, start_month)
    end_months_difference = userdefined_difference_in_months(end_year,end_month,start_year, start_month)
    # if start_year < current_date.year or start_month < current_date.month:
    if start_months_difference != 0:
        for i in range(0,start_months_difference):
            time.sleep(1)
            click(driver,left_calender_previous_btn)
    left_calender_date = (By.XPATH,f"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//td[contains(@class, 'available') and not(contains(@class, 'available off')) and text()='{start_day}']","left_calender_date")
    right_calender_date = (By.XPATH,f"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar first right']//div//td[contains(@class, 'available') and not(contains(@class, 'available off')) and text()='{end_day}']","right_calender_date")
    time.sleep(3)
    click(driver,left_calender_date)
    # if end_year < current_date.year or end_month < current_date.month:
    if end_months_difference != 0:
        for i in range(0,end_months_difference):
            time.sleep(1)
            click(driver,right_calender_forward_btn)
    time.sleep(1)
    click(driver, right_calender_date)
    time.sleep(1)
    datetime_apply_btn = (By.XPATH,"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//button[contains(text(),'Apply')]","datetime_apply_btn")
    click(driver, datetime_apply_btn)
    time.sleep(5)

def current_difference_in_months(year, month):
    from datetime import datetime
    # Get the current year and month
    current_year = datetime.now().year
    current_month = datetime.now().month

    # Calculate the difference in months
    difference = (current_year - year) * 12 + (current_month - month)

    return difference

def userdefined_difference_in_months(year1,month1,year2, month2):

    # Calculate the difference in months
    difference = (year1 - year2) * 12 + (month1 - month2)

    return difference
##############################################################################################################################################################
# def navigate_to_date(driver, start_date, end_date):
#     # Extract year, month, and day from start_date
#     start_year = start_date.year
#     start_month = start_date.month
#     start_day = start_date.day
#     # Assuming start_month is the month number, for example, 5 for May
#     start_month_number = start_month
#     # Convert the month number to its corresponding name
#     start_month_name = calendar.month_name[start_month_number][:3]
#
#
#
#     # Click on the button to open the date picker
#     date_picker_button_xpath = "//li[@id='calendar']//a[contains(text(),'Custom Date')]"
#     date_picker_button = driver.find_element(By.XPATH, date_picker_button_xpath)
#     date_picker_button.click()
#
#     # Loop until the start month and year are visible
#     while True:
#             try:
#                 try:
#                     # Get the currently displayed month and year
#                     current_month_year_xpath = "//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//th[@class='prev available']/following-sibling::th[@colspan='5'][@class='month']"  # Adjust this XPath as needed
#                         # # Wait for the current_month_year_element to contain text
#                     # current_month_year_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, current_month_year_xpath)))
#                     current_month_year_element = driver.find_element(By.XPATH, current_month_year_xpath)
#                     # current_month_year_xpath1.click()
#
#                     # Get the currently displayed month and year
#                     current_month_year = current_month_year_element.text
#
#                     # Parse the displayed month and year
#                     start_month_name, current_year = current_month_year.split()
#                 except Exception as e:
#                     print(Exception)
#                     pass
#                 start_month_name = list(calendar.month_abbr).index(start_month_name)
#
#                 # Check if the start month and year are visible
#                 if start_month_name == start_month and int(current_year) == start_year:
#                     break
#             except Exception as e:
#                 pass
#             # Find and click the navigation button to go to the previous or next month
#             if int(current_year) > start_year or (int(current_year) == start_year and start_month_name > start_month):
#                 # Click on the "<" button to go to the previous month
#                 navigation_button_xpath = "//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//th[@class ='prev available']"
#             else:
#                 # Click on the ">" button to go to the next month
#                 navigation_button_xpath = "//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//th[@class ='next available']"
#             navigation_button = driver.find_element(By.XPATH, navigation_button_xpath)
#             navigation_button.click()
#
#     # Find and click the element corresponding to the start day
#     start_day_element_xpath = f"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//td[contains(@class, 'available') and not(contains(@class, 'available off')) and text()='{start_day}']"
#     start_day_element = driver.find_element(By.XPATH, start_day_element_xpath)
#     start_day_element.click()
#
#     # # Extract year, month, and day from end_date
#     end_year = end_date.year
#     end_month = end_date.month
#     end_day = end_date.day
#     end_month_number = end_month
#     end_month_name = calendar.month_name[end_month_number][:3]
#     # Loop until the end month and year are visible
#     while True:
#         try:
#             try:
#                 # Get the currently displayed month and year
#                 current_month_year_xpath1 = "//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar first right']//div[@class='calendar-date']//th[@class='next available']/preceding-sibling::th[@colspan='5'][@class='month']"
#                 current_month_year_element = driver.find_element(By.XPATH, current_month_year_xpath1)
#                 current_month_year = current_month_year_element.text
#
#                 # Parse the displayed month and year
#                 end_month_name, current_year = current_month_year.split()
#             except Exception as e:
#                 pass
#             end_month_name = list(calendar.month_abbr).index(end_month_name)
#
#             # Check if the end month and year are visible
#             if end_month_name == end_month and int(current_year) == end_year:
#                 break
#         except Exception as e:
#             pass
#
#         # Find and click the navigation button to go to the previous or next month
#         if int(current_year) > end_year or (int(current_year) == end_year and end_month_name > end_month):
#             navigation_button_xpath = "//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar first right']//div[@class='calendar-date']//th[@class ='prev available']"
#         else:
#             navigation_button_xpath = "//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar first right']//div[@class='calendar-date']//th[@class ='next available']"
#         navigation_button = driver.find_element(By.XPATH, navigation_button_xpath)
#         navigation_button.click()
#
#     # Find and click the element corresponding to the end day
#     end_day_element_xpath = f"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar first right']//div//td[contains(@class, 'available') and not(contains(@class, 'available off')) and text()='{end_day}']"
#     end_day_element = driver.find_element(By.XPATH, end_day_element_xpath)
#     end_day_element.click()
###############################################################################################################################################################################################
def extract_table_column_data(driver,excelpath):
    try:
        # Wait for the table to be present
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='div-table-content-wrapper']"))
        )
    except Exception as e:
        print("Error waiting for table to be present:", e)
    individual_pop_loadcamp_element = driver.find_elements(*individual_pop_table.loaderCamp[:2])
    data = []
    for i in range(len(individual_pop_loadcamp_element)):
        i += 1
        individual_pop_table_loaderCamp = (By.XPATH, f"//tr[{i}]//*[@id='loaderCamp']/abbr/a", "individual_pop_table_loaderCamp")
        test_name_xpath = (By.XPATH, f"//tr[{i}]//*[@id='loaderCamp']/following-sibling::td[1]")
        Device_name_xpath = (By.XPATH, f"//tr[{i}]//*[@id='loaderCamp']/following-sibling::td[2]")
        individual_pop_table_loaderCamp_element = driver.find_element(*individual_pop_table_loaderCamp[:2])
        individual_pop_table_loaderCamp_name = individual_pop_table_loaderCamp_element.text
        test_name_element = driver.find_element(*test_name_xpath)
        test_name = test_name_element.text
        Device_name_element = driver.find_element(*Device_name_xpath)
        Device_name = Device_name_element.text
        data.append({
            'Operator Name': individual_pop_table_loaderCamp_name, 'Test Name': test_name, 'Device': Device_name # "Device" in the third column (index 2)
        })
        updatecomponentstatus("Date and Time query", f"'Operator Name', 'Test Name', 'Device'", "",f"'Operator Name': {individual_pop_table_loaderCamp_name}, 'Test Name': {test_name}, 'Device': {Device_name}",excelpath)
#######################################################################################################################################################################################################################
def date_and_time_main_function(driver,excelpath):
    datetime_runvalue = Testrun_mode(value="Date and Time")
    if "Yes".lower() == datetime_runvalue[-1].strip().lower():
        try:
            clickec(driver, date_time.date_and_time_click_button)
        except Exception as e:
            print("Could not able to click")
        df = pd.read_excel(config.test_data_path, sheet_name='date_time')
        timerequired =  False
        # Loop through each row in the DataFrame
        for index, row in df.iterrows():
            select_hours = row['Select_hours']
            execute_flag = row['Execute']
            if isinstance(execute_flag, str) and execute_flag.lower() == 'yes':
                if select_hours.lower() == 'custom date':
                    # Handle custom date range
                    start_date = row['Start Date']
                    end_date = row['End Date']

                    # code to select custom date range
                    navigate_to_date(driver, start_date,end_date)
                    timerequired = True
                else:
                    try:
                        radio_button_xpath = f"//a[normalize-space()='{select_hours}']"
                        radio_button = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, radio_button_xpath)))
                        radio_button.click()
                    except Exception as e:
                        print(f"Failed to select {select_hours}: {str(e)}")
        time.sleep(2)
        datetime_apply_btn = (By.XPATH,"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//button[contains(text(),'Apply')]","datetime_apply_btn")
        click(driver, datetime_apply_btn)
        Page_Down(driver)
        clickec(driver,date_time.expand_table_button)
        time.sleep(2)
        Page_up(driver)
        time.sleep(2)
        load_more_button_xpath = "//button[@class='btn btn-primary btn-sm ng-scope']"  # Example XPath, replace it with your actual XPath
        click_load_more(driver, load_more_button_xpath)
        if timerequired == True:
            time.sleep(5)
            click_load_more(driver, load_more_button_xpath)
        extract_table_column_data(driver,excelpath)
        clickec(driver, close_button.closeFullTableView)